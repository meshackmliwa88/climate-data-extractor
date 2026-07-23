"""Plotting, climate indices and cost-recovery helpers for CDE.

These helpers are intentionally self-contained so that the web routes can use the
Zarr stores already held under storage/zarr without moving, deleting, or opening other data formats.
"""
from __future__ import annotations

import json
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from contextvars import ContextVar
from contextlib import contextmanager
from functools import lru_cache
from typing import Any, Dict, Iterable, List, Tuple

import numpy as np
import pandas as pd
import xarray as xr

from cde_store import iter_data_stores, open_data_store, resolve_time_axis, select_nearest_time, slice_time_range, store_display_name, store_kind
from cde_variable_selection import choose_data_variable, requested_element_label, select_requested_statistic_dimension

import matplotlib
matplotlib.use("Agg")
# Render very long series in bounded path chunks.  All observations remain in
# the line; this only reduces renderer memory and applies pixel-level path
# simplification that is invisible at exported image resolution.
matplotlib.rcParams["agg.path.chunksize"] = 20000
matplotlib.rcParams["path.simplify"] = True
matplotlib.rcParams["path.simplify_threshold"] = 0.65
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon, Ellipse


PROJECT_ROOT = Path(__file__).resolve().parent
PARQUET_DB_DIR = PROJECT_ROOT / "storage" / "db" / "parquet"
TANZANIA_BOUNDS = {"lon_min": 28.0, "lon_max": 41.5, "lat_min": -12.0, "lat_max": 0.0}


# Request-local cache used by the unified products workspace.  A complete
# product suite often needs the same point series for several plots and the
# analytical workbook.  Keeping the already-selected pandas frame here means
# the source store is opened and sliced only once during that request.
_PRODUCT_DATA_CACHE: ContextVar[dict[str, Any] | None] = ContextVar("cde_product_data_cache", default=None)


@contextmanager
def product_data_cache():
    """Reuse selected point/grid slices within one product-generation request.

    The cache is deliberately request-local: it never keeps complete climate
    arrays in process memory after the request has finished.  Nested callers
    share the same cache, which also lets the analysis module reuse plot data.
    """
    existing = _PRODUCT_DATA_CACHE.get()
    if existing is not None:
        yield existing
        return
    state: dict[str, Any] = {
        "point": {},
        "grid": {},
        "stats": {"point_hits": 0, "point_misses": 0, "grid_hits": 0, "grid_misses": 0},
    }
    token = _PRODUCT_DATA_CACHE.set(state)
    try:
        yield state
    finally:
        _PRODUCT_DATA_CACHE.reset(token)


def current_product_cache_stats() -> dict[str, int]:
    state = _PRODUCT_DATA_CACHE.get()
    return dict((state or {}).get("stats", {}))


SEASON_DEFINITIONS = {
    "DJF": [12, 1, 2],
    "MAM": [3, 4, 5],
    "JJA": [6, 7, 8],
    "SON": [9, 10, 11],
    "OND": [10, 11, 12],
    "NDJ": [11, 12, 1],
    "DJFMA": [12, 1, 2, 3, 4],
    "NDJFMA": [11, 12, 1, 2, 3, 4],
}

PLOT_TYPES = [
    ("data_extraction", "Data Extraction Package"),
    ("time_series", "Time Series Line Plot"),
    ("bar", "Comparative Bar Chart"),
    ("area", "Accumulated Area Plot"),
    ("multi_line", "Multi-Element Comparison Plot"),
    ("monthly_climatology", "Monthly Climatology Plot"),
    ("seasonal_profile", "Seasonal Profile Plot"),
    ("annual_trend", "Annual Trend Plot"),
    ("anomaly", "Climate Anomaly Plot"),
    ("standardized_anomaly", "Standardized Anomaly (Z-Score) Plot"),
    ("spatial_map", "Spatial Distribution Map"),
    ("spatial_std_map", "Spatial Standard Deviation Map"),
    ("spatial_cv_map", "Spatial Coefficient of Variation Map"),
    ("heatmap", "Monthly–Annual Heat Map"),
    ("mean_std_band", "Mean with Standard Deviation Band"),
    ("std_error_bars", "Mean with Standard Deviation Error Bars"),
    ("standard_deviation", "Standard Deviation Profile"),
    ("coefficient_variation", "Coefficient of Variation Profile"),
    ("histogram", "Frequency Distribution Histogram"),
    ("box", "Monthly Box Plot"),
    ("extreme_value", "Annual Extreme Value Plot"),
    ("scatter", "Relationship Scatter Plot"),
    ("wind_rose", "Wind Rose"),
    ("statistical_summary", "Descriptive Statistical Summary"),
    ("climatology_profile", "Climatology and Seasonal Profile"),
    ("variability_analysis", "Standard Deviation and Variability Assessment"),
    ("trend_variability", "Trend and Variability Assessment"),
    ("extremes_analysis", "Extremes and Percentile Assessment"),
    ("comprehensive_analysis", "Comprehensive Climate Analysis"),
]

_ANALYTICAL_PRODUCTS = [
    "data_extraction", "statistical_summary", "climatology_profile",
    "variability_analysis", "trend_variability", "extremes_analysis", "comprehensive_analysis",
]

PLOT_FAMILIES = {
    "rainfall": [
        "time_series", "bar", "area", "monthly_climatology", "seasonal_profile",
        "annual_trend", "anomaly", "standardized_anomaly", "spatial_map",
        "spatial_std_map", "spatial_cv_map", "heatmap", "mean_std_band",
        "std_error_bars", "standard_deviation", "coefficient_variation",
        "histogram", "box", "extreme_value", *_ANALYTICAL_PRODUCTS,
    ],
    "temperature": [
        "time_series", "bar", "multi_line", "monthly_climatology", "seasonal_profile",
        "annual_trend", "anomaly", "standardized_anomaly", "spatial_map",
        "spatial_std_map", "spatial_cv_map", "heatmap", "mean_std_band",
        "std_error_bars", "standard_deviation", "coefficient_variation",
        "histogram", "box", "extreme_value", "scatter", *_ANALYTICAL_PRODUCTS,
    ],
    "humidity": [
        "time_series", "bar", "monthly_climatology", "seasonal_profile", "annual_trend",
        "anomaly", "standardized_anomaly", "spatial_map", "spatial_std_map",
        "spatial_cv_map", "heatmap", "mean_std_band", "std_error_bars",
        "standard_deviation", "coefficient_variation", "histogram", "box",
        "extreme_value", *_ANALYTICAL_PRODUCTS,
    ],
    "soil_moisture": [
        "time_series", "bar", "monthly_climatology", "seasonal_profile", "annual_trend",
        "anomaly", "standardized_anomaly", "spatial_map", "spatial_std_map",
        "spatial_cv_map", "heatmap", "mean_std_band", "std_error_bars",
        "standard_deviation", "coefficient_variation", "histogram", "box",
        "extreme_value", *_ANALYTICAL_PRODUCTS,
    ],
    "wind": [
        "time_series", "bar", "multi_line", "annual_trend", "spatial_map",
        "spatial_std_map", "histogram", "box", "scatter", "wind_rose",
        *_ANALYTICAL_PRODUCTS,
    ],
    "pressure_cloud": [
        "time_series", "bar", "multi_line", "monthly_climatology", "seasonal_profile",
        "annual_trend", "anomaly", "standardized_anomaly", "spatial_map",
        "spatial_std_map", "spatial_cv_map", "heatmap", "mean_std_band",
        "std_error_bars", "standard_deviation", "coefficient_variation",
        "histogram", "box", "extreme_value", "scatter", *_ANALYTICAL_PRODUCTS,
    ],
}

RAINFALL_INDICES = [
    ("total_rainfall", "Total Rainfall"),
    ("number_wet_days", "Number of Wet Days"),
    ("number_dry_days", "Number of Dry Days"),
    ("consecutive_dry_days", "Consecutive Dry Days"),
    ("consecutive_wet_days", "Consecutive Wet Days"),
    ("wet_spell_length", "Wet Spell Length"),
    ("dry_spell_length", "Dry Spell Length"),
    ("rainy_season_onset", "Rainy Season Onset"),
    ("rainy_season_cessation", "Rainy Season Cessation"),
    ("length_of_rainy_season", "Length of Rainy Season"),
    ("max_1day_rainfall", "Maximum 1-Day Rainfall"),
    ("max_5day_rainfall", "Maximum 5-Day Rainfall"),
    ("heavy_rainfall_days", "Heavy Rainfall Days"),
    ("very_heavy_rainfall_days", "Very Heavy Rainfall Days"),
    ("sdii", "Rainfall Intensity / SDII"),
    ("rainfall_anomaly", "Rainfall Anomaly"),
]

TEMPERATURE_INDICES = [
    ("mean_temperature", "Mean Temperature"),
    ("maximum_temperature", "Maximum Temperature"),
    ("minimum_temperature", "Minimum Temperature"),
    ("temperature_anomaly", "Temperature Anomaly"),
    ("hot_days", "Hot Days"),
    ("hot_nights", "Hot Nights"),
    ("cold_days", "Cold Days"),
    ("cold_nights", "Cold Nights"),
    ("heat_index", "Heat Index"),
    ("dtr", "Diurnal Temperature Range"),
]

OTHER_INDICES = [
    ("relative_humidity_index", "Relative Humidity Index"),
    ("soil_moisture_index", "Soil Moisture Index"),
    ("soil_moisture_anomaly", "Soil Moisture Anomaly"),
    ("wind_speed_index", "Wind Speed Index"),
]

ALL_INDICES = TEMPERATURE_INDICES + OTHER_INDICES

DATASETS = {
    "chirps_rainfall": {
        "label": "CHIRPS Precipitation",
        "family": "rainfall",
        "unit": "mm",
        "resolutions": {"daily": "Daily_Total", "monthly": "Monthly_Total", "annual": "Annual_Total", "seasonal": "Seasonal_{season}_Total"},
        "keywords": ["CDE_CHIRPS_Tanzania_Rainfall"],
        "variables": ["precip", "rainfall", "precipitation"],
    },
    "era5_total_precipitation": {
        "label": "ERA5 Precipitation",
        "family": "rainfall",
        "unit": "mm",
        "resolutions": {"hourly": "Hourly", "daily": "Daily_Total", "monthly": "Monthly_Total", "annual": "Annual_Total", "seasonal": "Seasonal_{season}_Total"},
        "keywords": ["CDE_ERA5_Tanzania_Total_Precipitation", "ERA5_Tanzania_Total_Precipitation_TP_Hourly"],
        "variables": ["tp", "precip", "precipitation", "total_precipitation"],
    },
    "era5_temperature": {
        "label": "Temperature Mean, Min and Max",
        "family": "temperature",
        "unit": "°C",
        "resolutions": {"hourly": "Hourly", "daily": "Daily_Mean", "monthly": "Monthly_Mean", "annual": "Annual_Mean", "seasonal": "Seasonal_{season}_Mean"},
        "keywords": ["CDE_ERA5_Tanzania_Temperature_Mean_Min_Max", "ERA5_Tanzania_Temperature _2M_Hourly", "ERA5_Tanzania_Temperature_2M_Hourly"],
        "variables": ["ta", "tmean", "temperature", "tmin", "tn", "tmax", "tx"],
    },
    "era5_dew_point": {
        "label": "Dew Point Temperature 2m",
        "family": "temperature",
        "unit": "°C",
        "resolutions": {"hourly": "Hourly", "daily": "Daily_Mean", "monthly": "Monthly_Mean", "annual": "Annual_Mean", "seasonal": "Seasonal_{season}_Mean"},
        "keywords": ["CDE_ERA5_Tanzania_Dew_Point_Temperature_2m", "ERA5_Tanzania_Dew_Point_Temperature_D2M_Hourly"],
        "variables": ["d2m", "dew_point_temperature", "dewpoint", "td"],
    },
    "era5_relative_humidity": {
        "label": "Relative Humidity",
        "family": "humidity",
        "unit": "%",
        "resolutions": {"hourly": "Hourly", "daily": "Daily_Mean", "monthly": "Monthly_Mean", "annual": "Annual_Mean", "seasonal": "Seasonal_{season}_Mean"},
        "keywords": ["CDE_ERA5_Tanzania_Relative_Humidity", "ERA5_Tanzania_Relative_Humidity_Hourly"],
        "variables": ["r", "rh", "relative_humidity"],
    },
    "era5_skin_temperature": {
        "label": "Skin Temperature",
        "family": "temperature",
        "unit": "°C",
        "resolutions": {"hourly": "Hourly", "daily": "Daily_Mean", "monthly": "Monthly_Mean", "annual": "Annual_Mean", "seasonal": "Seasonal_{season}_Mean"},
        "keywords": ["CDE_ERA5_Tanzania_Skin_Temperature", "ERA5_Tanzania_Skin_Temperature_SKT_Hourly"],
        "variables": ["skt", "skin_temperature"],
    },
    "era5_soil_temperature": {
        "label": "Soil Temperature Level 1",
        "family": "temperature",
        "unit": "°C",
        "resolutions": {"hourly": "Hourly", "daily": "Daily_Mean", "monthly": "Monthly_Mean", "annual": "Annual_Mean", "seasonal": "Seasonal_{season}_Mean"},
        "keywords": ["CDE_ERA5_Tanzania_Soil_Temperature_Level_1", "ERA5_Tanzania_Soil_Temperature_Level_1_STL1_Hourly"],
        "variables": ["stl1", "soil_temperature_level_1", "soil_temperature"],
    },
    "era5_pressure_cloud": {
        "label": "Surface Pressure and Total Cloud Cover",
        "family": "pressure_cloud",
        "unit": "mixed",
        "resolutions": {"hourly": "Hourly", "daily": "Daily_Mean", "monthly": "Monthly_Mean", "annual": "Annual_Mean", "seasonal": "Seasonal_{season}_Mean"},
        "keywords": ["CDE_ERA5_Tanzania_Surface_Pressure_and_Total_Cloud_Cover", "ERA5_Tanzania_Pressure_and_Total_Cloud_Cover_Hourly"],
        "variables": ["sp", "msl", "tcc", "surface_pressure", "total_cloud_cover"],
    },
    "era5_soil_water": {
        "label": "Volumetric Soil Moisture",
        "family": "soil_moisture",
        "unit": "m³/m³",
        "resolutions": {"hourly": "Hourly", "daily": "Daily_Mean", "monthly": "Monthly_Mean", "annual": "Annual_Mean", "seasonal": "Seasonal_{season}_Mean"},
        "keywords": ["CDE_ERA5_Tanzania_Volumetric_Soil_Water", "ERA5_Tanzania_Volumetric_Soil_Water_SWVL_Hourly"],
        "variables": ["swvl1", "swvl", "volumetric_soil_water"],
    },
    "era5_wind": {
        "label": "Wind Speed and Direction 10m",
        "family": "wind",
        "unit": "knots / degrees",
        "resolutions": {"hourly": "Hourly", "daily": "Daily_Mean", "monthly": "Monthly_Mean", "annual": "Annual_Mean", "seasonal": "Seasonal_{season}_Mean"},
        "keywords": ["CDE_ERA5_Tanzania_Wind_Speed_and_Direction_10m", "ERA5_Tanzania_Wind_Speed_Direction_10m_Hourly"],
        "variables": ["wind_speed", "speed", "ws", "si10", "wind_direction", "direction", "wd"],
    },
}


def human_money(value: float) -> str:
    return f"Tsh {value:,.0f}/="


def slugify(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(text).strip()).strip("_") or "output"


def ensure_output_dirs(export_dir: Path) -> Dict[str, Path]:
    paths = {
        "plots": export_dir / "plots",
        "indices": export_dir / "indices",
        "invoices": export_dir / "invoices",
    }
    for p in paths.values():
        p.mkdir(parents=True, exist_ok=True)
    return paths


def _normalize_store_name(value: str) -> str:
    """Normalize case, spacing and punctuation for resilient store discovery."""
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def _file_contains(path: Path, pieces: Iterable[str]) -> bool:
    name = _normalize_store_name(path.name)
    return all(_normalize_store_name(piece) in name for piece in pieces)


def available_datasets(data_dir: Path) -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}
    for key, meta in DATASETS.items():
        available_res = []
        for res in ["hourly", "daily", "monthly", "annual", "seasonal"]:
            if find_file(data_dir, key, res, season="MAM", silent=True) or (res != "seasonal" and find_file(data_dir, key, res, silent=True)):
                available_res.append(res)
        if not available_res:
            available_res = list(meta["resolutions"].keys())
        result[key] = {**meta, "available_resolutions": available_res, "available_plot_types": dataset_allowed_plots(key)}
    return result


@lru_cache(maxsize=512)
def _find_file_cached(
    data_dir_text: str,
    dataset_key: str,
    resolution: str,
    season: str,
    store_paths: tuple[str, ...],
) -> str | None:
    meta = DATASETS[dataset_key]
    token = meta["resolutions"].get(resolution)
    if token is None:
        return None
    if season:
        token = token.replace("{season}", season.upper())
    normalized_token = _normalize_store_name(token)
    candidates: list[Path] = []
    for path_text in store_paths:
        f = Path(path_text)
        normalized_name = _normalize_store_name(f.name)
        if any(_normalize_store_name(keyword) in normalized_name for keyword in meta["keywords"]):
            if normalized_token in normalized_name:
                candidates.append(f)
            elif resolution == "hourly" and "hourly" in normalized_name:
                candidates.append(f)
    if not candidates:
        return None
    return str(sorted(set(candidates), key=lambda path: (len(path.name), path.name))[0])


def find_file(data_dir: Path, dataset_key: str, resolution: str, season: str | None = None, silent: bool = False) -> Path | None:
    """Resolve one configured source store without rescanning the inventory."""
    data_dir = Path(data_dir).expanduser().resolve()
    if dataset_key not in DATASETS:
        if silent:
            return None
        raise ValueError(f"Unknown dataset: {dataset_key}")
    meta = DATASETS[dataset_key]
    if resolution not in meta["resolutions"]:
        if silent:
            return None
        raise ValueError(f"{meta['label']} does not support {resolution}")
    stores = tuple(str(path) for path in iter_data_stores(data_dir))
    if not stores:
        if silent:
            return None
        raise FileNotFoundError(f"No compatible climate data stores were found in {data_dir}")
    found = _find_file_cached(str(data_dir), dataset_key, resolution, str(season or ""), stores)
    if found:
        return Path(found)
    if silent:
        return None
    raise FileNotFoundError(
        f"No compatible data store was found for {meta['label']} / {resolution}"
        + (f" / {season}" if season else "") + f" in {data_dir}"
    )


def detect_coord(ds: xr.Dataset, names: Iterable[str]) -> str | None:
    for n in names:
        if n in ds.coords or n in ds.variables:
            return n
    return None


def detect_time_coord(ds: xr.Dataset) -> str:
    """Return the operational date coordinate, including non-indexed valid_time."""
    name, _dim = resolve_time_axis(ds, None)
    return name


def detect_lat_lon(ds: xr.Dataset) -> Tuple[str, str]:
    lat = detect_coord(ds, ["latitude", "lat", "y"])
    lon = detect_coord(ds, ["longitude", "lon", "x"])
    if not lat or not lon:
        raise ValueError("No latitude/longitude coordinates were found in the selected data store.")
    return lat, lon


def variable_score(var_name: str, da: xr.DataArray, candidates: Iterable[str]) -> int:
    text = " ".join([
        var_name.lower(),
        str(da.attrs.get("long_name", "")).lower(),
        str(da.attrs.get("standard_name", "")).lower(),
        str(da.attrs.get("short_name", "")).lower(),
    ])
    score = 0
    for idx, c in enumerate(candidates):
        c = c.lower()
        if var_name.lower() == c:
            score += 100 - idx
        if c in text:
            score += 40 - min(idx, 30)
    return score


def data_variables(ds: xr.Dataset) -> List[str]:
    return [v for v in ds.data_vars if ds[v].ndim >= 1 and np.issubdtype(ds[v].dtype, np.number)]


def pick_variable(ds: xr.Dataset, dataset_key: str, variable: str | None = None) -> str:
    """Select the requested variable without substituting an opposite statistic."""
    requested = variable if variable not in (None, "", "auto") else DATASETS[dataset_key]["variables"][0]
    return choose_data_variable(ds, requested, DATASETS[dataset_key].get("variables", []))


def _select_product_statistic_dimension(da: xr.DataArray, variable: str | None, *, keep_dims: set[str]) -> xr.DataArray:
    """Compatibility wrapper around the shared strict statistic selector."""
    return select_requested_statistic_dimension(da, variable, keep_dims=keep_dims)


def _point_dataarray_to_frame(point: xr.DataArray, time_name: str, variable: str | None) -> pd.DataFrame:
    point = _select_product_statistic_dimension(point, variable, keep_dims={time_name})
    if time_name not in point.coords and time_name not in point.dims:
        for candidate in ("time", "valid_time", "date", "datetime"):
            if candidate in point.coords or candidate in point.dims:
                time_name = candidate
                break
    if time_name not in point.coords and time_name not in point.dims:
        raise ValueError("No time coordinate was found after selecting the requested grid point.")
    times = np.asarray(point[time_name].values).reshape(-1)
    values = np.asarray(point.values)
    if values.size != times.size:
        time_dim = point[time_name].dims[0] if point[time_name].dims else None
        if time_dim and time_dim in point.dims:
            axis = point.get_axis_num(time_dim)
            values = np.moveaxis(values, axis, 0).reshape(len(times), -1)[:, 0]
        else:
            values = values.reshape(-1)[:len(times)]
    else:
        values = values.reshape(-1)
    frame = pd.DataFrame({
        "time": pd.to_datetime(pd.Series(times), errors="coerce").to_numpy(),
        "value": pd.to_numeric(pd.Series(values), errors="coerce").to_numpy(),
    })
    return frame.dropna(subset=["time"]).drop_duplicates(subset=["time"]).sort_values("time").reset_index(drop=True)


def _grid_dataarray_to_frame(da: xr.DataArray, lat_name: str, lon_name: str, variable: str | None) -> pd.DataFrame:
    da = _select_product_statistic_dimension(da, variable, keep_dims={lat_name, lon_name})
    # Remove any unexpected residual dimension deterministically.
    for dim in list(da.dims):
        if dim not in {lat_name, lon_name}:
            da = da.isel({dim: 0}, drop=True)
    lat_values = np.asarray(da[lat_name].values)
    lon_values = np.asarray(da[lon_name].values)
    values = np.asarray(da.values)
    try:
        ordered = da.transpose(lat_name, lon_name)
        values = np.asarray(ordered.values)
        lat_values = np.asarray(ordered[lat_name].values)
        lon_values = np.asarray(ordered[lon_name].values)
    except Exception:
        pass
    if lat_values.ndim == 1 and lon_values.ndim == 1 and values.shape == (len(lat_values), len(lon_values)):
        xx, yy = np.meshgrid(lon_values, lat_values)
        return pd.DataFrame({"latitude": yy.ravel(), "longitude": xx.ravel(), "value": values.ravel()})
    frame = da.to_series().rename("value").reset_index()
    rename = {}
    if lat_name != "latitude": rename[lat_name] = "latitude"
    if lon_name != "longitude": rename[lon_name] = "longitude"
    return frame.rename(columns=rename)


def list_file_variables(data_dir: Path, dataset_key: str, resolution: str, season: str | None = None) -> Dict[str, str]:
    try:
        file_path = find_file(data_dir, dataset_key, resolution, season=season)
        assert file_path
        with open_data_store(file_path, decode_times=True) as ds:
            out = {}
            for v in data_variables(ds):
                label = ds[v].attrs.get("long_name") or ds[v].attrs.get("standard_name") or v
                unit = ds[v].attrs.get("units", "")
                out[v] = f"{label} ({unit})" if unit else str(label)
            return out
    except Exception:
        # Fallback labels based on known files.
        key = dataset_key
        if key == "era5_temperature":
            return {"ta": "Mean Temperature", "tmin": "Minimum Temperature", "tmax": "Maximum Temperature"}
        if key == "era5_pressure_cloud":
            return {"msl": "Mean Sea Level Pressure (hPa)", "sp": "Surface Pressure (hPa)", "tcc": "Total Cloud Cover (octas)"}
        if key == "era5_wind":
            return {"wind_speed": "Wind Speed", "wind_direction": "Wind Direction"}
        return {"auto": "Auto-detect variable"}


def convert_series_units(series: pd.Series, units: str, family: str, var_name: str = "") -> Tuple[pd.Series, str]:
    u = str(units or "").strip().lower()
    out = series.astype(float)
    if family == "rainfall" and u in {"m", "meter", "metre", "meters", "metres"}:
        return out * 1000.0, "mm"
    if family == "temperature" and u in {"k", "kelvin"}:
        return out - 273.15, "°C"
    name_l = str(var_name or "").lower()
    if "pressure_cloud" == family and ("cloud" in name_l or name_l == "tcc") and out.max(skipna=True) <= 1.1:
        return out * 8.0, "octas"
    if (family == "pressure_cloud" and ("pressure" in name_l or name_l in {"sp", "msl", "surface_pressure", "mean_sea_level_pressure"})) or ("pressure" in name_l):
        try:
            med = float(pd.to_numeric(out, errors="coerce").dropna().median())
        except Exception:
            med = 0.0
        if u in {"pa", "pascal", "pascals"} or med > 2000:
            return out / 100.0, "hPa"
        return out, "hPa" if not units else units
    if ("cloud" in name_l or name_l == "tcc") and out.max(skipna=True) <= 1.1:
        return out * 8.0, "octas"
    return out, units or DATASETS.get(family, {}).get("unit", "")


def convert_dataarray_units(da: xr.DataArray, family: str, var_name: str = "") -> Tuple[xr.DataArray, str]:
    u = str(da.attrs.get("units", "") or "").strip().lower()
    out = da.astype(float)
    if family == "rainfall" and u in {"m", "meter", "metre", "meters", "metres"}:
        return out * 1000.0, "mm"
    if family == "temperature" and u in {"k", "kelvin"}:
        return out - 273.15, "°C"
    name_l = str(var_name or "").lower()
    if family == "pressure_cloud" and ("cloud" in name_l or name_l == "tcc"):
        try:
            mx = float(out.max(skipna=True).compute() if hasattr(out.max(skipna=True), "compute") else out.max(skipna=True))
        except Exception:
            mx = 999.0
        if mx <= 1.1:
            return out * 8.0, "octas"
        return out, "octas"
    if (family == "pressure_cloud" and ("pressure" in name_l or name_l in {"sp", "msl", "surface_pressure", "mean_sea_level_pressure"})) or ("pressure" in name_l):
        try:
            med_da = out.median(skipna=True)
            med = float(med_da.compute() if hasattr(med_da, "compute") else med_da)
        except Exception:
            med = 0.0
        if u in {"pa", "pascal", "pascals"} or med > 2000:
            return out / 100.0, "hPa"
        return out, "hPa" if not da.attrs.get("units") else da.attrs.get("units")
    return out, da.attrs.get("units", DATASETS.get(family, {}).get("unit", ""))


def extract_point_series(
    data_dir: Path,
    dataset_key: str,
    resolution: str,
    lat: float,
    lon: float,
    start_date: str | None,
    end_date: str | None,
    variable: str | None = None,
    season: str | None = None,
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    file_path = find_file(data_dir, dataset_key, resolution, season=season)
    assert file_path
    cache = _PRODUCT_DATA_CACHE.get()
    key = (
        "point", str(file_path.resolve()), dataset_key, resolution,
        round(float(lat), 8), round(float(lon), 8),
        str(start_date or ""), str(end_date or ""), str(variable or "auto"), str(season or ""),
    )
    if cache is not None and key in cache["point"]:
        cache["stats"]["point_hits"] += 1
        cached_df, cached_context = cache["point"][key]
        return cached_df.copy(deep=False), dict(cached_context)
    if cache is not None:
        cache["stats"]["point_misses"] += 1

    meta = DATASETS[dataset_key]
    with open_data_store(file_path, decode_times=True) as ds:
        time_name = detect_time_coord(ds)
        lat_name, lon_name = detect_lat_lon(ds)
        var_name = pick_variable(ds, dataset_key, variable)
        da = ds[var_name]
        actual_variable_label = str(da.attrs.get("long_name") or da.attrs.get("standard_name") or var_name).replace("_", " ").title()
        variable_label = requested_element_label(dataset_key, variable) or actual_variable_label
        source_units = da.attrs.get("units", meta.get("unit", ""))
        # Apply all selectors while the array is still lazy.  Only the nearest
        # grid cell and requested time chunks are materialised by to_dataframe.
        if start_date or end_date:
            da, time_name = slice_time_range(da, time_name, start_date, end_date)
        point = da.sel({lat_name: lat, lon_name: lon}, method="nearest")
        nearest_lat = float(np.asarray(point[lat_name].values).reshape(-1)[0])
        nearest_lon = float(np.asarray(point[lon_name].values).reshape(-1)[0])
        df = _point_dataarray_to_frame(point, time_name, variable or var_name)
        df["value"], unit = convert_series_units(df["value"], source_units, meta["family"], var_name)
        df = df[["time", "value"]].dropna(subset=["value"]).sort_values("time").reset_index(drop=True)
    context = {
        "file": file_path.name,
        "storage_format": store_kind(file_path),
        "source_path": str(file_path),
        "dataset_key": dataset_key,
        "dataset_label": meta["label"],
        "family": meta["family"],
        "variable": str(variable or var_name),
        "actual_variable": var_name,
        "variable_label": variable_label,
        "actual_variable_label": actual_variable_label,
        "unit": unit,
        "resolution": resolution,
        "season": season or "",
        "requested_latitude": lat,
        "requested_longitude": lon,
        "nearest_latitude": nearest_lat,
        "nearest_longitude": nearest_lon,
    }
    if cache is not None:
        cache["point"][key] = (df.copy(deep=False), dict(context))
    return df.copy(deep=False), context


def extract_grid_slice(
    data_dir: Path,
    dataset_key: str,
    resolution: str,
    date_value: str | None,
    variable: str | None = None,
    season: str | None = None,
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    file_path = find_file(data_dir, dataset_key, resolution, season=season)
    assert file_path
    cache = _PRODUCT_DATA_CACHE.get()
    key = (
        "grid", str(file_path.resolve()), dataset_key, resolution,
        str(date_value or ""), str(variable or "auto"), str(season or ""),
    )
    if cache is not None and key in cache["grid"]:
        cache["stats"]["grid_hits"] += 1
        cached_df, cached_context = cache["grid"][key]
        return cached_df.copy(deep=False), dict(cached_context)
    if cache is not None:
        cache["stats"]["grid_misses"] += 1

    meta = DATASETS[dataset_key]
    with open_data_store(file_path, decode_times=True) as ds:
        time_name = detect_time_coord(ds)
        lat_name, lon_name = detect_lat_lon(ds)
        var_name = pick_variable(ds, dataset_key, variable)
        da = ds[var_name]
        if date_value:
            da, time_name = select_nearest_time(da, time_name, pd.Timestamp(date_value))
        else:
            da = da.isel({time_name: 0}) if time_name in da.dims else da
        da, unit = convert_dataarray_units(da, meta["family"], var_name)
        selected_time = ""
        if time_name in da.coords:
            try:
                selected_time = str(pd.to_datetime(np.asarray(da[time_name].values).reshape(-1)[0]))
            except Exception:
                selected_time = ""
        df = _grid_dataarray_to_frame(da, lat_name, lon_name, variable or var_name)
        df = df[["latitude", "longitude", "value"]].dropna(subset=["value"]).reset_index(drop=True)
    context = {
        "file": file_path.name,
        "storage_format": store_kind(file_path),
        "source_path": str(file_path),
        "dataset_key": dataset_key,
        "dataset_label": meta["label"],
        "family": meta["family"],
        "variable": var_name,
        "unit": unit,
        "resolution": resolution,
        "selected_time": selected_time,
    }
    if cache is not None:
        cache["grid"][key] = (df.copy(deep=False), dict(context))
    return df.copy(deep=False), context


def _append_parquet_database(table: str, df: pd.DataFrame, stem: str) -> Path:
    """Store generated records as Parquet under storage/db/parquet.

    This acts as a lightweight file-based database and does not touch the Zarr
    source stores. Each generation creates a separate Parquet partition file.
    """
    db_dir = PARQUET_DB_DIR / table
    db_dir.mkdir(parents=True, exist_ok=True)
    safe = slugify(stem)
    db_path = db_dir / f"{safe}.parquet"
    df.to_parquet(db_path, index=False)
    return db_path


def _round_output_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Round exported CSV/parquet data to operational precision."""
    out = df.copy()
    zero_keywords = ["relative_humidity", "relative humidity", "humidity", "wind_speed", "wind speed", "wind_direction", "wind direction", "direction", "speed"]
    for col in out.columns:
        if not pd.api.types.is_numeric_dtype(out[col]):
            continue
        name = str(col).lower()
        if any(k in name for k in zero_keywords):
            out[col] = pd.to_numeric(out[col], errors="coerce").round(0).astype("Int64")
        elif name in {"lat", "latitude", "lon", "longitude", "year", "month", "day", "hour"}:
            continue
        else:
            out[col] = pd.to_numeric(out[col], errors="coerce").round(1)
    return out


def _compact_export_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Keep product CSVs focused on the important columns."""
    out = _round_output_dataframe(df)
    preferred = [c for c in ["time", "date", "year", "month", "season", "latitude", "longitude", "value"] if c in out.columns]
    if preferred:
        extra_value_cols = [c for c in out.columns if c not in preferred and any(k in c.lower() for k in ["rain", "temp", "humidity", "wind", "index", "anomaly", "total", "days", "speed", "direction", "precip"])]
        keep = []
        for c in preferred + extra_value_cols:
            if c not in keep:
                keep.append(c)
        return out[keep]
    return out


def _save_data(df: pd.DataFrame, out_dir: Path, stem: str, table: str = "products") -> Tuple[Path, Path, Path]:
    csv_path = out_dir / f"{stem}.csv"
    parquet_path = out_dir / f"{stem}.parquet"
    clean_df = _compact_export_dataframe(df)
    clean_df.to_csv(csv_path, index=False)
    clean_df.to_parquet(parquet_path, index=False)
    db_path = _append_parquet_database(table, clean_df, stem)
    return csv_path, parquet_path, db_path


def _finalize_plot(fig, out_path: Path) -> Path:
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return out_path


def apply_plot_grids(ax):
    try:
        ax.minorticks_on()
    except Exception:
        pass
    ax.grid(True, which="major", alpha=0.35, linewidth=0.8)
    ax.grid(True, which="minor", alpha=0.15, linewidth=0.5)


def clean_source_words(text: str) -> str:
    """Remove data-source/vendor words from plot titles while keeping variables clear."""
    text = str(text or "")
    patterns = [
        r"\bCHIRPS\b",
        r"\bERA5[-\s]?Land\b",
        r"\bERA5\b",
        r"\bECMWF\b",
        r"\bCopernicus\b",
        r"\bReanalysis\b",
        r"Climate Hazards Group Infra\s*red Precipitation With Stations",
        r"Climate Hazards Group Infra\s*red Precipitation With Station Data",
        r"Climate Hazards Group Infra\s*red Precipitation",
        r"Climate Hazards Group",
        r"With Station Data",
        r"With Stations",
    ]
    for pat in patterns:
        text = re.sub(pat, "", text, flags=re.IGNORECASE)
    text = text.replace("Rainfall", "Precipitation")
    text = re.sub(r"\s+", " ", text).strip(" -_")
    return text or "Weather Element"

def short_dataset_label(label: str) -> str:
    label = clean_source_words(label)
    label = label.replace("Temperature Mean, Min and Max", "Temperature")
    label = label.replace("Volumetric Soil Moisture", "Soil Moisture")
    label = label.replace("Total Precipitation", "Precipitation")
    label = label.replace("Rainfall", "Precipitation")
    return label or "Weather Element"




def source_prefix_for_dataset(dataset_key: str | None) -> str:
    key = str(dataset_key or "")
    if key == "chirps_rainfall":
        return "CHIRPS"
    if key in {"era5_soil_water", "era5_soil_temperature"}:
        return "ERA5-Land"
    if key.startswith("era5_"):
        return "ERA5"
    return ""


def element_with_source(dataset_key: str | None, element: str) -> str:
    prefix = source_prefix_for_dataset(dataset_key)
    element = str(element or "Weather Element").replace("Rainfall", "Precipitation")
    if prefix and not element.lower().startswith(prefix.lower()):
        return f"{prefix} {element}"
    return element


def y_axis_label(element: str | None, unit: str | None) -> str:
    element = str(element or "Value").replace("Rainfall", "Precipitation")
    # Keep the axis short; the title carries dataset/source names.
    for prefix in ["CHIRPS ", "ERA5-Land ", "ERA5 Land ", "ERA5 "]:
        if element.startswith(prefix):
            element = element[len(prefix):]
    unit = str(unit or "").strip()
    return f"{element} ({unit})" if unit else element



# Dataset/variable-specific colors for clear visual identity.
DATASET_COLORS = {
    "chirps_rainfall": "#1f78b4",          # blue precipitation
    "era5_total_precipitation": "#0b5cad", # darker blue precipitation
    "era5_temperature": "#d95f02",         # orange temperature
    "era5_dew_point": "#66a61e",           # green dew point
    "era5_relative_humidity": "#1b9e77",   # teal humidity
    "era5_skin_temperature": "#e7298a",    # magenta/pink temperature
    "era5_soil_temperature": "#a6761d",    # brown land temperature
    "era5_pressure_cloud": "#7570b3",      # purple pressure/cloud
    "era5_soil_water": "#2ca25f",          # green soil moisture
    "era5_wind": "#666666",                # grey wind
}
VARIABLE_COLORS = {
    "ta": "#d95f02",
    "tmean": "#d95f02",
    "mean_temperature": "#d95f02",
    "tmax": "#d73027",
    "tx": "#d73027",
    "maximum_temperature": "#d73027",
    "tmin": "#4575b4",
    "tn": "#4575b4",
    "minimum_temperature": "#4575b4",
    "wind_speed": "#525252",
    "wind_direction": "#969696",
    "relative_humidity": "#1b9e77",
    "soil_moisture": "#2ca25f",
    "precipitation": "#1f78b4",
    "rainfall": "#1f78b4",
}
DATASET_CMAPS = {
    "chirps_rainfall": "Blues",
    "era5_total_precipitation": "Blues",
    "era5_temperature": "coolwarm",
    "era5_dew_point": "YlGn",
    "era5_relative_humidity": "YlGnBu",
    "era5_skin_temperature": "coolwarm",
    "era5_soil_temperature": "YlOrBr",
    "era5_pressure_cloud": "viridis",
    "era5_soil_water": "YlGn",
    "era5_wind": "viridis",
}
INDEX_COLORS = {
    "mean_temperature": "#d95f02",
    "maximum_temperature": "#d73027",
    "minimum_temperature": "#4575b4",
    "temperature_anomaly": "#7570b3",
    "hot_days": "#d73027",
    "hot_nights": "#f46d43",
    "cold_days": "#4575b4",
    "cold_nights": "#313695",
    "heat_index": "#e7298a",
    "dtr": "#a6761d",
    "relative_humidity_index": "#1b9e77",
    "soil_moisture_index": "#2ca25f",
    "soil_moisture_anomaly": "#66a61e",
    "wind_speed_index": "#525252",
}


def plot_color_for(dataset_key: str | None = None, variable: str | None = None, index_type: str | None = None) -> str:
    if index_type and index_type in INDEX_COLORS:
        return INDEX_COLORS[index_type]
    v = str(variable or "").lower()
    if v in VARIABLE_COLORS:
        return VARIABLE_COLORS[v]
    return DATASET_COLORS.get(str(dataset_key or ""), "#1f77b4")


def plot_cmap_for(dataset_key: str | None = None) -> str:
    return DATASET_CMAPS.get(str(dataset_key or ""), "viridis")


def climate_index_csv_dataframe(summary: pd.DataFrame, selected_col: str | None = None, index_type: str | None = None) -> pd.DataFrame:
    """Clean climate-index CSVs only, without changing Data Extractor output.

    Keep just the columns users normally need: year/season/location, dataset/source,
    the selected index value, and key supporting context. Values are rounded for
    readable operational CSVs.
    """
    out = summary.copy()
    selected_col = selected_col if selected_col in out.columns else None
    fixed = [
        "year", "season", "season_year", "month", "location", "index_type",
        "rainfall_dataset", "source_resolution_used", "source_file_season",
        "selected_index_column",
    ]
    metric_cols = []
    if selected_col:
        metric_cols.append(selected_col)
    # Keep a few useful companion columns for rainfall anomaly/normal products.
    for c in [
        "rainfall_total_mm", "rainfall_anomaly_mm", "percentage_of_normal",
        "spi_normal_approx", "spei_normal_approx", "rainy_season_onset",
        "rainy_season_cessation",
    ]:
        if c in out.columns and c not in metric_cols:
            metric_cols.append(c)
    keep = []
    for c in fixed + metric_cols:
        if c in out.columns and c not in keep:
            keep.append(c)
    if not keep:
        keep = list(out.columns)
    out = out[keep].copy()
    return _round_output_dataframe(out)

def _month_labels(values):
    return [pd.Timestamp(2000, int(m), 1).strftime("b") if False else pd.Timestamp(2000, int(m), 1).strftime("%b") for m in values]


def resolution_text(resolution: str) -> str:
    mapping = {"hourly": "Hourly", "daily": "Daily", "monthly": "Monthly", "annual": "Annual", "seasonal": "Seasonal"}
    return mapping.get(str(resolution).lower(), str(resolution).replace("_", " ").title())


def variable_display_name(variable: str | None, ctx: Dict[str, Any] | None = None, dataset_key: str | None = None) -> str:
    key = dataset_key or (ctx or {}).get("dataset_key") or ""
    var = str(variable or "auto")
    mapping = {
        ("chirps_rainfall", "auto"): "CHIRPS Precipitation",
        ("era5_total_precipitation", "auto"): "ERA5 Precipitation",
        ("era5_temperature", "ta"): "Mean Temperature",
        ("era5_temperature", "tmean"): "Mean Temperature",
        ("era5_temperature", "tmax"): "Maximum Temperature",
        ("era5_temperature", "tx"): "Maximum Temperature",
        ("era5_temperature", "tmin"): "Minimum Temperature",
        ("era5_temperature", "tn"): "Minimum Temperature",
        ("era5_dew_point", "auto"): "Dew Point Temperature",
        ("era5_dew_point", "d2m"): "Dew Point Temperature at 2 m",
        ("era5_skin_temperature", "auto"): "Skin Temperature",
        ("era5_skin_temperature", "skt"): "Skin Temperature",
        ("era5_soil_temperature", "auto"): "Soil Temperature Level 1",
        ("era5_soil_temperature", "stl1"): "Soil Temperature Level 1",
        ("era5_relative_humidity", "auto"): "Relative Humidity",
        ("era5_relative_humidity", "r"): "Relative Humidity",
        ("era5_relative_humidity", "rh"): "Relative Humidity",
        ("era5_soil_water", "auto"): "Volumetric Soil Moisture",
        ("era5_soil_water", "swvl1"): "Volumetric Soil Moisture",
        ("era5_wind", "wind_speed"): "10 m Wind Speed",
        ("era5_wind", "wind_direction"): "10 m Wind Direction",
        ("era5_pressure_cloud", "msl"): "Mean Sea Level Pressure",
        ("era5_pressure_cloud", "sp"): "Surface Pressure",
        ("era5_pressure_cloud", "tcc"): "Total Cloud Cover",
    }
    if (key, var) in mapping:
        return mapping[(key, var)]
    if var in {"", "auto", "None"} and key in DATASETS:
        family_defaults = {
            "rainfall": "Precipitation",
            "temperature": short_dataset_label(DATASETS[key]["label"]),
            "humidity": "Relative Humidity",
            "soil_moisture": "Soil Moisture",
            "wind": "Wind Speed",
            "pressure_cloud": "Weather Element",
        }
        return family_defaults.get(DATASETS[key].get("family", ""), short_dataset_label(DATASETS[key]["label"]))
    if ctx and ctx.get("variable_label"):
        return clean_source_words(str(ctx.get("variable_label"))).replace("Rainfall", "Precipitation")
    return clean_source_words(str(variable or "Weather Element").replace("_", " ").title()).replace("Rainfall", "Precipitation")


def plot_period_text(resolution: str, start: str | None, end: str | None, map_date: str | None = None) -> str:
    if map_date:
        try:
            dt = pd.to_datetime(map_date)
            return dt.strftime('%Y') if str(resolution).lower() in {'annual', 'monthly'} else dt.strftime('%Y-%m-%d')
        except Exception:
            return str(map_date)[:10]
    sy = str(start or '')[:4]
    ey = str(end or '')[:4]
    if sy and ey:
        return f"{sy} - {ey}"
    return sy or ey or ''


def plot_title_for(plot_type: str, element: str, location: str, resolution: str, start: str | None = None, end: str | None = None, map_date: str | None = None) -> str:
    element = str(element or "Weather Element").replace(" And ", " and ").strip()
    if element.lower() in {"rainfall", "precip"}:
        element = "Precipitation"
    period = plot_period_text(resolution, start, end, map_date)
    rt = resolution_text(resolution)
    loc = location or "Selected Location"

    if plot_type == 'spatial_map':
        if str(resolution).lower() == 'monthly':
            return f"Monthly Average {element} over Tanzania ({period})"
        if str(resolution).lower() == 'annual':
            return f"Annual {element} over Tanzania ({period})"
        return f"{rt} {element} over Tanzania ({period})"

    if plot_type in {'time_series', 'bar', 'area', 'monthly_climatology'}:
        if str(resolution).lower() == 'monthly':
            return f"Monthly Average {element} for {loc} ({period})"
        if str(resolution).lower() == 'annual':
            return f"Annual {element} for {loc} ({period})"
        if str(resolution).lower() == 'daily':
            return f"Daily {element} for {loc} ({period})"
        if str(resolution).lower() == 'hourly':
            return f"Hourly {element} for {loc} ({period})"
        return f"{rt} {element} for {loc} ({period})"

    if plot_type == 'annual_trend':
        return f"Annual Trend of {element} for {loc} ({period})"
    if plot_type == 'anomaly':
        return f"Climate Anomaly of {element} for {loc} ({period})"
    if plot_type == 'heatmap':
        return f"Heat Map of {element} for {loc} ({period})"
    if plot_type == 'histogram':
        return f"Histogram of {element} for {loc} ({period})"
    if plot_type == 'wind_rose':
        return f"Wind Rose for {loc} ({period})"
    return f"{element} for {loc} ({period})"

def rainfall_stats_text(values: pd.Series) -> str:
    vals = pd.to_numeric(values, errors='coerce').dropna()
    if vals.empty:
        return ''
    return f"Average: {vals.mean():,.1f}  Total: {vals.sum():,.1f}"


def annual_trend_frame(df: pd.DataFrame, family: str) -> tuple[pd.DataFrame, dict[str, float | str]]:
    """Aggregate values annually and calculate a reproducible linear trend.

    Precipitation is summed by year; all other variables are averaged by year.
    The returned table includes the fitted value used in the plot and Excel
    export, together with slope, intercept and coefficient of determination.
    """
    work = df.copy()
    work["time"] = pd.to_datetime(work["time"], errors="coerce")
    work["value"] = pd.to_numeric(work["value"], errors="coerce")
    work = work.dropna(subset=["time", "value"])
    work["year"] = work["time"].dt.year
    aggregation = "sum" if str(family).lower() == "rainfall" else "mean"
    annual = work.groupby("year", as_index=False)["value"].agg(aggregation)
    annual = annual.rename(columns={"value": "annual_value"}).sort_values("year").reset_index(drop=True)
    if len(annual) < 2:
        annual["linear_trend"] = np.nan
        return annual, {
            "slope_per_year": np.nan, "intercept": np.nan, "r_squared": np.nan,
            "aggregation": aggregation,
        }
    x = annual["year"].to_numpy(dtype=float)
    y = annual["annual_value"].to_numpy(dtype=float)
    slope, intercept = np.polyfit(x, y, 1)
    fitted = slope * x + intercept
    residual = float(np.sum((y - fitted) ** 2))
    total = float(np.sum((y - np.mean(y)) ** 2))
    r_squared = 1.0 - residual / total if total > 0 else 0.0
    annual["linear_trend"] = fitted
    return annual, {
        "slope_per_year": float(slope),
        "intercept": float(intercept),
        "r_squared": float(r_squared),
        "aggregation": aggregation,
    }


def _annual_trend_labels(element_name: str, unit: str, family: str) -> tuple[str, str, str]:
    clean = re.sub(r"^(CHIRPS|ERA5(?:-Land)?)\s+", "", str(element_name or "Weather Element"), flags=re.I).strip()
    display_unit = _normalise_display_unit(unit) if '_normalise_display_unit' in globals() else str(unit or "")
    if str(family).lower() == "rainfall":
        title_element = "Rainfall"
        series_label = "Annual rainfall"
        # Annual precipitation is a yearly accumulation, so the axis unit is mm
        # even when the source time step is expressed as mm/day or mm/hour.
        y_label = "Annual Rainfall (mm)"
    else:
        title_element = clean
        series_label = f"Annual mean {clean}"
        y_label = f"Annual Mean {clean}" + (f" ({display_unit})" if display_unit else "")
    return title_element, series_label, y_label



def monthly_average_total_dataframe(df: pd.DataFrame, family: str) -> pd.DataFrame:
    """For monthly files and a selected year range, compute the average value for each calendar month.

    For precipitation/rainfall monthly files, each input value is already a monthly total,
    so the output is the average of those total monthly precipitation values across the selected years.
    Example: Jan = mean(Jan 1991, Jan 1992, ..., Jan 2020).
    """
    tmp = df.copy()
    tmp["year"] = tmp["time"].dt.year
    tmp["month"] = tmp["time"].dt.month
    grp = tmp.groupby("month", as_index=False)["value"].mean()
    grp["month_name"] = grp["month"].apply(lambda m: pd.Timestamp(2000, int(m), 1).strftime("%b"))
    if family == "rainfall":
        grp["statistic"] = "Average of total monthly precipitation across selected years"
    else:
        grp["statistic"] = "Monthly average across selected years"
    return grp[["month", "month_name", "value", "statistic"]]

def decorate_tanzania_map(ax):
    """Plain spatial maps: no borders, lakes or ocean overlays."""
    return


def dataset_allowed_plots(dataset_key: str) -> list[str]:
    meta = DATASETS.get(dataset_key, {})
    return list(PLOT_FAMILIES.get(meta.get("family", "rainfall"), []))


def _extract_multi_temperature_series(data_dir, resolution, lat, lon, start, end, season=None):
    mapping = [("ta", "Mean Temperature"), ("tmax", "Maximum Temperature"), ("tmin", "Minimum Temperature")]
    frames = []
    contexts = {}
    for var, label in mapping:
        df, ctx = extract_point_series(data_dir, "era5_temperature", resolution, lat, lon, start, end, variable=var, season=season)
        df = df.rename(columns={"value": var})
        frames.append(df[["time", var]])
        contexts[var] = {**ctx, "variable_label": label}
    out = frames[0]
    for df in frames[1:]:
        out = out.merge(df, on="time", how="outer")
    out = out.sort_values("time")
    return out, contexts


def _extract_multi_temperature_grid(data_dir, resolution, date_value, season=None):
    mapping = [("ta", "Mean Temperature"), ("tmax", "Maximum Temperature"), ("tmin", "Minimum Temperature")]
    grids = []
    contexts = {}
    for var, label in mapping:
        df, ctx = extract_grid_slice(data_dir, "era5_temperature", resolution, date_value, variable=var, season=season)
        df = df.rename(columns={"value": var})
        grids.append(df[[c for c in df.columns if c in ["latitude","longitude",var]]])
        contexts[var] = {**ctx, "variable_label": label}
    out = grids[0]
    for df in grids[1:]:
        out = out.merge(df, on=["latitude","longitude"], how="outer")
    return out, contexts



def _aggregate_temperature_plot_frame(df: pd.DataFrame, resolution: str) -> pd.DataFrame:
    tmp = df.copy().sort_values("time")
    tmp["year"] = tmp["time"].dt.year
    tmp["month"] = tmp["time"].dt.month
    if resolution == "monthly":
        # Average each calendar month across selected years.
        return tmp.groupby("month", as_index=False)[["ta", "tmax", "tmin"]].mean()
    if resolution == "annual":
        return tmp.groupby("year", as_index=False)[["ta", "tmax", "tmin"]].mean()
    # Daily/hourly: plot direct values.
    return tmp


def _plot_temperature_combined(df: pd.DataFrame, variables: list[str], resolution: str, location: str, start: str | None, end: str | None, out_path: Path, moving_average: bool = True) -> Path:
    labels = {"ta": "Mean Temperature", "tmax": "Maximum Temperature", "tmin": "Minimum Temperature"}
    colors = {"ta": "#d95f02", "tmax": "#d73027", "tmin": "#4575b4"}
    agg = _aggregate_temperature_plot_frame(df, resolution)
    fig, ax = plt.subplots(figsize=(10.8, 5.8))
    period = plot_period_text(resolution, start, end)
    try:
        sy, ey = int(str(start)[:4]), int(str(end)[:4])
    except Exception:
        sy = ey = 0
    many_years = (ey - sy + 1) > 20
    if resolution == "monthly" and "month" in agg.columns:
        x = np.arange(len(agg))
        xlabels = _month_labels(agg["month"])
        for v in variables:
            ax.plot(x, agg[v], marker="o", linewidth=2, label=labels[v], color=colors[v])
        ax.set_xticks(x); ax.set_xticklabels(xlabels); ax.set_xlabel("Month")
        title = f"Monthly Average Temperature for {location} ({period})"
    elif resolution == "annual" and "year" in agg.columns:
        for v in variables:
            ax.plot(agg["year"], agg[v], marker="o", linewidth=1.8, label=labels[v], color=colors[v])
            if moving_average and many_years:
                ma = pd.to_numeric(agg[v], errors="coerce").rolling(5, min_periods=3).mean()
                ax.plot(agg["year"], ma, linestyle="--", linewidth=2.2, color=colors[v], label=f"{labels[v]} 5-year moving average")
        ax.set_xlabel("Year")
        title = f"Annual Temperature for {location} ({period})"
    else:
        for v in variables:
            ax.plot(agg["time"], agg[v], linewidth=1.2, label=labels[v], color=colors[v])
            if moving_average and many_years:
                ma = pd.to_numeric(agg[v], errors="coerce").rolling(30, min_periods=10).mean()
                ax.plot(agg["time"], ma, linestyle="--", linewidth=1.8, color=colors[v], label=f"{labels[v]} moving average")
        ax.set_xlabel("Time")
        title = f"{resolution_text(resolution)} Temperature for {location} ({period})"
    ax.set_title(title)
    ax.set_ylabel("Temperature (°C)")
    ax.legend(loc="best", fontsize=8)
    apply_plot_grids(ax)
    return _finalize_plot(fig, out_path)

def generate_plot_product(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    dirs = ensure_output_dirs(export_dir)
    dataset_key = params.get("dataset")
    plot_type = params.get("plot_type", "time_series")
    if plot_type in {"monthly_climatology", "extreme_value"}:
        plot_type = "time_series"
    resolution = params.get("resolution", "monthly")
    variable = params.get("variable") or "auto"
    if plot_type == "wind_rose" and str(variable) in {"wind_speed_direction", "auto", ""}:
        variable = "wind_speed"
    season = (params.get("season") or "").strip().upper() or None
    lat = float(params.get("latitude") or 0)
    lon = float(params.get("longitude") or 0)
    location = params.get("location_name") or "Selected Location"
    start = params.get("start_date") or None
    end = params.get("end_date") or None
    date_value = params.get("map_date") or start or None
    baseline_start = int(params.get("baseline_start") or 1991)
    baseline_end = int(params.get("baseline_end") or 2020)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = slugify(f"CDE_{plot_type}_{dataset_key}_{resolution}_{location}_{timestamp}")
    family = DATASETS[dataset_key]["family"]

    if plot_type == "spatial_map":
        if dataset_key == "era5_temperature" and variable in {None, "", "auto", "all", "all_in_one", "max_min"}:
            grid, ctxs = _extract_multi_temperature_grid(data_dir, resolution, date_value, season=season)
            grid = grid[(grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])) &
                        (grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"]))].copy()
            csv_path, parquet_path, db_path = _save_data(grid, dirs["plots"], stem + "_grid_data", table="plot_products")
            fig, axes = plt.subplots(3, 1, figsize=(8.5, 14), sharex=True, sharey=True)
            for ax, var, title in zip(axes, ["ta","tmax","tmin"], ["Mean Temperature", "Maximum Temperature", "Minimum Temperature"]):
                pivot = grid.pivot_table(index="latitude", columns="longitude", values=var, aggfunc="mean").sort_index()
                x = pivot.columns.values.astype(float); y = pivot.index.values.astype(float)
                im = ax.pcolormesh(x, y, pivot.values, shading="auto", zorder=3, cmap=plot_cmap_for(dataset_key))
                decorate_tanzania_map(ax)
                ax.set_xlim(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
                ax.set_ylim(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
                ax.set_title(f"{title} ({resolution_text(resolution)})")
                ax.set_ylabel("Latitude")
                apply_plot_grids(ax)
                fig.colorbar(im, ax=ax, label=y_axis_label(title, '°C'))
            axes[-1].set_xlabel("Longitude")
            if TANZANIA_BOUNDS["lon_min"] <= lon <= TANZANIA_BOUNDS["lon_max"] and TANZANIA_BOUNDS["lat_min"] <= lat <= TANZANIA_BOUNDS["lat_max"]:
                for ax in axes:
                    ax.scatter([lon], [lat], marker="x", s=60, color="black")
            fig.suptitle(f"Spatial Distribution of Temperature over Tanzania ({plot_period_text(resolution, start, end, date_value)})", fontsize=14, y=0.995)
            plot_path = _finalize_plot(fig, dirs["plots"] / f"{stem}.png")
            return {"plot_path": plot_path, "csv_path": csv_path, "parquet_path": parquet_path, "db_path": db_path, "context": {"dataset_label": DATASETS[dataset_key]["label"], "file": ctxs['ta']['file']}, "rows": len(grid)}
        grid, ctx = extract_grid_slice(data_dir, dataset_key, resolution, date_value, variable, season=season)
        grid = grid[(grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])) &
                    (grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"]))].copy()
        csv_path, parquet_path, db_path = _save_data(grid, dirs["plots"], stem + "_grid_data", table="plot_products")
        fig, ax = plt.subplots(figsize=(8.5, 7))
        pivot = grid.pivot_table(index="latitude", columns="longitude", values="value", aggfunc="mean").sort_index()
        x = pivot.columns.values.astype(float); y = pivot.index.values.astype(float)
        im = ax.pcolormesh(x, y, pivot.values, shading="auto", zorder=3, cmap=plot_cmap_for(dataset_key))
        decorate_tanzania_map(ax)
        ax.set_xlim(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
        ax.set_ylim(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
        ax.set_xlabel("Longitude")
        ax.set_ylabel("Latitude")
        element_name = variable_display_name(variable, ctx, dataset_key)
        ax.set_title(plot_title_for("spatial_map", element_with_source(dataset_key, element_name), location, resolution, start, end, date_value))
        cb = fig.colorbar(im, ax=ax); cb.set_label(y_axis_label(element_name, ctx.get("unit", "")))
        if TANZANIA_BOUNDS["lon_min"] <= lon <= TANZANIA_BOUNDS["lon_max"] and TANZANIA_BOUNDS["lat_min"] <= lat <= TANZANIA_BOUNDS["lat_max"]:
            ax.scatter([lon], [lat], marker="x", s=60, color="black")
        apply_plot_grids(ax)
        plot_path = _finalize_plot(fig, dirs["plots"] / f"{stem}.png")
        return {"plot_path": plot_path, "csv_path": csv_path, "parquet_path": parquet_path, "db_path": db_path, "context": ctx, "rows": len(grid)}

    # time-series style products
    if dataset_key == "era5_temperature" and variable in {None, "", "auto", "all", "all_in_one", "max_min"}:
        df, ctxs = _extract_multi_temperature_series(data_dir, resolution, lat, lon, start, end, season=season)
        if df.empty:
            raise ValueError("No data found for the selected options/date range.")
        df["year"] = pd.to_datetime(df["time"]).dt.year
        df["month"] = pd.to_datetime(df["time"]).dt.month
        csv_path, parquet_path, db_path = _save_data(df, dirs["plots"], stem + "_data", table="plot_products")
        if variable in {"all_in_one", "all"} and plot_type in {"time_series", "annual_trend"}:
            plot_path = _plot_temperature_combined(df, ["ta", "tmax", "tmin"], resolution, location, start, end, dirs["plots"] / f"{stem}.png", moving_average=True)
            return {"plot_path": plot_path, "csv_path": csv_path, "parquet_path": parquet_path, "db_path": db_path, "context": {"dataset_label": DATASETS[dataset_key]["label"], "file": ctxs["ta"]["file"]}, "rows": len(df)}
        if variable == "max_min" and plot_type in {"time_series", "annual_trend"}:
            plot_path = _plot_temperature_combined(df, ["tmax", "tmin"], resolution, location, start, end, dirs["plots"] / f"{stem}.png", moving_average=True)
            return {"plot_path": plot_path, "csv_path": csv_path, "parquet_path": parquet_path, "db_path": db_path, "context": {"dataset_label": DATASETS[dataset_key]["label"], "file": ctxs["ta"]["file"]}, "rows": len(df)}
        fig, axes = plt.subplots(3, 1, figsize=(10, 13), sharex=False)
        temp_vars = [("ta", "Mean Temperature"), ("tmax", "Maximum Temperature"), ("tmin", "Minimum Temperature")]
        for ax, (v, label) in zip(axes, temp_vars):
            sdf = df[["time","year","month",v]].dropna().rename(columns={v:"value"})
            if plot_type in {"time_series"}:
                ax.plot(sdf["time"], sdf["value"], linewidth=1.4, marker="o" if resolution in {"monthly", "annual", "seasonal"} else None, markersize=2.8, color=plot_color_for(dataset_key, v))
                ax.set_xlabel("Time" if resolution not in {"annual", "seasonal"} else "Year")
            elif plot_type == "bar":
                if resolution == "monthly":
                    grp = sdf.groupby("month")["value"].mean()
                    ax.bar(_month_labels(grp.index), grp.values, color=plot_color_for(dataset_key, v))
                    ax.set_xlabel("Month")
                else:
                    grp = sdf.groupby("year")["value"].mean().reset_index() if resolution == "annual" else sdf.resample("MS", on="time")["value"].mean().reset_index()
                    x = grp["year"] if "year" in grp.columns else grp["time"].dt.strftime('%Y-%m')
                    ax.bar(x.astype(str), grp["value"], color=plot_color_for(dataset_key, v))
                    ax.set_xlabel("Year" if resolution == "annual" else "Time")
                    ax.tick_params(axis='x', rotation=45)
            elif plot_type == "heatmap":
                hm = sdf.pivot_table(index="year", columns="month", values="value", aggfunc="mean")
                im = ax.imshow(hm.values, aspect="auto")
                ax.set_xticks(np.arange(len(hm.columns))); ax.set_xticklabels(_month_labels(hm.columns))
                ax.set_yticks(np.arange(len(hm.index))); ax.set_yticklabels(hm.index)
                fig.colorbar(im, ax=ax, label=y_axis_label(label, '°C'))
            elif plot_type == "monthly_climatology":
                grp = sdf.groupby("month")["value"].mean(); ax.plot(_month_labels(grp.index), grp.values, marker='o', color=plot_color_for(dataset_key, v))
                ax.set_xlabel("Month")
            elif plot_type == "annual_trend":
                grp = sdf.groupby("year")["value"].mean().reset_index(); ax.plot(grp["year"], grp["value"], marker='o', color=plot_color_for(dataset_key, v))
                if len(grp)>1:
                    z=np.polyfit(grp['year'],grp['value'],1); ax.plot(grp['year'], np.poly1d(z)(grp['year']), linestyle='--')
                ax.set_xlabel("Year")
            elif plot_type == "seasonal":
                if not season: season='MAM'
                months = SEASON_DEFINITIONS.get(season,[3,4,5])
                sdf2 = sdf[sdf['month'].isin(months)].copy()
                sdf2['season_year']=pd.to_datetime(sdf2['time']).apply(lambda x: season_year(pd.Timestamp(x), months))
                grp=sdf2.groupby('season_year')['value'].mean().reset_index(); ax.plot(grp['season_year'], grp['value'], marker='o')
                ax.set_xlabel('Year')
            elif plot_type == 'anomaly':
                grp=sdf.groupby('year')['value'].mean().reset_index(); base=grp[(grp['year']>=baseline_start)&(grp['year']<=baseline_end)]
                if base.empty: base=grp
                normal=base['value'].mean(); grp['anom']=grp['value']-normal; colors=np.where(grp['anom']>=0, '#d73027', '#4575b4'); ax.bar(grp['year'], grp['anom'], color=colors); ax.axhline(0,color='black',linewidth=1)
                ax.set_xlabel('Year')
            elif plot_type == 'box':
                data=[sdf[sdf['month']==m]['value'].dropna().values for m in range(1,13)]; _boxplot_with_labels(ax, data, _month_labels(range(1,13)), showfliers=False); ax.set_xlabel('Month')
            elif plot_type == 'histogram':
                ax.hist(sdf['value'].dropna(), bins=30); ax.set_xlabel('°C'); ax.set_ylabel('Frequency')
            elif plot_type == 'area':
                xvals = pd.to_datetime(sdf['time'])
                ax.fill_between(xvals, sdf['value'], alpha=0.4)
                ax.plot(xvals, sdf['value'])
                ax.set_xlabel('Time' if resolution not in {'annual','seasonal'} else 'Year')
            elif plot_type == 'extreme_value':
                grp=sdf.groupby('year')['value'].max().reset_index(); ax.bar(grp['year'], grp['value']); ax.set_xlabel('Year')
            ax.set_title(plot_title_for(plot_type, element_with_source(dataset_key, label), location, resolution, start, end))
            ax.set_ylabel(y_axis_label(label if 'label' in locals() else 'Temperature', '°C'))
            apply_plot_grids(ax)
        fig.suptitle(f"Temperature for {location} ({plot_period_text(resolution, start, end)}, {resolution_text(resolution)})", fontsize=14, y=0.995)
        plot_path = _finalize_plot(fig, dirs['plots']/f"{stem}.png")
        return {"plot_path": plot_path, "csv_path": csv_path, "parquet_path": parquet_path, "db_path": db_path, "context": {"dataset_label": DATASETS[dataset_key]['label'], "file": ctxs['ta']['file']}, "rows": len(df)}

    df, ctx = extract_point_series(data_dir, dataset_key, resolution, lat, lon, start, end, variable, season=season)
    if df.empty:
        raise ValueError("No data found for the selected options/date range.")
    df = df.sort_values("time")
    df["year"] = df["time"].dt.year
    df["month"] = df["time"].dt.month
    csv_path = parquet_path = db_path = None
    data_to_save = df.copy()
    unit = ctx.get('unit', '')
    fig, ax = plt.subplots(figsize=(10, 5.8))
    title = short_dataset_label(ctx.get('dataset_label'))
    element_name = variable_display_name(variable, ctx, dataset_key)
    plot_element_name = element_with_source(dataset_key, element_name)
    if plot_type == 'wind_rose':
        plot_element_name = 'Wind Speed and Direction'
    if plot_type in {'time_series'}:
        ax.plot(df['time'], df['value'], linewidth=1.5, marker='o' if resolution in {'monthly','annual','seasonal'} else None, markersize=2.8, color=plot_color_for(dataset_key, variable))
        if family == 'rainfall' and len(df):
            avg = pd.to_numeric(df['value'], errors='coerce').mean()
            ax.axhline(avg, linestyle='--', linewidth=1.2)
        ax.set_xlabel('Time' if resolution not in {'annual','seasonal'} else 'Year')
        ax.set_title(plot_title_for('time_series', plot_element_name, location, resolution, start, end))
        ax.set_ylabel(y_axis_label(element_name, unit))
    elif plot_type == 'bar':
        x = pd.to_datetime(df['time'])
        if resolution in {'annual', 'seasonal'}:
            x = df['year'].astype(str)
        ax.bar(x, df['value'], color=plot_color_for(dataset_key, variable))
        ax.tick_params(axis='x', rotation=45)
        ax.set_xlabel('Year' if resolution in {'annual','seasonal'} else 'Time')
        ax.set_title(plot_title_for('bar', plot_element_name, location, resolution, start, end)); ax.set_ylabel(y_axis_label(element_name, unit))
    elif plot_type == 'heatmap':
        hm = df.pivot_table(index='year', columns='month', values='value', aggfunc='sum' if family=='rainfall' else 'mean')
        im = ax.imshow(hm.values, aspect='auto'); ax.set_xticks(np.arange(len(hm.columns))); ax.set_xticklabels(_month_labels(hm.columns)); ax.set_yticks(np.arange(len(hm.index))); ax.set_yticklabels(hm.index); fig.colorbar(im, ax=ax, label=y_axis_label(element_name, unit)); ax.set_title(plot_title_for('heatmap', plot_element_name, location, resolution, start, end))
    elif plot_type == 'monthly_climatology':
        grp = monthly_average_total_dataframe(df, family)
        data_to_save = grp.copy()
        ax.plot(grp['month_name'], grp['value'], marker='o', color=plot_color_for(dataset_key, variable))
        ax.set_xlabel('Month'); ax.set_title(plot_title_for('monthly_climatology', plot_element_name, location, resolution, start, end)); ax.set_ylabel(y_axis_label(element_name, unit))
        if family == 'rainfall':
            ax.text(0.98, 0.98, rainfall_stats_text(grp['value']), transform=ax.transAxes, ha='right', va='top', fontsize=8, bbox=dict(boxstyle='round,pad=0.25', facecolor='white', alpha=0.8))
    elif plot_type == 'annual_trend':
        grp, trend_stats = annual_trend_frame(df, family)
        data_to_save = grp.copy()
        title_element, series_label, annual_y_label = _annual_trend_labels(element_name, unit, family)
        line_colour = plot_color_for(dataset_key, variable)
        marker = 'o' if family == 'rainfall' else None
        ax.plot(
            grp['year'], grp['annual_value'], marker=marker, linewidth=1.8,
            markersize=5 if marker else 0, color=line_colour, label=series_label,
        )
        if len(grp) > 1 and np.isfinite(float(trend_stats.get('slope_per_year', np.nan))):
            ax.plot(
                grp['year'], grp['linear_trend'], linestyle=':', linewidth=2.2,
                color='#ff7f0e', label='Linear trend',
            )
            slope = float(trend_stats['slope_per_year'])
            intercept = float(trend_stats['intercept'])
            sign = '+' if intercept >= 0 else '-'
            equation = f"y = {slope:.4f}x {sign} {abs(intercept):.3f}"
            annotation = f"{equation}\nR² = {float(trend_stats['r_squared']):.4f}"
            ax.text(
                0.98, 0.955, annotation, transform=ax.transAxes, ha='right', va='top',
                fontsize=10.5, bbox=dict(boxstyle='round,pad=0.28', facecolor='white', edgecolor='none', alpha=0.88),
            )
        start_year = int(grp['year'].min()) if not grp.empty else str(start)[:4]
        end_year = int(grp['year'].max()) if not grp.empty else str(end)[:4]
        custom_title = ' '.join(str(params.get('custom_plot_title') or '').split())[:180]
        automatic_title = f"Long-Term Annual {title_element} Trend for {location} ({start_year}–{end_year})"
        ax.set_xlabel('Years')
        ax.set_ylabel(annual_y_label)
        ax.set_title(custom_title or automatic_title, fontsize=14, pad=13)
        ax.legend(loc='upper left')
        ctx = dict(ctx)
        ctx.update({
            'annual_trend_aggregation': trend_stats.get('aggregation'),
            'slope_per_year': trend_stats.get('slope_per_year'),
            'trend_intercept': trend_stats.get('intercept'),
            'r_squared': trend_stats.get('r_squared'),
            'annual_trend_title': custom_title or automatic_title,
            'annual_trend_y_label': annual_y_label,
        })
    elif plot_type == 'seasonal':
        if not season: season='MAM'
        months = SEASON_DEFINITIONS.get(season,[3,4,5]); sdf=df[df['month'].isin(months)].copy(); sdf['season_year']=sdf['time'].apply(lambda x: season_year(pd.Timestamp(x), months)); grp=sdf.groupby('season_year')['value'].agg('sum' if family=='rainfall' else 'mean').reset_index(); (ax.bar if family=='rainfall' else ax.plot)(grp['season_year'], grp['value'], marker='o' if family!='rainfall' else None); ax.set_xlabel('Year'); ax.set_title(f"{title} - {season} Seasonal Plot"); ax.set_ylabel(y_axis_label(element_name, unit))
    elif plot_type == 'anomaly':
        grp = df.groupby('year')['value'].agg('sum' if family=='rainfall' else 'mean').reset_index(); base=grp[(grp['year']>=baseline_start)&(grp['year']<=baseline_end)];
        if base.empty: base=grp
        normal=base['value'].mean(); grp['anomaly']=grp['value']-normal; colors=np.where(grp['anomaly']>=0, '#d73027', '#4575b4'); ax.bar(grp['year'], grp['anomaly'], color=colors); ax.axhline(0,color='black',linewidth=1); ax.set_xlabel('Year'); ax.set_title(plot_title_for('anomaly', plot_element_name, location, resolution, start, end)); ax.set_ylabel(y_axis_label(element_name + ' Anomaly', unit))
    elif plot_type == 'wind_rose':
        plt.close(fig)
        speed, direction, ctx = wind_speed_direction_series(data_dir, dataset_key, resolution, lat, lon, start, end, season=season)
        wdf = pd.DataFrame({'Wind Speed': pd.to_numeric(speed, errors='coerce'), 'Wind Direction': pd.to_numeric(direction, errors='coerce')}).dropna()
        csv_path, parquet_path, db_path = _save_data(wdf, dirs['plots'], stem + '_wind_rose_data', table='plot_products')
        fig = plt.figure(figsize=(9.5, 8.8))
        ax = fig.add_subplot(111, polar=True)
        # 18 wind directions, WRPLOT-like sectors.
        dir_bins = np.arange(0, 361, 20)
        theta = np.deg2rad((dir_bins[:-1] + dir_bins[1:]) / 2)
        width = np.deg2rad(20)
        unit_label = ctx.get('unit', '') or 'm/s'
        spd = wdf['Wind Speed']
        finite_spd = spd[np.isfinite(spd)]
        if finite_spd.empty:
            speed_bins = [0, 2, 4, 6, 8, 10, np.inf]
        else:
            mx = float(finite_spd.max())
            if mx <= 12:
                speed_bins = [0, 2, 4, 6, 8, 10, np.inf]
            else:
                upper = max(12, math.ceil(mx / 3.0) * 3.0)
                speed_bins = [0, 2, 4, 6, 8, 10, upper, np.inf]
        speed_labels = []
        for a, b in zip(speed_bins[:-1], speed_bins[1:]):
            speed_labels.append(f">= {a:g}" if np.isinf(b) else f"{a:g} - {b:g}")
        wdf['_dir_bin'] = pd.cut(wdf['Wind Direction'] % 360, bins=dir_bins, include_lowest=True, labels=False)
        wdf['_speed_bin'] = pd.cut(wdf['Wind Speed'], bins=speed_bins, include_lowest=True, right=False, labels=speed_labels)
        table = pd.crosstab(wdf['_dir_bin'], wdf['_speed_bin']).reindex(index=range(len(theta)), columns=speed_labels, fill_value=0)
        total_count = table.to_numpy().sum()
        values_pct = table / total_count * 100.0 if total_count else table.astype(float)
        # WRPLOT-style bright colours by magnitude: calm/low near centre warm, high speeds green/blue.
        wind_colors = ['#ff8c00', '#ffff00', '#00e5ff', '#7d7dff', '#0000c8', '#00c800', '#008000'][:len(speed_labels)]
        bottom = np.zeros(len(theta))
        for i, label in enumerate(speed_labels):
            vals = values_pct[label].to_numpy(dtype=float)
            ax.bar(theta, vals, width=width, bottom=bottom, color=wind_colors[i], edgecolor='black', linewidth=0.35, alpha=0.98, label=f"{label} {unit_label}".strip())
            bottom += vals
        ax.set_theta_zero_location('N')
        ax.set_theta_direction(-1)
        ax.set_thetagrids([0, 90, 180, 270], labels=['NORTH', 'EAST', 'SOUTH', 'WEST'])
        ax.grid(True, linestyle=':', linewidth=0.75, color='#0033cc', alpha=0.85)
        ax.spines['polar'].set_color('#0033cc')
        ax.set_rlabel_position(45)
        ax.set_title(plot_title_for("wind_rose", "Wind Speed and Direction", location, resolution, start, end), pad=22, fontsize=14)
        ax.legend(title=f'WIND SPEED\n({unit_label})', loc='center left', bbox_to_anchor=(1.08, 0.18), fontsize=8, frameon=False)
        plot_path = _finalize_plot(fig, dirs['plots']/f"{stem}.png")
        return {'plot_path': plot_path, 'csv_path': csv_path, 'parquet_path': parquet_path, 'db_path': db_path, 'context': ctx, 'rows': len(wdf)}
    elif plot_type == 'box':
        data=[df[df['month']==m]['value'].dropna().values for m in range(1,13)]; _boxplot_with_labels(ax, data, _month_labels(range(1,13)), showfliers=False); ax.set_xlabel('Month'); ax.set_title(f"{title} - Box Plot"); ax.set_ylabel(y_axis_label(element_name, unit))
    elif plot_type == 'histogram':
        ax.hist(df['value'].dropna(), bins=30, color=plot_color_for(dataset_key, variable), alpha=0.85); ax.set_title(plot_title_for('histogram', plot_element_name, location, resolution, start, end)); ax.set_xlabel(unit); ax.set_ylabel('Frequency')
    elif plot_type == 'area':
        xvals = pd.to_datetime(df['time'])
        ax.fill_between(xvals, df['value'], alpha=0.25, color=plot_color_for(dataset_key, variable))
        ax.plot(xvals, df['value'], color=plot_color_for(dataset_key, variable))
        ax.set_xlabel('Time' if resolution not in {'annual','seasonal'} else 'Year')
        ax.set_title(plot_title_for('bar', plot_element_name, location, resolution, start, end)); ax.set_ylabel(y_axis_label(element_name, unit))
    elif plot_type == 'extreme_value':
        grp=df.groupby('year')['value'].max().reset_index(); ax.bar(grp['year'], grp['value'], color=plot_color_for(dataset_key, variable)); ax.set_title(plot_title_for('extreme_value', plot_element_name, location, resolution, start, end)); ax.set_xlabel('Year'); ax.set_ylabel(y_axis_label(element_name, unit))
    else:
        ax.plot(df['time'], df['value'], color=plot_color_for(dataset_key, variable)); ax.set_title(title)
    apply_plot_grids(ax)
    plot_path = _finalize_plot(fig, dirs['plots']/f"{stem}.png")
    if csv_path is None or parquet_path is None or db_path is None:
        csv_path, parquet_path, db_path = _save_data(data_to_save, dirs['plots'], stem + '_data', table='plot_products')
    return {"plot_path": plot_path, "csv_path": csv_path, "parquet_path": parquet_path, "db_path": db_path, "context": ctx, "rows": len(data_to_save)}

def season_year(ts: pd.Timestamp, months: List[int]) -> int:
    # For seasons crossing calendar years, Jan-Apr belong to the year in which the season ends.
    crosses = max(months) - min(months) > 6 or (12 in months and 1 in months)
    if crosses and ts.month <= max([m for m in months if m < 7] or [0]):
        return ts.year
    if crosses and ts.month >= min([m for m in months if m > 7] or [13]):
        return ts.year + 1
    return ts.year


def consecutive_count(mask: Iterable[bool]) -> int:
    best = cur = 0
    for val in mask:
        if bool(val):
            cur += 1
            best = max(best, cur)
        else:
            cur = 0
    return best


def parse_custom_months(season: str, custom_months: str | None = None) -> Tuple[str, List[int]]:
    if season == "CUSTOM":
        parts = [p for p in re.split(r"[\s,;|/]+", custom_months or "") if p]
        months = []
        for p in parts:
            m = int(p)
            if m < 1 or m > 12:
                raise ValueError("Custom season months must be between 1 and 12.")
            if m not in months:
                months.append(m)
        if not months:
            raise ValueError("Enter custom season months, for example 1,2,3.")
        return "CUSTOM_" + "_".join(map(str, months)), months
    return season, SEASON_DEFINITIONS.get(season, list(range(1, 13)))



def spell_max_length(mask: Iterable[bool]) -> int:
    return consecutive_count(mask)


def spell_duration_days(mask: Iterable[bool], minimum_length: int = 6) -> int:
    total = 0
    cur = 0
    for val in list(mask) + [False]:
        if bool(val):
            cur += 1
        else:
            if cur >= minimum_length:
                total += cur
            cur = 0
    return int(total)


def _tma_rainy_season_dates(grp: pd.DataFrame, rainy_threshold: float = 1.0) -> Tuple[pd.Timestamp | None, pd.Timestamp | None, int | None]:
    """Rainfall-only TMA-style onset/cessation rule.

    Onset: first 3-day period with at least 20 mm and at least two wet days,
    not followed by a dry spell of 10 or more consecutive days in the next 30 days.
    Cessation: first long dry spell after onset; the cessation date is the day before
    that dry spell starts. If no long dry spell is found, the last wet day is used.
    """
    if grp.empty:
        return None, None, None
    g = grp.sort_values("time").copy()
    r = pd.to_numeric(g["rainfall_mm"], errors="coerce").fillna(0.0).reset_index(drop=True)
    t = pd.to_datetime(g["time"]).reset_index(drop=True)
    onset_idx = None
    for i in range(0, max(0, len(g) - 2)):
        window = r.iloc[i:i+3]
        if float(window.sum()) >= 20.0 and int((window >= rainy_threshold).sum()) >= 2:
            look = r.iloc[i+3:i+33] if i + 3 < len(r) else pd.Series([], dtype=float)
            if spell_max_length((look < rainy_threshold).tolist()) < 10:
                onset_idx = i
                break
    if onset_idx is None:
        return None, None, None
    onset = pd.Timestamp(t.iloc[onset_idx])
    post = r.iloc[onset_idx:].reset_index(drop=True)
    post_dates = t.iloc[onset_idx:].reset_index(drop=True)
    cessation = None
    dry_count = 0
    for j, val in enumerate(post):
        if val < rainy_threshold:
            dry_count += 1
            if dry_count >= 10:
                dry_start = j - dry_count + 1
                cessation_idx = max(dry_start - 1, 0)
                cessation = pd.Timestamp(post_dates.iloc[cessation_idx])
                break
        else:
            dry_count = 0
    if cessation is None:
        wet_after = np.where(post.values >= rainy_threshold)[0]
        cessation = pd.Timestamp(post_dates.iloc[int(wet_after[-1])]) if len(wet_after) else pd.Timestamp(post_dates.iloc[-1])
    length = int((cessation - onset).days + 1) if cessation >= onset else None
    return onset, cessation, length


def _heat_index_celsius(temp_c: pd.Series, rh: pd.Series) -> pd.Series:
    t_f = (pd.to_numeric(temp_c, errors="coerce") * 9.0 / 5.0) + 32.0
    r = pd.to_numeric(rh, errors="coerce")
    hi_f = (-42.379 + 2.04901523*t_f + 10.14333127*r - 0.22475541*t_f*r
            - 0.00683783*(t_f**2) - 0.05481717*(r**2)
            + 0.00122874*(t_f**2)*r + 0.00085282*t_f*(r**2)
            - 0.00000199*(t_f**2)*(r**2))
    hi_c = (hi_f - 32.0) * 5.0 / 9.0
    return hi_c.where((temp_c >= 26.7) & (r >= 40.0), temp_c)


def _simple_pet_mm_day(tmean: pd.Series, tmin: pd.Series | None = None, tmax: pd.Series | None = None) -> pd.Series:
    tm = pd.to_numeric(tmean, errors="coerce")
    if tmin is not None and tmax is not None:
        tr = (pd.to_numeric(tmax, errors="coerce") - pd.to_numeric(tmin, errors="coerce")).clip(lower=0)
        pet = 0.0023 * (tm + 17.8).clip(lower=0) * np.sqrt(tr) * 15.0
    else:
        pet = 0.16 * (tm + 5.0).clip(lower=0)
    return pd.Series(pet).fillna(0.0).clip(lower=0)

def _best_index_resolution(index_type: str, season_label: str) -> tuple[str, str | None]:
    """Choose the lightest suitable file for each index."""
    daily_needed = {"hot_days", "hot_nights", "cold_days", "cold_nights", "heat_index"}
    if index_type in daily_needed:
        return "daily", None
    if season_label == "ANNUAL":
        return "annual", None
    if season_label in SEASON_DEFINITIONS:
        return "seasonal", season_label
    return "monthly", None


def _extract_index_series_best(
    data_dir: Path,
    dataset_key: str,
    index_type: str,
    season_label: str,
    lat: float,
    lon: float,
    start_year: int,
    end_year: int,
    variable: str = "auto",
    custom_months: list[int] | None = None,
):
    resolution, file_season = _best_index_resolution(index_type, season_label)
    df, ctx = extract_point_series(
        data_dir, dataset_key, resolution, lat, lon,
        f"{start_year}-01-01", f"{end_year}-12-31",
        variable=variable, season=file_season
    )
    df["time"] = pd.to_datetime(df["time"])
    df["year"] = df["time"].dt.year
    df["month"] = df["time"].dt.month
    if resolution == "monthly" and custom_months:
        df = df[df["month"].isin(custom_months)].copy()
    df["source_resolution_used"] = resolution
    df["source_file_season"] = file_season or ""
    return df, {**ctx, "source_resolution_used": resolution, "source_file_season": file_season or ""}


def _save_index_excel(summary: pd.DataFrame, metadata: pd.DataFrame, daily_used: pd.DataFrame | None, out_path: Path):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Index Summary", index=False)
        metadata.to_excel(writer, sheet_name="Metadata", index=False)
        if daily_used is not None and not daily_used.empty:
            daily_used.to_excel(writer, sheet_name="Data Used", index=False)

    wb = writer.book if False else None
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        thin = Side(style="thin", color="A7B7C7")
        header_fill = PatternFill("solid", fgColor="0B5E7A")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(vertical="top")
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 1, 9), 18)
        ws.freeze_panes = None
    wb.save(out_path)


def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate non-rainfall climate indices as Excel only.

    Rainfall indices are intentionally disabled from the Climate Indices page.
    File choice is automatic: daily for threshold-day indices, annual for annual
    mean/index products, seasonal for configured seasons, and monthly for custom seasons.
    """
    dirs = ensure_output_dirs(export_dir)
    index_type = params.get("index_type")
    if index_type in set(dict(RAINFALL_INDICES)) or str(index_type or "").lower().find("rain") >= 0:
        raise ValueError("Rainfall-related climate indices are disabled on this page. Use data extraction/products for rainfall data.")

    season_label, months = parse_custom_months(params.get("season") or "ANNUAL", params.get("custom_months"))
    start_year = int(params.get("start_year") or 1991)
    end_year = int(params.get("end_year") or start_year)
    lat = float(params.get("latitude") or 0)
    lon = float(params.get("longitude") or 0)
    location = params.get("location_name") or "Selected Location"
    heat_threshold = float(params.get("heat_threshold") or 35.0)
    warm_threshold = float(params.get("warm_threshold") or 30.0)
    cold_threshold = float(params.get("cold_threshold") or 15.0)
    wind_threshold = float(params.get("wind_threshold") or 10.0)
    baseline_start = int(params.get("baseline_start") or 1991)
    baseline_end = int(params.get("baseline_end") or 2020)

    temp_keys = set(dict(TEMPERATURE_INDICES))
    other_keys = set(dict(OTHER_INDICES))
    if index_type not in temp_keys and index_type not in other_keys:
        raise ValueError("Unsupported climate index selected.")

    ctx: Dict[str, Any] = {}
    data_used = None
    summary: pd.DataFrame
    selected_col = "index_value"

    def _standardize_by_baseline(frame: pd.DataFrame, value_col: str, out_col: str):
        base = frame[(frame["year"] >= baseline_start) & (frame["year"] <= baseline_end)]
        if base.empty:
            base = frame
        mean = pd.to_numeric(base[value_col], errors="coerce").mean()
        sd = pd.to_numeric(base[value_col], errors="coerce").std(ddof=0)
        frame[out_col] = (pd.to_numeric(frame[value_col], errors="coerce") - mean) / sd if sd and not np.isnan(sd) else np.nan
        return mean, sd

    if index_type in temp_keys:
        if index_type in {"hot_days", "hot_nights", "cold_days", "cold_nights", "heat_index"}:
            ta, ctx = _extract_index_series_best(data_dir, "era5_temperature", index_type, season_label, lat, lon, start_year, end_year, variable="ta", custom_months=months)
            tmin, _ = _extract_index_series_best(data_dir, "era5_temperature", index_type, season_label, lat, lon, start_year, end_year, variable="tmin", custom_months=months)
            tmax, _ = _extract_index_series_best(data_dir, "era5_temperature", index_type, season_label, lat, lon, start_year, end_year, variable="tmax", custom_months=months)
            df = ta.rename(columns={"value":"mean_temperature_c"}).merge(tmin[["time","value"]].rename(columns={"value":"minimum_temperature_c"}), on="time", how="left").merge(tmax[["time","value"]].rename(columns={"value":"maximum_temperature_c"}), on="time", how="left")
            df["year"] = df["time"].dt.year
            df["month"] = df["time"].dt.month
            if season_label != "ANNUAL" and months:
                df = df[df["month"].isin(months)].copy()
            if index_type == "heat_index":
                try:
                    rh, _ = _extract_index_series_best(data_dir, "era5_relative_humidity", index_type, season_label, lat, lon, start_year, end_year, variable="auto", custom_months=months)
                    df = df.merge(rh[["time","value"]].rename(columns={"value":"relative_humidity"}), on="time", how="left")
                except Exception:
                    df["relative_humidity"] = 60.0
                df["heat_index_c"] = _heat_index_celsius(df["mean_temperature_c"], df["relative_humidity"])
            rows=[]
            for year, grp in df.groupby("year"):
                row={"year": int(year), "season": season_label, "days_used": int(len(grp)), "source_resolution_used": "daily"}
                row["mean_temperature_c"] = float(pd.to_numeric(grp["mean_temperature_c"], errors="coerce").mean())
                row["maximum_temperature_c"] = float(pd.to_numeric(grp["maximum_temperature_c"], errors="coerce").max())
                row["minimum_temperature_c"] = float(pd.to_numeric(grp["minimum_temperature_c"], errors="coerce").min())
                row["hot_days"] = int((pd.to_numeric(grp["maximum_temperature_c"], errors="coerce") >= heat_threshold).sum())
                row["hot_nights"] = int((pd.to_numeric(grp["minimum_temperature_c"], errors="coerce") >= warm_threshold).sum())
                row["cold_days"] = int((pd.to_numeric(grp["maximum_temperature_c"], errors="coerce") <= cold_threshold).sum())
                row["cold_nights"] = int((pd.to_numeric(grp["minimum_temperature_c"], errors="coerce") <= cold_threshold).sum())
                if index_type == "heat_index":
                    row["mean_heat_index_c"] = float(pd.to_numeric(grp["heat_index_c"], errors="coerce").mean())
                    row["maximum_heat_index_c"] = float(pd.to_numeric(grp["heat_index_c"], errors="coerce").max())
                rows.append(row)
            summary = pd.DataFrame(rows).sort_values("year")
            selected_col = {"hot_days":"hot_days", "hot_nights":"hot_nights", "cold_days":"cold_days", "cold_nights":"cold_nights", "heat_index":"mean_heat_index_c"}.get(index_type, "mean_temperature_c")
            data_used = df
        else:
            var = {"mean_temperature":"ta", "maximum_temperature":"tmax", "minimum_temperature":"tmin", "temperature_anomaly":"ta", "dtr":"ta"}.get(index_type, "ta")
            df, ctx = _extract_index_series_best(data_dir, "era5_temperature", index_type, season_label, lat, lon, start_year, end_year, variable=var, custom_months=months)
            if index_type == "dtr":
                tmin, _ = _extract_index_series_best(data_dir, "era5_temperature", index_type, season_label, lat, lon, start_year, end_year, variable="tmin", custom_months=months)
                tmax, _ = _extract_index_series_best(data_dir, "era5_temperature", index_type, season_label, lat, lon, start_year, end_year, variable="tmax", custom_months=months)
                tmp = tmax[["time","year","value"]].rename(columns={"value":"maximum_temperature_c"}).merge(tmin[["time","value"]].rename(columns={"value":"minimum_temperature_c"}), on="time", how="left")
                tmp["diurnal_temperature_range_c"] = tmp["maximum_temperature_c"] - tmp["minimum_temperature_c"]
                summary = tmp.groupby("year", as_index=False)["diurnal_temperature_range_c"].mean()
                selected_col = "diurnal_temperature_range_c"
                data_used = tmp
            else:
                source_resolution = str(ctx.get("source_resolution_used", ""))
                agg = "mean"
                if index_type == "maximum_temperature":
                    agg = "max"
                elif index_type == "minimum_temperature":
                    agg = "min"
                grouped = df.groupby("year")["value"].agg(agg).reset_index()
                col = {"mean_temperature":"mean_temperature_c", "maximum_temperature":"maximum_temperature_c", "minimum_temperature":"minimum_temperature_c", "temperature_anomaly":"mean_temperature_c"}.get(index_type, "mean_temperature_c")
                summary = grouped.rename(columns={"value": col})
                summary["source_resolution_used"] = source_resolution
                selected_col = col
                if index_type == "temperature_anomaly":
                    base = summary[(summary["year"] >= baseline_start) & (summary["year"] <= baseline_end)]
                    if base.empty:
                        base = summary
                    normal = pd.to_numeric(base[col], errors="coerce").mean()
                    summary["temperature_anomaly_c"] = summary[col] - normal
                    selected_col = "temperature_anomaly_c"
                data_used = df
    elif index_type == "relative_humidity_index":
        df, ctx = _extract_index_series_best(data_dir, "era5_relative_humidity", index_type, season_label, lat, lon, start_year, end_year, variable="auto", custom_months=months)
        summary = df.groupby("year", as_index=False)["value"].mean().rename(columns={"value":"mean_relative_humidity_percent"})
        _standardize_by_baseline(summary, "mean_relative_humidity_percent", "relative_humidity_index")
        summary["source_resolution_used"] = ctx.get("source_resolution_used", "")
        selected_col = "relative_humidity_index"
        data_used = df
    elif index_type in {"soil_moisture_index", "soil_moisture_anomaly"}:
        df, ctx = _extract_index_series_best(data_dir, "era5_soil_water", index_type, season_label, lat, lon, start_year, end_year, variable="auto", custom_months=months)
        summary = df.groupby("year", as_index=False)["value"].mean().rename(columns={"value":"mean_soil_moisture"})
        base = summary[(summary["year"] >= baseline_start) & (summary["year"] <= baseline_end)]
        if base.empty:
            base = summary
        mean = pd.to_numeric(base["mean_soil_moisture"], errors="coerce").mean()
        sd = pd.to_numeric(base["mean_soil_moisture"], errors="coerce").std(ddof=0)
        summary["soil_moisture_anomaly"] = summary["mean_soil_moisture"] - mean
        summary["soil_moisture_index"] = summary["soil_moisture_anomaly"] / sd if sd and not np.isnan(sd) else np.nan
        summary["source_resolution_used"] = ctx.get("source_resolution_used", "")
        selected_col = "soil_moisture_anomaly" if index_type == "soil_moisture_anomaly" else "soil_moisture_index"
        data_used = df
    elif index_type in {"wind_speed_index", "windy_days"}:
        df, ctx = _extract_index_series_best(data_dir, "era5_wind", index_type, season_label, lat, lon, start_year, end_year, variable="wind_speed", custom_months=months)
        if index_type == "windy_days":
            if season_label != "ANNUAL" and months:
                df = df[df["month"].isin(months)].copy()
            rows=[]
            for year, grp in df.groupby("year"):
                v = pd.to_numeric(grp["value"], errors="coerce")
                rows.append({"year": int(year), "season": season_label, "days_used": int(len(grp)), "windy_days": int((v >= wind_threshold).sum()), "wind_threshold": wind_threshold, "source_resolution_used": "daily"})
            summary = pd.DataFrame(rows).sort_values("year")
            selected_col = "windy_days"
        else:
            summary = df.groupby("year", as_index=False)["value"].mean().rename(columns={"value":"mean_wind_speed"})
            _standardize_by_baseline(summary, "mean_wind_speed", "wind_speed_index")
            summary["source_resolution_used"] = ctx.get("source_resolution_used", "")
            selected_col = "wind_speed_index"
        data_used = df

    if summary.empty:
        raise ValueError("No data found for the selected index and period.")

    summary = summary[(summary["year"] >= start_year) & (summary["year"] <= end_year)].copy()
    summary.insert(1, "location", location)
    summary.insert(2, "latitude", lat)
    summary.insert(3, "longitude", lon)
    summary.insert(4, "index_type", dict(ALL_INDICES).get(index_type, index_type))
    summary.insert(5, "selected_index_column", selected_col)

    metadata_df = pd.DataFrame([{
        "index_type": dict(ALL_INDICES).get(index_type, index_type),
        "location": location,
        "latitude": lat,
        "longitude": lon,
        "nearest_latitude": ctx.get("nearest_latitude"),
        "nearest_longitude": ctx.get("nearest_longitude"),
        "data_file": ctx.get("file"),
        "source_resolution_used": ctx.get("source_resolution_used"),
        "source_file_season": ctx.get("source_file_season"),
        "season": season_label,
        "months": ",".join(map(str, months)),
        "start_year": start_year,
        "end_year": end_year,
        "baseline_start": baseline_start,
        "baseline_end": baseline_end,
        "note": "Rainfall-related indices are disabled. Excel output only; no plot is generated.",
    }])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = slugify(f"CDE_indices_{index_type}_{location}_{season_label}_{start_year}_{end_year}_{timestamp}")
    excel_path = dirs["indices"] / f"{stem}.xlsx"
    _save_index_excel(summary, metadata_df, data_used, excel_path)
    db_summary_path = _append_parquet_database("climate_indices", summary.assign(raw_index_type=index_type), stem)

    return {"excel_path": excel_path, "db_path": db_summary_path, "context": ctx, "rows": len(summary), "selected_column": selected_col}


def wind_speed_direction_series(data_dir: Path, dataset_key: str, resolution: str, lat: float, lon: float, start: str | None, end: str | None, season: str | None = None):
    file_path = find_file(data_dir, "era5_wind", resolution, season=season)
    assert file_path
    with open_data_store(file_path, decode_times=True) as ds:
        time_name = detect_time_coord(ds)
        lat_name, lon_name = detect_lat_lon(ds)
        vars_ = data_variables(ds)
        speed_var = None
        dir_var = None
        for v in vars_:
            t = (v + " " + str(ds[v].attrs.get("long_name", ""))).lower()
            if speed_var is None and any(x in t for x in ["speed", "si10", "wind_speed", "ws"]):
                speed_var = v
            if dir_var is None and any(x in t for x in ["direction", "wind_direction", "wd"]):
                dir_var = v
        if speed_var is None:
            speed_var = pick_variable(ds, "era5_wind", "wind_speed")
        if dir_var is None:
            # Try u/v components.
            u = next((v for v in vars_ if v.lower() in {"u10", "u", "u_component_of_wind"}), None)
            vv = next((v for v in vars_ if v.lower() in {"v10", "v", "v_component_of_wind"}), None)
            if u and vv:
                da_u = ds[u]
                da_v = ds[vv]
                if start or end:
                    da_u, resolved_u_time = slice_time_range(da_u, time_name, start, end)
                    da_v, resolved_v_time = slice_time_range(da_v, time_name, start, end)
                    time_name = resolved_u_time or resolved_v_time
                pu = da_u.sel({lat_name: lat, lon_name: lon}, method="nearest")
                pv = da_v.sel({lat_name: lat, lon_name: lon}, method="nearest")
                spd = np.sqrt(pu.values ** 2 + pv.values ** 2)
                direction = (270 - np.rad2deg(np.arctan2(pv.values, pu.values))) % 360
                return pd.Series(spd), pd.Series(direction), {"file": file_path.name, "storage_format": store_kind(file_path), "dataset_label": DATASETS["era5_wind"]["label"], "unit": ds[u].attrs.get("units", "")}
            # Fallback: no direction.
            dir_var = speed_var
        speed_da = ds[speed_var]
        dir_da = ds[dir_var]
        if start or end:
            speed_da, resolved_speed_time = slice_time_range(speed_da, time_name, start, end)
            dir_da, resolved_direction_time = slice_time_range(dir_da, time_name, start, end)
            time_name = resolved_speed_time or resolved_direction_time
        ps = speed_da.sel({lat_name: lat, lon_name: lon}, method="nearest")
        pd_ = dir_da.sel({lat_name: lat, lon_name: lon}, method="nearest")
        return pd.Series(ps.values.ravel()), pd.Series(pd_.values.ravel()), {"file": file_path.name, "storage_format": store_kind(file_path), "dataset_label": DATASETS["era5_wind"]["label"], "unit": speed_da.attrs.get("units", "")}


def calculate_cost_recovery(values: Dict[str, Any]) -> Dict[str, Any]:
    categories = {
        "undergraduate": {"label": "Diploma / Undergraduate Student", "n": 20},
        "postgraduate": {"label": "Masters / PhD Student", "n": 18},
        "government_local": {"label": "Government / Local Private Firm", "n": 6},
        "international": {"label": "International Firm / Foreigner", "n": 2},
        "student": {"label": "Diploma / Undergraduate Student", "n": 20},
        "researcher": {"label": "Masters / PhD Student", "n": 18},
        "government": {"label": "Government / Local Private Firm", "n": 6},
        "contractor": {"label": "Government / Local Private Firm", "n": 6},
        "private": {"label": "International Firm / Foreigner", "n": 2},
    }
    resolution_rules = {
        "hourly": {"label": "Hourly", "ms": 60.0, "d": 4.0, "m": 12.0},
        "daily": {"label": "Daily / Dekadal", "ms": 60.0, "d": 4.0, "m": 12.0},
        "dekadal": {"label": "Daily / Dekadal", "ms": 60.0, "d": 4.0, "m": 12.0},
        "monthly": {"label": "Monthly", "ms": 20.0, "d": 2.0, "m": 12.0},
        "seasonal": {"label": "Monthly", "ms": 20.0, "d": 2.0, "m": 12.0},
        "annual": {"label": "Annual", "ms": 10.0, "d": 1.0, "m": 1.0},
    }
    category = str(values.get("customer_category") or "government_local").strip().lower()
    if category not in categories:
        category = "government_local"
    resolution = str(values.get("temporal_resolution") or values.get("resolution") or "monthly").strip().lower()
    if resolution not in resolution_rules:
        resolution = "monthly"
    stations = max(1, int(float(values.get("stations") or 1)))
    parameters = max(1, int(float(values.get("parameters") or values.get("variables") or 1)))
    years = max(1, int(float(values.get("years") or 1)))
    usd_rate = float(values.get("usd_rate") or values.get("rate") or 2650)
    rule = resolution_rules[resolution]
    ms = float(rule["ms"])
    d = float(rule["d"])
    months = float(rule["m"])
    n = float(categories[category]["n"])
    workload = d * months * years
    ps_factor = parameters if resolution in {"daily", "dekadal", "hourly"} else (parameters + stations)
    sc_usd = ms + ((workload / n) * ps_factor)
    total = sc_usd * usd_rate
    return {
        "customer_category": category,
        "customer_category_label": categories[category]["label"],
        "temporal_resolution": resolution,
        "temporal_resolution_label": rule["label"],
        "stations": stations,
        "variables": parameters,
        "parameters": parameters,
        "periods": int(months),
        "months": int(months),
        "years": years,
        "products": 0,
        "ms": ms,
        "d": d,
        "n": n,
        "usd_rate": usd_rate,
        "workload": workload,
        "ps_factor": ps_factor,
        "sc_usd": round(sc_usd, 6),
        "unit_rate": usd_rate,
        "data_amount": total,
        "product_amount": 0,
        "subtotal": total,
        "discount_percent": 0,
        "discount_amount": 0,
        "total": total,
        "total_plain": (f"{total:.0f}" if abs(total - round(total)) < 0.005 else f"{total:.10f}".rstrip("0").rstrip(".")),
        "total_text": human_money(total),
        "rate_matrix": categories,
    }


# ============================================================
# PROPOSED COST-RECOVERY MODEL
# This remains separate from calculate_cost_recovery(), so the current CDE
# calculator and its historical records remain unchanged.
#
# The proposed model deliberately uses the current formula as its baseline and
# adds a modest, fixed 10% CDE cost-recovery adjustment. This makes the new fee
# simple to explain, simple to audit and consistently a little higher than the
# current fee before any optional professional service is added.
# ============================================================
PROPOSED_COST_RECOVERY_CATEGORIES = {
    "undergraduate": {
        "label": "Academic – Diploma / Undergraduate",
        "former_category": "student",
        "former_label": "Diploma / Undergraduate Student",
        "example": "Diploma or undergraduate student using data for a non-commercial academic project.",
    },
    "postgraduate": {
        "label": "Academic – Masters / PhD",
        "former_category": "researcher",
        "former_label": "Masters / PhD Student",
        "example": "Masters or PhD student using data for a thesis, dissertation or academic research.",
    },
    "government": {
        "label": "Government / Public Institution",
        "former_category": "government_local",
        "former_label": "Government / Local Private Firm",
        "example": "Ministry, local government authority, public university or government research institution.",
    },
    "local_private": {
        "label": "Local Private / Commercial Institution",
        "former_category": "government_local",
        "former_label": "Government / Local Private Firm",
        "example": "Tanzanian company, consultant, insurer, agricultural firm or energy company.",
    },
    "international": {
        "label": "International Institution / Foreign Applicant",
        "former_category": "private",
        "former_label": "International Firm / Foreigner",
        "example": "Foreign company, international consultant, overseas applicant or commercial contractor.",
    },
}

PROPOSED_COST_RECOVERY_RESOLUTIONS = {
    "hourly": {"label": "Hourly", "records_per_year": 8760},
    "daily": {"label": "Daily", "records_per_year": 365},
    "dekadal": {"label": "Dekadal (10-day)", "records_per_year": 36},
    "monthly": {"label": "Monthly", "records_per_year": 12},
    "annual": {"label": "Annual", "records_per_year": 1},
}

PROPOSED_COST_RECOVERY_DEFAULTS = {
    "usd_rate": 2650.0,
    "recovery_uplift_percent": 10.0,
}


def _proposed_non_negative_number(value: Any, default: float, *, minimum: float = 0.0) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        parsed = float(default)
    if not math.isfinite(parsed):
        parsed = float(default)
    return max(minimum, parsed)


def calculate_proposed_cost_recovery(values: Dict[str, Any]) -> Dict[str, Any]:
    """Calculate the simplified proposed CDE fee in Tanzanian shillings.

    Current fee = existing CDE formula using the same category, resolution,
                  station, parameter, year and exchange-rate inputs.
    Proposed fee = current fee + 10% CDE recovery adjustment
                   + optional additional professional services.

    The fixed 10% adjustment ensures an ordinary proposed fee is moderately
    higher than the current fee while keeping the difference predictable.
    """
    aliases = {
        "student": "undergraduate",
        "researcher": "postgraduate",
        "government_local": "government",
        "contractor": "local_private",
        "private": "local_private",
        "foreigner": "international",
    }
    category = str(values.get("customer_category") or "government").strip().lower()
    category = aliases.get(category, category)
    if category not in PROPOSED_COST_RECOVERY_CATEGORIES:
        category = "government"

    resolution = str(values.get("temporal_resolution") or values.get("resolution") or "monthly").strip().lower()
    if resolution == "seasonal":
        resolution = "monthly"
    if resolution not in PROPOSED_COST_RECOVERY_RESOLUTIONS:
        resolution = "monthly"

    stations = int(_proposed_non_negative_number(values.get("stations"), 1, minimum=1))
    parameters = int(_proposed_non_negative_number(values.get("parameters") or values.get("variables"), 1, minimum=1))
    years = int(_proposed_non_negative_number(values.get("years"), 1, minimum=1))
    usd_rate = _proposed_non_negative_number(
        values.get("usd_rate"), PROPOSED_COST_RECOVERY_DEFAULTS["usd_rate"], minimum=1
    )
    additional_service_fee = _proposed_non_negative_number(values.get("additional_service_fee"), 0.0)
    uplift_percent = float(PROPOSED_COST_RECOVERY_DEFAULTS["recovery_uplift_percent"])

    category_rule = PROPOSED_COST_RECOVERY_CATEGORIES[category]
    resolution_rule = PROPOSED_COST_RECOVERY_RESOLUTIONS[resolution]
    records_per_year = int(resolution_rule["records_per_year"])
    data_values = stations * parameters * years * records_per_year

    current_values = {
        "customer_category": category_rule["former_category"],
        "temporal_resolution": resolution,
        "stations": stations,
        "parameters": parameters,
        "years": years,
        "usd_rate": usd_rate,
    }
    current_result = calculate_cost_recovery(current_values)
    current_fee = round(float(current_result.get("total") or 0.0))
    recovery_uplift_amount = round(current_fee * uplift_percent / 100.0)
    proposed_before_services = current_fee + recovery_uplift_amount
    total = round(proposed_before_services + additional_service_fee)
    difference_amount = total - current_fee
    difference_percent = (difference_amount / current_fee * 100.0) if current_fee else 0.0

    return {
        "customer_category": category,
        "customer_category_label": category_rule["label"],
        "customer_example": category_rule["example"],
        "former_customer_category": category_rule["former_category"],
        "former_customer_category_label": category_rule["former_label"],
        "temporal_resolution": resolution,
        "temporal_resolution_label": resolution_rule["label"],
        "records_per_year": records_per_year,
        "stations": stations,
        "variables": parameters,
        "parameters": parameters,
        "years": years,
        "data_values": data_values,
        "usd_rate": usd_rate,
        "current_fee": current_fee,
        "current_fee_text": f"TZS {current_fee:,.0f}",
        "recovery_uplift_percent": uplift_percent,
        "recovery_uplift_amount": recovery_uplift_amount,
        "recovery_uplift_text": f"TZS {recovery_uplift_amount:,.0f}",
        "proposed_before_services": proposed_before_services,
        "additional_service_fee": additional_service_fee,
        "difference_amount": difference_amount,
        "difference_text": f"TZS {difference_amount:,.0f}",
        "difference_percent": difference_percent,
        "total": total,
        "total_plain": f"{total:.0f}",
        "total_text": f"TZS {total:,.0f}",
        "formula_text": "Current formula fee + 10% CDE cost-recovery adjustment + optional additional services",
        "comparison_note": "Without additional services, the proposed fee is 10% higher than the current fee.",
        # Compatibility fields retained for older stored-record readers.
        "customer_factor": 1.0 + uplift_percent / 100.0,
        "minimum_fee": 0.0,
        "minimum_fee_text": "Not applicable",
        "minimum_applied": False,
        "access_fee": 0.0,
        "processing_fee": recovery_uplift_amount,
        "volume_block_size": 1000,
        "volume_blocks": int(math.ceil(data_values / 1000)) if data_values else 0,
        "volume_rate": 0.0,
        "volume_fee": 0.0,
        "standard_cost": current_fee,
        "adjusted_cost": proposed_before_services,
        "data_amount": current_fee,
        "product_amount": additional_service_fee,
        "subtotal": proposed_before_services,
        "discount_percent": 0.0,
        "discount_amount": 0.0,
        "rate_matrix": PROPOSED_COST_RECOVERY_CATEGORIES,
    }


# ============================================================
# USER REQUEST OVERRIDES - clear dataset sources, no map borders/lakes/ocean,
# climate indices with CSV + plots, rainfall indices CSV only.
# ============================================================

# Clear source labels for Plots and Products. Keep source names visible.
DATASETS["chirps_rainfall"]["label"] = "CHIRPS Precipitation"
DATASETS["era5_total_precipitation"]["label"] = "ERA5 Precipitation"
DATASETS["era5_temperature"]["label"] = "ERA5 Temperature Mean, Min and Max"
DATASETS["era5_dew_point"]["label"] = "ERA5 Dew Point Temperature 2m"
DATASETS["era5_relative_humidity"]["label"] = "ERA5 Relative Humidity"
DATASETS["era5_skin_temperature"]["label"] = "ERA5 Skin Temperature"
DATASETS["era5_soil_temperature"]["label"] = "ERA5-Land Soil Temperature Level 1"
DATASETS["era5_pressure_cloud"]["label"] = "ERA5 Surface Pressure and Total Cloud Cover"
DATASETS["era5_soil_water"]["label"] = "ERA5-Land Volumetric Soil Moisture"
DATASETS["era5_wind"]["label"] = "ERA5 Wind Speed and Direction 10m"

# Spatial maps should show only the gridded values; no borders, oceans or lakes.
def decorate_tanzania_map(ax):
    return ax

# Restore rainfall-related indices on the Climate Indices page. Rainfall indices
# are generated as CSV only; non-rainfall indices generate plot + CSV + Excel.
RAINFALL_INDICES = [
    ("total_rainfall", "Total Rainfall"),
    ("number_wet_days", "Number of Wet Days"),
    ("number_dry_days", "Number of Dry Days"),
    ("consecutive_dry_days", "Consecutive Dry Days"),
    ("consecutive_wet_days", "Consecutive Wet Days"),
    ("wet_spell_length", "Wet Spell Length"),
    ("dry_spell_length", "Dry Spell Length"),
    ("rainy_season_onset", "Rainy Season Onset"),
    ("rainy_season_cessation", "Rainy Season Cessation"),
    ("length_of_rainy_season", "Length of Rainy Season"),
    ("max_1day_rainfall", "Maximum 1-Day Rainfall"),
    ("max_5day_rainfall", "Maximum 5-Day Rainfall"),
    ("heavy_rainfall_days", "Heavy Rainfall Days"),
    ("very_heavy_rainfall_days", "Very Heavy Rainfall Days"),
    ("sdii", "Simple Daily Intensity Index"),
    ("r95p", "R95p Very Wet Days"),
    ("r99p", "R99p Extremely Wet Days"),
    ("rainfall_anomaly", "Rainfall Anomaly"),
    ("spi", "Standardized Precipitation Index"),
    ("spei", "Standardized Precipitation Evapotranspiration Index"),
]
OTHER_INDICES = [
    ("relative_humidity_index", "Relative Humidity Index"),
    ("soil_moisture_index", "Soil Moisture Index"),
    ("soil_moisture_anomaly", "Soil Moisture Anomaly"),
    ("wind_speed_index", "Wind Speed Index"),
]
ALL_INDICES = RAINFALL_INDICES + TEMPERATURE_INDICES + OTHER_INDICES

# Save the previous best-resolution non-rainfall Excel generator.
_generate_indices_best_excel_only = generate_indices


def _best_rainfall_resolution(index_type: str, season_label: str) -> tuple[str, str | None]:
    daily_needed = {
        "number_wet_days", "number_dry_days", "consecutive_dry_days", "consecutive_wet_days",
        "wet_spell_length", "dry_spell_length", "rainy_season_onset", "rainy_season_cessation",
        "length_of_rainy_season", "max_1day_rainfall", "max_5day_rainfall", "heavy_rainfall_days",
        "very_heavy_rainfall_days", "sdii", "r95p", "r99p",
    }
    if index_type in daily_needed:
        return "daily", None
    if season_label == "ANNUAL":
        return "annual", None
    if season_label in SEASON_DEFINITIONS:
        return "seasonal", season_label
    return "monthly", None


def _rainfall_summary_from_aggregated(df: pd.DataFrame, index_type: str, season_label: str,
                                      start_year: int, end_year: int,
                                      baseline_start: int, baseline_end: int,
                                      months: list[int]) -> tuple[pd.DataFrame, str]:
    tmp = df.copy()
    tmp["time"] = pd.to_datetime(tmp["time"])
    tmp["year"] = tmp["time"].dt.year
    tmp["month"] = tmp["time"].dt.month
    if season_label == "CUSTOM":
        tmp = tmp[tmp["month"].isin(months)].copy()
        summary = tmp.groupby("year", as_index=False)["value"].sum().rename(columns={"value": "rainfall_total_mm"})
    else:
        summary = tmp.groupby("year", as_index=False)["value"].sum().rename(columns={"value": "rainfall_total_mm"})
    summary = summary[(summary["year"] >= start_year) & (summary["year"] <= end_year)].copy()
    base = summary[(summary["year"] >= baseline_start) & (summary["year"] <= baseline_end)]
    if base.empty:
        base = summary
    normal = pd.to_numeric(base["rainfall_total_mm"], errors="coerce").mean()
    sd = pd.to_numeric(base["rainfall_total_mm"], errors="coerce").std(ddof=0)
    summary["rainfall_anomaly_mm"] = summary["rainfall_total_mm"] - normal
    summary["percentage_of_normal"] = (summary["rainfall_total_mm"] / normal) * 100 if normal else np.nan
    summary["spi_normal_approx"] = (summary["rainfall_total_mm"] - normal) / sd if sd and not np.isnan(sd) else np.nan
    # SPEI approximation is left as normal approximation when PET is not part of the selected source file.
    summary["spei_normal_approx"] = summary["spi_normal_approx"]
    selected_col = {
        "total_rainfall": "rainfall_total_mm",
        "rainfall_anomaly": "rainfall_anomaly_mm",
        "spi": "spi_normal_approx",
        "spei": "spei_normal_approx",
    }.get(index_type, "rainfall_total_mm")
    return summary, selected_col


def _generate_rainfall_index_csv_only(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    dirs = ensure_output_dirs(export_dir)
    index_type = params.get("index_type")
    rainfall_keys = set(dict(RAINFALL_INDICES))
    if index_type not in rainfall_keys:
        raise ValueError("Unsupported rainfall index selected.")
    dataset_key = params.get("dataset") if params.get("dataset") in {"chirps_rainfall", "era5_total_precipitation"} else "chirps_rainfall"
    season_label, months = parse_custom_months(params.get("season") or "ANNUAL", params.get("custom_months"))
    start_year = int(params.get("start_year") or 1991)
    end_year = int(params.get("end_year") or start_year)
    lat = float(params.get("latitude") or 0)
    lon = float(params.get("longitude") or 0)
    location = params.get("location_name") or "Selected Location"
    rainy_threshold = float(params.get("rainy_threshold") or 1.0)
    heavy_threshold = float(params.get("heavy_threshold") or 50.0)
    very_heavy_threshold = float(params.get("very_heavy_threshold") or 100.0)
    baseline_start = int(params.get("baseline_start") or 1991)
    baseline_end = int(params.get("baseline_end") or 2020)

    resolution, file_season = _best_rainfall_resolution(index_type, season_label)
    # Use aggregate files for total/anomaly/SPI-like rainfall totals. Use daily for event/threshold indices.
    if resolution in {"annual", "seasonal", "monthly"}:
        df, ctx = extract_point_series(data_dir, dataset_key, resolution, lat, lon,
                                       f"{start_year}-01-01", f"{end_year}-12-31",
                                       variable="auto", season=file_season)
        df["source_resolution_used"] = resolution
        df["source_file_season"] = file_season or ""
        summary, selected_col = _rainfall_summary_from_aggregated(df, index_type, season_label, start_year, end_year, baseline_start, baseline_end, months)
        data_used = df.copy()
    else:
        df, ctx = extract_point_series(data_dir, dataset_key, "daily", lat, lon,
                                       f"{start_year-1}-01-01", f"{end_year}-12-31",
                                       variable="auto")
        df = df.rename(columns={"value": "rainfall_mm"})
        df["time"] = pd.to_datetime(df["time"])
        df["year"] = df["time"].dt.year
        df["month"] = df["time"].dt.month
        sdf = df[df["month"].isin(months)].copy()
        sdf["season_year"] = sdf["time"].apply(lambda x: season_year(pd.Timestamp(x), months))
        sdf = sdf[(sdf["season_year"] >= start_year) & (sdf["season_year"] <= end_year)].copy()
        if sdf.empty:
            raise ValueError("No daily rainfall data found for the selected season/year range.")
        baseline_daily = sdf[(sdf["season_year"] >= baseline_start) & (sdf["season_year"] <= baseline_end)].copy()
        if baseline_daily.empty:
            baseline_daily = sdf.copy()
        wet_baseline = pd.to_numeric(baseline_daily["rainfall_mm"], errors="coerce").fillna(0)
        wet_baseline = wet_baseline[wet_baseline >= rainy_threshold]
        r95_threshold = float(wet_baseline.quantile(0.95)) if len(wet_baseline) else np.nan
        r99_threshold = float(wet_baseline.quantile(0.99)) if len(wet_baseline) else np.nan
        rows = []
        for sy, grp in sdf.groupby("season_year"):
            grp = grp.sort_values("time")
            r = pd.to_numeric(grp["rainfall_mm"], errors="coerce").fillna(0)
            rainy = r >= rainy_threshold
            dry = r < rainy_threshold
            onset, cessation, season_len = _tma_rainy_season_dates(grp, rainy_threshold=rainy_threshold)
            rows.append({
                "year": int(sy),
                "season": season_label,
                "days_used": int(len(grp)),
                "rainfall_total_mm": float(r.sum()),
                "number_wet_days": int(rainy.sum()),
                "number_dry_days": int(dry.sum()),
                "consecutive_dry_days": consecutive_count(dry),
                "consecutive_wet_days": consecutive_count(rainy),
                "wet_spell_length_days": spell_max_length(rainy),
                "dry_spell_length_days": spell_max_length(dry),
                "rainy_season_onset": onset.strftime("%Y-%m-%d") if onset is not None else "",
                "rainy_season_onset_doy": int(onset.dayofyear) if onset is not None else np.nan,
                "rainy_season_cessation": cessation.strftime("%Y-%m-%d") if cessation is not None else "",
                "rainy_season_cessation_doy": int(cessation.dayofyear) if cessation is not None else np.nan,
                "length_of_rainy_season_days": season_len if season_len is not None else np.nan,
                "max_1day_rainfall_mm": float(r.max()),
                "max_5day_rainfall_mm": float(r.rolling(5, min_periods=1).sum().max()),
                "heavy_rainfall_days": int((r >= heavy_threshold).sum()),
                "very_heavy_rainfall_days": int((r >= very_heavy_threshold).sum()),
                "sdii_mm_per_wet_day": float(r[rainy].sum() / max(1, rainy.sum())),
                "r95p_mm": float(r[r > r95_threshold].sum()) if not np.isnan(r95_threshold) else np.nan,
                "r99p_mm": float(r[r > r99_threshold].sum()) if not np.isnan(r99_threshold) else np.nan,
                "source_resolution_used": "daily",
                "source_file_season": "",
            })
        summary = pd.DataFrame(rows).sort_values("year")
        base = summary[(summary["year"] >= baseline_start) & (summary["year"] <= baseline_end)]
        if base.empty:
            base = summary
        normal = pd.to_numeric(base["rainfall_total_mm"], errors="coerce").mean()
        sd = pd.to_numeric(base["rainfall_total_mm"], errors="coerce").std(ddof=0)
        summary["rainfall_anomaly_mm"] = summary["rainfall_total_mm"] - normal
        summary["percentage_of_normal"] = (summary["rainfall_total_mm"] / normal) * 100 if normal else np.nan
        summary["spi_normal_approx"] = (summary["rainfall_total_mm"] - normal) / sd if sd and not np.isnan(sd) else np.nan
        summary["spei_normal_approx"] = summary["spi_normal_approx"]
        selected_col = {
            "total_rainfall": "rainfall_total_mm",
            "number_wet_days": "number_wet_days",
            "number_dry_days": "number_dry_days",
            "consecutive_dry_days": "consecutive_dry_days",
            "consecutive_wet_days": "consecutive_wet_days",
            "wet_spell_length": "wet_spell_length_days",
            "dry_spell_length": "dry_spell_length_days",
            "rainy_season_onset": "rainy_season_onset_doy",
            "rainy_season_cessation": "rainy_season_cessation_doy",
            "length_of_rainy_season": "length_of_rainy_season_days",
            "max_1day_rainfall": "max_1day_rainfall_mm",
            "max_5day_rainfall": "max_5day_rainfall_mm",
            "heavy_rainfall_days": "heavy_rainfall_days",
            "very_heavy_rainfall_days": "very_heavy_rainfall_days",
            "sdii": "sdii_mm_per_wet_day",
            "r95p": "r95p_mm",
            "r99p": "r99p_mm",
            "rainfall_anomaly": "rainfall_anomaly_mm",
            "spi": "spi_normal_approx",
            "spei": "spei_normal_approx",
        }.get(index_type, "rainfall_total_mm")
        data_used = sdf.copy()

    if summary.empty:
        raise ValueError("No rainfall index data found for the selected options.")
    summary.insert(1, "location", location)
    summary.insert(2, "latitude", lat)
    summary.insert(3, "longitude", lon)
    summary.insert(4, "index_type", dict(RAINFALL_INDICES).get(index_type, index_type))
    summary.insert(5, "selected_index_column", selected_col)
    summary["rainfall_dataset"] = DATASETS[dataset_key]["label"]
    summary["nearest_latitude"] = ctx.get("nearest_latitude")
    summary["nearest_longitude"] = ctx.get("nearest_longitude")
    summary["data_file"] = ctx.get("file")
    summary["source_resolution_used"] = summary.get("source_resolution_used", resolution)
    summary["source_file_season"] = summary.get("source_file_season", file_season or "")
    summary["note"] = "Rainfall-related climate indices are CSV-download only; no plot is generated."

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = slugify(f"CDE_rainfall_index_{index_type}_{location}_{season_label}_{start_year}_{end_year}_{timestamp}")
    csv_path = dirs["indices"] / f"{stem}.csv"
    climate_index_csv_dataframe(summary, selected_col, index_type).to_csv(csv_path, index=False)
    data_used_path = dirs["indices"] / f"{stem}_data_used.csv"
    try:
        _compact_export_dataframe(data_used).to_csv(data_used_path, index=False)
    except Exception:
        data_used_path = None
    db_summary_path = _append_parquet_database("climate_indices", summary.assign(raw_index_type=index_type), stem)
    return {
        "csv_path": csv_path,
        "data_used_csv_path": data_used_path,
        "db_path": db_summary_path,
        "context": {**ctx, "dataset_label": DATASETS[dataset_key]["label"], "source_resolution_used": resolution, "source_file_season": file_season or ""},
        "rows": len(summary),
        "selected_column": selected_col,
        "rainfall_csv_only": True,
    }



# ============================================================
# FORCE FIX - Climate index headings, no old "for Location (years)" style
# ============================================================
def _cde_force_year_range(start_year, end_year):
    try:
        sy, ey = int(start_year), int(end_year)
    except Exception:
        return f"{start_year}–{end_year}"
    return str(sy) if sy == ey else f"{sy}–{ey}"


def _cde_force_season_period(season_label, start_year, end_year):
    season = str(season_label or "ANNUAL").upper().strip()
    try:
        sy, ey = int(start_year), int(end_year)
    except Exception:
        return f"{start_year}–{end_year}"
    cross_year = {"DJF", "NDJ", "DJFMA", "NDJFMA"}
    if season in cross_year:
        if sy == ey:
            return f"{sy-1}/{sy} Season"
        return f"{sy-1}/{sy}–{ey-1}/{ey} Seasons"
    return _cde_force_year_range(sy, ey)


def _cde_force_index_heading(index_type, location, season_label=None, start_year=1991, end_year=2020, baseline_start=1991, baseline_end=2020, custom_months=None):
    labels = dict(ALL_INDICES)
    index_name = labels.get(index_type, str(index_type or "Climate Index").replace("_", " ").title())
    area = str(location or "Selected Area").strip() or "Selected Area"
    season = str(season_label or "ANNUAL").upper().strip()
    month_names = {1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"}
    if season in {"", "ANNUAL", "WHOLE YEAR", "ANNUAL / WHOLE YEAR"}:
        prefix = "Annual"
        period = _cde_force_year_range(start_year, end_year)
    elif season.startswith("CUSTOM"):
        months = []
        for m in (custom_months or []):
            try:
                mi = int(m)
                if mi in month_names:
                    months.append(mi)
            except Exception:
                pass
        prefix = month_names[months[0]] if len(months) == 1 else "–".join(month_names[m] for m in months) if months else "Monthly"
        period = _cde_force_year_range(start_year, end_year)
    elif season in SEASON_DEFINITIONS:
        prefix = f"{season} Seasonal"
        period = _cde_force_season_period(season, start_year, end_year)
    else:
        prefix = season.title()
        period = _cde_force_year_range(start_year, end_year)
    title = f"{prefix} {index_name} over {area}, {period}"
    if str(index_type) in {"rainfall_anomaly", "temperature_anomaly", "spi", "spei", "relative_humidity_index", "soil_moisture_index", "soil_moisture_anomaly", "wind_speed_index", "r95p", "r99p"}:
        title += f" Relative to {int(baseline_start)}–{int(baseline_end)} Baseline"
    return title


def _cde_force_season_from_summary(summary, default="ANNUAL"):
    try:
        if "season" in summary.columns and len(summary["season"].dropna()):
            return str(summary["season"].dropna().iloc[0])
    except Exception:
        pass
    return default

def _plot_non_rainfall_index(summary: pd.DataFrame, selected_col: str, index_type: str, location: str,
                             start_year: int, end_year: int, out_path: Path) -> Path:
    fig, ax = plt.subplots(figsize=(10.5, 5.8))
    values = pd.to_numeric(summary[selected_col], errors="coerce")
    labels = dict(ALL_INDICES)
    if "year" in summary.columns:
        x = pd.to_numeric(summary["year"], errors="coerce")
    else:
        x = np.arange(len(summary)) + 1
    if "anomaly" in selected_col or "index" in selected_col or "standardized" in selected_col:
        colors = np.where(values >= 0, '#d73027', '#4575b4')
        ax.bar(x, values, color=colors)
        ax.axhline(0, linewidth=1, color='black')
    elif "days" in selected_col or index_type in {"hot_days", "hot_nights", "cold_days", "cold_nights"}:
        ax.bar(x, values, color=plot_color_for(index_type=index_type))
    else:
        ax.plot(x, values, marker="o", color=plot_color_for(index_type=index_type))
    ax.set_title(_cde_force_index_heading(index_type, location, _cde_force_season_from_summary(summary), start_year, end_year), fontsize=13)
    ax.set_xlabel("Year")
    unit_map = {
        "mean_temperature": "°C", "maximum_temperature": "°C", "minimum_temperature": "°C",
        "temperature_anomaly": "°C", "hot_days": "days", "hot_nights": "days",
        "cold_days": "days", "cold_nights": "days", "heat_index": "°C", "dtr": "°C",
        "relative_humidity_index": "index", "soil_moisture_index": "index",
        "soil_moisture_anomaly": "m³/m³", "wind_speed_index": "index",
    }
    ax.set_ylabel(y_axis_label(labels.get(index_type, selected_col.replace("_", " ").title()), unit_map.get(index_type, "")))
    apply_plot_grids(ax)
    return _finalize_plot(fig, out_path)


def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate climate indices.

    Rainfall indices: CSV only, no plot.
    Temperature/other indices: Excel + CSV + plot.
    """
    index_type = params.get("index_type")
    rainfall_keys = set(dict(RAINFALL_INDICES))
    if index_type in rainfall_keys:
        return _generate_rainfall_index_csv_only(params, data_dir, export_dir)

    # Use the best-resolution Excel generator already present in this project.
    result = _generate_indices_best_excel_only(params, data_dir, export_dir)
    excel_path = Path(result["excel_path"])
    try:
        summary = pd.read_excel(excel_path, sheet_name="Index Summary")
    except Exception:
        summary = pd.read_excel(excel_path)
    selected_col = result.get("selected_column")
    if not selected_col or selected_col not in summary.columns:
        numeric_cols = [c for c in summary.columns if c not in {"year", "latitude", "longitude"} and pd.api.types.is_numeric_dtype(summary[c])]
        selected_col = numeric_cols[-1] if numeric_cols else summary.columns[-1]
    dirs = ensure_output_dirs(export_dir)
    start_year = int(params.get("start_year") or (summary["year"].min() if "year" in summary.columns else 0))
    end_year = int(params.get("end_year") or (summary["year"].max() if "year" in summary.columns else start_year))
    location = params.get("location_name") or "Selected Location"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = slugify(f"CDE_index_{index_type}_{location}_{start_year}_{end_year}_{timestamp}")
    csv_path = dirs["indices"] / f"{stem}.csv"
    climate_index_csv_dataframe(summary, selected_col, index_type).to_csv(csv_path, index=False)
    plot_path = _plot_non_rainfall_index(summary, selected_col, index_type, location, start_year, end_year, dirs["indices"] / f"{stem}.png")
    return {**result, "csv_path": csv_path, "plot_path": plot_path, "rows": len(summary), "selected_column": selected_col, "rainfall_csv_only": False}

# ============================================================
# FINAL USER REQUEST OVERRIDES - July 2026
# ============================================================

# Official clear names for ERA5-Land datasets across the whole system.
DATASETS["era5_soil_temperature"]["label"] = "ERA5-Land Soil Temperature Level 1"
DATASETS["era5_soil_water"]["label"] = "ERA5-Land Volumetric Soil Moisture"


def climate_index_csv_dataframe(summary: pd.DataFrame, selected_col: str | None = None, index_type: str | None = None) -> pd.DataFrame:
    """Standard CDE climate-index data layout.

    Output columns follow: Location, Coordinates, Date, Year, Month, Day, Hour,
    Season where relevant, and Data. This keeps CSV/Excel products compact and
    consistent across hourly, daily, monthly and annual/seasonal index outputs.
    """
    out = summary.copy()
    if selected_col not in out.columns:
        numeric_cols = [c for c in out.columns if pd.api.types.is_numeric_dtype(out[c]) and c not in {"latitude", "longitude", "year", "month", "day", "hour"}]
        selected_col = numeric_cols[-1] if numeric_cols else (out.columns[-1] if len(out.columns) else None)

    location = out["location"] if "location" in out.columns else "Selected Location"
    lat = out["latitude"] if "latitude" in out.columns else out.get("nearest_latitude", "")
    lon = out["longitude"] if "longitude" in out.columns else out.get("nearest_longitude", "")
    coords = [f"{a}, {b}" if a not in (None, "") and b not in (None, "") else "" for a, b in zip(list(lat) if hasattr(lat, '__iter__') and not isinstance(lat, str) else [lat]*len(out), list(lon) if hasattr(lon, '__iter__') and not isinstance(lon, str) else [lon]*len(out))]

    # Build temporal fields from available columns.
    if "time" in out.columns:
        t = pd.to_datetime(out["time"], errors="coerce")
    elif "date" in out.columns:
        t = pd.to_datetime(out["date"], errors="coerce")
    elif "year" in out.columns:
        month_vals = pd.to_numeric(out["month"], errors="coerce").fillna(1).astype(int) if "month" in out.columns else 1
        day_vals = pd.to_numeric(out["day"], errors="coerce").fillna(1).astype(int) if "day" in out.columns else 1
        t = pd.to_datetime(dict(year=pd.to_numeric(out["year"], errors="coerce").fillna(1900).astype(int), month=month_vals, day=day_vals), errors="coerce")
    else:
        t = pd.Series([pd.NaT] * len(out))

    result = pd.DataFrame({
        "Location": location,
        "Coordinates": coords,
        "Date": pd.to_datetime(t, errors="coerce").dt.strftime("%Y-%m-%d"),
        "Year": pd.to_datetime(t, errors="coerce").dt.year,
        "Month": pd.to_datetime(t, errors="coerce").dt.month,
        "Day": pd.to_datetime(t, errors="coerce").dt.day,
        "Hour": pd.to_datetime(t, errors="coerce").dt.hour,
    })
    if "season" in out.columns:
        result["Season"] = out["season"]
    if "index_type" in out.columns:
        result["Index"] = out["index_type"]
    elif index_type:
        result["Index"] = dict(ALL_INDICES).get(index_type, str(index_type).replace("_", " ").title())
    result["Data"] = pd.to_numeric(out[selected_col], errors="coerce") if selected_col else np.nan
    # Remove temporal fields that are not relevant to the selected resolution/source summary.
    original_cols = {str(c).lower() for c in out.columns}
    if "hour" not in original_cols and not ("time" in original_cols and (pd.to_datetime(t, errors="coerce").dt.hour.fillna(0) != 0).any()):
        result.drop(columns=["Hour"], inplace=True, errors="ignore")
    if "day" not in original_cols and "hour" not in original_cols and "time" not in original_cols:
        result.drop(columns=["Day"], inplace=True, errors="ignore")
    if "month" not in original_cols and "day" not in original_cols and "hour" not in original_cols and "time" not in original_cols:
        result.drop(columns=["Month"], inplace=True, errors="ignore")
    return _round_output_dataframe(result)


_generate_indices_before_final_override = generate_indices


def _plot_rainfall_index(summary: pd.DataFrame, selected_col: str, index_type: str, location: str, start_year: int, end_year: int, out_path: Path) -> Path:
    labels = dict(ALL_INDICES)
    fig, ax = plt.subplots(figsize=(10.5, 5.8))
    values = pd.to_numeric(summary[selected_col], errors="coerce")
    x = pd.to_numeric(summary["year"], errors="coerce") if "year" in summary.columns else np.arange(len(summary)) + 1
    if "anomaly" in selected_col or index_type in {"rainfall_anomaly", "spi", "spei"}:
        colors = np.where(values >= 0, "#1f78b4", "#d73027")
        ax.bar(x, values, color=colors)
        ax.axhline(0, color="black", linewidth=1)
    else:
        ax.bar(x.astype(str), values, color="#1f78b4")
    ax.set_title(_cde_force_index_heading(index_type, location, _cde_force_season_from_summary(summary), start_year, end_year), fontsize=13)
    ax.set_xlabel("Year")
    unit = "mm" if any(k in selected_col for k in ["rainfall", "r95", "r99", "sdii"]) else ("days" if "days" in selected_col or "length" in selected_col else "")
    ax.set_ylabel(y_axis_label(labels.get(index_type, selected_col.replace("_", " ").title()), unit))
    apply_plot_grids(ax)
    return _finalize_plot(fig, out_path)


def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate climate indices with compact CSV, visible rainfall plots and two-year rainfall limit."""
    index_type = params.get("index_type")
    rainfall_keys = set(dict(RAINFALL_INDICES))
    start_year = int(params.get("start_year") or 1991)
    end_year = int(params.get("end_year") or start_year)

    result = _generate_indices_before_final_override(params, data_dir, export_dir)

    # Rainfall helper above used to create CSV only; now also create a plot for the selected index.
    if index_type in rainfall_keys:
        csv_path = Path(result.get("csv_path")) if result.get("csv_path") else None
        selected_col = result.get("selected_column")
        try:
            # Use detailed DB summary where possible, otherwise use the compact CSV.
            db_path = result.get("db_path")
            if db_path:
                summary = pd.read_parquet(db_path)
            elif csv_path and csv_path.exists():
                summary = pd.read_csv(csv_path)
            else:
                summary = pd.DataFrame()
            if selected_col not in summary.columns:
                numeric_cols = [c for c in summary.columns if pd.api.types.is_numeric_dtype(summary[c]) and c not in {"latitude", "longitude", "year"}]
                selected_col = numeric_cols[-1] if numeric_cols else None
            if selected_col:
                dirs = ensure_output_dirs(export_dir)
                location = params.get("location_name") or "Selected Location"
                stem = slugify(f"CDE_rainfall_index_plot_{index_type}_{location}_{start_year}_{end_year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
                result["plot_path"] = _plot_rainfall_index(summary, selected_col, index_type, location, start_year, end_year, dirs["indices"] / f"{stem}.png")
                result["rainfall_csv_only"] = False
        except Exception:
            pass
    return result


# ============================================================
# FINAL HOTFIX - Delivery report, rainfall indices, pressure/cloud and titles
# ============================================================

# Keep plot titles clean: do not inject source names into titles.
def element_with_source(dataset_key: str | None, element: str) -> str:
    return clean_source_words(str(element or "Weather Element")).replace("Rainfall", "Precipitation")

# Make sure pressure/cloud products are always averaged/taken as-is, never summed.
if "era5_pressure_cloud" in DATASETS:
    DATASETS["era5_pressure_cloud"]["family"] = "pressure_cloud"
    DATASETS["era5_pressure_cloud"]["unit"] = "mixed"

# Rainfall climate indices are selectable, but CSV-only: no plot is created.
def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    index_type = params.get("index_type")
    rainfall_keys = set(dict(RAINFALL_INDICES))
    result = _generate_indices_before_final_override(params, data_dir, export_dir)
    if index_type in rainfall_keys:
        result.pop("plot_path", None)
        result["rainfall_csv_only"] = True
        return result
    return result


# ============================================================
# FINAL HOTFIX - 2026-07-07 requested output/pressure/parquet behaviour
# ============================================================

def _pressure_is_hpa_variable(var_name: str) -> bool:
    name_l = str(var_name or "").lower()
    return "pressure" in name_l or name_l in {"sp", "msl", "surface_pressure", "mean_sea_level_pressure"}

# Keep rainfall indices CSV-only: no plot output is returned.
_generate_indices_before_no_rain_plot = generate_indices
def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    result = _generate_indices_before_no_rain_plot(params, data_dir, export_dir)
    if params.get("index_type") in set(dict(RAINFALL_INDICES)):
        result.pop("plot_path", None)
        result["rainfall_csv_only"] = True
    return result


# Do not create new Parquet files for downloads/products. CSV/Excel/PDF remain user-facing outputs.
def _append_parquet_database(table: str, df: pd.DataFrame, stem: str):
    return None

def _save_data(df: pd.DataFrame, out_dir: Path, stem: str, table: str = "products"):
    out_dir.mkdir(parents=True, exist_ok=True)
    clean_df = _round_output_dataframe(_compact_export_dataframe(df))
    csv_path = out_dir / f"{stem}.csv"
    clean_df.to_csv(csv_path, index=False)
    return csv_path, None, None

# ============================================================
# HOTFIX - 2026-07-08 Climate-index heading format
# ============================================================
# Climate-index plots should clearly show index + time scale/season/month + area + period/baseline.

_MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
    7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December",
}

_CROSS_YEAR_SEASONS = {"DJF", "NDJ", "DJFMA", "NDJFMA"}


def _year_range_text(start_year: int, end_year: int) -> str:
    start_year = int(start_year)
    end_year = int(end_year)
    return str(start_year) if start_year == end_year else f"{start_year}–{end_year}"


def _season_year_text(season_label: str, start_year: int, end_year: int) -> str:
    season_label = str(season_label or "ANNUAL").upper()
    start_year = int(start_year)
    end_year = int(end_year)
    if season_label in _CROSS_YEAR_SEASONS:
        if start_year == end_year:
            return f"{start_year - 1}/{start_year} Season"
        return f"{start_year - 1}/{start_year}–{end_year - 1}/{end_year} Seasons"
    return _year_range_text(start_year, end_year)


def _month_prefix_from_months(months: list[int] | None) -> str:
    months = [int(m) for m in (months or []) if int(m) in _MONTH_NAMES]
    if not months:
        return "Monthly"
    if len(months) == 1:
        return _MONTH_NAMES[months[0]]
    return "–".join(_MONTH_NAMES[m] for m in months)


def climate_index_plot_heading(index_type: str, location: str, season_label: str | None,
                               start_year: int, end_year: int,
                               baseline_start: int = 1991, baseline_end: int = 2020,
                               custom_months: list[int] | None = None) -> str:
    return _cde_force_index_heading(index_type, location, season_label, start_year, end_year, baseline_start, baseline_end, custom_months)


def _plot_climate_index_with_heading(summary: pd.DataFrame, selected_col: str, index_type: str,
                                     location: str, season_label: str, start_year: int, end_year: int,
                                     baseline_start: int, baseline_end: int,
                                     out_path: Path, custom_months: list[int] | None = None) -> Path:
    labels = dict(ALL_INDICES)
    fig, ax = plt.subplots(figsize=(10.5, 5.8))
    values = pd.to_numeric(summary[selected_col], errors="coerce")
    x = pd.to_numeric(summary["year"], errors="coerce") if "year" in summary.columns else np.arange(len(summary)) + 1

    if "anomaly" in str(selected_col).lower() or index_type in {"rainfall_anomaly", "temperature_anomaly", "spi", "spei", "relative_humidity_index", "soil_moisture_index", "soil_moisture_anomaly", "wind_speed_index"}:
        colors = np.where(values >= 0, "#d73027", "#4575b4")
        ax.bar(x, values, color=colors)
        ax.axhline(0, linewidth=1, color="black")
    elif "days" in str(selected_col).lower() or index_type in {"hot_days", "hot_nights", "cold_days", "cold_nights", "number_wet_days", "number_dry_days", "consecutive_dry_days", "consecutive_wet_days", "heavy_rainfall_days", "very_heavy_rainfall_days"}:
        ax.bar(x, values, color=plot_color_for(index_type=index_type))
    elif index_type in set(dict(RAINFALL_INDICES)):
        ax.bar(x.astype(str) if hasattr(x, "astype") else x, values, color="#1f78b4")
    else:
        ax.plot(x, values, marker="o", color=plot_color_for(index_type=index_type))

    unit_map = {
        "total_rainfall": "mm", "rainfall_anomaly": "mm", "max_1day_rainfall": "mm",
        "max_5day_rainfall": "mm", "sdii": "mm/wet day", "r95p": "mm", "r99p": "mm",
        "mean_temperature": "°C", "maximum_temperature": "°C", "minimum_temperature": "°C",
        "temperature_anomaly": "°C", "hot_days": "days", "hot_nights": "days",
        "cold_days": "days", "cold_nights": "days", "heat_index": "°C", "dtr": "°C",
        "relative_humidity_index": "index", "soil_moisture_index": "index",
        "soil_moisture_anomaly": "m³/m³", "wind_speed_index": "index",
        "number_wet_days": "days", "number_dry_days": "days", "consecutive_dry_days": "days",
        "consecutive_wet_days": "days", "wet_spell_length": "days", "dry_spell_length": "days",
        "length_of_rainy_season": "days", "heavy_rainfall_days": "days", "very_heavy_rainfall_days": "days",
    }
    ax.set_title(climate_index_plot_heading(index_type, location, season_label, start_year, end_year, baseline_start, baseline_end, custom_months), fontsize=13)
    ax.set_xlabel("Year")
    ax.set_ylabel(y_axis_label(labels.get(index_type, str(selected_col).replace("_", " ").title()), unit_map.get(index_type, "")))
    apply_plot_grids(ax)
    return _finalize_plot(fig, out_path)


# Override the non-rainfall index plotter used by the existing generator. When only the old
# signature is available, infer season from summary or default to Annual.
def _plot_non_rainfall_index(summary: pd.DataFrame, selected_col: str, index_type: str, location: str,
                             start_year: int, end_year: int, out_path: Path) -> Path:
    season_label = "ANNUAL"
    if "season" in summary.columns and len(summary["season"].dropna()):
        season_label = str(summary["season"].dropna().iloc[0])
    return _plot_climate_index_with_heading(summary, selected_col, index_type, location, season_label,
                                            start_year, end_year, 1991, 2020, out_path)


# Final wrapper: regenerate climate-index plot titles using the form season/months/baseline.
_generate_indices_before_heading_hotfix = generate_indices

def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    result = _generate_indices_before_heading_hotfix(params, data_dir, export_dir)
    index_type = params.get("index_type")
    # Rainfall indices remain CSV-only where configured by earlier hotfixes.
    if index_type in set(dict(RAINFALL_INDICES)):
        return result
    try:
        plot_path = result.get("plot_path")
        if not plot_path:
            return result
        excel_path = Path(result.get("excel_path")) if result.get("excel_path") else None
        csv_path = Path(result.get("csv_path")) if result.get("csv_path") else None
        if excel_path and excel_path.exists():
            summary = pd.read_excel(excel_path, sheet_name="Index Summary")
        elif csv_path and csv_path.exists():
            summary = pd.read_csv(csv_path)
        else:
            return result
        selected_col = result.get("selected_column")
        if not selected_col or selected_col not in summary.columns:
            numeric_cols = [c for c in summary.columns if pd.api.types.is_numeric_dtype(summary[c]) and c not in {"latitude", "longitude", "year", "month", "day", "hour"}]
            selected_col = numeric_cols[-1] if numeric_cols else None
        if not selected_col:
            return result
        season_label, months = parse_custom_months(params.get("season") or "ANNUAL", params.get("custom_months"))
        start_year = int(params.get("start_year") or (summary["year"].min() if "year" in summary.columns else 1991))
        end_year = int(params.get("end_year") or (summary["year"].max() if "year" in summary.columns else start_year))
        baseline_start = int(params.get("baseline_start") or 1991)
        baseline_end = int(params.get("baseline_end") or 2020)
        location = params.get("location_name") or "Selected Area"
        _plot_climate_index_with_heading(summary, selected_col, index_type, location, season_label,
                                         start_year, end_year, baseline_start, baseline_end, Path(plot_path), months)
    except Exception:
        # Never fail the data product just because of a title/plot-refresh issue.
        pass
    return result


# ============================================================
# ABSOLUTE FINAL OVERRIDE - 2026-07-08 force climate-index plot headings
# ============================================================
_CDE_generate_indices_before_force_heading_final = generate_indices

def _plot_non_rainfall_index(summary: pd.DataFrame, selected_col: str, index_type: str, location: str,
                             start_year: int, end_year: int, out_path: Path) -> Path:
    labels = dict(ALL_INDICES)
    fig, ax = plt.subplots(figsize=(10.5, 5.8))
    values = pd.to_numeric(summary[selected_col], errors="coerce")
    x = pd.to_numeric(summary["year"], errors="coerce") if "year" in summary.columns else np.arange(len(summary)) + 1
    if "anomaly" in str(selected_col).lower() or "index" in str(selected_col).lower() or "standardized" in str(selected_col).lower():
        colors = np.where(values >= 0, "#d73027", "#4575b4")
        ax.bar(x, values, color=colors)
        ax.axhline(0, linewidth=1, color="black")
    elif "days" in str(selected_col).lower() or index_type in {"hot_days", "hot_nights", "cold_days", "cold_nights", "number_wet_days", "number_dry_days"}:
        ax.bar(x, values, color=plot_color_for(index_type=index_type))
    else:
        ax.plot(x, values, marker="o", color=plot_color_for(index_type=index_type))
    season_label = _cde_force_season_from_summary(summary)
    ax.set_title(_cde_force_index_heading(index_type, location, season_label, start_year, end_year), fontsize=13)
    ax.set_xlabel("Year")
    unit_map = {
        "total_rainfall": "mm", "rainfall_anomaly": "mm", "max_1day_rainfall": "mm", "max_5day_rainfall": "mm", "sdii": "mm/wet day",
        "mean_temperature": "°C", "maximum_temperature": "°C", "minimum_temperature": "°C",
        "temperature_anomaly": "°C", "hot_days": "days", "hot_nights": "days",
        "cold_days": "days", "cold_nights": "days", "heat_index": "°C", "dtr": "°C",
        "relative_humidity_index": "index", "soil_moisture_index": "index",
        "soil_moisture_anomaly": "m³/m³", "wind_speed_index": "index",
    }
    ax.set_ylabel(y_axis_label(labels.get(index_type, str(selected_col).replace("_", " ").title()), unit_map.get(index_type, "")))
    apply_plot_grids(ax)
    return _finalize_plot(fig, out_path)


def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    result = _CDE_generate_indices_before_force_heading_final(params, data_dir, export_dir)
    index_type = params.get("index_type")
    # Rainfall indices remain CSV-only if earlier settings require it.
    if index_type in set(dict(RAINFALL_INDICES)):
        result.pop("plot_path", None)
        result["rainfall_csv_only"] = True
        return result
    plot_path = result.get("plot_path")
    if not plot_path:
        return result
    try:
        summary = None
        if result.get("excel_path") and Path(result.get("excel_path")).exists():
            try:
                summary = pd.read_excel(Path(result.get("excel_path")), sheet_name="Index Summary")
            except Exception:
                summary = pd.read_excel(Path(result.get("excel_path")))
        elif result.get("csv_path") and Path(result.get("csv_path")).exists():
            summary = pd.read_csv(Path(result.get("csv_path")))
        if summary is None or summary.empty:
            return result
        selected_col = result.get("selected_column")
        if not selected_col or selected_col not in summary.columns:
            numeric_cols = [c for c in summary.columns if pd.api.types.is_numeric_dtype(summary[c]) and str(c).lower() not in {"latitude", "longitude", "year", "month", "day", "hour"}]
            selected_col = numeric_cols[-1] if numeric_cols else None
        if not selected_col:
            return result
        start_year = int(params.get("start_year") or (summary["year"].min() if "year" in summary.columns else 1991))
        end_year = int(params.get("end_year") or (summary["year"].max() if "year" in summary.columns else start_year))
        location = params.get("location_name") or params.get("station_name") or "Selected Area"
        _plot_non_rainfall_index(summary, selected_col, index_type, location, start_year, end_year, Path(plot_path))
    except Exception as exc:
        print(f"Climate index heading redraw skipped: {exc}", flush=True)
    return result

# ============================================================
# FINAL PATCH - 2026-07-08 use selected season from form in climate index headings
# This fixes cases where the summary file does not carry a season column.
# ============================================================
_CDE_generate_indices_before_selected_season_heading_patch = generate_indices


def _cde_selected_season_from_params(params: Dict[str, Any]):
    raw = (
        params.get("season")
        or params.get("season_label")
        or params.get("season_name")
        or params.get("period")
        or "ANNUAL"
    )
    custom_raw = params.get("custom_months") or params.get("months") or params.get("custom_month")
    try:
        season_label, months = parse_custom_months(raw, custom_raw)
        return season_label, months
    except Exception:
        pass

    text = str(raw or "ANNUAL").strip()
    upper = text.upper()
    if "ANNUAL" in upper or "WHOLE" in upper:
        return "ANNUAL", []
    known = ["NDJFMA", "DJFMA", "MAM", "OND", "NDJ", "DJF", "JJA", "SON"]
    for code in known:
        if upper.startswith(code) or f"{code} " in upper or f"{code}(" in upper:
            return code, []
    return upper or "ANNUAL", []


def _cde_final_redraw_index_plot_with_selected_season(summary: pd.DataFrame, selected_col: str, index_type: str,
                                                       location: str, season_label: str,
                                                       start_year: int, end_year: int,
                                                       baseline_start: int, baseline_end: int,
                                                       out_path: Path, custom_months=None) -> Path:
    labels = dict(ALL_INDICES)
    fig, ax = plt.subplots(figsize=(10.5, 5.8))
    values = pd.to_numeric(summary[selected_col], errors="coerce")
    x = pd.to_numeric(summary["year"], errors="coerce") if "year" in summary.columns else np.arange(len(summary)) + 1

    lower_col = str(selected_col).lower()
    if "anomaly" in lower_col or "index" in lower_col or "standardized" in lower_col or index_type in {"temperature_anomaly", "rainfall_anomaly", "spi", "spei", "relative_humidity_index", "soil_moisture_index", "soil_moisture_anomaly", "wind_speed_index"}:
        colors = np.where(values >= 0, "#d73027", "#4575b4")
        ax.bar(x, values, color=colors)
        ax.axhline(0, linewidth=1, color="black")
    elif "days" in lower_col or index_type in {"hot_days", "hot_nights", "cold_days", "cold_nights", "number_wet_days", "number_dry_days", "consecutive_dry_days", "consecutive_wet_days", "heavy_rainfall_days", "very_heavy_rainfall_days"}:
        ax.bar(x, values, color=plot_color_for(index_type=index_type))
    else:
        ax.plot(x, values, marker="o", color=plot_color_for(index_type=index_type))

    unit_map = {
        "total_rainfall": "mm", "rainfall_anomaly": "mm", "max_1day_rainfall": "mm", "max_5day_rainfall": "mm", "sdii": "mm/wet day", "r95p": "mm", "r99p": "mm",
        "mean_temperature": "°C", "maximum_temperature": "°C", "minimum_temperature": "°C", "temperature_anomaly": "°C", "hot_days": "days", "hot_nights": "days", "cold_days": "days", "cold_nights": "days", "heat_index": "°C", "dtr": "°C",
        "relative_humidity_index": "index", "soil_moisture_index": "index", "soil_moisture_anomaly": "m³/m³", "wind_speed_index": "index",
        "number_wet_days": "days", "number_dry_days": "days", "consecutive_dry_days": "days", "consecutive_wet_days": "days", "wet_spell_length": "days", "dry_spell_length": "days", "length_of_rainy_season": "days", "heavy_rainfall_days": "days", "very_heavy_rainfall_days": "days",
    }
    title = _cde_force_index_heading(index_type, location, season_label, start_year, end_year, baseline_start, baseline_end, custom_months)
    ax.set_title(title, fontsize=13)
    ax.set_xlabel("Year")
    ax.set_ylabel(y_axis_label(labels.get(index_type, str(selected_col).replace("_", " ").title()), unit_map.get(index_type, "")))
    apply_plot_grids(ax)
    return _finalize_plot(fig, out_path)


def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    result = _CDE_generate_indices_before_selected_season_heading_patch(params, data_dir, export_dir)
    index_type = params.get("index_type")

    # Rainfall indices stay CSV-only. No plot is returned.
    if index_type in set(dict(RAINFALL_INDICES)):
        result.pop("plot_path", None)
        result["rainfall_csv_only"] = True
        return result

    plot_path = result.get("plot_path")
    if not plot_path:
        return result

    try:
        summary = None
        if result.get("excel_path") and Path(result.get("excel_path")).exists():
            try:
                summary = pd.read_excel(Path(result.get("excel_path")), sheet_name="Index Summary")
            except Exception:
                summary = pd.read_excel(Path(result.get("excel_path")))
        elif result.get("csv_path") and Path(result.get("csv_path")).exists():
            summary = pd.read_csv(Path(result.get("csv_path")))

        if summary is None or summary.empty:
            return result

        selected_col = result.get("selected_column")
        if not selected_col or selected_col not in summary.columns:
            numeric_cols = [
                c for c in summary.columns
                if pd.api.types.is_numeric_dtype(summary[c]) and str(c).lower() not in {"latitude", "longitude", "year", "month", "day", "hour"}
            ]
            selected_col = numeric_cols[-1] if numeric_cols else None
        if not selected_col:
            return result

        season_label, custom_months = _cde_selected_season_from_params(params)
        start_year = int(params.get("start_year") or (summary["year"].min() if "year" in summary.columns else 1991))
        end_year = int(params.get("end_year") or (summary["year"].max() if "year" in summary.columns else start_year))
        baseline_start = int(params.get("baseline_start") or 1991)
        baseline_end = int(params.get("baseline_end") or 2020)
        location = params.get("location_name") or params.get("station_name") or params.get("location") or "Selected Area"

        _cde_final_redraw_index_plot_with_selected_season(
            summary, selected_col, index_type, location, season_label,
            start_year, end_year, baseline_start, baseline_end, Path(plot_path), custom_months
        )
    except Exception as exc:
        print(f"Climate index selected-season heading redraw skipped: {exc}", flush=True)

    return result

# ============================================================
# PRECOMPUTED ZARR CLIMATE-INDEX SUPPORT
# Prefer an existing store under storage/zarr/climate_indices when its name
# matches the requested index and season. If it cannot be read or matched, the
# established source-data calculation workflow remains the automatic fallback.
# ============================================================
_CDE_generate_indices_from_source_zarr = generate_indices

_PRECOMPUTED_INDEX_TOKENS = {
    "total_rainfall": ["Total_Rainfall"],
    "number_wet_days": ["Number_of_Wet_Days"],
    "number_dry_days": ["Number_of_Dry_Days"],
    "consecutive_dry_days": ["Consecutive_Dry_Days"],
    "consecutive_wet_days": ["Consecutive_Wet_Days"],
    "wet_spell_length": ["Wet_Spell_Length"],
    "dry_spell_length": ["Dry_Spell_Length"],
    "rainy_season_onset": ["Rainy_Season_Onset"],
    "rainy_season_cessation": ["Rainy_Season_Cessation"],
    "max_1day_rainfall": ["Maximum_1_Day_Rainfall"],
    "max_5day_rainfall": ["Maximum_5_Day_Rainfall"],
    "heavy_rainfall_days": ["Heavy_Rainfall_Days"],
    "very_heavy_rainfall_days": ["Very_Heavy_Rainfall_Days"],
    "rainfall_anomaly": ["Rainfall_Anomaly"],
    "r95p": ["R95p_Very_Wet_Rainfall", "R95p_Very_Wet_Days"],
    "r99p": ["R99p_Extremely_Wet_Rainfall", "R99p_Extremely_Wet_Days"],
    "mean_temperature": ["Mean_Temperature"],
    "maximum_temperature": ["Maximum_Temperature"],
    "minimum_temperature": ["Minimum_Temperature"],
    "temperature_anomaly": ["Temperature_Anomaly"],
    "hot_days": ["Hot_Days"],
    "hot_nights": ["Hot_Nights"],
    "cold_days": ["Cold_Days"],
    "cold_nights": ["Cold_Nights"],
    "heat_index": ["Heat_Index"],
    "dtr": ["Diurnal_Temperature_Range"],
    "relative_humidity_index": ["Relative_Humidity"],
    "soil_moisture_index": ["Volumetric_Soil_Water"],
    "soil_moisture_anomaly": ["Volumetric_Soil_Water"],
    "wind_speed_index": ["Wind_Speed_10m"],
}
_PRECOMPUTED_SEASONS = {"DJF", "MAM", "JJA", "SON", "OND", "NDJ", "DJFMA", "NDJFMA"}


def _find_precomputed_index_store(data_dir: Path, params: Dict[str, Any]) -> Path | None:
    index_type = str(params.get("index_type") or "")
    tokens = _PRECOMPUTED_INDEX_TOKENS.get(index_type, [])
    if not tokens:
        return None
    index_root = Path(data_dir) / "climate_indices"
    if not index_root.exists():
        return None

    season_label, _ = _cde_selected_season_from_params(params)
    season = str(season_label or "ANNUAL").upper()
    dataset = str(params.get("dataset") or "")
    stores = sorted(index_root.glob("*.zarr"))
    candidates: list[Path] = []
    for store in stores:
        normalized = _normalize_store_name(store.name)
        if not any(_normalize_store_name(token) in normalized for token in tokens):
            continue
        if season in _PRECOMPUTED_SEASONS:
            if _normalize_store_name(f"CDE_INDEX_{season}_") not in normalized:
                continue
        else:
            # Prefer the annual/non-seasonal store rather than a seasonal store.
            if any(_normalize_store_name(f"CDE_INDEX_{code}_") in normalized for code in _PRECOMPUTED_SEASONS):
                continue
        if index_type in set(dict(RAINFALL_INDICES)) | {"r95p", "r99p"}:
            if dataset == "chirps_rainfall" and "chirps" not in normalized:
                continue
            if dataset == "era5_total_precipitation" and "era5tp" not in normalized:
                continue
        candidates.append(store)
    return sorted(candidates, key=lambda p: (len(p.name), p.name))[0] if candidates else None


def _precomputed_index_unit(index_type: str, attrs: Dict[str, Any]) -> str:
    unit = str(attrs.get("units") or "").strip()
    if unit:
        return unit
    if index_type in {"total_rainfall", "rainfall_anomaly", "max_1day_rainfall", "max_5day_rainfall", "r95p", "r99p"}:
        return "mm"
    if index_type in {"mean_temperature", "maximum_temperature", "minimum_temperature", "temperature_anomaly", "heat_index", "dtr"}:
        return "°C"
    if index_type in {
        "number_wet_days", "number_dry_days", "consecutive_dry_days", "consecutive_wet_days",
        "wet_spell_length", "dry_spell_length", "heavy_rainfall_days", "very_heavy_rainfall_days",
        "hot_days", "hot_nights", "cold_days", "cold_nights",
    }:
        return "days"
    if index_type == "relative_humidity_index":
        return "%"
    if index_type in {"soil_moisture_index", "soil_moisture_anomaly"}:
        return "m³/m³"
    return "index"


def _extract_precomputed_index_series(store: Path, params: Dict[str, Any]) -> tuple[pd.DataFrame, str, str, Dict[str, Any]]:
    index_type = str(params.get("index_type") or "index_value")
    start_year = int(params.get("start_year") or 1900)
    end_year = int(params.get("end_year") or 2100)
    latitude = float(params.get("latitude"))
    longitude = float(params.get("longitude"))

    with open_data_store(store, decode_times=True) as ds:
        lat_name = detect_coord(ds, ["latitude", "lat", "y"])
        lon_name = detect_coord(ds, ["longitude", "lon", "x"])
        selected = ds
        nearest_lat = latitude
        nearest_lon = longitude
        if lat_name and lon_name:
            selected = selected.sel({lat_name: latitude, lon_name: longitude}, method="nearest")
            nearest_lat = float(np.asarray(selected[lat_name]).reshape(-1)[0])
            nearest_lon = float(np.asarray(selected[lon_name]).reshape(-1)[0])

        variables = data_variables(selected)
        if not variables:
            raise ValueError(f"No numeric index variable was found in {store.name}.")
        desired = _PRECOMPUTED_INDEX_TOKENS.get(index_type, [index_type])
        scored = []
        for name in variables:
            score = variable_score(name, selected[name], desired + [index_type, "index", "value"])
            scored.append((score, name))
        scored.sort(reverse=True)
        var_name = scored[0][1]
        da = selected[var_name]

        time_name = detect_coord(selected, ["time", "valid_time", "datetime", "date"])
        year_name = detect_coord(selected, ["year", "season_year", "hydrological_year"])
        keep_dim = time_name or year_name
        for dim in list(da.dims):
            if dim == keep_dim:
                continue
            da = da.isel({dim: 0}, drop=True)

        values = np.asarray(da.values).reshape(-1)
        if time_name and time_name in da.coords:
            dates = pd.to_datetime(np.asarray(da[time_name].values).reshape(-1), errors="coerce")
            years = pd.Series(dates).dt.year.to_numpy()
        elif year_name and year_name in da.coords:
            years = pd.to_numeric(pd.Series(np.asarray(da[year_name].values).reshape(-1)), errors="coerce").to_numpy()
        elif len(values) == end_year - start_year + 1:
            years = np.arange(start_year, end_year + 1)
        else:
            raise ValueError(f"No time or year coordinate was found in {store.name}.")

        size = min(len(years), len(values))
        frame = pd.DataFrame({"year": years[:size], index_type: pd.to_numeric(values[:size], errors="coerce")})
        frame = frame.dropna(subset=["year"]).copy()
        frame["year"] = frame["year"].astype(int)
        frame = frame[(frame["year"] >= start_year) & (frame["year"] <= end_year)].sort_values("year")
        if frame.empty:
            raise ValueError("The precomputed index store has no values for the selected years.")

        unit = _precomputed_index_unit(index_type, dict(da.attrs))
        context = {
            "dataset_label": dict(ALL_INDICES).get(index_type, index_type.replace("_", " ").title()),
            "variable_label": da.attrs.get("long_name") or da.attrs.get("standard_name") or var_name,
            "variable": var_name,
            "unit": unit,
            "file": store.name,
            "source_path": str(store),
            "storage_format": store_kind(store),
            "source_mode": "Precomputed Zarr climate-index store",
            "nearest_latitude": nearest_lat,
            "nearest_longitude": nearest_lon,
        }
    return frame, index_type, unit, context


def _generate_precomputed_index_product(params: Dict[str, Any], store: Path, export_dir: Path) -> Dict[str, Any]:
    summary, selected_col, unit, context = _extract_precomputed_index_series(store, params)
    index_type = str(params.get("index_type") or selected_col)
    location = str(params.get("location_name") or "Selected Location")
    start_year = int(params.get("start_year") or summary["year"].min())
    end_year = int(params.get("end_year") or summary["year"].max())
    baseline_start = int(params.get("baseline_start") or 1991)
    baseline_end = int(params.get("baseline_end") or 2020)
    season_label, custom_months = _cde_selected_season_from_params(params)

    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = slugify(f"CDE_precomputed_index_{index_type}_{location}_{season_label}_{start_year}_{end_year}_{stamp}")
    csv_path = dirs["indices"] / f"{stem}.csv"
    excel_path = dirs["indices"] / f"{stem}.xlsx"
    data_used_csv = dirs["indices"] / f"{stem}_data_used.csv"
    summary.to_csv(csv_path, index=False)
    summary.to_csv(data_used_csv, index=False)

    metadata = pd.DataFrame([{
        **context,
        "location": location,
        "requested_latitude": float(params.get("latitude")),
        "requested_longitude": float(params.get("longitude")),
        "start_year": start_year,
        "end_year": end_year,
        "season": season_label,
    }])
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        metadata.to_excel(writer, sheet_name="Metadata", index=False)
        summary.to_excel(writer, sheet_name="Index Summary", index=False)

    result: Dict[str, Any] = {
        "context": context,
        "rows": len(summary),
        "csv_path": csv_path,
        "excel_path": excel_path,
        "data_used_csv_path": data_used_csv,
        "selected_column": selected_col,
        "precomputed_store": str(store),
    }
    if index_type not in set(dict(RAINFALL_INDICES)):
        plot_path = dirs["indices"] / f"{stem}.png"
        _cde_final_redraw_index_plot_with_selected_season(
            summary, selected_col, index_type, location, season_label,
            start_year, end_year, baseline_start, baseline_end, plot_path, custom_months,
        )
        result["plot_path"] = plot_path
    else:
        result["rainfall_csv_only"] = True
    return result


def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    prefer_precomputed = str(os.environ.get("CDE_PREFER_PRECOMPUTED_INDICES", "1")).strip().lower() not in {"0", "false", "no"}
    if prefer_precomputed:
        try:
            store = _find_precomputed_index_store(Path(data_dir), params)
            if store:
                return _generate_precomputed_index_product(params, store, Path(export_dir))
        except Exception as exc:
            print(f"Precomputed Zarr index fallback: {exc}", flush=True)
    return _CDE_generate_indices_from_source_zarr(params, Path(data_dir), Path(export_dir))


# ============================================================
# FINAL SYSTEM-WIDE EXPORT RULES - July 2026
# Unlimited periods, Excel-only data downloads, QR verification,
# CDE_ filenames and operational precision.
# ============================================================
from openpyxl import load_workbook as _cde_load_workbook
from openpyxl.styles import Alignment as _CDEAlignment, Border as _CDEBorder, Font as _CDEFont, PatternFill as _CDEPatternFill, Side as _CDESide
from openpyxl.utils import get_column_letter as _cde_get_column_letter
from scripts.extractor import add_qr_codes_to_workbook as _cde_add_qr_codes_to_workbook, default_download_context as _cde_default_download_context
from cde_excel import write_single_sheet_workbook as _cde_write_single_sheet_workbook


def _cde_prefixed_stem(stem: str) -> str:
    cleaned = slugify(stem)
    for prefix in ("CDE_", "CDE-", "CDE_"):
        if cleaned.startswith(prefix):
            cleaned = cleaned[len(prefix):]
            break
    return f"CDE_{cleaned}" if cleaned else "CDE_Product"


def _cde_prefixed_path(path: Path | str | None) -> Path | None:
    if not path:
        return None
    old = Path(path)
    if not old.exists():
        return old
    if old.name.startswith("CDE_"):
        return old
    new = old.with_name(_cde_prefixed_stem(old.stem) + old.suffix)
    if new != old:
        if new.exists():
            new.unlink()
        old.replace(new)
    return new


def _cde_zero_decimal_hint(text: str) -> bool:
    value = str(text or "").lower()
    return any(token in value for token in (
        "relative_humidity", "relative humidity", "wind_speed", "wind speed",
        "wind_direction", "wind direction", "wind_rose", "era5_wind",
    ))


def _cde_round_export_frame(df: pd.DataFrame, hint: str = "") -> pd.DataFrame:
    """Apply the requested precision without altering identifiers or coordinates."""
    out = df.copy()
    zero_hint = _cde_zero_decimal_hint(hint)
    coordinate_tokens = ("latitude", "longitude", "_lat", "_lon")
    structural = {"year", "month", "day", "hour", "season_year", "count", "frequency", "observations", "records", "n", "s"}
    for column in out.columns:
        if not pd.api.types.is_numeric_dtype(out[column]):
            continue
        name = str(column).strip().lower()
        values = pd.to_numeric(out[column], errors="coerce")
        if name in {"lat", "lon"} or any(token in name for token in coordinate_tokens):
            out[column] = values.round(4)
            continue
        is_structural = name in structural or any(token in name for token in ("count", "records", "observations"))
        is_day_count = any(token in name for token in ("days", "day_count", "wet_observations", "dry_observations"))
        if is_structural or is_day_count:
            finite = values.dropna()
            if finite.empty or finite.apply(lambda x: float(x).is_integer()).all():
                out[column] = values.round(0).astype("Int64")
            else:
                out[column] = values.round(1)
            continue
        zero_column = zero_hint or any(token in name for token in (
            "relative_humidity", "relative humidity", "wind_speed", "wind speed",
            "wind_direction", "wind direction",
        ))
        if zero_column:
            out[column] = values.round(0).astype("Int64")
        else:
            out[column] = values.round(1)
    return out


def _cde_style_excel(path: Path, zero_hint: bool = False) -> None:
    wb = _cde_load_workbook(path)
    header_fill = _CDEPatternFill("solid", fgColor="E7E6E6")
    thin = _CDESide(style="thin", color="D9E2EC")
    border = _CDEBorder(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        if ws.max_row:
            for cell in ws[1]:
                cell.font = _CDEFont(bold=True, color="000000")
                cell.fill = header_fill
                cell.alignment = _CDEAlignment(horizontal="center", vertical="center", wrap_text=True)
        headers = {cell.column: str(cell.value or "").strip().lower() for cell in ws[1]}
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = _CDEAlignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, (int, float, np.integer, np.floating)) and not isinstance(cell.value, bool):
                    header = headers.get(cell.column, "")
                    if header in {"year", "month", "day", "hour", "count", "records", "observations", "season year"} or "days" in header:
                        cell.number_format = "0"
                    elif "latitude" in header or "longitude" in header or header in {"lat", "lon"}:
                        cell.number_format = "0.0000"
                    elif zero_hint or any(token in header for token in ("relative humidity", "wind speed", "wind direction")):
                        cell.number_format = "0"
                    else:
                        cell.number_format = "0.0"
        for col_idx, cells in enumerate(ws.columns, start=1):
            width = max((len(str(c.value or "")) for c in cells), default=10)
            ws.column_dimensions[_cde_get_column_letter(col_idx)].width = min(max(width + 1, 9), 18)
        ws.freeze_panes = None
    wb.save(path)


def _cde_add_qr_to_excel(path: Path, payload: Dict[str, Any]) -> None:
    context = _cde_default_download_context(path)
    context.update({
        "document_type": "Climate Data Product",
        "file_name": path.name,
        "source": payload.get("dataset_label") or payload.get("source") or payload.get("dataset") or "",
        "element": payload.get("variable_label") or payload.get("element") or payload.get("index_label") or "",
        "data_type": payload.get("resolution") or payload.get("season") or "",
        "period": payload.get("period") or "",
        "station_name": payload.get("location") or "",
        "latitude": payload.get("latitude") or payload.get("requested_latitude") or "",
        "longitude": payload.get("longitude") or payload.get("requested_longitude") or "",
        "units": payload.get("unit") or "",
        "description": payload.get("description") or "",
    })
    wb = _cde_load_workbook(path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    _cde_add_qr_codes_to_workbook(path, {name: dict(context) for name in names}, context)


_CDE_SUPPRESS_DATA_EXPORT: ContextVar[bool] = ContextVar("cde_suppress_data_export", default=False)


def _save_data(df: pd.DataFrame, out_dir: Path, stem: str, table: str = "products"):
    """Save complete product data as Excel only; no row or period limit is applied."""
    if _CDE_SUPPRESS_DATA_EXPORT.get():
        return None, None, None
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    final_stem = _cde_prefixed_stem(stem)
    clean_df = _cde_round_export_frame(_compact_export_dataframe(df), final_stem)
    excel_path = out_dir / f"{final_stem}.xlsx"
    metadata = pd.DataFrame([{
        "Product": final_stem.replace("CDE_", "").replace("_", " "),
        "Records": len(clean_df),
        "Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Download Format": "Excel",
    }])
    zero_hint = _cde_zero_decimal_hint(final_stem)
    qr_payload = _cde_default_download_context(excel_path)
    qr_payload.update({
        "document_type": "Climate Data Product",
        "file_name": excel_path.name,
        "description": "Extracted data used to generate the selected climate product.",
        "element": final_stem.replace("CDE_", "").replace("_", " "),
        "period": "Complete selected period",
    })
    _cde_write_single_sheet_workbook(
        excel_path,
        [("Product Information", metadata), ("Extracted Data", clean_df)],
        qr_payload=qr_payload,
        zero_decimal=zero_hint,
        sheet_name="Data",
        workbook_title=final_stem.replace("CDE_", "").replace("_", " "),
    )
    return excel_path, None, None


_CDE_GENERATE_PLOT_EXCEL_ONLY_BASE = generate_plot_product


def generate_plot_product(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    token = _CDE_SUPPRESS_DATA_EXPORT.set(bool(params.get("_skip_data_export")))
    try:
        result = _CDE_GENERATE_PLOT_EXCEL_ONLY_BASE(params, data_dir, export_dir)
    finally:
        _CDE_SUPPRESS_DATA_EXPORT.reset(token)
    # The legacy internal key represented the accompanying data file. It is now Excel.
    legacy_data_path = result.pop("csv_path", None)
    if legacy_data_path:
        result["excel_path"] = Path(legacy_data_path)
    if result.get("excel_path"):
        result["excel_path"] = _cde_prefixed_path(result["excel_path"])
    if result.get("plot_path"):
        result["plot_path"] = _cde_prefixed_path(result["plot_path"])
    result.pop("parquet_path", None)
    result.pop("db_path", None)
    return result


_CDE_GENERATE_INDEX_EXCEL_ONLY_BASE = generate_indices


def _cde_read_index_sheets(result: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    sheets: Dict[str, pd.DataFrame] = {}
    excel_path = Path(result["excel_path"]) if result.get("excel_path") else None
    if excel_path and excel_path.exists():
        try:
            sheets = pd.read_excel(excel_path, sheet_name=None)
        except Exception:
            sheets = {}
    csv_path = Path(result["csv_path"]) if result.get("csv_path") else None
    if not sheets and csv_path and csv_path.exists():
        sheets["Index Summary"] = pd.read_csv(csv_path)
    data_used = Path(result["data_used_csv_path"]) if result.get("data_used_csv_path") else None
    if data_used and data_used.exists():
        try:
            sheets["Data Used"] = pd.read_csv(data_used)
        except Exception:
            pass
    return sheets


def _cde_selected_index_frame(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    for name in ("Index Summary", "Data", "Summary"):
        if name in sheets and not sheets[name].empty:
            return sheets[name]
    for name, frame in sheets.items():
        if name.lower() != "metadata" and frame is not None and not frame.empty:
            return frame
    return pd.DataFrame()


def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate every selected index for the full requested period with Excel data and a plot."""
    result = _CDE_GENERATE_INDEX_EXCEL_ONLY_BASE(params, data_dir, export_dir)
    index_type = str(params.get("index_type") or "total_rainfall")
    index_label = dict(ALL_INDICES).get(index_type, index_type.replace("_", " ").title())
    location = str(params.get("location_name") or "Selected Location")
    start_year = int(params.get("start_year") or 0)
    end_year = int(params.get("end_year") or start_year)
    season_label, months = parse_custom_months(params.get("season") or "ANNUAL", params.get("custom_months"))
    sheets = _cde_read_index_sheets(result)
    summary = _cde_selected_index_frame(sheets)
    if summary.empty:
        raise ValueError("No index values were produced for the selected period.")

    hint = f"{index_type} {index_label}"
    rounded_sheets = {name[:31]: _cde_round_export_frame(frame, hint) for name, frame in sheets.items() if frame is not None}
    if "Index Summary" not in rounded_sheets:
        rounded_sheets["Index Summary"] = _cde_round_export_frame(summary, hint)

    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"index_{index_type}_{location}_{season_label}_{start_year}_{end_year}_{stamp}")
    excel_path = dirs["indices"] / f"{stem}.xlsx"
    qr_payload = _cde_default_download_context(excel_path)
    qr_payload.update({
        "document_type": "Climate Index Product",
        "file_name": excel_path.name,
        "index_label": index_label,
        "element": index_label,
        "station_name": location,
        "location": location,
        "latitude": params.get("latitude"),
        "longitude": params.get("longitude"),
        "data_type": season_label,
        "season": season_label,
        "period": f"{start_year}-{end_year}",
        "description": "Climate index assessment for the complete selected period.",
    })
    _cde_write_single_sheet_workbook(
        excel_path,
        [(name, frame) for name, frame in rounded_sheets.items()],
        qr_payload=qr_payload,
        zero_decimal=_cde_zero_decimal_hint(hint),
        sheet_name="Data",
        workbook_title=f"{index_label} for {location}",
    )

    selected_col = result.get("selected_column")
    summary_for_plot = rounded_sheets.get("Index Summary", summary)
    if not selected_col or selected_col not in summary_for_plot.columns:
        excluded = {"year", "month", "day", "hour", "latitude", "longitude"}
        numeric_cols = [c for c in summary_for_plot.columns if str(c).lower() not in excluded and pd.api.types.is_numeric_dtype(summary_for_plot[c])]
        selected_col = numeric_cols[-1] if numeric_cols else None
    if selected_col:
        plot_path = dirs["indices"] / f"{stem}.png"
        baseline_start = int(params.get("baseline_start") or 1991)
        baseline_end = int(params.get("baseline_end") or 2020)
        _cde_final_redraw_index_plot_with_selected_season(
            summary_for_plot, selected_col, index_type, location, season_label,
            start_year, end_year, baseline_start, baseline_end, plot_path, months,
        )
        result["plot_path"] = plot_path

    # Remove all CSV artifacts created by legacy layers and expose Excel only.
    for key in ("csv_path", "data_used_csv_path"):
        old = result.pop(key, None)
        if old:
            try:
                path = Path(old)
                if path.exists() and path.suffix.lower() == ".csv":
                    path.unlink()
            except Exception:
                pass
    old_excel = result.get("excel_path")
    if old_excel and Path(old_excel) != excel_path:
        try:
            old_path = Path(old_excel)
            if old_path.exists():
                old_path.unlink()
        except Exception:
            pass
    result.update({
        "excel_path": excel_path,
        "rows": len(summary_for_plot),
        "selected_column": selected_col,
        "rainfall_csv_only": False,
    })
    result.pop("db_path", None)
    return result

# ---------------------------------------------------------------------------
# 2026-07-17: request-only custom spatial maps and monthly heat maps
# ---------------------------------------------------------------------------

_CDE_REQUEST_ONLY_PLOT_BASE = generate_plot_product


def _truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _parse_month_numbers(value: Any, *, default: Iterable[int] | None = None) -> list[int]:
    if isinstance(value, (list, tuple, set)):
        pieces = list(value)
    else:
        pieces = re.split(r"[,;\s]+", str(value or "").strip()) if str(value or "").strip() else []
    months: list[int] = []
    for piece in pieces:
        try:
            month = int(str(piece).strip())
        except Exception:
            continue
        if 1 <= month <= 12 and month not in months:
            months.append(month)
    if not months and default is not None:
        months = [int(month) for month in default if 1 <= int(month) <= 12]
    return sorted(set(months))


def _parse_year_numbers(value: Any, *, start: int, end: int) -> list[int]:
    text = str(value or "").strip()
    years: set[int] = set()
    if text:
        for token in re.split(r"[,;\s]+", text):
            token = token.strip()
            if not token:
                continue
            match = re.fullmatch(r"(\d{4})\s*[-:]\s*(\d{4})", token)
            if match:
                first, last = int(match.group(1)), int(match.group(2))
                if first > last:
                    first, last = last, first
                years.update(range(first, last + 1))
                continue
            if token.isdigit() and len(token) == 4:
                years.add(int(token))
    if not years:
        years.update(range(int(start), int(end) + 1))
    return sorted(years)


def _month_names(months: Iterable[int]) -> list[str]:
    return [pd.Timestamp(2000, int(month), 1).strftime("%B") for month in months]


def _compact_year_label(years: list[int]) -> str:
    if not years:
        return "No years"
    if years == list(range(years[0], years[-1] + 1)):
        return str(years[0]) if len(years) == 1 else f"{years[0]}–{years[-1]}"
    if len(years) <= 8:
        return ", ".join(str(year) for year in years)
    return f"{len(years)} selected years ({years[0]}–{years[-1]})"


def _spatial_period_definition(params: Dict[str, Any]) -> dict[str, Any]:
    mode = str(params.get("map_period_mode") or "month").strip().lower()
    if mode not in {"month", "year", "season", "custom"}:
        mode = "month"
    start = int(params.get("start_year") or 1991)
    end = int(params.get("end_year") or start)
    map_year = int(params.get("map_year") or end)
    if mode == "month":
        months = _parse_month_numbers(params.get("map_month"), default=[1])
        years = [map_year]
        period_label = f"{_month_names(months)[0]} {map_year}"
        preferred_resolution = "monthly"
        season = None
    elif mode == "year":
        months = list(range(1, 13))
        years = [map_year]
        period_label = str(map_year)
        preferred_resolution = "annual"
        season = None
    elif mode == "season":
        season = str(params.get("map_season") or "MAM").strip().upper()
        months = list(SEASON_DEFINITIONS.get(season, SEASON_DEFINITIONS["MAM"]))
        years = [map_year]
        period_label = f"{season} season {map_year} ({', '.join(_month_names(months))})"
        preferred_resolution = "seasonal"
    else:
        months = _parse_month_numbers(params.get("map_custom_months"), default=range(1, 13))
        years = _parse_year_numbers(params.get("map_custom_years"), start=start, end=end)
        period_label = f"{', '.join(_month_names(months))} · {_compact_year_label(years)}"
        preferred_resolution = "monthly"
        season = None
    return {
        "mode": mode,
        "months": months,
        "years": years,
        "season": season,
        "period_label": period_label,
        "preferred_resolution": preferred_resolution,
    }


def _select_time_period(da: xr.DataArray, time_name: str, years: list[int], months: list[int] | None = None) -> xr.DataArray:
    coord = da[time_name]
    try:
        year_values = np.asarray(coord.dt.year.values, dtype=int)
        month_values = np.asarray(coord.dt.month.values, dtype=int)
    except Exception:
        raw = np.asarray(coord.values)
        year_values = pd.to_numeric(pd.Series(raw.ravel()), errors="coerce").fillna(-9999).astype(int).to_numpy().reshape(raw.shape)
        month_values = np.ones_like(year_values, dtype=int)
    mask = np.isin(year_values, np.asarray(years, dtype=int))
    if months:
        mask &= np.isin(month_values, np.asarray(months, dtype=int))
    selector = xr.DataArray(mask, coords=coord.coords, dims=coord.dims)
    return da.where(selector, drop=True)


def _extract_spatial_period(data_dir: Path, params: Dict[str, Any]) -> tuple[pd.DataFrame, Dict[str, Any]]:
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    variable = str(params.get("variable") or "auto")
    if dataset_key == "era5_temperature" and variable in {"all", "all_in_one", "max_min"}:
        variable = "ta"
    if dataset_key == "era5_wind" and variable in {"auto", "wind_speed_direction", ""}:
        variable = "wind_speed"
    period = _spatial_period_definition(params)
    preferred = period["preferred_resolution"]
    season = period["season"] if preferred == "seasonal" else None
    try:
        file_path = find_file(data_dir, dataset_key, preferred, season=season)
        source_resolution = preferred
    except Exception:
        source_resolution = str(params.get("resolution") or "monthly")
        file_path = find_file(data_dir, dataset_key, source_resolution, season=season if source_resolution == "seasonal" else None)
    assert file_path
    meta = DATASETS[dataset_key]
    with open_data_store(file_path, decode_times=True) as ds:
        time_name = detect_time_coord(ds)
        lat_name, lon_name = detect_lat_lon(ds)
        var_name = pick_variable(ds, dataset_key, variable)
        da = _select_product_statistic_dimension(ds[var_name], variable or var_name, keep_dims={time_name, lat_name, lon_name})
        actual_variable_label = str(da.attrs.get("long_name") or da.attrs.get("standard_name") or var_name).replace("_", " ").title()
        variable_label = requested_element_label(dataset_key, variable) or actual_variable_label
        source_units = str(da.attrs.get("units") or meta.get("unit") or "")
        da, unit = convert_dataarray_units(da, meta["family"], var_name)
        selected = _select_time_period(
            da,
            time_name,
            period["years"],
            None if source_resolution in {"annual", "seasonal"} else period["months"],
        )
        count = int(selected.sizes.get(time_name, 1)) if time_name in selected.dims else 1
        if count < 1:
            raise ValueError(f"No values were found for {period['period_label']}.")
        aggregation = "Selected period"
        if time_name in selected.dims:
            if count == 1:
                aggregated = selected.isel({time_name: 0}, drop=True)
            elif meta["family"] == "rainfall":
                if source_resolution in {"annual", "seasonal"}:
                    aggregated = selected.mean(time_name, skipna=True)
                    aggregation = "Mean total across selected years"
                elif period["mode"] == "custom" and len(period["years"]) > 1:
                    try:
                        by_year = selected.groupby(f"{time_name}.year").sum(time_name, skipna=True)
                        aggregated = by_year.mean("year", skipna=True)
                    except Exception:
                        aggregated = selected.sum(time_name, skipna=True) / max(1, len(period["years"]))
                    aggregation = "Mean selected-month total across years"
                else:
                    aggregated = selected.sum(time_name, skipna=True)
                    aggregation = "Selected-period total"
            else:
                aggregated = selected.mean(time_name, skipna=True)
                aggregation = "Selected-period mean"
        else:
            aggregated = selected
        frame = aggregated.to_dataframe(name="value").reset_index()
        if lat_name != "latitude":
            frame = frame.rename(columns={lat_name: "latitude"})
        if lon_name != "longitude":
            frame = frame.rename(columns={lon_name: "longitude"})
        frame = frame[["latitude", "longitude", "value"]].dropna(subset=["value"]).reset_index(drop=True)
    context = {
        "file": file_path.name,
        "source_path": str(file_path),
        "dataset_key": dataset_key,
        "dataset_label": meta["label"],
        "family": meta["family"],
        "variable": variable,
        "actual_variable": var_name,
        "variable_label": variable_label,
        "actual_variable_label": actual_variable_label,
        "unit": unit or source_units,
        "resolution": source_resolution,
        "period_label": period["period_label"],
        "aggregation": aggregation,
        "months": period["months"],
        "years": period["years"],
    }
    return frame, context


def _geometry_polygons(geometry: dict[str, Any]) -> list[np.ndarray]:
    kind = str(geometry.get("type") or "")
    coords = geometry.get("coordinates") or []
    polygons: list[np.ndarray] = []
    if kind == "Polygon":
        for ring in coords:
            arr = np.asarray(ring, dtype=float)
            if arr.ndim == 2 and arr.shape[0] >= 3:
                polygons.append(arr[:, :2])
    elif kind == "MultiPolygon":
        for polygon in coords:
            for ring in polygon:
                arr = np.asarray(ring, dtype=float)
                if arr.ndim == 2 and arr.shape[0] >= 3:
                    polygons.append(arr[:, :2])
    elif kind == "GeometryCollection":
        for item in geometry.get("geometries") or []:
            polygons.extend(_geometry_polygons(item))
    return polygons


@lru_cache(maxsize=12)
def _cached_geojson_polygons(path_text: str, modified_ns: int) -> tuple[tuple[tuple[float, float], ...], ...]:
    path = Path(path_text)
    payload = json.loads(path.read_text(encoding="utf-8"))
    features = payload.get("features", []) if payload.get("type") == "FeatureCollection" else [{"geometry": payload}]
    polygons: list[tuple[tuple[float, float], ...]] = []
    for feature in features:
        for array in _geometry_polygons(feature.get("geometry") or {}):
            polygons.append(tuple((float(x), float(y)) for x, y in array))
    return tuple(polygons)


def _admin_polygons(data_dir: Path, level: int) -> tuple[list[np.ndarray], Path]:
    level = max(1, min(3, int(level)))
    path = Path(data_dir) / "shapefiles" / f"gadm41_TZA_{level}.json"
    if not path.is_file():
        raise FileNotFoundError(f"Administrative boundary file not found: {path}")
    cached = _cached_geojson_polygons(str(path.resolve()), path.stat().st_mtime_ns)
    return [np.asarray(points, dtype=float) for points in cached], path


def _mask_grid_to_tanzania(values: np.ndarray, longitudes: np.ndarray, latitudes: np.ndarray, polygons: list[np.ndarray]) -> np.ndarray:
    from matplotlib.path import Path as MplPath
    xx, yy = np.meshgrid(longitudes, latitudes)
    points = np.column_stack([xx.ravel(), yy.ravel()])
    inside = np.zeros(len(points), dtype=bool)
    for polygon in polygons:
        if len(polygon) >= 3:
            inside |= MplPath(polygon).contains_points(points, radius=1e-9)
    return np.where(inside.reshape(values.shape), values, np.nan)


def _draw_admin_boundaries(ax, polygons: list[np.ndarray], level: int) -> None:
    from matplotlib.collections import LineCollection
    widths = {1: 0.85, 2: 0.48, 3: 0.25}
    alpha = {1: 0.95, 2: 0.8, 3: 0.62}
    lines = LineCollection(polygons, colors="#263746", linewidths=widths.get(level, 0.5), alpha=alpha.get(level, 0.8), zorder=8)
    ax.add_collection(lines)


def _draw_hydrology(ax, *, ocean: bool, lakes: bool, rivers: bool) -> None:
    from matplotlib.lines import Line2D
    legend_items = []
    if ocean:
        ax.set_facecolor("#dff3fb")
        ax.text(40.15, -8.2, "Indian Ocean", fontsize=8.5, color="#347c98", rotation=90, ha="center", va="center", alpha=0.9, zorder=1)
        legend_items.append(Line2D([0], [0], color="#8bc8df", lw=5, label="Ocean"))
    else:
        ax.set_facecolor("white")
    if lakes:
        lake_polygons = {
            "Lake Victoria": [(31.25,-0.90),(33.25,-0.90),(34.15,-1.45),(34.00,-2.35),(32.55,-2.65),(31.45,-2.05)],
            "Lake Tanganyika": [(29.05,-3.25),(29.75,-3.55),(30.45,-5.60),(30.80,-8.35),(30.35,-8.75),(29.45,-6.75)],
            "Lake Nyasa": [(34.30,-9.00),(34.85,-9.25),(35.65,-11.65),(35.15,-11.95),(34.55,-10.65)],
            "Lake Rukwa": [(31.70,-7.10),(32.45,-7.30),(32.45,-8.15),(31.85,-8.20)],
        }
        for name, vertices in lake_polygons.items():
            patch = Polygon(vertices, closed=True, facecolor="#bfe7f5", edgecolor="#4f9ebd", linewidth=0.7, zorder=9)
            ax.add_patch(patch)
        ax.text(32.55, -1.65, "Victoria", fontsize=6.5, color="#327a98", ha="center", zorder=10)
        ax.text(29.90, -6.20, "Tanganyika", fontsize=6.2, color="#327a98", rotation=74, ha="center", zorder=10)
        ax.text(35.05, -10.60, "Nyasa", fontsize=6.2, color="#327a98", rotation=70, ha="center", zorder=10)
        legend_items.append(Line2D([0], [0], color="#4f9ebd", lw=4, label="Major lakes"))
    if rivers:
        river_lines = {
            "Rufiji": [(34.6,-7.6),(36.0,-7.8),(37.2,-8.0),(38.3,-7.7),(39.3,-7.8)],
            "Ruvuma": [(34.8,-10.6),(36.2,-10.8),(37.6,-10.7),(39.2,-10.4)],
            "Pangani": [(36.6,-3.2),(37.0,-3.8),(37.6,-4.4),(38.5,-5.1),(39.1,-5.4)],
            "Wami": [(35.8,-5.8),(36.8,-5.9),(37.8,-6.1),(38.8,-6.2)],
            "Malagarasi": [(30.2,-4.1),(31.0,-4.3),(31.7,-4.9),(32.5,-5.1)],
        }
        for points in river_lines.values():
            arr = np.asarray(points, dtype=float)
            ax.plot(arr[:,0], arr[:,1], color="#2d86b3", linewidth=0.8, alpha=0.9, zorder=10)
        legend_items.append(Line2D([0], [0], color="#2d86b3", lw=1.5, label="Major rivers"))
    if legend_items:
        ax.legend(handles=legend_items, loc="lower left", fontsize=7.5, frameon=True, framealpha=0.9)


def _generate_requested_spatial_map(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    grid, context = _extract_spatial_period(Path(data_dir), params)
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    level = max(1, min(3, int(params.get("map_admin_level") or 1)))
    selected_boundaries, boundary_path = _admin_polygons(Path(data_dir), level)
    national_polygons, _ = _admin_polygons(Path(data_dir), 1)
    admin_label = {1: "Admin 1 — Regions", 2: "Admin 2 — Districts", 3: "Admin 3 — Wards"}[level]
    context.update({
        "administrative_level": level,
        "administrative_level_label": admin_label,
        "boundary_file": boundary_path.name,
    })
    grid = grid[
        grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
        & grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    ].copy()
    if grid.empty:
        raise ValueError("No spatial values were found within the Tanzania map extent.")
    pivot = grid.pivot_table(index="latitude", columns="longitude", values="value", aggfunc="mean").sort_index()
    x = pivot.columns.to_numpy(dtype=float)
    y = pivot.index.to_numpy(dtype=float)
    masked_values = _mask_grid_to_tanzania(pivot.to_numpy(dtype=float), x, y, national_polygons)

    element = element_with_source(dataset_key, variable_display_name(params.get("variable"), context, dataset_key))
    period_label = context["period_label"]
    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"spatial_map_{dataset_key}_{slugify(period_label)}_admin{level}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"

    fig, ax = plt.subplots(figsize=(11.0, 8.6))
    _draw_hydrology(
        ax,
        ocean=_truthy(params.get("show_ocean", True)),
        lakes=_truthy(params.get("show_lakes", True)),
        rivers=_truthy(params.get("show_rivers", True)),
    )
    im = ax.pcolormesh(x, y, masked_values, shading="auto", cmap=plot_cmap_for(dataset_key), zorder=4)
    _draw_admin_boundaries(ax, selected_boundaries, level)
    ax.set_xlim(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
    ax.set_ylim(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    ax.set_xlabel("Longitude")
    ax.set_ylabel("Latitude")
    ax.set_title(f"{element} Spatial Distribution over Tanzania\n{period_label}", fontsize=14, fontweight="bold", pad=14)
    ax.text(0.5, 1.005, f"{context['aggregation']} · {admin_label}", transform=ax.transAxes, ha="center", va="bottom", fontsize=9, color="#4b5d6b")
    apply_plot_grids(ax)
    fig.colorbar(im, ax=ax, shrink=0.82, pad=0.025, label=y_axis_label(element, context.get("unit")))
    _finalize_plot(fig, plot_path)

    export_grid = grid.copy()
    export_grid.insert(0, "period", period_label)
    export_grid.insert(1, "administrative_level", admin_label)
    excel_path, _, _ = _save_data(export_grid, dirs["plots"], stem + "_data", table="plot_products")
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "rows": len(export_grid),
        "context": context,
        "period_label": period_label,
        "summary_cards": [
            {"label": "Period", "value": period_label, "note": context["aggregation"]},
            {"label": "Boundary Level", "value": admin_label, "note": boundary_path.name},
            {"label": "Grid Cells", "value": f"{len(export_grid):,}", "note": "Tanzania land grid cells"},
        ],
    }


def _heatmap_period_definition(params: Dict[str, Any]) -> tuple[list[int], list[int], str]:
    start = int(params.get("start_year") or 1991)
    end = int(params.get("end_year") or start)
    year_mode = str(params.get("heatmap_year_mode") or "range").lower()
    years = _parse_year_numbers(params.get("heatmap_custom_years") if year_mode == "custom" else "", start=start, end=end)
    month_mode = str(params.get("heatmap_month_mode") or "all").lower()
    if month_mode == "season":
        season = str(params.get("heatmap_season") or "MAM").upper()
        months = list(SEASON_DEFINITIONS.get(season, SEASON_DEFINITIONS["MAM"]))
        month_label = f"{season} ({', '.join(_month_names(months))})"
    elif month_mode == "custom":
        months = _parse_month_numbers(params.get("heatmap_custom_months"), default=range(1, 13))
        month_label = ", ".join(_month_names(months))
    else:
        months = list(range(1, 13))
        month_label = "All months (January–December)"
    return months, years, f"{month_label} · {_compact_year_label(years)}"


def _generate_requested_heatmap(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    variable = str(params.get("variable") or "auto")
    if dataset_key == "era5_temperature" and variable in {"all", "all_in_one", "max_min"}:
        variable = "ta"
    if dataset_key == "era5_wind" and variable in {"auto", "wind_speed_direction", ""}:
        variable = "wind_speed"
    months, years, period_label = _heatmap_period_definition(params)
    if not years:
        raise ValueError("Select at least one heat-map year.")
    frame, context = extract_point_series(
        Path(data_dir), dataset_key, "monthly",
        float(params.get("latitude") or 0), float(params.get("longitude") or 0),
        f"{min(years)}-01-01", f"{max(years)}-12-31", variable=variable,
    )
    frame = frame.copy()
    frame["year"] = frame["time"].dt.year.astype(int)
    frame["month"] = frame["time"].dt.month.astype(int)
    selected = frame[frame["year"].isin(years) & frame["month"].isin(months)].copy()
    if selected.empty:
        raise ValueError(f"No monthly values were found for {period_label}.")
    pivot = selected.pivot_table(index="year", columns="month", values="value", aggfunc="mean")
    pivot = pivot.reindex(index=years, columns=months)
    context.update({"period_label": period_label, "months": months, "years": years, "resolution": "monthly"})

    element = element_with_source(dataset_key, variable_display_name(variable, context, dataset_key))
    location = str(params.get("location_name") or "Selected Location")
    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"heatmap_{dataset_key}_{location}_{slugify(period_label)}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"

    fig_height = min(15.0, max(6.0, 4.5 + 0.13 * len(years)))
    fig, ax = plt.subplots(figsize=(12.0, fig_height))
    im = ax.imshow(pivot.to_numpy(dtype=float), aspect="auto", cmap=plot_cmap_for(dataset_key), interpolation="nearest")
    ax.set_xticks(np.arange(len(months)))
    ax.set_xticklabels([pd.Timestamp(2000, month, 1).strftime("%b") for month in months])
    tick_step = max(1, math.ceil(len(years) / 24))
    tick_positions = np.arange(0, len(years), tick_step)
    ax.set_yticks(tick_positions)
    ax.set_yticklabels([str(years[pos]) for pos in tick_positions])
    ax.set_xlabel("Month")
    ax.set_ylabel("Year")
    ax.set_title(f"{element} Monthly–Annual Heat Map for {location}\n{period_label}", fontsize=14, fontweight="bold", pad=14)
    fig.colorbar(im, ax=ax, shrink=0.86, pad=0.02, label=y_axis_label(element, context.get("unit")))
    _finalize_plot(fig, plot_path)

    export_frame = pivot.copy()
    export_frame.columns = [pd.Timestamp(2000, int(month), 1).strftime("January") if False else pd.Timestamp(2000, int(month), 1).strftime("%B") for month in months]
    export_frame.index.name = "Year"
    export_frame = export_frame.reset_index()
    excel_path, _, _ = _save_data(export_frame, dirs["plots"], stem + "_data", table="plot_products")
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "rows": int(selected["value"].notna().sum()),
        "context": context,
        "period_label": period_label,
        "summary_cards": [
            {"label": "Months", "value": str(len(months)), "note": ", ".join(pd.Timestamp(2000, month, 1).strftime("%b") for month in months)},
            {"label": "Years", "value": str(len(years)), "note": _compact_year_label(years)},
            {"label": "Observations", "value": f"{int(selected['value'].notna().sum()):,}", "note": "Monthly values represented"},
        ],
    }


def generate_plot_product(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate only the explicitly requested product.

    Spatial maps and heat maps use their own period-aware, memory-bounded paths.
    Every other product delegates to the established single-product generator.
    """
    plot_type = str(params.get("plot_type") or "time_series")
    if plot_type == "spatial_map":
        return _generate_requested_spatial_map(params, Path(data_dir), Path(export_dir))
    if plot_type == "heatmap":
        return _generate_requested_heatmap(params, Path(data_dir), Path(export_dir))
    return _CDE_REQUEST_ONLY_PLOT_BASE(params, Path(data_dir), Path(export_dir))

# ---------------------------------------------------------------------------
# FINAL SIMPLE-WORKSPACE RULES - 2026-07-17
# ---------------------------------------------------------------------------
# These three indices are intentionally removed from the user interface and
# rejected server-side.  They remain in historical code only for backwards
# compatibility with previously generated files.
_REMOVED_RAINY_SEASON_INDICES = {
    "rainy_season_onset",
    "rainy_season_cessation",
    "length_of_rainy_season",
}
RAINFALL_INDICES = [item for item in RAINFALL_INDICES if item[0] not in _REMOVED_RAINY_SEASON_INDICES]
ALL_INDICES = RAINFALL_INDICES + TEMPERATURE_INDICES + OTHER_INDICES

_CDE_SIMPLE_INDEX_BASE = generate_indices

def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate one requested index and create a chart only when requested/useful."""
    index_type = str(params.get("index_type") or "total_rainfall")
    if index_type in _REMOVED_RAINY_SEASON_INDICES:
        raise ValueError("The selected rainy-season index is no longer available.")
    result = _CDE_SIMPLE_INDEX_BASE(params, Path(data_dir), Path(export_dir))
    include_plot = _truthy(params.get("include_plot", False))
    rows = int(result.get("rows") or 0)
    if not include_plot or rows < 2:
        old_plot = result.pop("plot_path", None)
        if old_plot:
            try:
                path = Path(old_plot)
                if path.exists():
                    path.unlink()
            except Exception:
                pass
    return result

# ---------------------------------------------------------------------------
# 2026-07-17: complete user-facing CDE plot catalogue and variability products
# ---------------------------------------------------------------------------

_CDE_COMPLETE_PLOT_BASE = generate_plot_product
_CDE_COMPLETE_FIND_FILE_BASE = find_file
_CDE_COMPLETE_PLOT_TITLE_BASE = plot_title_for

_HOURLY_TEMPERATURE_STORE = "ERA5_Tanzania_Temperature _2M_Hourly_1940_2026.zarr"
_HOURLY_TEMPERATURE_STORE_ALIASES = (
    _HOURLY_TEMPERATURE_STORE,
    "ERA5_Tanzania_Temperature_2M_Hourly_1940_2026.zarr",
)


def find_file(data_dir: Path, dataset_key: str, resolution: str, season: str | None = None, silent: bool = False) -> Path | None:
    """Resolve stores, with one authoritative source for hourly 2m temperature.

    Hourly mean/minimum/maximum temperature must come only from the consolidated
    1940–2026 2m temperature store. This prevents an older or similarly named
    hourly Zarr store from being selected accidentally.
    """
    if str(dataset_key) == "era5_temperature" and str(resolution).lower() == "hourly":
        root = Path(data_dir).expanduser().resolve()
        stores = list(iter_data_stores(root))
        for expected_name in _HOURLY_TEMPERATURE_STORE_ALIASES:
            match = next((path for path in stores if path.name == expected_name), None)
            if match is not None:
                return match
        if silent:
            return None
        raise FileNotFoundError(
            f"Hourly ERA5 temperature requires the consolidated 1940–2026 store {_HOURLY_TEMPERATURE_STORE} under {root}."
        )
    return _CDE_COMPLETE_FIND_FILE_BASE(data_dir, dataset_key, resolution, season=season, silent=silent)


def plot_title_for(plot_type: str, element: str, location: str, resolution: str, start: str | None = None, end: str | None = None, map_date: str | None = None) -> str:
    """Create explicit titles that always name the weather element and period."""
    element = str(element or "Weather Element").replace("Rainfall", "Precipitation")
    loc = location or "Selected Location"
    period = plot_period_text(resolution, start, end, map_date)
    titles = {
        "multi_line": f"Comparison of {element} for {loc} ({period})",
        "monthly_climatology": f"Monthly Climatology of {element} for {loc} ({period})",
        "seasonal_profile": f"Seasonal Profile of {element} for {loc} ({period})",
        "standardized_anomaly": f"Standardized Anomaly (Z-Score) of {element} for {loc} ({period})",
        "mean_std_band": f"Monthly Mean ± Standard Deviation of {element} for {loc} ({period})",
        "std_error_bars": f"Monthly Mean with Standard Deviation Error Bars: {element} for {loc} ({period})",
        "standard_deviation": f"Monthly Standard Deviation of {element} for {loc} ({period})",
        "coefficient_variation": f"Monthly Coefficient of Variation of {element} for {loc} ({period})",
        "box": f"Monthly Distribution Box Plot of {element} for {loc} ({period})",
        "extreme_value": f"Annual Extreme Values of {element} for {loc} ({period})",
        "scatter": f"Relationship between {element} for {loc} ({period})",
        "spatial_std_map": f"Spatial Standard Deviation of {element} over Tanzania ({period})",
        "spatial_cv_map": f"Spatial Coefficient of Variation of {element} over Tanzania ({period})",
    }
    return titles.get(plot_type) or _CDE_COMPLETE_PLOT_TITLE_BASE(plot_type, element, loc, resolution, start, end, map_date)


def _monthly_variability_frame(df: pd.DataFrame) -> pd.DataFrame:
    tmp = df.copy()
    tmp["month"] = pd.to_datetime(tmp["time"]).dt.month
    out = tmp.groupby("month", as_index=False)["value"].agg(
        observations="count",
        mean="mean",
        standard_deviation=lambda values: pd.to_numeric(values, errors="coerce").std(ddof=0),
        minimum="min",
        maximum="max",
    )
    out["coefficient_of_variation_percent"] = np.where(
        out["mean"].abs() > 1e-12,
        out["standard_deviation"] / out["mean"].abs() * 100.0,
        np.nan,
    )
    out["month_name"] = out["month"].map(lambda month: pd.Timestamp(2000, int(month), 1).strftime("%b"))
    return out


def _annual_values_for_plot(df: pd.DataFrame, family: str) -> pd.DataFrame:
    tmp = df.copy()
    tmp["year"] = pd.to_datetime(tmp["time"]).dt.year
    aggregation = "sum" if family == "rainfall" else "mean"
    return tmp.groupby("year", as_index=False)["value"].agg(aggregation)


def _seasonal_profile_frame(df: pd.DataFrame, family: str) -> pd.DataFrame:
    tmp = df.copy()
    tmp["time"] = pd.to_datetime(tmp["time"])
    tmp["year"] = tmp["time"].dt.year
    tmp["month"] = tmp["time"].dt.month
    rows: list[dict[str, Any]] = []
    for season_name, months in [("DJF", [12, 1, 2]), ("MAM", [3, 4, 5]), ("JJA", [6, 7, 8]), ("SON", [9, 10, 11])]:
        selected = tmp[tmp["month"].isin(months)].copy()
        if selected.empty:
            continue
        selected["season_year"] = selected["time"].map(lambda value: season_year(pd.Timestamp(value), months))
        annual = selected.groupby("season_year")["value"].agg("sum" if family == "rainfall" else "mean")
        rows.append({
            "season": season_name,
            "mean": annual.mean(),
            "standard_deviation": annual.std(ddof=0),
            "minimum": annual.min(),
            "maximum": annual.max(),
            "years": int(annual.count()),
        })
    return pd.DataFrame(rows)


def _extract_multiple_variables(params: Dict[str, Any], data_dir: Path) -> tuple[pd.DataFrame, list[tuple[str, str, str]], Dict[str, Any]]:
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    resolution = str(params.get("resolution") or "monthly")
    season = str(params.get("season") or "").strip().upper() or None
    variables = [str(value) for value in params.get("variables", []) if str(value).strip()]
    if len(variables) < 2:
        raise ValueError("Select at least two weather elements for this plot.")
    merged: pd.DataFrame | None = None
    descriptors: list[tuple[str, str, str]] = []
    first_context: Dict[str, Any] = {}
    for variable in variables:
        frame, context = extract_point_series(
            data_dir, dataset_key, resolution,
            float(params.get("latitude")), float(params.get("longitude")),
            str(params.get("start_date") or ""), str(params.get("end_date") or ""),
            variable=variable, season=season,
        )
        label = variable_display_name(variable, context, dataset_key)
        column = slugify(label).lower()
        part = frame[["time", "value"]].rename(columns={"value": column})
        merged = part if merged is None else merged.merge(part, on="time", how="inner")
        descriptors.append((column, label, str(context.get("unit") or "")))
        if not first_context:
            first_context = dict(context)
    if merged is None or merged.empty:
        raise ValueError("No overlapping observations were found for the selected weather elements.")
    return merged.sort_values("time"), descriptors, first_context


def _generate_multi_variable_plot(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    plot_type = str(params.get("plot_type") or "multi_line")
    dataset_key = str(params.get("dataset") or "era5_temperature")
    resolution = str(params.get("resolution") or "monthly")
    location = str(params.get("location_name") or "Selected Location")
    start = str(params.get("start_date") or "")
    end = str(params.get("end_date") or "")
    frame, descriptors, context = _extract_multiple_variables(params, data_dir)
    labels = [label for _, label, _ in descriptors]
    units = {unit for _, _, unit in descriptors if unit}
    element_text = " and ".join(labels)
    dirs = ensure_output_dirs(export_dir)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"{plot_type}_{dataset_key}_{location}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    fig, ax = plt.subplots(figsize=(11.2, 6.2))
    if plot_type == "scatter":
        x_col, x_label, x_unit = descriptors[0]
        y_col, y_label, y_unit = descriptors[1]
        points = frame[[x_col, y_col]].dropna()
        ax.scatter(points[x_col], points[y_col], alpha=0.65, s=24)
        correlation = float(points[x_col].corr(points[y_col])) if len(points) > 1 else np.nan
        if len(points) > 1 and points[x_col].nunique() > 1:
            slope, intercept = np.polyfit(points[x_col], points[y_col], 1)
            x_values = np.linspace(points[x_col].min(), points[x_col].max(), 100)
            ax.plot(x_values, slope * x_values + intercept, linestyle="--", linewidth=1.4, label=f"Linear fit · r={correlation:.2f}")
            ax.legend()
        ax.set_xlabel(y_axis_label(x_label, x_unit))
        ax.set_ylabel(y_axis_label(y_label, y_unit))
        ax.set_title(plot_title_for("scatter", f"{element_with_source(dataset_key, x_label)} and {y_label}", location, resolution, start, end))
        export_frame = points
    else:
        for column, label, _unit in descriptors:
            ax.plot(frame["time"], frame[column], linewidth=1.5, label=label)
        ax.set_xlabel("Time")
        ax.set_ylabel("Value" if len(units) != 1 else f"Value ({next(iter(units))})")
        ax.set_title(plot_title_for("multi_line", element_with_source(dataset_key, element_text), location, resolution, start, end))
        ax.legend(ncol=2)
        export_frame = frame
    apply_plot_grids(ax)
    _finalize_plot(fig, plot_path)
    excel_path, _, _ = _save_data(export_frame, dirs["plots"], stem + "_data", table="plot_products")
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "rows": len(export_frame),
        "context": context,
        "summary_cards": [
            {"label": "Weather Elements", "value": str(len(descriptors)), "note": ", ".join(labels)},
            {"label": "Observations", "value": f"{len(export_frame):,}", "note": "Overlapping records"},
            {"label": "Resolution", "value": resolution_text(resolution), "note": f"{start[:4]}–{end[:4]}"},
        ],
    }


def _generate_variability_plot(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    plot_type = str(params.get("plot_type") or "standard_deviation")
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    variable = str(params.get("variable") or "auto")
    resolution = str(params.get("resolution") or "monthly")
    season = str(params.get("season") or "").strip().upper() or None
    location = str(params.get("location_name") or "Selected Location")
    start = str(params.get("start_date") or "")
    end = str(params.get("end_date") or "")
    frame, context = extract_point_series(
        data_dir, dataset_key, resolution,
        float(params.get("latitude")), float(params.get("longitude")),
        start, end, variable=variable, season=season,
    )
    if frame.empty:
        raise ValueError("No values were found for the selected period.")
    family = str(context.get("family") or DATASETS[dataset_key].get("family") or "")
    unit = str(context.get("unit") or DATASETS[dataset_key].get("unit") or "")
    element = element_with_source(dataset_key, variable_display_name(variable, context, dataset_key))
    dirs = ensure_output_dirs(export_dir)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"{plot_type}_{dataset_key}_{variable}_{location}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    fig, ax = plt.subplots(figsize=(11.0, 6.0))

    if plot_type == "standardized_anomaly":
        export_frame = _annual_values_for_plot(frame, family)
        baseline_start = int(params.get("baseline_start") or 1991)
        baseline_end = int(params.get("baseline_end") or 2020)
        baseline = export_frame[export_frame["year"].between(baseline_start, baseline_end)]
        if baseline.empty:
            baseline = export_frame
        mean = float(baseline["value"].mean())
        sd = float(baseline["value"].std(ddof=0))
        if not np.isfinite(sd) or sd == 0:
            raise ValueError("Standardized anomaly cannot be calculated because the baseline standard deviation is zero or unavailable.")
        export_frame["standardized_anomaly_z_score"] = (export_frame["value"] - mean) / sd
        values = export_frame["standardized_anomaly_z_score"]
        ax.bar(export_frame["year"], values)
        ax.axhline(0, color="black", linewidth=1)
        ax.axhline(1, linestyle="--", linewidth=0.9)
        ax.axhline(-1, linestyle="--", linewidth=0.9)
        ax.set_xlabel("Year")
        ax.set_ylabel("Standardized anomaly (z-score)")
        ax.set_title(plot_title_for(plot_type, element, location, resolution, start, end))
        context.update({"baseline_mean": mean, "baseline_standard_deviation": sd, "baseline_period": f"{baseline_start}–{baseline_end}"})
    elif plot_type == "seasonal_profile":
        export_frame = _seasonal_profile_frame(frame, family)
        ax.bar(export_frame["season"], export_frame["mean"], yerr=export_frame["standard_deviation"], capsize=5)
        ax.set_xlabel("Season")
        ax.set_ylabel(y_axis_label(element, unit))
        ax.set_title(plot_title_for(plot_type, element, location, resolution, start, end))
    else:
        export_frame = _monthly_variability_frame(frame)
        x = np.arange(len(export_frame))
        labels = export_frame["month_name"]
        means = export_frame["mean"].to_numpy(dtype=float)
        sd = export_frame["standard_deviation"].to_numpy(dtype=float)
        if plot_type == "mean_std_band":
            ax.plot(x, means, marker="o", linewidth=2, label="Mean")
            ax.fill_between(x, means - sd, means + sd, alpha=0.22, label="Mean ± 1 SD")
            ax.set_ylabel(y_axis_label(element, unit))
            ax.legend()
        elif plot_type == "std_error_bars":
            ax.bar(x, means, yerr=sd, capsize=5)
            ax.set_ylabel(y_axis_label(element, unit))
        elif plot_type == "standard_deviation":
            ax.plot(x, sd, marker="o", linewidth=2)
            ax.set_ylabel(f"Standard deviation ({unit})" if unit else "Standard deviation")
        elif plot_type == "coefficient_variation":
            ax.plot(x, export_frame["coefficient_of_variation_percent"], marker="o", linewidth=2)
            ax.set_ylabel("Coefficient of variation (%)")
        else:
            raise ValueError("Unsupported variability plot.")
        ax.set_xticks(x)
        ax.set_xticklabels(labels)
        ax.set_xlabel("Month")
        ax.set_title(plot_title_for(plot_type, element, location, resolution, start, end))
    apply_plot_grids(ax)
    _finalize_plot(fig, plot_path)
    excel_path, _, _ = _save_data(export_frame, dirs["plots"], stem + "_data", table="plot_products")
    values = pd.to_numeric(frame["value"], errors="coerce").dropna()
    overall_sd = float(values.std(ddof=0)) if not values.empty else np.nan
    overall_cv = float(overall_sd / abs(values.mean()) * 100.0) if not values.empty and abs(values.mean()) > 1e-12 else np.nan
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "rows": len(export_frame),
        "context": context,
        "summary_cards": [
            {"label": "Standard Deviation", "value": f"{overall_sd:,.1f} {unit}".strip() if np.isfinite(overall_sd) else "N/A", "note": "Selected-period variability"},
            {"label": "Coefficient of Variation", "value": f"{overall_cv:,.1f}%" if np.isfinite(overall_cv) else "N/A", "note": "Standard deviation relative to the mean"},
            {"label": "Observations", "value": f"{len(values):,}", "note": resolution_text(resolution)},
        ],
    }


def _extract_spatial_variability(data_dir: Path, params: Dict[str, Any], statistic: str) -> tuple[pd.DataFrame, Dict[str, Any]]:
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    variable = str(params.get("variable") or "auto")
    if dataset_key == "era5_wind" and variable in {"auto", "wind_speed_direction", ""}:
        variable = "wind_speed"
    period = _spatial_period_definition(params)
    preferred = period["preferred_resolution"]
    season = period["season"] if preferred == "seasonal" else None
    try:
        file_path = find_file(data_dir, dataset_key, preferred, season=season)
        source_resolution = preferred
    except Exception:
        source_resolution = str(params.get("resolution") or "monthly")
        file_path = find_file(data_dir, dataset_key, source_resolution, season=season if source_resolution == "seasonal" else None)
    assert file_path
    meta = DATASETS[dataset_key]
    with open_data_store(file_path, decode_times=True) as ds:
        time_name = detect_time_coord(ds)
        lat_name, lon_name = detect_lat_lon(ds)
        var_name = pick_variable(ds, dataset_key, variable)
        da = _select_product_statistic_dimension(ds[var_name], variable or var_name, keep_dims={time_name, lat_name, lon_name})
        da, unit = convert_dataarray_units(da, meta["family"], var_name)
        selected = _select_time_period(da, time_name, period["years"], None if source_resolution in {"annual", "seasonal"} else period["months"])
        count = int(selected.sizes.get(time_name, 1)) if time_name in selected.dims else 1
        if count < 2:
            raise ValueError("At least two time steps are required for a spatial variability map.")
        sd = selected.std(time_name, skipna=True)
        if statistic == "cv":
            mean = selected.mean(time_name, skipna=True)
            aggregated = xr.where(abs(mean) > 1e-12, sd / abs(mean) * 100.0, np.nan)
            out_unit = "%"
            aggregation = "Coefficient of variation across selected time steps"
        else:
            aggregated = sd
            out_unit = unit
            aggregation = "Standard deviation across selected time steps"
        frame = aggregated.to_dataframe(name="value").reset_index()
        if lat_name != "latitude":
            frame = frame.rename(columns={lat_name: "latitude"})
        if lon_name != "longitude":
            frame = frame.rename(columns={lon_name: "longitude"})
        frame = frame[["latitude", "longitude", "value"]].dropna(subset=["value"]).reset_index(drop=True)
    return frame, {
        "file": file_path.name,
        "source_path": str(file_path),
        "dataset_key": dataset_key,
        "dataset_label": meta["label"],
        "family": meta["family"],
        "variable": var_name,
        "unit": out_unit,
        "resolution": source_resolution,
        "period_label": period["period_label"],
        "aggregation": aggregation,
        "selected_time_steps": count,
    }


def _generate_spatial_variability_map(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    plot_type = str(params.get("plot_type") or "spatial_std_map")
    statistic = "cv" if plot_type == "spatial_cv_map" else "std"
    grid, context = _extract_spatial_variability(data_dir, params, statistic)
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    level = max(1, min(3, int(params.get("map_admin_level") or 1)))
    selected_boundaries, boundary_path = _admin_polygons(data_dir, level)
    national_polygons, _ = _admin_polygons(data_dir, 1)
    admin_label = {1: "Admin 1 — Regions", 2: "Admin 2 — Districts", 3: "Admin 3 — Wards"}[level]
    grid = grid[
        grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
        & grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    ].copy()
    if grid.empty:
        raise ValueError("No spatial values were found within Tanzania.")
    pivot = grid.pivot_table(index="latitude", columns="longitude", values="value", aggfunc="mean").sort_index()
    x = pivot.columns.to_numpy(dtype=float)
    y = pivot.index.to_numpy(dtype=float)
    values = _mask_grid_to_tanzania(pivot.to_numpy(dtype=float), x, y, national_polygons)
    element = element_with_source(dataset_key, variable_display_name(params.get("variable"), context, dataset_key))
    period_label = context["period_label"]
    dirs = ensure_output_dirs(export_dir)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"{plot_type}_{dataset_key}_{slugify(period_label)}_admin{level}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    fig, ax = plt.subplots(figsize=(11.0, 8.6))
    _draw_hydrology(ax, ocean=_truthy(params.get("show_ocean", True)), lakes=_truthy(params.get("show_lakes", True)), rivers=_truthy(params.get("show_rivers", True)))
    im = ax.pcolormesh(x, y, values, shading="auto", cmap=plot_cmap_for(dataset_key), zorder=4)
    _draw_admin_boundaries(ax, selected_boundaries, level)
    ax.set_xlim(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
    ax.set_ylim(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    ax.set_xlabel("Longitude")
    ax.set_ylabel("Latitude")
    variability_name = "Coefficient of Variation" if statistic == "cv" else "Standard Deviation"
    ax.set_title(
        f"Spatial {variability_name} of {element} over Tanzania ({period_label})",
        fontsize=14,
        fontweight="bold",
        pad=14,
    )
    ax.text(0.5, 1.005, f"{context['aggregation']} · {admin_label}", transform=ax.transAxes, ha="center", va="bottom", fontsize=9)
    apply_plot_grids(ax)
    fig.colorbar(im, ax=ax, shrink=0.82, pad=0.025, label=y_axis_label(element if statistic == "std" else "Coefficient of Variation", context.get("unit")))
    _finalize_plot(fig, plot_path)
    export_grid = grid.copy()
    export_grid.insert(0, "period", period_label)
    export_grid.insert(1, "statistic", "Coefficient of Variation" if statistic == "cv" else "Standard Deviation")
    export_grid.insert(2, "administrative_level", admin_label)
    excel_path, _, _ = _save_data(export_grid, dirs["plots"], stem + "_data", table="plot_products")
    context.update({"administrative_level_label": admin_label, "boundary_file": boundary_path.name})
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "rows": len(export_grid),
        "context": context,
        "period_label": period_label,
        "summary_cards": [
            {"label": "Statistic", "value": "Coefficient of Variation" if statistic == "cv" else "Standard Deviation", "note": context["aggregation"]},
            {"label": "Period", "value": period_label, "note": f"{context['selected_time_steps']} time steps"},
            {"label": "Boundary Level", "value": admin_label, "note": boundary_path.name},
        ],
    }


def _generate_cumulative_area(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    if DATASETS.get(dataset_key, {}).get("family") != "rainfall":
        raise ValueError("The accumulated area plot is available only for precipitation datasets.")
    variable = str(params.get("variable") or "auto")
    resolution = str(params.get("resolution") or "monthly")
    season = str(params.get("season") or "").strip().upper() or None
    location = str(params.get("location_name") or "Selected Location")
    start = str(params.get("start_date") or "")
    end = str(params.get("end_date") or "")
    frame, context = extract_point_series(data_dir, dataset_key, resolution, float(params.get("latitude")), float(params.get("longitude")), start, end, variable=variable, season=season)
    frame = frame.sort_values("time").copy()
    frame["cumulative_precipitation"] = pd.to_numeric(frame["value"], errors="coerce").fillna(0).cumsum()
    element = element_with_source(dataset_key, variable_display_name(variable, context, dataset_key))
    dirs = ensure_output_dirs(export_dir)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"area_{dataset_key}_{location}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    fig, ax = plt.subplots(figsize=(11.0, 6.0))
    ax.fill_between(frame["time"], frame["cumulative_precipitation"], alpha=0.25)
    ax.plot(frame["time"], frame["cumulative_precipitation"], linewidth=1.6)
    ax.set_xlabel("Time")
    ax.set_ylabel("Cumulative precipitation (mm)")
    ax.set_title(f"Cumulative {element} for {location} ({plot_period_text(resolution, start, end)})")
    apply_plot_grids(ax)
    _finalize_plot(fig, plot_path)
    excel_path, _, _ = _save_data(frame, dirs["plots"], stem + "_data", table="plot_products")
    return {"plot_path": plot_path, "excel_path": excel_path, "rows": len(frame), "context": context}


def generate_plot_product(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate one relevant visual product from the complete CDE catalogue."""
    plot_type = str(params.get("plot_type") or "time_series")
    data_dir = Path(data_dir)
    export_dir = Path(export_dir)
    if plot_type in {"spatial_std_map", "spatial_cv_map"}:
        return _generate_spatial_variability_map(params, data_dir, export_dir)
    if plot_type in {"multi_line", "scatter"}:
        return _generate_multi_variable_plot(params, data_dir, export_dir)
    if plot_type in {"mean_std_band", "std_error_bars", "standard_deviation", "coefficient_variation", "standardized_anomaly", "seasonal_profile"}:
        return _generate_variability_plot(params, data_dir, export_dir)
    if plot_type == "area":
        return _generate_cumulative_area(params, data_dir, export_dir)
    if plot_type == "wind_rose":
        # The wind-rose generator reads speed and direction together; run it once.
        single = dict(params)
        single["variable"] = "wind_speed"
        return _CDE_COMPLETE_PLOT_BASE(single, data_dir, export_dir)
    return _CDE_COMPLETE_PLOT_BASE(params, data_dir, export_dir)

# ---------------------------------------------------------------------------
# 2026-07-17: climate-index time scales, custom periods and mandatory plots
# ---------------------------------------------------------------------------
# ERA5 2m temperature is intentionally exposed as one raw hourly dataset.
# Daily, monthly, annual and seasonal temperature indices are calculated from
# that authoritative hourly source rather than selecting prepared temperature
# stores with different statistics.
DATASETS["era5_temperature"].update({
    "label": "ERA5 2m Temperature",
    "resolutions": {"hourly": "Hourly"},
    "keywords": [
        "ERA5_Tanzania_Temperature_2M_Hourly_1940_2026",
        "ERA5_Tanzania_Temperature _2M_Hourly_1940_2026",
    ],
    "variables": ["ta", "t2m", "temperature_2m", "2m_temperature", "temperature"],
})

INDEX_RESOLUTION_LABELS = {
    "hourly": "Hourly",
    "daily": "Daily",
    "monthly": "Monthly",
    "annual": "Annual",
    "seasonal": "Seasonal / Custom Months",
}

INDEX_RESOLUTION_RULES: dict[str, list[str]] = {
    "total_rainfall": ["hourly", "daily", "monthly", "annual", "seasonal"],
    "number_wet_days": ["daily", "monthly", "annual", "seasonal"],
    "number_dry_days": ["daily", "monthly", "annual", "seasonal"],
    "consecutive_dry_days": ["daily", "monthly", "annual", "seasonal"],
    "consecutive_wet_days": ["daily", "monthly", "annual", "seasonal"],
    "wet_spell_length": ["daily", "monthly", "annual", "seasonal"],
    "dry_spell_length": ["daily", "monthly", "annual", "seasonal"],
    "max_1day_rainfall": ["daily", "monthly", "annual", "seasonal"],
    "max_5day_rainfall": ["daily", "monthly", "annual", "seasonal"],
    "heavy_rainfall_days": ["daily", "monthly", "annual", "seasonal"],
    "very_heavy_rainfall_days": ["daily", "monthly", "annual", "seasonal"],
    "sdii": ["daily", "monthly", "annual", "seasonal"],
    "rainfall_anomaly": ["daily", "monthly", "annual", "seasonal"],
    "mean_temperature": ["hourly", "daily", "monthly", "annual", "seasonal"],
    "maximum_temperature": ["daily", "monthly", "annual", "seasonal"],
    "minimum_temperature": ["daily", "monthly", "annual", "seasonal"],
    "temperature_anomaly": ["hourly", "daily", "monthly", "annual", "seasonal"],
    "hot_days": ["daily", "monthly", "annual", "seasonal"],
    "hot_nights": ["daily", "monthly", "annual", "seasonal"],
    "cold_days": ["daily", "monthly", "annual", "seasonal"],
    "cold_nights": ["daily", "monthly", "annual", "seasonal"],
    "heat_index": ["hourly", "daily", "monthly", "annual", "seasonal"],
    "dtr": ["daily", "monthly", "annual", "seasonal"],
    "relative_humidity_index": ["hourly", "daily", "monthly", "annual", "seasonal"],
    "soil_moisture_index": ["hourly", "daily", "monthly", "annual", "seasonal"],
    "soil_moisture_anomaly": ["hourly", "daily", "monthly", "annual", "seasonal"],
    "wind_speed_index": ["hourly", "daily", "monthly", "annual", "seasonal"],
}

INDEX_PLOT_TYPES = [
    ("auto", "Automatic — best plot for the selected time scale"),
    ("line", "Time-Series Line Plot"),
    ("bar", "Period Comparison Bar Chart"),
    ("heatmap", "Time Heat Map"),
]

INDEX_PLOT_RULES: dict[str, list[str]] = {
    "hourly": ["auto", "line", "heatmap"],
    "daily": ["auto", "line", "bar", "heatmap"],
    "monthly": ["auto", "line", "bar", "heatmap"],
    "annual": ["auto", "line", "bar"],
    "seasonal": ["auto", "line", "bar"],
}

_INDEX_LABELS = dict(RAINFALL_INDICES + TEMPERATURE_INDICES + OTHER_INDICES)
_INDEX_COUNT_TYPES = {
    "number_wet_days", "number_dry_days", "heavy_rainfall_days",
    "very_heavy_rainfall_days", "hot_days", "hot_nights", "cold_days", "cold_nights",
}
_INDEX_RUN_TYPES = {"consecutive_dry_days", "consecutive_wet_days", "wet_spell_length", "dry_spell_length"}


def _index_parse_years(value: Any, start_year: int, end_year: int) -> list[int]:
    return _parse_year_numbers(value, start=start_year, end=end_year)


def _index_selected_months(params: Dict[str, Any]) -> tuple[str, list[int]]:
    selection = str(params.get("season") or params.get("index_season") or "ANNUAL").strip().upper()
    if selection in {"", "ANNUAL", "ALL"}:
        return "All months", list(range(1, 13))
    if selection == "CUSTOM":
        months = _parse_month_numbers(params.get("custom_months"), default=range(1, 13))
        return "Custom months " + ", ".join(_month_names(months)), months
    months = list(SEASON_DEFINITIONS.get(selection, []))
    if not months:
        months = _parse_month_numbers(params.get("custom_months"), default=range(1, 13))
        return "Custom months " + ", ".join(_month_names(months)), months
    return f"{selection} ({', '.join(_month_names(months))})", months


def _index_filter_time(frame: pd.DataFrame, params: Dict[str, Any]) -> tuple[pd.DataFrame, list[int], list[int], str]:
    out = frame.copy()
    out["time"] = pd.to_datetime(out["time"])
    start_year = int(params.get("start_year") or pd.Timestamp(out["time"].min()).year)
    end_year = int(params.get("end_year") or pd.Timestamp(out["time"].max()).year)
    years = _index_parse_years(params.get("index_custom_years"), start_year, end_year)
    season_label, months = _index_selected_months(params)
    out = out[out["time"].dt.year.isin(years) & out["time"].dt.month.isin(months)].copy()
    start_hour = max(0, min(23, int(params.get("index_start_hour") or 0)))
    end_hour = max(0, min(23, int(params.get("index_end_hour") or 23)))
    if start_hour > end_hour:
        start_hour, end_hour = end_hour, start_hour
    if str(params.get("index_resolution") or "annual").lower() == "hourly":
        out = out[out["time"].dt.hour.between(start_hour, end_hour)].copy()
    if out.empty:
        raise ValueError("No observations were found for the selected dates, years, months and hours.")
    return out.sort_values("time"), years, months, season_label


def _index_period_key(frame: pd.DataFrame, resolution: str, months: list[int]) -> pd.DataFrame:
    out = frame.copy()
    out["time"] = pd.to_datetime(out["time"])
    if resolution == "hourly":
        out["period"] = out["time"].dt.floor("h")
    elif resolution == "daily":
        out["period"] = out["time"].dt.floor("D")
    elif resolution == "monthly":
        out["period"] = out["time"].dt.to_period("M").dt.to_timestamp()
    elif resolution == "annual":
        out["period"] = pd.to_datetime(out["time"].dt.year.astype(str) + "-01-01")
    else:
        # December-led seasons are attached to the following year.
        crosses_year = 12 in months and any(month < 12 for month in months)
        out["season_year"] = out["time"].dt.year + ((out["time"].dt.month == 12) & crosses_year).astype(int)
        out["period"] = pd.to_datetime(out["season_year"].astype(str) + "-01-01")
    return out


def _index_group_scalar(frame: pd.DataFrame, resolution: str, months: list[int], agg: str) -> pd.DataFrame:
    keyed = _index_period_key(frame, resolution, months)
    if agg == "sum":
        grouped = keyed.groupby("period", as_index=False)["value"].sum(min_count=1)
    elif agg == "max":
        grouped = keyed.groupby("period", as_index=False)["value"].max()
    elif agg == "min":
        grouped = keyed.groupby("period", as_index=False)["value"].min()
    else:
        grouped = keyed.groupby("period", as_index=False)["value"].mean()
    return grouped.rename(columns={"period": "time"})


def _index_daily_temperature(hourly: pd.DataFrame) -> pd.DataFrame:
    tmp = hourly.copy()
    tmp["date"] = pd.to_datetime(tmp["time"]).dt.floor("D")
    return tmp.groupby("date", as_index=False)["value"].agg(
        daily_mean="mean", daily_max="max", daily_min="min"
    ).rename(columns={"date": "time"})


def _index_run_lengths(flags: pd.Series) -> pd.Series:
    values = flags.fillna(False).astype(bool).to_numpy()
    result = np.zeros(len(values), dtype=int)
    current = 0
    for idx, flag in enumerate(values):
        current = current + 1 if flag else 0
        result[idx] = current
    return pd.Series(result, index=flags.index)


def _index_climatology_anomaly(values: pd.DataFrame, resolution: str, baseline_start: int, baseline_end: int, standardized: bool = False) -> pd.DataFrame:
    out = values.copy()
    out["time"] = pd.to_datetime(out["time"])
    baseline = out[out["time"].dt.year.between(baseline_start, baseline_end)].copy()
    if baseline.empty:
        baseline = out.copy()
    if resolution == "hourly":
        key_cols = [out["time"].dt.month.rename("month"), out["time"].dt.day.rename("day"), out["time"].dt.hour.rename("hour")]
        base_keys = [baseline["time"].dt.month.rename("month"), baseline["time"].dt.day.rename("day"), baseline["time"].dt.hour.rename("hour")]
        out_keys = pd.concat(key_cols, axis=1)
        base_work = pd.concat([baseline.reset_index(drop=True), pd.concat(base_keys, axis=1).reset_index(drop=True)], axis=1)
        stats = base_work.groupby(["month", "day", "hour"])["value"].agg(["mean", "std"]).reset_index()
        joined = pd.concat([out.reset_index(drop=True), out_keys.reset_index(drop=True)], axis=1).merge(stats, on=["month", "day", "hour"], how="left")
    elif resolution == "daily":
        out_keys = pd.DataFrame({"day_of_year": out["time"].dt.dayofyear})
        base_work = baseline.copy(); base_work["day_of_year"] = base_work["time"].dt.dayofyear
        stats = base_work.groupby("day_of_year")["value"].agg(["mean", "std"]).reset_index()
        joined = pd.concat([out.reset_index(drop=True), out_keys.reset_index(drop=True)], axis=1).merge(stats, on="day_of_year", how="left")
    elif resolution == "monthly":
        out_keys = pd.DataFrame({"month": out["time"].dt.month})
        base_work = baseline.copy(); base_work["month"] = base_work["time"].dt.month
        stats = base_work.groupby("month")["value"].agg(["mean", "std"]).reset_index()
        joined = pd.concat([out.reset_index(drop=True), out_keys.reset_index(drop=True)], axis=1).merge(stats, on="month", how="left")
    else:
        mean_value = float(pd.to_numeric(baseline["value"], errors="coerce").mean())
        std_value = float(pd.to_numeric(baseline["value"], errors="coerce").std(ddof=0))
        joined = out.copy(); joined["mean"] = mean_value; joined["std"] = std_value
    anomaly = pd.to_numeric(joined["value"], errors="coerce") - pd.to_numeric(joined["mean"], errors="coerce")
    if standardized:
        denom = pd.to_numeric(joined["std"], errors="coerce").replace(0, np.nan)
        joined["value"] = anomaly / denom
    else:
        joined["value"] = anomaly
    return joined[["time", "value"]]


def _heat_index_celsius(temp_c: pd.Series, rh: pd.Series) -> pd.Series:
    """NOAA Rothfusz heat-index approximation, returned in degrees Celsius."""
    t_f = pd.to_numeric(temp_c, errors="coerce") * 9.0 / 5.0 + 32.0
    r = pd.to_numeric(rh, errors="coerce")
    hi_f = (
        -42.379 + 2.04901523 * t_f + 10.14333127 * r
        - 0.22475541 * t_f * r - 0.00683783 * t_f.pow(2)
        - 0.05481717 * r.pow(2) + 0.00122874 * t_f.pow(2) * r
        + 0.00085282 * t_f * r.pow(2) - 0.00000199 * t_f.pow(2) * r.pow(2)
    )
    simple_f = 0.5 * (t_f + 61.0 + ((t_f - 68.0) * 1.2) + (r * 0.094))
    hi_f = hi_f.where((t_f >= 80.0) & (r >= 40.0), simple_f)
    return (hi_f - 32.0) * 5.0 / 9.0


def _index_source_frame(params: Dict[str, Any], data_dir: Path, index_type: str, resolution: str) -> tuple[pd.DataFrame, Dict[str, Any], str, str]:
    start_date = str(params.get("index_start_date") or params.get("start_date") or "")
    end_date = str(params.get("index_end_date") or params.get("end_date") or "")
    # Anomaly and standardized-index products need the baseline observations as
    # well as the requested display period. Expand the lazy source slice only
    # to the required baseline bounds; final output is filtered afterwards.
    if index_type in {"rainfall_anomaly", "temperature_anomaly", "soil_moisture_index", "soil_moisture_anomaly"}:
        baseline_start = int(params.get("baseline_start") or 1991)
        baseline_end = int(params.get("baseline_end") or 2020)
        try:
            start_date = min(pd.Timestamp(start_date), pd.Timestamp(f"{baseline_start}-01-01")).strftime("%Y-%m-%d")
            end_date = max(pd.Timestamp(end_date), pd.Timestamp(f"{baseline_end}-12-31 23:59:59")).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            start_date = f"{baseline_start}-01-01"
            end_date = f"{baseline_end}-12-31 23:59:59"
    latitude = float(params.get("latitude")); longitude = float(params.get("longitude"))
    rainfall = index_type in dict(RAINFALL_INDICES)
    temperature = index_type in dict(TEMPERATURE_INDICES)
    if rainfall:
        dataset_key = str(params.get("dataset") or "chirps_rainfall")
        if resolution == "hourly" and dataset_key != "era5_total_precipitation":
            raise ValueError("Hourly precipitation indices require ERA5 Precipitation. CHIRPS begins at daily resolution.")
        source_resolution = "hourly" if resolution == "hourly" else "daily"
        variable = "tp" if dataset_key == "era5_total_precipitation" else "precip"
    elif temperature:
        dataset_key = "era5_temperature"
        source_resolution = "hourly"
        variable = "ta"
    elif index_type == "relative_humidity_index":
        dataset_key, variable = "era5_relative_humidity", "r"
        source_resolution = resolution if resolution in DATASETS[dataset_key]["resolutions"] else "hourly"
    elif index_type in {"soil_moisture_index", "soil_moisture_anomaly"}:
        dataset_key, variable = "era5_soil_water", "swvl1"
        source_resolution = resolution if resolution in DATASETS[dataset_key]["resolutions"] else "hourly"
    else:
        dataset_key, variable = "era5_wind", "wind_speed"
        source_resolution = resolution if resolution in DATASETS[dataset_key]["resolutions"] else "hourly"
    frame, context = extract_point_series(
        Path(data_dir), dataset_key, source_resolution, latitude, longitude,
        start_date, end_date, variable=variable,
    )
    return frame, context, dataset_key, source_resolution


def _calculate_index_series(params: Dict[str, Any], data_dir: Path) -> tuple[pd.DataFrame, Dict[str, Any], str, str, str]:
    index_type = str(params.get("index_type") or "total_rainfall")
    resolution = str(params.get("index_resolution") or "annual").lower()
    allowed = INDEX_RESOLUTION_RULES.get(index_type, ["annual"])
    if resolution not in allowed:
        raise ValueError(f"{_INDEX_LABELS.get(index_type, index_type)} is not calculated at {resolution} resolution.")
    source_all, context, dataset_key, source_resolution = _index_source_frame(params, data_dir, index_type, resolution)
    source, years, months, season_label = _index_filter_time(source_all, params)
    baseline_start = int(params.get("baseline_start") or 1991)
    baseline_end = int(params.get("baseline_end") or 2020)
    unit = str(context.get("unit") or "")

    if index_type in dict(RAINFALL_INDICES):
        daily = source.copy() if source_resolution == "daily" else _index_group_scalar(source, "daily", months, "sum")
        daily = daily.sort_values("time").reset_index(drop=True)
        threshold = float(params.get("rainy_threshold") or 1.0)
        heavy = float(params.get("heavy_threshold") or 50.0)
        very_heavy = float(params.get("very_heavy_threshold") or 100.0)
        if index_type == "total_rainfall":
            values = source[["time", "value"]].copy() if resolution == "hourly" else _index_group_scalar(daily, resolution, months, "sum")
            unit = "mm"
        elif index_type in {"number_wet_days", "number_dry_days", "heavy_rainfall_days", "very_heavy_rainfall_days"}:
            if index_type == "number_wet_days": flag = daily["value"] >= threshold
            elif index_type == "number_dry_days": flag = daily["value"] < threshold
            elif index_type == "heavy_rainfall_days": flag = daily["value"] >= heavy
            else: flag = daily["value"] >= very_heavy
            base = daily[["time"]].copy(); base["value"] = flag.astype(int)
            values = base if resolution == "daily" else _index_group_scalar(base, resolution, months, "sum")
            unit = "days"
        elif index_type in _INDEX_RUN_TYPES:
            wet = index_type in {"consecutive_wet_days", "wet_spell_length"}
            flag = daily["value"] >= threshold if wet else daily["value"] < threshold
            base = daily[["time"]].copy(); base["value"] = _index_run_lengths(flag)
            values = base if resolution == "daily" else _index_group_scalar(base, resolution, months, "max")
            unit = "days"
        elif index_type == "max_1day_rainfall":
            values = daily[["time", "value"]].copy() if resolution == "daily" else _index_group_scalar(daily, resolution, months, "max")
            unit = "mm"
        elif index_type == "max_5day_rainfall":
            base = daily[["time"]].copy(); base["value"] = daily["value"].rolling(5, min_periods=1).sum()
            values = base if resolution == "daily" else _index_group_scalar(base, resolution, months, "max")
            unit = "mm"
        elif index_type == "sdii":
            wet = daily[daily["value"] >= threshold].copy()
            if resolution == "daily":
                values = daily[["time", "value"]].copy(); values.loc[values["value"] < threshold, "value"] = np.nan
            else:
                keyed = _index_period_key(daily, resolution, months)
                keyed["wet_value"] = keyed["value"].where(keyed["value"] >= threshold)
                grouped = keyed.groupby("period")["wet_value"].agg(["sum", "count"]).reset_index()
                grouped["value"] = grouped["sum"] / grouped["count"].replace(0, np.nan)
                values = grouped[["period", "value"]].rename(columns={"period": "time"})
            unit = "mm/wet day"
        else:  # rainfall anomaly
            daily_all = source_all.copy() if source_resolution == "daily" else _index_group_scalar(source_all, "daily", months, "sum")
            daily_all = daily_all[pd.to_datetime(daily_all["time"]).dt.month.isin(months)].copy()
            totals_all = _index_group_scalar(daily_all, resolution, months, "sum")
            values_all = _index_climatology_anomaly(totals_all, resolution, baseline_start, baseline_end)
            values, _y, _m, _s = _index_filter_time(values_all, params)
            unit = "mm"

    elif index_type in dict(TEMPERATURE_INDICES):
        hourly = source[["time", "value"]].sort_values("time").copy()
        daily = _index_daily_temperature(hourly)
        if index_type == "mean_temperature":
            if resolution == "hourly": values = hourly
            elif resolution == "daily": values = daily[["time", "daily_mean"]].rename(columns={"daily_mean": "value"})
            else: values = _index_group_scalar(daily[["time", "daily_mean"]].rename(columns={"daily_mean": "value"}), resolution, months, "mean")
            unit = "°C"
        elif index_type in {"maximum_temperature", "minimum_temperature"}:
            column = "daily_max" if index_type == "maximum_temperature" else "daily_min"
            agg = "max" if index_type == "maximum_temperature" else "min"
            base = daily[["time", column]].rename(columns={column: "value"})
            values = base if resolution == "daily" else _index_group_scalar(base, resolution, months, agg)
            unit = "°C"
        elif index_type == "dtr":
            base = daily[["time"]].copy(); base["value"] = daily["daily_max"] - daily["daily_min"]
            values = base if resolution == "daily" else _index_group_scalar(base, resolution, months, "mean")
            unit = "°C"
        elif index_type in {"hot_days", "hot_nights", "cold_days", "cold_nights"}:
            heat_threshold = float(params.get("heat_threshold") or 35.0)
            warm_threshold = float(params.get("warm_threshold") or 30.0)
            cold_threshold = float(params.get("cold_threshold") or 15.0)
            if index_type == "hot_days": flag = daily["daily_max"] >= heat_threshold
            elif index_type == "hot_nights": flag = daily["daily_min"] >= warm_threshold
            elif index_type == "cold_days": flag = daily["daily_max"] <= cold_threshold
            else: flag = daily["daily_min"] <= cold_threshold
            base = daily[["time"]].copy(); base["value"] = flag.astype(int)
            values = base if resolution == "daily" else _index_group_scalar(base, resolution, months, "sum")
            unit = "days"
        elif index_type == "temperature_anomaly":
            hourly_all = source_all[["time", "value"]].sort_values("time").copy()
            hourly_all = hourly_all[pd.to_datetime(hourly_all["time"]).dt.month.isin(months)].copy()
            daily_all = _index_daily_temperature(hourly_all)
            if resolution == "hourly": base_all = hourly_all
            elif resolution == "daily": base_all = daily_all[["time", "daily_mean"]].rename(columns={"daily_mean": "value"})
            else: base_all = _index_group_scalar(daily_all[["time", "daily_mean"]].rename(columns={"daily_mean": "value"}), resolution, months, "mean")
            values_all = _index_climatology_anomaly(base_all, resolution, baseline_start, baseline_end)
            values, _y, _m, _s = _index_filter_time(values_all, params)
            unit = "°C"
        else:  # heat index
            try:
                rh, _rh_context = extract_point_series(
                    Path(data_dir), "era5_relative_humidity", "hourly",
                    float(params.get("latitude")), float(params.get("longitude")),
                    str(params.get("index_start_date") or params.get("start_date") or ""),
                    str(params.get("index_end_date") or params.get("end_date") or ""), variable="r",
                )
                merged = hourly.merge(rh.rename(columns={"value": "rh"}), on="time", how="inner")
                merged["value"] = _heat_index_celsius(merged["value"], merged["rh"])
                merged = merged[["time", "value"]]
            except Exception as exc:
                raise ValueError(f"Heat Index requires matching hourly ERA5 relative humidity: {exc}") from exc
            merged, _years, _months, _label = _index_filter_time(merged, params)
            if resolution == "hourly": values = merged
            else: values = _index_group_scalar(merged, resolution, months, "mean")
            unit = "°C"
    else:
        base = source[["time", "value"]].copy()
        if source_resolution != resolution:
            base = _index_group_scalar(base, resolution, months, "mean")
        if index_type in {"soil_moisture_index", "soil_moisture_anomaly"}:
            base_all = source_all[["time", "value"]].copy()
            base_all = base_all[pd.to_datetime(base_all["time"]).dt.month.isin(months)].copy()
            if source_resolution != resolution:
                base_all = _index_group_scalar(base_all, resolution, months, "mean")
            values_all = _index_climatology_anomaly(
                base_all, resolution, baseline_start, baseline_end,
                standardized=index_type == "soil_moisture_index",
            )
            values, _y, _m, _s = _index_filter_time(values_all, params)
            if index_type == "soil_moisture_index": unit = "standard deviations"
        else:
            values = base

    values = values.dropna(subset=["value"]).sort_values("time").reset_index(drop=True)
    if values.empty:
        raise ValueError("The selected climate index produced no valid values.")
    return values, context, dataset_key, source_resolution, season_label


def _index_plot(values: pd.DataFrame, params: Dict[str, Any], output: Path, title: str, unit: str, resolution: str) -> str:
    plot_type = str(params.get("index_plot_type") or "auto").lower()
    allowed = INDEX_PLOT_RULES.get(resolution, ["auto", "line"])
    if plot_type not in allowed:
        plot_type = "auto"
    if plot_type == "auto":
        plot_type = "bar" if resolution in {"annual", "seasonal"} else "line"
    fig, ax = plt.subplots(figsize=(11.4, 6.2))
    frame = values.copy(); frame["time"] = pd.to_datetime(frame["time"])
    if plot_type == "heatmap":
        if resolution == "hourly":
            frame["date"] = frame["time"].dt.strftime("%Y-%m-%d")
            frame["bucket"] = frame["time"].dt.hour
            pivot = frame.pivot_table(index="date", columns="bucket", values="value", aggfunc="mean")
            x_label = "Hour of day"
        elif resolution == "daily":
            frame["year"] = frame["time"].dt.year
            frame["bucket"] = frame["time"].dt.dayofyear
            pivot = frame.pivot_table(index="year", columns="bucket", values="value", aggfunc="mean")
            x_label = "Day of year"
        else:
            frame["year"] = frame["time"].dt.year
            frame["bucket"] = frame["time"].dt.month
            pivot = frame.pivot_table(index="year", columns="bucket", values="value", aggfunc="mean")
            x_label = "Month"
        im = ax.imshow(pivot.to_numpy(dtype=float), aspect="auto", interpolation="nearest", cmap="viridis")
        if len(pivot.columns) <= 24:
            ax.set_xticks(np.arange(len(pivot.columns))); ax.set_xticklabels([str(v) for v in pivot.columns])
        if len(pivot.index) <= 30:
            ax.set_yticks(np.arange(len(pivot.index))); ax.set_yticklabels([str(v) for v in pivot.index])
        ax.set_xlabel(x_label); ax.set_ylabel("Date / Year")
        fig.colorbar(im, ax=ax, shrink=0.84, pad=0.02, label=unit or "Index value")
    elif plot_type == "bar":
        labels = frame["time"].dt.strftime("%Y") if resolution in {"annual", "seasonal"} else frame["time"].dt.strftime("%Y-%m-%d")
        ax.bar(labels, frame["value"])
        if len(frame) > 20:
            step = max(1, math.ceil(len(frame) / 20)); ax.set_xticks(np.arange(0, len(frame), step)); ax.set_xticklabels(labels.iloc[::step], rotation=45, ha="right")
        else:
            ax.tick_params(axis="x", rotation=45)
        ax.set_xlabel("Period"); ax.set_ylabel(unit or "Index value")
    else:
        ax.plot(frame["time"], frame["value"], linewidth=1.4)
        ax.set_xlabel("Time"); ax.set_ylabel(unit or "Index value")
    ax.set_title(title, fontsize=14, fontweight="bold", pad=13)
    apply_plot_grids(ax)
    _finalize_plot(fig, output)
    return plot_type


def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Calculate and plot one climate index at its selected temporal resolution."""
    index_type = str(params.get("index_type") or "total_rainfall")
    if index_type in _REMOVED_RAINY_SEASON_INDICES:
        raise ValueError("Rainy-season onset, cessation and length indices are not available.")
    resolution = str(params.get("index_resolution") or "annual").lower()
    values, context, dataset_key, source_resolution, season_label = _calculate_index_series(params, Path(data_dir))
    index_label = _INDEX_LABELS.get(index_type, index_type.replace("_", " ").title())
    location = str(params.get("location_name") or "Selected Location")
    start_date = str(params.get("index_start_date") or params.get("start_date") or "")
    end_date = str(params.get("index_end_date") or params.get("end_date") or "")
    unit = "days" if index_type in _INDEX_COUNT_TYPES | _INDEX_RUN_TYPES else str(context.get("unit") or "")
    if index_type in {"mean_temperature", "maximum_temperature", "minimum_temperature", "temperature_anomaly", "heat_index", "dtr"}:
        unit = "°C"
    if index_type == "total_rainfall" or index_type in {"max_1day_rainfall", "max_5day_rainfall", "rainfall_anomaly"}:
        unit = "mm"
    if index_type == "sdii": unit = "mm/wet day"
    if index_type == "soil_moisture_index": unit = "standard deviations"

    export_frame = values.copy()
    export_frame["Date / Time"] = pd.to_datetime(export_frame["time"])
    export_frame["Year"] = export_frame["Date / Time"].dt.year
    export_frame["Month"] = export_frame["Date / Time"].dt.month
    export_frame["Day"] = export_frame["Date / Time"].dt.day
    export_frame["Hour"] = export_frame["Date / Time"].dt.hour if resolution == "hourly" else pd.NA
    export_frame["Temporal Resolution"] = INDEX_RESOLUTION_LABELS.get(resolution, resolution.title())
    export_frame["Climate Index"] = index_label
    export_frame["Value"] = pd.to_numeric(export_frame["value"], errors="coerce")
    export_frame["Unit"] = unit
    export_frame["Weather Element / Dataset"] = DATASETS.get(dataset_key, {}).get("label", dataset_key)
    export_frame["Source Resolution"] = source_resolution.title()
    export_frame["Selected Months / Season"] = season_label
    export_frame["Location"] = location
    export_frame["Latitude"] = float(params.get("latitude"))
    export_frame["Longitude"] = float(params.get("longitude"))
    columns = ["Date / Time", "Year", "Month", "Day"]
    if resolution == "hourly": columns.append("Hour")
    columns += ["Temporal Resolution", "Climate Index", "Value", "Unit", "Weather Element / Dataset", "Source Resolution", "Selected Months / Season", "Location", "Latitude", "Longitude"]
    export_frame = export_frame[columns]

    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"index_{index_type}_{resolution}_{location}_{stamp}")
    excel_path = dirs["indices"] / f"{stem}.xlsx"
    csv_path = dirs["indices"] / f"{stem}.csv"
    plot_path = dirs["indices"] / f"{stem}.png"
    export_frame.to_csv(csv_path, index=False)
    qr_payload = _cde_default_download_context(excel_path)
    qr_payload.update({
        "document_type": "Climate Index Product",
        "index": index_label,
        "dataset": DATASETS.get(dataset_key, {}).get("label", dataset_key),
        "temporal_resolution": INDEX_RESOLUTION_LABELS.get(resolution, resolution.title()),
        "period": f"{start_date} to {end_date}",
        "months_or_season": season_label,
        "station_name": location,
        "latitude": params.get("latitude"),
        "longitude": params.get("longitude"),
    })
    _cde_write_single_sheet_workbook(
        excel_path,
        [(f"{index_label} — {INDEX_RESOLUTION_LABELS.get(resolution, resolution.title())}", export_frame)],
        qr_payload=qr_payload,
        zero_decimal=index_type in _INDEX_COUNT_TYPES | _INDEX_RUN_TYPES,
        sheet_name="Data",
        workbook_title=f"{index_label} for {location}",
    )
    source_label = DATASETS.get(dataset_key, {}).get("label", dataset_key)
    period_text = f"{start_date[:10]} to {end_date[:10]}" if start_date and end_date else plot_period_text(resolution, params.get("start_date"), params.get("end_date"))
    title = f"{INDEX_RESOLUTION_LABELS.get(resolution, resolution.title())} {index_label} from {source_label} for {location} ({period_text})"
    actual_plot_type = _index_plot(values, params, plot_path, title, unit, resolution)
    context.update({
        "dataset_key": dataset_key,
        "dataset_label": source_label,
        "resolution": resolution,
        "source_resolution": source_resolution,
        "season_label": season_label,
        "unit": unit,
        "plot_type": actual_plot_type,
    })
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "csv_path": csv_path,
        "rows": len(export_frame),
        "selected_column": "Value",
        "context": context,
        "summary_cards": [
            {"label": "Climate Index", "value": index_label, "note": source_label},
            {"label": "Time Scale", "value": INDEX_RESOLUTION_LABELS.get(resolution, resolution.title()), "note": season_label},
            {"label": "Records", "value": f"{len(export_frame):,}", "note": f"Derived from {source_resolution} data"},
        ],
    }

# ---------------------------------------------------------------------------
# 2026-07-18: guided extraction, full temperature resolutions and cartographic maps
# ---------------------------------------------------------------------------
# ERA5 2m Temperature remains the simple single-element dataset.  Its hourly
# values come only from the authoritative 1940-2026 store, while prepared CDE
# stores provide the longer time scales.  A second dataset exposes the prepared
# mean/minimum/maximum variables explicitly.
DATASETS["era5_temperature"].update({
    "label": "ERA5 2m Temperature",
    "family": "temperature",
    "unit": "°C",
    "resolutions": {
        "hourly": "Hourly",
        "daily": "Daily",
        "monthly": "Monthly",
        "annual": "Annual",
        "seasonal": "Seasonal",
    },
    "keywords": [
        "ERA5_Tanzania_Temperature_2M_Hourly_1940_2026",
        "ERA5_Tanzania_Temperature _2M_Hourly_1940_2026",
        "CDE_ERA5_Tanzania_Temperature_Mean_Min_Max_1940_2025",
    ],
    "variables": ["ta", "t2m", "tmean", "mean_temperature", "temperature"],
})
DATASETS["era5_temperature_stats"] = {
    "label": "ERA5 Temperature Mean, Minimum and Maximum",
    "family": "temperature",
    "unit": "°C",
    "resolutions": {
        "daily": "Daily",
        "monthly": "Monthly",
        "annual": "Annual",
        "seasonal": "Seasonal",
    },
    "keywords": ["CDE_ERA5_Tanzania_Temperature_Mean_Min_Max_1940_2025"],
    "variables": [
        "ta", "tmean", "mean_temperature",
        "tmin", "tn", "minimum_temperature",
        "tmax", "tx", "maximum_temperature",
    ],
}

# Make the January-December product name explain the aggregation directly.
PLOT_TYPES = [
    (key, "Average Monthly Total / Mean Profile (Jan–Dec)" if key == "monthly_climatology" else label)
    for key, label in PLOT_TYPES
]

# More useful climate-index display choices.
INDEX_PLOT_TYPES = [
    ("auto", "Automatic — best plot for the selected time scale"),
    ("line", "Time-Series Line Plot"),
    ("bar", "Period Comparison Bar Chart"),
    ("heatmap", "Time Heat Map"),
    ("box", "Distribution Box Plot"),
    ("histogram", "Frequency Distribution Histogram"),
]
INDEX_PLOT_RULES = {
    "hourly": ["auto", "line", "heatmap", "box", "histogram"],
    "daily": ["auto", "line", "bar", "heatmap", "box", "histogram"],
    "monthly": ["auto", "line", "bar", "heatmap", "box", "histogram"],
    "annual": ["auto", "line", "bar", "box", "histogram"],
    "seasonal": ["auto", "line", "bar", "box", "histogram"],
}

_CDE_20260718_FIND_FILE_BASE = find_file


def _store_is_in_resolution_folder(path: Path, resolution: str) -> bool:
    resolution = str(resolution or "").lower()
    return resolution in {part.lower() for part in Path(path).parts}


def find_file(data_dir: Path, dataset_key: str, resolution: str, season: str | None = None, silent: bool = False) -> Path | None:
    """Resolve temperature stores despite harmless filename suffix differences."""
    dataset_key = str(dataset_key)
    resolution = str(resolution).lower()
    root = Path(data_dir).expanduser().resolve()
    if dataset_key == "era5_temperature" and resolution == "hourly":
        for path in iter_data_stores(root):
            if path.name in _HOURLY_TEMPERATURE_STORE_ALIASES:
                return path
        if silent:
            return None
        raise FileNotFoundError(
            "Hourly ERA5 2m Temperature requires "
            "ERA5_Tanzania_Temperature_2M_Hourly_1940_2026.zarr under storage/zarr/hourly."
        )
    if dataset_key in {"era5_temperature", "era5_temperature_stats"} and resolution in {"daily", "monthly", "annual", "seasonal"}:
        token = _normalize_store_name("CDE_ERA5_Tanzania_Temperature_Mean_Min_Max_1940_2025")
        season_token = _normalize_store_name(season or "")
        candidates: list[Path] = []
        for path in iter_data_stores(root):
            if not _store_is_in_resolution_folder(path, resolution):
                continue
            normalized = _normalize_store_name(path.name)
            if token not in normalized:
                continue
            if resolution == "seasonal" and season_token and season_token not in normalized:
                continue
            candidates.append(path)
        if candidates:
            # Prefer a filename that explicitly names the requested resolution,
            # but accept the exact unsuffixed operational store named by users.
            return sorted(
                candidates,
                key=lambda path: (
                    0 if _normalize_store_name(resolution) in _normalize_store_name(path.name) else 1,
                    len(path.name),
                    path.name,
                ),
            )[0]
        if silent:
            return None
        raise FileNotFoundError(
            f"No prepared ERA5 temperature Mean/Minimum/Maximum Zarr store was found for {resolution}"
            + (f" / {season}" if season else "")
            + f" under {root}. Expected a name beginning CDE_ERA5_Tanzania_Temperature_Mean_Min_Max_1940_2025."
        )
    return _CDE_20260718_FIND_FILE_BASE(root, dataset_key, resolution, season=season, silent=silent)


@lru_cache(maxsize=12)
def _cached_admin_feature_labels(path_text: str, modified_ns: int) -> tuple[tuple[str, float, float], ...]:
    """Return labels and stable visual centres from a GADM GeoJSON file."""
    path = Path(path_text)
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows: list[tuple[str, float, float]] = []
    for feature in payload.get("features", []):
        props = feature.get("properties") or {}
        name = ""
        for key in ("NAME_3", "NAME_2", "NAME_1", "NAME_0", "name", "Name"):
            if props.get(key):
                name = str(props[key]).strip()
                break
        polygons = _geometry_polygons(feature.get("geometry") or {})
        if not name or not polygons:
            continue
        # Use the centre of the largest ring's bounding box.  This is more
        # stable for irregular administrative polygons than averaging vertices.
        polygon = max(polygons, key=lambda arr: abs(float(np.ptp(arr[:, 0]) * np.ptp(arr[:, 1]))))
        x = float((np.nanmin(polygon[:, 0]) + np.nanmax(polygon[:, 0])) / 2.0)
        y = float((np.nanmin(polygon[:, 1]) + np.nanmax(polygon[:, 1])) / 2.0)
        rows.append((name, x, y))
    return tuple(rows)


def _admin_feature_labels(data_dir: Path, level: int) -> list[tuple[str, float, float]]:
    path = Path(data_dir) / "shapefiles" / f"gadm41_TZA_{int(level)}.json"
    if not path.is_file():
        return []
    return list(_cached_admin_feature_labels(str(path.resolve()), path.stat().st_mtime_ns))


def _draw_hydrology(ax, *, ocean: bool, lakes: bool, rivers: bool) -> None:
    """Draw unobtrusive ocean, major lakes and river context around Tanzania."""
    from matplotlib.lines import Line2D
    legend_items = []
    ax.set_facecolor("#dcebf7" if ocean else "white")
    if ocean:
        ax.text(
            40.25, -7.6, "INDIAN OCEAN", fontsize=8.3, fontweight="bold",
            color="#557d9b", rotation=90, ha="center", va="center", alpha=0.82, zorder=10,
        )
        legend_items.append(Line2D([0], [0], color="#b8d1e7", lw=6, label="Ocean"))
    if lakes:
        # Refined simplified outlines.  Operational GeoJSON lake layers may be
        # added later without changing the map renderer.
        lake_polygons = {
            "Lake Victoria": [(31.25,-0.78),(32.10,-0.63),(33.20,-0.82),(34.15,-1.33),(34.05,-2.22),(33.40,-2.52),(32.35,-2.55),(31.45,-2.02),(31.18,-1.35)],
            "Lake Tanganyika": [(29.10,-3.30),(29.55,-3.42),(30.12,-4.45),(30.42,-5.75),(30.72,-7.15),(30.55,-8.35),(30.18,-8.74),(29.72,-7.72),(29.42,-6.18),(29.12,-4.55)],
            "Lake Nyasa": [(34.42,-9.05),(34.82,-9.24),(35.18,-10.15),(35.62,-11.52),(35.28,-11.92),(34.86,-11.18),(34.58,-10.12)],
            "Lake Rukwa": [(31.72,-7.05),(32.20,-7.12),(32.48,-7.43),(32.38,-8.10),(31.90,-8.20),(31.70,-7.72)],
            "Lake Eyasi": [(34.62,-3.32),(35.05,-3.35),(35.18,-3.72),(34.83,-3.90),(34.52,-3.66)],
            "Lake Natron": [(35.72,-1.92),(36.02,-1.96),(36.08,-2.50),(35.82,-2.67),(35.66,-2.32)],
        }
        for vertices in lake_polygons.values():
            ax.add_patch(Polygon(vertices, closed=True, facecolor="#bfd9ee", edgecolor="#6b9fbe", linewidth=0.7, zorder=9))
        ax.text(32.65, -1.60, "Victoria", fontsize=6.0, color="#4f7893", ha="center", zorder=10)
        ax.text(29.92, -6.10, "Tanganyika", fontsize=5.8, color="#4f7893", rotation=76, ha="center", zorder=10)
        ax.text(35.03, -10.55, "Nyasa", fontsize=5.8, color="#4f7893", rotation=72, ha="center", zorder=10)
        legend_items.append(Line2D([0], [0], color="#6b9fbe", lw=4, label="Major lakes"))
    if rivers:
        river_lines = {
            "Rufiji": [(34.4,-7.55),(35.2,-7.68),(36.0,-7.80),(36.8,-8.02),(37.6,-8.02),(38.5,-7.78),(39.30,-7.82)],
            "Ruvuma": [(34.9,-10.58),(35.8,-10.72),(36.8,-10.86),(37.8,-10.72),(38.8,-10.48)],
            "Pangani": [(36.55,-3.10),(36.83,-3.62),(37.20,-4.16),(37.70,-4.62),(38.25,-5.02),(39.05,-5.42)],
            "Wami": [(35.55,-5.65),(36.35,-5.80),(37.20,-6.00),(38.10,-6.10),(38.85,-6.18)],
            "Malagarasi": [(30.15,-4.05),(30.85,-4.26),(31.45,-4.63),(32.05,-4.92),(32.58,-5.05)],
        }
        for points in river_lines.values():
            arr = np.asarray(points, dtype=float)
            ax.plot(arr[:, 0], arr[:, 1], color="#4d95bf", linewidth=0.62, alpha=0.78, zorder=10)
        legend_items.append(Line2D([0], [0], color="#4d95bf", lw=1.4, label="Major rivers"))
    if legend_items:
        ax.legend(handles=legend_items, loc="lower left", fontsize=6.7, frameon=True, framealpha=0.90, borderpad=0.35)


def _draw_north_arrow(ax) -> None:
    ax.annotate(
        "N", xy=(0.91, 0.90), xytext=(0.91, 0.78), xycoords="axes fraction",
        ha="center", va="center", fontsize=10, fontweight="bold",
        arrowprops=dict(facecolor="#111111", edgecolor="#111111", width=3.2, headwidth=11, headlength=12),
        zorder=20,
    )


def _draw_scale_bar(ax, length_km: int = 200) -> None:
    latitude = -11.55
    x0 = 28.55
    degrees = float(length_km) / (111.32 * math.cos(math.radians(latitude)))
    segment = degrees / 4.0
    y0 = latitude
    for idx in range(4):
        ax.plot([x0 + idx * segment, x0 + (idx + 1) * segment], [y0, y0],
                color="#111111" if idx % 2 == 0 else "white", linewidth=6.0,
                solid_capstyle="butt", zorder=20)
        ax.plot([x0 + idx * segment, x0 + (idx + 1) * segment], [y0, y0],
                color="#111111", linewidth=0.8, zorder=21)
    for idx, label in enumerate((0, 50, 100, 150, 200)):
        x = x0 + idx * segment
        ax.plot([x, x], [y0 - 0.07, y0 + 0.07], color="#111111", linewidth=0.8, zorder=21)
        if idx in {0, 2, 4}:
            ax.text(x, y0 + 0.13, str(label), fontsize=5.8, ha="center", va="bottom", zorder=21)
    ax.text(x0, y0 + 0.34, "Scale", fontsize=6.1, fontweight="bold", ha="left", zorder=21)
    ax.text(x0 + degrees / 2.0, y0 - 0.24, "km", fontsize=5.8, ha="center", zorder=21)


def _render_tanzania_map_axis(
    ax,
    *,
    x: np.ndarray,
    y: np.ndarray,
    values: np.ndarray,
    data_dir: Path,
    selected_boundaries: list[np.ndarray],
    level: int,
    dataset_key: str,
    show_ocean: bool,
    show_lakes: bool,
    show_rivers: bool,
    title: str,
    norm=None,
    show_labels: bool = True,
    show_cartographic_elements: bool = True,
):
    _draw_hydrology(ax, ocean=show_ocean, lakes=show_lakes, rivers=show_rivers, data_dir=data_dir)
    im = ax.pcolormesh(x, y, values, shading="auto", cmap=plot_cmap_for(dataset_key), norm=norm, zorder=4)
    _draw_admin_boundaries(ax, selected_boundaries, level)
    # A stronger national-looking edge is obtained by overlaying Admin-1 rings.
    try:
        national_like, _ = _admin_polygons(data_dir, 1)
        from matplotlib.collections import LineCollection
        ax.add_collection(LineCollection(national_like, colors="#111820", linewidths=1.05, alpha=0.95, zorder=11))
    except Exception:
        pass
    if show_labels and level == 1:
        for name, lon, lat in _admin_feature_labels(data_dir, 1):
            if TANZANIA_BOUNDS["lon_min"] <= lon <= TANZANIA_BOUNDS["lon_max"] and TANZANIA_BOUNDS["lat_min"] <= lat <= TANZANIA_BOUNDS["lat_max"]:
                ax.text(lon, lat, name, fontsize=5.7, fontweight="bold", color="#16232d", ha="center", va="center", zorder=12)
    ax.set_xlim(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
    ax.set_ylim(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    ax.set_title(title, fontsize=9.2 if not show_cartographic_elements else 13.0, fontweight="bold", pad=8)
    if show_cartographic_elements:
        ax.set_xlabel("Longitude")
        ax.set_ylabel("Latitude")
        _draw_north_arrow(ax)
        _draw_scale_bar(ax)
        apply_plot_grids(ax)
    else:
        ax.set_xticks([])
        ax.set_yticks([])
    return im


def _spatial_panel_jobs(params: Dict[str, Any]) -> list[tuple[Dict[str, Any], str]]:
    period = _spatial_period_definition(params)
    mode = period["mode"]
    years = list(period["years"])
    months = list(period["months"])
    jobs: list[tuple[Dict[str, Any], str]] = []
    if mode == "custom" and len(years) > 1:
        for year in years:
            item = dict(params)
            item.update({
                "map_period_mode": "custom",
                "map_custom_years": str(year),
                "map_custom_months": ",".join(str(month) for month in months),
                "start_year": str(year),
                "end_year": str(year),
            })
            label = str(year) if months == list(range(1, 13)) else f"{year} · {', '.join(pd.Timestamp(2000,m,1).strftime('%b') for m in months)}"
            jobs.append((item, label))
    elif mode == "custom" and len(years) == 1 and len(months) > 1:
        for month in months:
            item = dict(params)
            item.update({"map_period_mode": "month", "map_year": str(years[0]), "map_month": f"{month:02d}"})
            jobs.append((item, f"{pd.Timestamp(2000, month, 1).strftime('%B')} {years[0]}"))
    else:
        jobs.append((dict(params), period["period_label"]))
    if len(jobs) > 36:
        raise ValueError("The all-maps figure is limited to 36 panels. Reduce the selected years or generate a single aggregated map.")
    return jobs


def _clean_map_element_name(element: str) -> str:
    """Remove source-resolution words when the represented period already says them."""
    value = re.sub(r"^(Hourly|Daily|Monthly|Annual|Seasonal)\s+", "", str(element or "Weather Element"), flags=re.I).strip()
    return value.replace("Rainfall", "Precipitation")


def _spatial_map_heading(element: str, context: Dict[str, Any]) -> str:
    """Create a period-aware map heading such as Total Annual Precipitation — 1998."""
    clean = _clean_map_element_name(element)
    period = str(context.get("period_label") or "Selected Period")
    aggregation = str(context.get("aggregation") or "").lower()
    family = str(context.get("family") or "")
    years = [int(v) for v in context.get("years", []) if str(v).strip()]
    months = [int(v) for v in context.get("months", []) if str(v).strip()]
    if family == "rainfall":
        if len(years) == 1 and months == list(range(1, 13)) and "total" in aggregation:
            return f"Total Annual {clean} — {years[0]}"
        if len(years) == 1 and len(months) == 1:
            month_name = pd.Timestamp(2000, months[0], 1).strftime("%B")
            return f"Total Monthly {clean} — {month_name} {years[0]}"
        if "mean selected-month total" in aggregation:
            return f"Mean Selected-Month Total {clean} — {period}"
        if "mean total" in aggregation:
            return f"Mean Total {clean} — {period}"
        if "total" in aggregation:
            return f"Total {clean} — {period}"
    if "mean" in aggregation:
        return f"Mean {clean} — {period}"
    return f"{clean} — {period}"


_CDE_20260718_SPATIAL_MAP_BASE = _generate_requested_spatial_map


def _generate_requested_spatial_map(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate a publication-style single map or a shared-scale multi-panel map."""
    layout = str(params.get("map_output_layout") or "single").strip().lower()
    if layout not in {"single", "panel"}:
        layout = "single"
    if layout == "single":
        grid, context = _extract_spatial_period(Path(data_dir), params)
        jobs = [(grid, context, context.get("period_label") or "Selected period")]
    else:
        jobs = []
        for job_params, label in _spatial_panel_jobs(params):
            grid, context = _extract_spatial_period(Path(data_dir), job_params)
            jobs.append((grid, context, label))
        if len(jobs) <= 1:
            layout = "single"

    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    level = max(1, min(3, int(params.get("map_admin_level") or 1)))
    selected_boundaries, boundary_path = _admin_polygons(Path(data_dir), level)
    national_polygons, _ = _admin_polygons(Path(data_dir), 1)
    admin_label = {1: "Admin 1 — Regions", 2: "Admin 2 — Districts", 3: "Admin 3 — Wards"}[level]
    prepared: list[tuple[np.ndarray, np.ndarray, np.ndarray, pd.DataFrame, Dict[str, Any], str]] = []
    all_values: list[np.ndarray] = []
    for grid, context, label in jobs:
        work = grid[
            grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
            & grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
        ].copy()
        if work.empty:
            continue
        if work.duplicated(subset=["latitude", "longitude"]).any():
            raise ValueError(
                "The selected spatial slice contains duplicate native grid coordinates. "
                "CDE will not average or resample spatial cells; inspect the source dimensions."
            )
        pivot = work.pivot(index="latitude", columns="longitude", values="value").sort_index()
        x = pivot.columns.to_numpy(dtype=float)
        y = pivot.index.to_numpy(dtype=float)
        values = _mask_grid_to_tanzania(pivot.to_numpy(dtype=float), x, y, national_polygons)
        prepared.append((x, y, values, work, context, label))
        all_values.append(values[np.isfinite(values)])
    if not prepared:
        raise ValueError("No spatial values were found within the Tanzania map extent.")

    finite = np.concatenate([value for value in all_values if value.size]) if any(value.size for value in all_values) else np.asarray([0.0, 1.0])
    from matplotlib.colors import Normalize
    vmin = float(np.nanmin(finite)); vmax = float(np.nanmax(finite))
    if np.isclose(vmin, vmax):
        vmax = vmin + 1.0
    norm = Normalize(vmin=vmin, vmax=vmax)
    element = element_with_source(dataset_key, variable_display_name(params.get("variable"), prepared[0][4], dataset_key))
    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    period_stem = "all_selected_periods" if layout == "panel" else slugify(prepared[0][5])
    stem = _cde_prefixed_stem(f"spatial_map_{dataset_key}_{period_stem}_admin{level}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"

    show_ocean = _truthy(params.get("show_ocean", True))
    show_lakes = _truthy(params.get("show_lakes", True))
    show_rivers = _truthy(params.get("show_rivers", True))
    if layout == "panel":
        count = len(prepared)
        columns = min(6, max(2, int(math.ceil(math.sqrt(count * 1.15)))))
        rows = int(math.ceil(count / columns))
        fig, axes = plt.subplots(rows, columns, figsize=(3.45 * columns, 3.55 * rows), squeeze=False)
        flat = axes.ravel()
        image = None
        for idx, (x, y, values, _grid, _context, label) in enumerate(prepared):
            image = _render_tanzania_map_axis(
                flat[idx], x=x, y=y, values=values, data_dir=Path(data_dir), selected_boundaries=selected_boundaries,
                level=level, dataset_key=dataset_key, show_ocean=show_ocean, show_lakes=show_lakes,
                show_rivers=False, title=label, norm=norm, show_labels=False, show_cartographic_elements=False,
            )
        for ax in flat[len(prepared):]:
            ax.axis("off")
        fig.suptitle(f"{element} Spatial Distribution over Tanzania\nAll Selected Periods · {admin_label}", fontsize=15, fontweight="bold", y=0.995)
        if image is not None:
            cbar = fig.colorbar(image, ax=list(flat[:len(prepared)]), fraction=0.018, pad=0.012, aspect=35)
            cbar.set_label(y_axis_label(element, prepared[0][4].get("unit")))
        fig.subplots_adjust(left=0.025, right=0.94, bottom=0.025, top=0.93, wspace=0.06, hspace=0.16)
        fig.savefig(plot_path, dpi=210, bbox_inches="tight", facecolor="white")
        plt.close(fig)
    else:
        x, y, values, _grid, context, label = prepared[0]
        fig, ax = plt.subplots(figsize=(10.8, 9.0))
        image = _render_tanzania_map_axis(
            ax, x=x, y=y, values=values, data_dir=Path(data_dir), selected_boundaries=selected_boundaries,
            level=level, dataset_key=dataset_key, show_ocean=show_ocean, show_lakes=show_lakes,
            show_rivers=show_rivers,
            title=f"{_spatial_map_heading(element, context)}\n{admin_label}", norm=norm,
            show_labels=True, show_cartographic_elements=True,
        )
        cbar = fig.colorbar(image, ax=ax, shrink=0.82, pad=0.025, extend="both")
        represented_measure = _spatial_map_heading(element, context).split("—", 1)[0].strip()
        cbar.set_label(y_axis_label(represented_measure, context.get("unit")))
        fig.tight_layout()
        fig.savefig(plot_path, dpi=165, bbox_inches="tight", facecolor="white")
        plt.close(fig)

    export_parts = []
    for _x, _y, _values, grid, context, label in prepared:
        part = grid.copy()
        part.insert(0, "map_period", label)
        part.insert(1, "administrative_level", admin_label)
        export_parts.append(part)
    export_grid = pd.concat(export_parts, ignore_index=True)
    excel_path, _, _ = _save_data(export_grid, dirs["plots"], stem + "_data", table="plot_products")
    context = dict(prepared[0][4])
    context.update({
        "administrative_level": level,
        "administrative_level_label": admin_label,
        "boundary_file": boundary_path.name,
        "map_output_layout": layout,
        "map_count": len(prepared),
    })
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "rows": len(export_grid),
        "context": context,
        "period_label": "All selected periods" if layout == "panel" else prepared[0][5],
        "summary_cards": [
            {"label": "Map Layout", "value": "All maps in one figure" if layout == "panel" else "One map", "note": f"{len(prepared)} map panel(s)"},
            {"label": "Boundary Level", "value": admin_label, "note": boundary_path.name},
            {"label": "Grid Cells", "value": f"{len(export_grid):,}", "note": "Tanzania land grid cells"},
        ],
    }


_CDE_20260718_SPATIAL_VARIABILITY_BASE = _generate_spatial_variability_map


def _generate_spatial_variability_map(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Apply the same cartographic styling to standard-deviation and CV maps."""
    plot_type = str(params.get("plot_type") or "spatial_std_map")
    statistic = "cv" if plot_type == "spatial_cv_map" else "std"
    grid, context = _extract_spatial_variability(Path(data_dir), params, statistic)
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    level = max(1, min(3, int(params.get("map_admin_level") or 1)))
    boundaries, boundary_path = _admin_polygons(Path(data_dir), level)
    national, _ = _admin_polygons(Path(data_dir), 1)
    work = grid[
        grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
        & grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    ].copy()
    pivot = work.pivot_table(index="latitude", columns="longitude", values="value", aggfunc="mean").sort_index()
    x = pivot.columns.to_numpy(dtype=float); y = pivot.index.to_numpy(dtype=float)
    values = _mask_grid_to_tanzania(pivot.to_numpy(dtype=float), x, y, national)
    element = element_with_source(dataset_key, variable_display_name(params.get("variable"), context, dataset_key))
    name = "Coefficient of Variation" if statistic == "cv" else "Standard Deviation"
    admin_label = {1: "Admin 1 — Regions", 2: "Admin 2 — Districts", 3: "Admin 3 — Wards"}[level]
    dirs = ensure_output_dirs(Path(export_dir)); stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"{plot_type}_{dataset_key}_{slugify(context['period_label'])}_admin{level}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    fig, ax = plt.subplots(figsize=(10.8, 9.0))
    image = _render_tanzania_map_axis(
        ax, x=x, y=y, values=values, data_dir=Path(data_dir), selected_boundaries=boundaries,
        level=level, dataset_key=dataset_key, show_ocean=_truthy(params.get("show_ocean", True)),
        show_lakes=_truthy(params.get("show_lakes", True)), show_rivers=_truthy(params.get("show_rivers", True)),
        title=f"Spatial {name} of {_clean_map_element_name(element)} — {context['period_label']}\n{admin_label}", show_labels=True,
    )
    cbar = fig.colorbar(image, ax=ax, shrink=0.82, pad=0.025, extend="max")
    cbar.set_label(y_axis_label(name if statistic == "cv" else element, context.get("unit")))
    fig.tight_layout(); fig.savefig(plot_path, dpi=165, bbox_inches="tight", facecolor="white"); plt.close(fig)
    export_grid = work.copy(); export_grid.insert(0, "period", context["period_label"]); export_grid.insert(1, "statistic", name); export_grid.insert(2, "administrative_level", admin_label)
    excel_path, _, _ = _save_data(export_grid, dirs["plots"], stem + "_data", table="plot_products")
    context.update({"administrative_level_label": admin_label, "boundary_file": boundary_path.name})
    return {"plot_path": plot_path, "excel_path": excel_path, "rows": len(export_grid), "context": context,
            "summary_cards": [{"label": "Variability", "value": name, "note": context["period_label"]}, {"label": "Boundary Level", "value": admin_label, "note": boundary_path.name}, {"label": "Grid Cells", "value": f"{len(export_grid):,}", "note": "Tanzania land grid cells"}]}


# Support prepared daily/longer-resolution temperature indices when that dataset
# is selected.  Heat Index remains on the hourly 2m temperature dataset because
# it requires matching hourly relative humidity.
_CDE_20260718_CALCULATE_INDEX_BASE = _calculate_index_series


def _calculate_prepared_temperature_index(params: Dict[str, Any], data_dir: Path):
    index_type = str(params.get("index_type") or "mean_temperature")
    resolution = str(params.get("index_resolution") or "annual").lower()
    if resolution == "hourly":
        raise ValueError("The prepared Mean/Minimum/Maximum dataset starts at daily resolution. Select ERA5 2m Temperature for hourly indices.")
    start = str(params.get("index_start_date") or params.get("start_date") or "")
    end = str(params.get("index_end_date") or params.get("end_date") or "")
    lat = float(params.get("latitude")); lon = float(params.get("longitude"))
    months_label, months = _index_selected_months(params)
    baseline_start = int(params.get("baseline_start") or 1991); baseline_end = int(params.get("baseline_end") or 2020)

    def read(variable: str, source_resolution: str = "daily"):
        frame, context = extract_point_series(data_dir, "era5_temperature_stats", source_resolution, lat, lon, start, end, variable=variable,
                                              season=str(params.get("index_season") or "").upper() if source_resolution == "seasonal" else None)
        filtered, years, selected_months, season_label = _index_filter_time(frame, params)
        return filtered, context, years, selected_months, season_label

    source_resolution = resolution if resolution in {"daily", "monthly", "annual", "seasonal"} else "daily"
    simple_variable = {
        "mean_temperature": "ta", "maximum_temperature": "tmax", "minimum_temperature": "tmin",
        "temperature_anomaly": "ta", "hot_days": "tmax", "cold_days": "tmax", "hot_nights": "tmin", "cold_nights": "tmin",
    }.get(index_type)
    if index_type == "dtr":
        maximum, context, _years, selected_months, season_label = read("tmax", "daily")
        minimum, _context2, _years2, _months2, _season2 = read("tmin", "daily")
        merged = maximum.merge(minimum.rename(columns={"value": "minimum"}), on="time", how="inner")
        merged["value"] = pd.to_numeric(merged["value"], errors="coerce") - pd.to_numeric(merged["minimum"], errors="coerce")
        values = merged[["time", "value"]] if resolution == "daily" else _index_group_scalar(merged[["time", "value"]], resolution, selected_months, "mean")
        return values, context, "era5_temperature_stats", "daily", season_label
    if not simple_variable:
        raise ValueError("This index requires hourly 2m temperature. Select the ERA5 2m Temperature dataset.")
    frame, context, _years, selected_months, season_label = read(simple_variable, source_resolution)
    values = frame[["time", "value"]].copy()
    if index_type == "temperature_anomaly":
        values = _index_climatology_anomaly(values, resolution, baseline_start, baseline_end)
    elif index_type in {"hot_days", "cold_days", "hot_nights", "cold_nights"}:
        heat = float(params.get("heat_threshold") or 35.0); warm = float(params.get("warm_threshold") or 30.0); cold = float(params.get("cold_threshold") or 15.0)
        if index_type == "hot_days": flag = values["value"] >= heat
        elif index_type == "hot_nights": flag = values["value"] >= warm
        elif index_type in {"cold_days", "cold_nights"}: flag = values["value"] <= cold
        values["value"] = flag.astype(int)
        if resolution != source_resolution:
            values = _index_group_scalar(values, resolution, selected_months, "sum")
    return values.dropna(subset=["value"]).sort_values("time"), context, "era5_temperature_stats", source_resolution, season_label


def _calculate_index_series(params: Dict[str, Any], data_dir: Path):
    if str(params.get("dataset") or "") == "era5_temperature_stats" and str(params.get("index_type") or "") in dict(TEMPERATURE_INDICES):
        return _calculate_prepared_temperature_index(params, Path(data_dir))
    return _CDE_20260718_CALCULATE_INDEX_BASE(params, Path(data_dir))


def _index_plot(values: pd.DataFrame, params: Dict[str, Any], output: Path, title: str, unit: str, resolution: str) -> str:
    """Render line, bar, heat-map, box or histogram climate-index products."""
    plot_type = str(params.get("index_plot_type") or "auto").lower()
    allowed = INDEX_PLOT_RULES.get(resolution, ["auto", "line"])
    if plot_type not in allowed:
        plot_type = "auto"
    if plot_type == "auto":
        plot_type = "bar" if resolution in {"annual", "seasonal"} else "line"
    frame = values.copy(); frame["time"] = pd.to_datetime(frame["time"]); numeric = pd.to_numeric(frame["value"], errors="coerce")
    fig, ax = plt.subplots(figsize=(11.4, 6.2))
    if plot_type == "heatmap":
        if resolution == "hourly":
            frame["row"] = frame["time"].dt.strftime("%Y-%m-%d"); frame["bucket"] = frame["time"].dt.hour; x_label = "Hour of day"
        elif resolution == "daily":
            frame["row"] = frame["time"].dt.year; frame["bucket"] = frame["time"].dt.dayofyear; x_label = "Day of year"
        else:
            frame["row"] = frame["time"].dt.year; frame["bucket"] = frame["time"].dt.month; x_label = "Month"
        pivot = frame.pivot_table(index="row", columns="bucket", values="value", aggfunc="mean")
        im = ax.imshow(pivot.to_numpy(dtype=float), aspect="auto", interpolation="nearest", cmap="viridis")
        if len(pivot.columns) <= 24:
            ax.set_xticks(np.arange(len(pivot.columns))); ax.set_xticklabels([str(v) for v in pivot.columns])
        if len(pivot.index) <= 30:
            ax.set_yticks(np.arange(len(pivot.index))); ax.set_yticklabels([str(v) for v in pivot.index])
        ax.set_xlabel(x_label); ax.set_ylabel("Date / Year"); fig.colorbar(im, ax=ax, shrink=0.84, pad=0.02, label=unit or "Index value")
    elif plot_type == "bar":
        labels = frame["time"].dt.strftime("%Y") if resolution in {"annual", "seasonal"} else frame["time"].dt.strftime("%Y-%m-%d")
        ax.bar(np.arange(len(frame)), numeric)
        step = max(1, math.ceil(len(frame) / 20)); ticks = np.arange(0, len(frame), step)
        ax.set_xticks(ticks); ax.set_xticklabels(labels.iloc[::step], rotation=45, ha="right"); ax.set_xlabel("Period"); ax.set_ylabel(unit or "Index value")
    elif plot_type == "box":
        if resolution in {"hourly", "daily", "monthly"}:
            groups = [numeric[frame["time"].dt.month == month].dropna().to_numpy() for month in range(1, 13)]
            _boxplot_with_labels(ax, groups, [pd.Timestamp(2000, month, 1).strftime("%b") for month in range(1, 13)], showfliers=False)
            ax.set_xlabel("Month")
        else:
            _boxplot_with_labels(ax, [numeric.dropna().to_numpy()], [INDEX_RESOLUTION_LABELS.get(resolution, resolution.title())], showfliers=True)
        ax.set_ylabel(unit or "Index value")
    elif plot_type == "histogram":
        ax.hist(numeric.dropna(), bins=min(30, max(8, int(math.sqrt(max(1, numeric.notna().sum()))))))
        ax.set_xlabel(unit or "Index value"); ax.set_ylabel("Frequency")
    else:
        ax.plot(frame["time"], numeric, linewidth=1.4); ax.set_xlabel("Time"); ax.set_ylabel(unit or "Index value")
    ax.set_title(title, fontsize=14, fontweight="bold", pad=13); apply_plot_grids(ax); _finalize_plot(fig, output)
    return plot_type

# ---------------------------------------------------------------------------
# FINAL 2026-07-18 PRODUCT FIX: true January-December monthly climatology
# ---------------------------------------------------------------------------
def _generate_average_monthly_profile(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Plot Jan-Dec climatology using monthly totals for rainfall and means otherwise."""
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    resolution = str(params.get("resolution") or "monthly").lower()
    variable = str(params.get("variable") or "auto")
    season = str(params.get("season") or "").strip().upper() or None
    location = str(params.get("location_name") or "Selected Location")
    start = str(params.get("start_date") or "")
    end = str(params.get("end_date") or "")
    frame, context = extract_point_series(
        Path(data_dir), dataset_key, resolution,
        float(params.get("latitude")), float(params.get("longitude")),
        start, end, variable=variable, season=season,
    )
    work = frame[["time", "value"]].copy()
    work["time"] = pd.to_datetime(work["time"])
    work["value"] = pd.to_numeric(work["value"], errors="coerce")
    work = work.dropna(subset=["time", "value"])
    if work.empty:
        raise ValueError("No values were found for the selected monthly profile period.")
    family = str(DATASETS.get(dataset_key, {}).get("family") or "")
    # Convert hourly/daily values into one value per calendar month before
    # averaging like months across years. Monthly rainfall stores already hold
    # total monthly precipitation and must not be summed a second time.
    work["year"] = work["time"].dt.year
    work["month"] = work["time"].dt.month
    if resolution in {"hourly", "daily"}:
        aggregation = "sum" if family == "rainfall" else "mean"
        monthly = work.groupby(["year", "month"], as_index=False)["value"].agg(aggregation)
    else:
        monthly = work.groupby(["year", "month"], as_index=False)["value"].mean()
    profile = monthly.groupby("month", as_index=False)["value"].mean()
    full = pd.DataFrame({"month": range(1, 13)})
    profile = full.merge(profile, on="month", how="left")
    profile["month_name"] = profile["month"].map(lambda value: pd.Timestamp(2000, int(value), 1).strftime("%b"))
    profile["statistic"] = (
        "Average of total monthly precipitation across selected years"
        if family == "rainfall" else "Average monthly value across selected years"
    )
    profile["years_represented"] = int(monthly["year"].nunique())

    element = _clean_map_element_name(element_with_source(dataset_key, variable_display_name(variable, context, dataset_key)))
    start_year = int(work["year"].min()); end_year = int(work["year"].max())
    year_text = str(start_year) if start_year == end_year else f"{start_year}–{end_year}"
    if family == "rainfall":
        title = f"Average Total Monthly {element} for {location} ({year_text})"
    else:
        title = f"Average Monthly {element} for {location} ({year_text})"

    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"monthly_climatology_{dataset_key}_{variable}_{location}_{start_year}_{end_year}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    fig, ax = plt.subplots(figsize=(10.8, 6.0))
    x = np.arange(1, 13)
    if family == "rainfall":
        ax.bar(x, profile["value"], width=0.72)
    else:
        ax.plot(x, profile["value"], marker="o", linewidth=2.0)
    ax.set_xticks(x)
    ax.set_xticklabels(profile["month_name"])
    ax.set_xlabel("Month")
    ax.set_ylabel(y_axis_label(element, context.get("unit")))
    ax.set_title(title, fontsize=14, fontweight="bold", pad=13)
    apply_plot_grids(ax)
    _finalize_plot(fig, plot_path)
    excel_path, _, _ = _save_data(profile, dirs["plots"], stem + "_data", table="plot_products")
    context = dict(context)
    context.update({
        "period_label": year_text,
        "aggregation": profile["statistic"].iloc[0],
        "months": list(range(1, 13)),
        "years": sorted(monthly["year"].astype(int).unique().tolist()),
    })
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "rows": int(profile["value"].notna().sum()),
        "context": context,
        "period_label": year_text,
        "summary_cards": [
            {"label": "Months", "value": "12", "note": "January to December"},
            {"label": "Years", "value": str(monthly["year"].nunique()), "note": year_text},
            {"label": "Statistic", "value": "Average monthly total" if family == "rainfall" else "Average monthly mean", "note": element},
        ],
    }


_CDE_20260718_FINAL_PLOT_BASE = generate_plot_product


def generate_plot_product(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    if str(params.get("plot_type") or "") == "monthly_climatology":
        return _generate_average_monthly_profile(params, Path(data_dir), Path(export_dir))
    return _CDE_20260718_FINAL_PLOT_BASE(params, Path(data_dir), Path(export_dir))

# Prefer operational natural-feature layers when available under
# storage/zarr/shapefiles; retain refined built-in major features as fallback.
def _geometry_lines(geometry: dict[str, Any]) -> list[np.ndarray]:
    kind = str(geometry.get("type") or "")
    coords = geometry.get("coordinates") or []
    lines: list[np.ndarray] = []
    if kind == "LineString":
        array = np.asarray(coords, dtype=float)
        if array.ndim == 2 and len(array) >= 2:
            lines.append(array[:, :2])
    elif kind == "MultiLineString":
        for values in coords:
            array = np.asarray(values, dtype=float)
            if array.ndim == 2 and len(array) >= 2:
                lines.append(array[:, :2])
    elif kind in {"Polygon", "MultiPolygon", "GeometryCollection"}:
        if kind == "GeometryCollection":
            for item in geometry.get("geometries") or []:
                lines.extend(_geometry_lines(item))
        else:
            lines.extend(_geometry_polygons(geometry))
    return lines


def _natural_layer_files(data_dir: Path | None, kind: str) -> list[Path]:
    if not data_dir:
        return []
    root = Path(data_dir) / "shapefiles"
    if not root.is_dir():
        return []
    tokens = {
        "lakes": ("lake", "lakes", "waterbody", "inland_water", "inlandwater"),
        "rivers": ("river", "rivers", "waterway", "hydrograph", "stream"),
    }.get(kind, ())
    found: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in {".json", ".geojson", ".shp"}:
            continue
        name = path.stem.lower()
        if name.startswith("gadm"):
            continue
        if any(token in name for token in tokens):
            found.append(path)
    return sorted(found, key=lambda path: (0 if path.suffix.lower() in {".json", ".geojson"} else 1, len(path.name), path.name))


def _read_natural_layer(path: Path, *, polygons: bool) -> list[np.ndarray]:
    output: list[np.ndarray] = []
    try:
        if path.suffix.lower() in {".json", ".geojson"}:
            payload = json.loads(path.read_text(encoding="utf-8"))
            features = payload.get("features", []) if payload.get("type") == "FeatureCollection" else [{"geometry": payload}]
            for feature in features:
                geometry = feature.get("geometry") or {}
                output.extend(_geometry_polygons(geometry) if polygons else _geometry_lines(geometry))
        elif path.suffix.lower() == ".shp":
            import shapefile  # pyshp; lightweight and installed by requirements.txt
            reader = shapefile.Reader(str(path))
            for shape in reader.shapes():
                points = np.asarray(shape.points, dtype=float)
                if points.ndim != 2 or len(points) < (3 if polygons else 2):
                    continue
                starts = list(shape.parts) + [len(points)]
                for start, stop in zip(starts[:-1], starts[1:]):
                    part = points[start:stop, :2]
                    if len(part) >= (3 if polygons else 2):
                        output.append(part)
    except Exception:
        return []
    return output


def _draw_hydrology(ax, *, ocean: bool, lakes: bool, rivers: bool, data_dir: Path | None = None) -> None:
    """Draw ocean plus available lake/river layers, with safe built-in fallback."""
    from matplotlib.lines import Line2D
    from matplotlib.collections import LineCollection, PatchCollection
    legend_items = []
    ax.set_facecolor("#dcebf7" if ocean else "white")
    if ocean:
        ax.text(40.30, -7.5, "INDIAN OCEAN", fontsize=8.0, fontweight="bold", color="#557d9b", rotation=90,
                ha="center", va="center", alpha=0.90, zorder=13)
        legend_items.append(Line2D([0], [0], color="#b8d1e7", lw=6, label="Indian Ocean"))

    lake_polygons: list[np.ndarray] = []
    if lakes:
        for path in _natural_layer_files(data_dir, "lakes"):
            lake_polygons.extend(_read_natural_layer(path, polygons=True))
        if lake_polygons:
            patches = [Polygon(array, closed=True) for array in lake_polygons if len(array) >= 3]
            if patches:
                ax.add_collection(PatchCollection(patches, facecolor="#b9d9ee", edgecolor="#5e96b6", linewidths=0.55, zorder=9))
        else:
            built_in = {
                "Victoria": [(31.25,-0.78),(32.10,-0.63),(33.20,-0.82),(34.15,-1.33),(34.05,-2.22),(33.40,-2.52),(32.35,-2.55),(31.45,-2.02),(31.18,-1.35)],
                "Tanganyika": [(29.10,-3.30),(29.55,-3.42),(30.12,-4.45),(30.42,-5.75),(30.72,-7.15),(30.55,-8.35),(30.18,-8.74),(29.72,-7.72),(29.42,-6.18),(29.12,-4.55)],
                "Nyasa": [(34.42,-9.05),(34.82,-9.24),(35.18,-10.15),(35.62,-11.52),(35.28,-11.92),(34.86,-11.18),(34.58,-10.12)],
                "Rukwa": [(31.72,-7.05),(32.20,-7.12),(32.48,-7.43),(32.38,-8.10),(31.90,-8.20),(31.70,-7.72)],
                "Eyasi": [(34.62,-3.32),(35.05,-3.35),(35.18,-3.72),(34.83,-3.90),(34.52,-3.66)],
                "Natron": [(35.72,-1.92),(36.02,-1.96),(36.08,-2.50),(35.82,-2.67),(35.66,-2.32)],
            }
            for vertices in built_in.values():
                ax.add_patch(Polygon(vertices, closed=True, facecolor="#b9d9ee", edgecolor="#5e96b6", linewidth=0.62, zorder=9))
            ax.text(32.65, -1.60, "Victoria", fontsize=5.8, color="#4f7893", ha="center", zorder=10)
            ax.text(29.92, -6.10, "Tanganyika", fontsize=5.6, color="#4f7893", rotation=76, ha="center", zorder=10)
            ax.text(35.03, -10.55, "Nyasa", fontsize=5.6, color="#4f7893", rotation=72, ha="center", zorder=10)
        legend_items.append(Line2D([0], [0], color="#5e96b6", lw=4, label="Lakes"))

    river_lines: list[np.ndarray] = []
    if rivers:
        for path in _natural_layer_files(data_dir, "rivers"):
            river_lines.extend(_read_natural_layer(path, polygons=False))
        if not river_lines:
            river_lines = [np.asarray(points, dtype=float) for points in {
                "Rufiji": [(34.4,-7.55),(35.2,-7.68),(36.0,-7.80),(36.8,-8.02),(37.6,-8.02),(38.5,-7.78),(39.30,-7.82)],
                "Ruvuma": [(34.9,-10.58),(35.8,-10.72),(36.8,-10.86),(37.8,-10.72),(38.8,-10.48)],
                "Pangani": [(36.55,-3.10),(36.83,-3.62),(37.20,-4.16),(37.70,-4.62),(38.25,-5.02),(39.05,-5.42)],
                "Wami": [(35.55,-5.65),(36.35,-5.80),(37.20,-6.00),(38.10,-6.10),(38.85,-6.18)],
                "Malagarasi": [(30.15,-4.05),(30.85,-4.26),(31.45,-4.63),(32.05,-4.92),(32.58,-5.05)],
            }.values()]
        if river_lines:
            ax.add_collection(LineCollection(river_lines, colors="#3f91bc", linewidths=0.55, alpha=0.76, zorder=10))
        legend_items.append(Line2D([0], [0], color="#3f91bc", lw=1.4, label="Rivers"))
    if legend_items:
        ax.legend(handles=legend_items, loc="lower left", fontsize=6.5, frameon=True, framealpha=0.92, borderpad=0.35)

# ---------------------------------------------------------------------------
# FINAL 2026-07-18 UPDATE: independent map periods, map surfaces and products
# ---------------------------------------------------------------------------
_NEW_SPATIAL_PRODUCTS = [
    ("temperature_monthly_profile", "Combined Mean, Minimum and Maximum Monthly Profile"),
    ("spatial_climatology", "Spatial Climatology Map"),
    ("spatial_monthly_climatology", "Monthly Spatial Climatology Maps (Jan–Dec)"),
    ("spatial_annual_series", "Annual Spatial Maps for Selected Years"),
    ("spatial_seasonal_climatology", "Seasonal Spatial Climatology Maps"),
]
_existing_plot_keys = {key for key, _label in PLOT_TYPES}
PLOT_TYPES.extend((key, label) for key, label in _NEW_SPATIAL_PRODUCTS if key not in _existing_plot_keys)
for _family, _items in PLOT_FAMILIES.items():
    for _key in ("spatial_climatology", "spatial_monthly_climatology", "spatial_annual_series", "spatial_seasonal_climatology"):
        if _key not in _items:
            _items.append(_key)
if "temperature_monthly_profile" not in PLOT_FAMILIES.get("temperature", []):
    PLOT_FAMILIES["temperature"].append("temperature_monthly_profile")


def _new_map_period_definition(params: Dict[str, Any]) -> dict[str, Any]:
    """Read month and year selections independently while retaining old requests."""
    start = int(params.get("start_year") or 1991)
    end = int(params.get("end_year") or start)
    if start > end:
        start, end = end, start

    month_mode = str(params.get("map_month_selection") or "").strip().lower()
    year_mode = str(params.get("map_year_selection") or "").strip().lower()
    old_mode = str(params.get("map_period_mode") or "").strip().lower()
    if not month_mode:
        month_mode = {"month": "single", "season": "season", "year": "all", "custom": "custom"}.get(old_mode, "all")
    if not year_mode:
        year_mode = "custom" if old_mode == "custom" else "single"

    if month_mode == "single":
        months = _parse_month_numbers(params.get("map_month"), default=[1])[:1]
        season = None
    elif month_mode == "season":
        season = str(params.get("map_season") or "MAM").strip().upper()
        months = list(SEASON_DEFINITIONS.get(season, SEASON_DEFINITIONS["MAM"]))
    elif month_mode == "custom":
        season = None
        months = _parse_month_numbers(params.get("map_custom_months"), default=range(1, 13))
    else:
        month_mode = "all"
        season = None
        months = list(range(1, 13))

    map_year = int(params.get("map_year") or end)
    if year_mode == "range":
        first = int(params.get("map_start_year") or start)
        last = int(params.get("map_end_year") or end)
        if first > last:
            first, last = last, first
        years = list(range(first, last + 1))
    elif year_mode == "custom":
        years = _parse_year_numbers(params.get("map_custom_years"), start=start, end=end)
    else:
        year_mode = "single"
        years = [map_year]

    month_label = (
        "All months" if months == list(range(1, 13))
        else season if season
        else ", ".join(_month_names(months))
    )
    period_label = f"{month_label} · {_compact_year_label(years)}"
    preferred_resolution = "annual" if len(years) == 1 and months == list(range(1, 13)) else "monthly"
    if season and len(years) == 1:
        preferred_resolution = "seasonal"
    return {
        "mode": "custom",
        "month_mode": month_mode,
        "year_mode": year_mode,
        "months": months,
        "years": years,
        "season": season,
        "period_label": period_label,
        "preferred_resolution": preferred_resolution,
    }


def _spatial_period_definition(params: Dict[str, Any]) -> dict[str, Any]:
    return _new_map_period_definition(params)


def _spatial_panel_jobs(params: Dict[str, Any]) -> list[tuple[Dict[str, Any], str]]:
    period = _new_map_period_definition(params)
    years = list(period["years"])
    months = list(period["months"])
    basis = str(params.get("map_panel_basis") or "auto").strip().lower()
    if basis not in {"auto", "year", "month", "season"}:
        basis = "auto"
    if basis == "auto":
        basis = "year" if len(years) > 1 else ("month" if len(months) > 1 and months != list(range(1, 13)) else "year")

    jobs: list[tuple[Dict[str, Any], str]] = []
    if basis == "month":
        panel_months = list(range(1, 13)) if period["month_mode"] == "all" else months
        for month in panel_months:
            item = dict(params)
            item.update({
                "map_month_selection": "single", "map_month": str(month),
                "map_year_selection": "custom", "map_custom_years": ",".join(map(str, years)),
                "map_output_layout": "single",
            })
            jobs.append((item, pd.Timestamp(2000, month, 1).strftime("%B")))
    elif basis == "season":
        selected = params.get("map_panel_seasons") or params.get("seasons") or ["MAM", "JJA", "SON", "DJF"]
        if isinstance(selected, str):
            selected = [part.strip().upper() for part in re.split(r"[,;\s]+", selected) if part.strip()]
        seasons = [value for value in selected if value in SEASON_DEFINITIONS] or ["MAM", "JJA", "SON", "DJF"]
        for season in seasons:
            item = dict(params)
            item.update({
                "map_month_selection": "season", "map_season": season,
                "map_year_selection": "custom", "map_custom_years": ",".join(map(str, years)),
                "map_output_layout": "single",
            })
            jobs.append((item, season))
    else:
        for year in years:
            item = dict(params)
            item.update({
                "map_year_selection": "single", "map_year": str(year),
                "map_month_selection": "custom" if months != list(range(1, 13)) else "all",
                "map_custom_months": ",".join(map(str, months)),
                "map_output_layout": "single",
            })
            jobs.append((item, str(year)))
    if not jobs:
        jobs = [(dict(params), period["period_label"])]
    if len(jobs) > 30:
        raise ValueError("To protect the server from overload, one multi-panel figure is limited to 30 maps. Use Custom Years or generate the remaining years in a second request.")
    return jobs


def _explicit_weather_element(dataset_key: str, variable: Any, context: Dict[str, Any]) -> str:
    label = _clean_map_element_name(variable_display_name(variable, context, dataset_key))
    generic = {"", "auto", "weather element", "value", "variable"}
    if label.strip().lower() in generic:
        family = str(DATASETS.get(dataset_key, {}).get("family") or context.get("family") or "")
        label = {
            "rainfall": "Precipitation",
            "temperature": "Temperature",
            "humidity": "Relative Humidity",
            "soil_moisture": "Volumetric Soil Moisture",
            "wind": "Wind Speed",
            "pressure_cloud": "Surface Pressure or Total Cloud Cover",
        }.get(family, DATASETS.get(dataset_key, {}).get("label", "Climate Variable"))
    label = re.sub(r"\s+\([^)]*source[^)]*\)$", "", str(label), flags=re.I).strip()
    return label.replace("Rainfall", "Precipitation")


def _draw_north_arrow(ax) -> None:
    """Publication-style four-point compass rose matching the supplied maps."""
    from matplotlib.patches import Circle, Polygon as MplPolygon
    cx, cy = 0.865, 0.805
    size = 0.075
    transform = ax.transAxes
    # Alternating black/white spearheads provide a readable compass on any map.
    triangles = [
        ([(cx, cy + size), (cx - 0.018, cy + 0.012), (cx + 0.018, cy + 0.012)], "black"),
        ([(cx, cy - size), (cx - 0.018, cy - 0.012), (cx + 0.018, cy - 0.012)], "white"),
        ([(cx + size, cy), (cx + 0.012, cy - 0.018), (cx + 0.012, cy + 0.018)], "black"),
        ([(cx - size, cy), (cx - 0.012, cy - 0.018), (cx - 0.012, cy + 0.018)], "white"),
    ]
    for vertices, face in triangles:
        ax.add_patch(MplPolygon(vertices, closed=True, transform=transform, facecolor=face, edgecolor="black", linewidth=1.0, zorder=25, clip_on=False))
    ax.add_patch(Circle((cx, cy), 0.020, transform=transform, facecolor="white", edgecolor="black", linewidth=1.0, zorder=26, clip_on=False))
    ax.text(cx, cy + size + 0.018, "N", transform=transform, ha="center", va="center", fontsize=9, fontweight="bold", zorder=27)
    ax.text(cx + size + 0.020, cy, "E", transform=transform, ha="center", va="center", fontsize=8, fontweight="bold", zorder=27)
    ax.text(cx, cy - size - 0.018, "S", transform=transform, ha="center", va="center", fontsize=8, fontweight="bold", zorder=27)
    ax.text(cx - size - 0.020, cy, "W", transform=transform, ha="center", va="center", fontsize=8, fontweight="bold", zorder=27)


def _render_tanzania_map_axis(
    ax, *, x: np.ndarray, y: np.ndarray, values: np.ndarray, data_dir: Path,
    selected_boundaries: list[np.ndarray], level: int, dataset_key: str,
    show_ocean: bool, show_lakes: bool, show_rivers: bool, title: str,
    norm=None, show_labels: bool = True, show_cartographic_elements: bool = True,
    render_style: str = "grid",
):
    _draw_hydrology(ax, ocean=show_ocean, lakes=show_lakes, rivers=show_rivers, data_dir=data_dir)
    # Preserve every native 0.25-degree source grid cell. Smoothing and
    # spatial interpolation are intentionally disabled for all map requests.
    render_style = "grid"
    cmap = plot_cmap_for(dataset_key)
    im = ax.pcolormesh(x, y, values, shading="nearest", cmap=cmap, norm=norm, zorder=4)
    _draw_admin_boundaries(ax, selected_boundaries, level)
    try:
        national_like, _ = _admin_polygons(data_dir, 1)
        from matplotlib.collections import LineCollection
        ax.add_collection(LineCollection(national_like, colors="#111820", linewidths=1.05, alpha=0.95, zorder=11))
    except Exception:
        pass
    if show_labels and level == 1:
        for name, lon, lat in _admin_feature_labels(data_dir, 1):
            if TANZANIA_BOUNDS["lon_min"] <= lon <= TANZANIA_BOUNDS["lon_max"] and TANZANIA_BOUNDS["lat_min"] <= lat <= TANZANIA_BOUNDS["lat_max"]:
                ax.text(lon, lat, name, fontsize=5.7, fontweight="bold", color="#16232d", ha="center", va="center", zorder=12)
    ax.set_xlim(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
    ax.set_ylim(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    ax.set_title(title, fontsize=9.0 if not show_cartographic_elements else 13.0, fontweight="bold", pad=8)
    if show_cartographic_elements:
        ax.set_xlabel("Longitude"); ax.set_ylabel("Latitude")
        _draw_north_arrow(ax); _draw_scale_bar(ax); apply_plot_grids(ax)
    else:
        ax.set_xticks([]); ax.set_yticks([])
    return im


def _prepare_spatial_jobs(params: Dict[str, Any], data_dir: Path):
    layout = str(params.get("map_output_layout") or "single").lower()
    if layout == "panel":
        raw_jobs = []
        for job_params, label in _spatial_panel_jobs(params):
            grid, context = _extract_spatial_period(data_dir, job_params)
            raw_jobs.append((grid, context, label))
        if len(raw_jobs) <= 1:
            layout = "single"
    else:
        grid, context = _extract_spatial_period(data_dir, params)
        raw_jobs = [(grid, context, context.get("period_label") or "Selected period")]
    return layout, raw_jobs


def _generate_requested_spatial_map(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    layout, jobs = _prepare_spatial_jobs(params, Path(data_dir))
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    level = max(1, min(3, int(params.get("map_admin_level") or 1)))
    boundaries, boundary_path = _admin_polygons(Path(data_dir), level)
    national, _ = _admin_polygons(Path(data_dir), 1)
    admin_label = {1: "Admin 1 — Regions", 2: "Admin 2 — Districts", 3: "Admin 3 — Wards"}[level]
    prepared = []
    finite_parts = []
    for grid, context, label in jobs:
        work = grid[grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"]) & grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])].copy()
        if work.empty:
            continue
        pivot = work.pivot_table(index="latitude", columns="longitude", values="value", aggfunc="mean").sort_index()
        x = pivot.columns.to_numpy(float); y = pivot.index.to_numpy(float)
        values = _mask_grid_to_tanzania(pivot.to_numpy(float), x, y, national)
        prepared.append((x, y, values, work, context, label))
        finite = values[np.isfinite(values)]
        if finite.size: finite_parts.append(finite)
    if not prepared:
        raise ValueError("No spatial values were found within the Tanzania map extent.")
    from matplotlib.colors import Normalize
    finite = np.concatenate(finite_parts) if finite_parts else np.asarray([0.0, 1.0])
    vmin, vmax = float(np.nanmin(finite)), float(np.nanmax(finite))
    if np.isclose(vmin, vmax): vmax = vmin + 1.0
    norm = Normalize(vmin=vmin, vmax=vmax)
    element = _explicit_weather_element(dataset_key, params.get("variable"), prepared[0][4])
    dirs = ensure_output_dirs(Path(export_dir)); stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    style = str(params.get("map_render_style") or "grid").lower()
    stem = _cde_prefixed_stem(f"spatial_{style}_{dataset_key}_{'panels' if layout == 'panel' else slugify(prepared[0][5])}_admin{level}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    show_ocean = _truthy(params.get("show_ocean", True)); show_lakes = _truthy(params.get("show_lakes", True)); show_rivers = _truthy(params.get("show_rivers", True))
    image = None
    if layout == "panel":
        count = len(prepared); columns = min(6, max(2, int(math.ceil(math.sqrt(count * 1.15))))); rows = int(math.ceil(count / columns))
        fig, axes = plt.subplots(rows, columns, figsize=(2.55 * columns + 0.7, 2.60 * rows), squeeze=False)
        flat = axes.ravel()
        for idx, (x, y, values, _work, _context, label) in enumerate(prepared):
            image = _render_tanzania_map_axis(flat[idx], x=x, y=y, values=values, data_dir=Path(data_dir), selected_boundaries=boundaries, level=level, dataset_key=dataset_key, show_ocean=show_ocean, show_lakes=show_lakes, show_rivers=False, title=label, norm=norm, show_labels=False, show_cartographic_elements=False, render_style=style)
            if idx == 0:
                _draw_north_arrow(flat[idx])
        for ax in flat[len(prepared):]: ax.axis("off")
        fig.suptitle(f"{element} Spatial Distribution over Tanzania\n{_compact_year_label(prepared[0][4].get('years', []))} · {admin_label}", fontsize=15, fontweight="bold", y=0.995)
        if image is not None:
            cax = fig.add_axes([0.948, 0.12, 0.014, 0.75])
            cbar = fig.colorbar(image, cax=cax, extend="both")
            cbar.set_label(y_axis_label(element, prepared[0][4].get("unit")))
        fig.subplots_adjust(left=0.025, right=0.925, bottom=0.025, top=0.92, wspace=0.06, hspace=0.16)
        fig.savefig(plot_path, dpi=165, bbox_inches="tight", facecolor="white"); plt.close(fig)
    else:
        x, y, values, _work, context, label = prepared[0]
        fig = plt.figure(figsize=(12.1, 9.0))
        gs = fig.add_gridspec(1, 2, width_ratios=[1.0, 0.040], wspace=0.10)
        ax = fig.add_subplot(gs[0, 0]); cax = fig.add_subplot(gs[0, 1])
        heading = _spatial_map_heading(element, context)
        image = _render_tanzania_map_axis(ax, x=x, y=y, values=values, data_dir=Path(data_dir), selected_boundaries=boundaries, level=level, dataset_key=dataset_key, show_ocean=show_ocean, show_lakes=show_lakes, show_rivers=show_rivers, title=f"{heading}\n{admin_label}", norm=norm, show_labels=True, show_cartographic_elements=True, render_style=style)
        cbar = fig.colorbar(image, cax=cax, extend="both")
        cbar.set_label(y_axis_label(element, context.get("unit")))
        fig.savefig(plot_path, dpi=200, bbox_inches="tight", facecolor="white"); plt.close(fig)
    export_parts=[]
    for _x,_y,_values,grid,context,label in prepared:
        part=grid.copy(); part.insert(0,"map_period",label); part.insert(1,"administrative_level",admin_label); part.insert(2,"map_surface",style.title()); export_parts.append(part)
    export_grid=pd.concat(export_parts,ignore_index=True)
    excel_path,_,_=_save_data(export_grid,dirs["plots"],stem+"_data",table="plot_products")
    context=dict(prepared[0][4]); context.update({"administrative_level":level,"administrative_level_label":admin_label,"boundary_file":boundary_path.name,"map_output_layout":layout,"map_count":len(prepared),"map_render_style":style,"variable_label":element})
    return {"plot_path":plot_path,"excel_path":excel_path,"rows":len(export_grid),"context":context,"period_label":"All selected periods" if layout=="panel" else prepared[0][5],"summary_cards":[{"label":"Weather Element","value":element,"note":context.get("period_label")},{"label":"Map Surface","value":"Smooth" if style=="smooth" else "Grid cells","note":f"{len(prepared)} map panel(s)"},{"label":"Boundary Level","value":admin_label,"note":boundary_path.name}]}


def _generate_spatial_variability_map(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    plot_type = str(params.get("plot_type") or "spatial_std_map")
    statistic = "cv" if plot_type == "spatial_cv_map" else "std"
    grid, context = _extract_spatial_variability(Path(data_dir), params, statistic)
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    level = max(1, min(3, int(params.get("map_admin_level") or 1)))
    boundaries, boundary_path = _admin_polygons(Path(data_dir), level); national,_ = _admin_polygons(Path(data_dir),1)
    work=grid[grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"],TANZANIA_BOUNDS["lon_max"]) & grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"],TANZANIA_BOUNDS["lat_max"])].copy()
    pivot=work.pivot_table(index="latitude",columns="longitude",values="value",aggfunc="mean").sort_index(); x=pivot.columns.to_numpy(float); y=pivot.index.to_numpy(float); values=_mask_grid_to_tanzania(pivot.to_numpy(float),x,y,national)
    element=_explicit_weather_element(dataset_key,params.get("variable"),context); name="Coefficient of Variation" if statistic=="cv" else "Standard Deviation"; admin_label={1:"Admin 1 — Regions",2:"Admin 2 — Districts",3:"Admin 3 — Wards"}[level]
    dirs=ensure_output_dirs(Path(export_dir)); stamp=datetime.now().strftime("%Y%m%d_%H%M%S"); style=str(params.get("map_render_style") or "grid").lower(); stem=_cde_prefixed_stem(f"{plot_type}_{style}_{dataset_key}_{slugify(context['period_label'])}_admin{level}_{stamp}"); plot_path=dirs["plots"]/f"{stem}.png"
    fig=plt.figure(figsize=(12.1,9.0)); gs=fig.add_gridspec(1,2,width_ratios=[1.0,0.025],wspace=0.055); ax=fig.add_subplot(gs[0,0]); cax=fig.add_subplot(gs[0,1])
    image=_render_tanzania_map_axis(ax,x=x,y=y,values=values,data_dir=Path(data_dir),selected_boundaries=boundaries,level=level,dataset_key=dataset_key,show_ocean=_truthy(params.get("show_ocean",True)),show_lakes=_truthy(params.get("show_lakes",True)),show_rivers=_truthy(params.get("show_rivers",True)),title=f"Spatial {name} of {element} — {context['period_label']}\n{admin_label}",show_labels=True,render_style=style)
    cbar=fig.colorbar(image,cax=cax,extend="max"); cbar.set_label("Percent (%)" if statistic=="cv" else y_axis_label(element,context.get("unit"))); fig.savefig(plot_path,dpi=200,bbox_inches="tight",facecolor="white"); plt.close(fig)
    export_grid=work.copy(); export_grid.insert(0,"period",context["period_label"]); export_grid.insert(1,"statistic",name); export_grid.insert(2,"administrative_level",admin_label); export_grid.insert(3,"map_surface",style.title()); excel_path,_,_=_save_data(export_grid,dirs["plots"],stem+"_data",table="plot_products")
    context.update({"administrative_level_label":admin_label,"boundary_file":boundary_path.name,"variable_label":element,"map_render_style":style})
    return {"plot_path":plot_path,"excel_path":excel_path,"rows":len(export_grid),"context":context,"summary_cards":[{"label":"Weather Element","value":element,"note":name},{"label":"Map Surface","value":"Smooth" if style=="smooth" else "Grid cells","note":context["period_label"]},{"label":"Boundary Level","value":admin_label,"note":boundary_path.name}]}


def _generate_combined_temperature_monthly_profile(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    dataset_key = str(params.get("dataset") or "era5_temperature_stats")
    if dataset_key not in {"era5_temperature", "era5_temperature_stats"}:
        raise ValueError("The combined mean, minimum and maximum monthly profile is available only for temperature datasets.")
    location=str(params.get("location_name") or "Selected Location"); resolution=str(params.get("resolution") or "monthly").lower(); start=str(params.get("start_date") or ""); end=str(params.get("end_date") or "")
    rows=[]; context=None
    for variable,label in (("ta","Mean Temperature"),("tmin","Minimum Temperature"),("tmax","Maximum Temperature")):
        frame,ctx=extract_point_series(Path(data_dir),dataset_key,resolution,float(params.get("latitude")),float(params.get("longitude")),start,end,variable=variable,season=params.get("season")); context=context or ctx
        work=frame[["time","value"]].copy(); work["time"]=pd.to_datetime(work["time"]); work["value"]=pd.to_numeric(work["value"],errors="coerce"); work=work.dropna(); work["year"]=work["time"].dt.year; work["month"]=work["time"].dt.month
        monthly=work.groupby(["year","month"],as_index=False)["value"].mean(); profile=monthly.groupby("month",as_index=False)["value"].mean().rename(columns={"value":variable}); rows.append(profile)
    profile=pd.DataFrame({"month":range(1,13)})
    for row in rows: profile=profile.merge(row,on="month",how="left")
    profile["month_name"]=profile["month"].map(lambda month:pd.Timestamp(2000,int(month),1).strftime("Jan") if False else pd.Timestamp(2000,int(month),1).strftime("%b"))
    years=pd.date_range(start=start or "1991-01-01",end=end or "2020-12-31",freq="YS").year.tolist(); year_text=_compact_year_label(years)
    dirs=ensure_output_dirs(Path(export_dir)); stamp=datetime.now().strftime("%Y%m%d_%H%M%S"); stem=_cde_prefixed_stem(f"combined_temperature_monthly_{slugify(location)}_{stamp}"); plot_path=dirs["plots"]/f"{stem}.png"
    fig,ax=plt.subplots(figsize=(11.2,6.3)); x=np.arange(1,13); ax.plot(x,profile["ta"],marker="o",linewidth=2,label="Mean Temperature"); ax.plot(x,profile["tmin"],marker="o",linewidth=2,label="Minimum Temperature"); ax.plot(x,profile["tmax"],marker="o",linewidth=2,label="Maximum Temperature"); ax.set_xticks(x); ax.set_xticklabels(profile["month_name"]); ax.set_xlabel("Month"); ax.set_ylabel("Temperature (°C)"); ax.set_title(f"Average Monthly Mean, Minimum and Maximum Temperature ({year_text}) — {location}",fontsize=14,fontweight="bold",pad=13); ax.legend(); apply_plot_grids(ax); _finalize_plot(fig,plot_path)
    excel_path,_,_=_save_data(profile,dirs["plots"],stem+"_data",table="plot_products"); context=dict(context or {}); context.update({"variable_label":"Mean, Minimum and Maximum Temperature","period_label":year_text})
    return {"plot_path":plot_path,"excel_path":excel_path,"rows":12,"context":context,"period_label":year_text,"summary_cards":[{"label":"Months","value":"12","note":"January to December"},{"label":"Weather Elements","value":"3","note":"Mean, minimum and maximum"},{"label":"Period","value":year_text,"note":location}]}


_CDE_20260718_MAP_SURFACE_BASE = generate_plot_product


def generate_plot_product(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    plot_type = str(params.get("plot_type") or "time_series")
    prepared = dict(params)
    if plot_type == "temperature_monthly_profile":
        return _generate_combined_temperature_monthly_profile(prepared, Path(data_dir), Path(export_dir))
    if plot_type == "spatial_climatology":
        prepared.update({"plot_type":"spatial_map","map_output_layout":"single","map_month_selection":prepared.get("map_month_selection") or "all","map_year_selection":prepared.get("map_year_selection") or "range"})
        return _generate_requested_spatial_map(prepared,Path(data_dir),Path(export_dir))
    if plot_type == "spatial_monthly_climatology":
        prepared.update({"plot_type":"spatial_map","map_output_layout":"panel","map_panel_basis":"month","map_month_selection":"all","map_year_selection":prepared.get("map_year_selection") or "range"})
        return _generate_requested_spatial_map(prepared,Path(data_dir),Path(export_dir))
    if plot_type == "spatial_annual_series":
        prepared.update({"plot_type":"spatial_map","map_output_layout":"panel","map_panel_basis":"year","map_month_selection":"all","map_year_selection":prepared.get("map_year_selection") or "range"})
        return _generate_requested_spatial_map(prepared,Path(data_dir),Path(export_dir))
    if plot_type == "spatial_seasonal_climatology":
        prepared.update({"plot_type":"spatial_map","map_output_layout":"panel","map_panel_basis":"season","map_year_selection":prepared.get("map_year_selection") or "range"})
        return _generate_requested_spatial_map(prepared,Path(data_dir),Path(export_dir))
    if plot_type == "spatial_map":
        return _generate_requested_spatial_map(prepared,Path(data_dir),Path(export_dir))
    if plot_type in {"spatial_std_map","spatial_cv_map"}:
        return _generate_spatial_variability_map(prepared,Path(data_dir),Path(export_dir))
    return _CDE_20260718_MAP_SURFACE_BASE(prepared,Path(data_dir),Path(export_dir))




# Matplotlib 3.9+ renamed ``labels`` to ``tick_labels``.  This helper supports
# both the newer deployment environment and older installations.
def _boxplot_with_labels(ax, data, labels, **kwargs):
    import inspect
    parameter = "tick_labels" if "tick_labels" in inspect.signature(ax.boxplot).parameters else "labels"
    return ax.boxplot(data, **{parameter: labels}, **kwargs)

# ---------------------------------------------------------------------------
# 2026-07-18 FINAL UPDATE: expanded plots/indices, lakes-only maps, Parquet-only
# storage, standard variable colours, and explicit weather-element names.
# ---------------------------------------------------------------------------

# User-facing dataset names and variables must always be explicit.
DATASETS["era5_temperature"]["label"] = "ERA5 2m Temperature"
DATASETS["era5_temperature_stats"]["label"] = "ERA5 Temperature Mean, Minimum and Maximum"
DATASETS["era5_dew_point"]["label"] = "ERA5 Dew Point Temperature at 2m"
DATASETS["era5_relative_humidity"]["label"] = "ERA5 Relative Humidity"
DATASETS["era5_skin_temperature"]["label"] = "ERA5 Skin Temperature"
DATASETS["era5_soil_temperature"]["label"] = "ERA5-Land Soil Temperature Level 1"
DATASETS["era5_pressure_cloud"]["label"] = "ERA5 Surface Pressure and Total Cloud Cover"
DATASETS["era5_soil_water"]["label"] = "ERA5-Land Volumetric Soil Moisture"
DATASETS["era5_wind"]["label"] = "ERA5 Wind Speed and Direction at 10m"

_EXPLICIT_ELEMENT_NAMES = {
    ("chirps_rainfall", "precip"): "CHIRPS Precipitation",
    ("chirps_rainfall", "auto"): "CHIRPS Precipitation",
    ("era5_total_precipitation", "tp"): "ERA5 Precipitation",
    ("era5_total_precipitation", "auto"): "ERA5 Precipitation",
    ("era5_temperature", "ta"): "Mean 2m Temperature",
    ("era5_temperature", "t2m"): "Mean 2m Temperature",
    ("era5_temperature", "auto"): "Mean 2m Temperature",
    ("era5_temperature_stats", "ta"): "Mean Temperature",
    ("era5_temperature_stats", "tmean"): "Mean Temperature",
    ("era5_temperature_stats", "tmin"): "Minimum Temperature",
    ("era5_temperature_stats", "tn"): "Minimum Temperature",
    ("era5_temperature_stats", "tmax"): "Maximum Temperature",
    ("era5_temperature_stats", "tx"): "Maximum Temperature",
    ("era5_dew_point", "d2m"): "Dew Point Temperature at 2m",
    ("era5_dew_point", "auto"): "Dew Point Temperature at 2m",
    ("era5_relative_humidity", "r"): "Relative Humidity",
    ("era5_relative_humidity", "rh"): "Relative Humidity",
    ("era5_relative_humidity", "auto"): "Relative Humidity",
    ("era5_skin_temperature", "skt"): "Skin Temperature",
    ("era5_skin_temperature", "auto"): "Skin Temperature",
    ("era5_soil_temperature", "stl1"): "Soil Temperature Level 1",
    ("era5_soil_temperature", "auto"): "Soil Temperature Level 1",
    ("era5_soil_water", "swvl1"): "Volumetric Soil Moisture",
    ("era5_soil_water", "auto"): "Volumetric Soil Moisture",
    ("era5_pressure_cloud", "sp"): "Surface Pressure",
    ("era5_pressure_cloud", "msl"): "Mean Sea-Level Pressure",
    ("era5_pressure_cloud", "tcc"): "Total Cloud Cover",
    ("era5_pressure_cloud", "auto"): "Surface Pressure",
    ("era5_wind", "wind_speed"): "Wind Speed at 10m",
    ("era5_wind", "wind_direction"): "Wind Direction at 10m",
    ("era5_wind", "auto"): "Wind Speed at 10m",
}


def variable_display_name(variable: str | None, ctx: Dict[str, Any] | None = None, dataset_key: str | None = None) -> str:
    key = str(dataset_key or (ctx or {}).get("dataset_key") or "")
    var = str(variable or (ctx or {}).get("variable") or "auto").lower()
    if (key, var) in _EXPLICIT_ELEMENT_NAMES:
        return _EXPLICIT_ELEMENT_NAMES[(key, var)]
    context_label = str((ctx or {}).get("variable_label") or "").strip()
    if context_label and context_label.lower() not in {"weather element", "variable", "value"}:
        return context_label.replace("Rainfall", "Precipitation")
    return DATASETS.get(key, {}).get("label", str(variable or "Climate Variable")).replace("Rainfall", "Precipitation")


# Operationally familiar colours: precipitation blue, temperature cool-to-warm,
# humidity/soil water blue-green, cloud grey, wind viridis, pressure spectral.
DATASET_CMAPS.update({
    "chirps_rainfall": "Blues",
    "era5_total_precipitation": "Blues",
    "era5_temperature": "RdYlBu_r",
    "era5_temperature_stats": "RdYlBu_r",
    "era5_dew_point": "coolwarm",
    "era5_relative_humidity": "YlGnBu",
    "era5_skin_temperature": "inferno",
    "era5_soil_temperature": "coolwarm",
    "era5_pressure_cloud": "Spectral_r",
    "era5_soil_water": "YlGnBu",
    "era5_wind": "viridis",
})
VARIABLE_COLORS.update({
    "precip": "#1f78b4", "tp": "#1f78b4", "rainfall": "#1f78b4", "precipitation": "#1f78b4",
    "ta": "#f28e2b", "t2m": "#f28e2b", "tmean": "#f28e2b",
    "tmax": "#d73027", "tx": "#d73027", "maximum_temperature": "#d73027",
    "tmin": "#4575b4", "tn": "#4575b4", "minimum_temperature": "#4575b4",
    "d2m": "#7b3294", "r": "#2c7fb8", "rh": "#2c7fb8",
    "skt": "#e6550d", "stl1": "#a6611a", "swvl1": "#238b45",
    "sp": "#756bb1", "tcc": "#636363", "wind_speed": "#2c7fb8", "wind_direction": "#6a51a3",
})


def plot_color_for(dataset_key: str | None = None, variable: str | None = None, index_type: str | None = None) -> str:
    if index_type and index_type in INDEX_COLORS:
        return INDEX_COLORS[index_type]
    return VARIABLE_COLORS.get(str(variable or "").lower(), DATASET_COLORS.get(str(dataset_key or ""), "#1f77b4"))


def plot_cmap_for(dataset_key: str | None = None) -> str:
    return DATASET_CMAPS.get(str(dataset_key or ""), "viridis")


# Restore Parquet database persistence after earlier compatibility layers.
def _append_parquet_database(table: str, df: pd.DataFrame, stem: str) -> Path:
    db_dir = PARQUET_DB_DIR / slugify(table)
    db_dir.mkdir(parents=True, exist_ok=True)
    db_path = db_dir / f"{slugify(stem)}.parquet"
    df.to_parquet(db_path, index=False)
    return db_path


# Save every generated table as Excel for download and Parquet for persistence.
# No CSV file is created by this final implementation.
def _save_data(df: pd.DataFrame, out_dir: Path, stem: str, table: str = "products") -> Tuple[Path, Path, Path]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    clean_df = _compact_export_dataframe(df)
    excel_path = out_dir / f"{stem}.xlsx"
    parquet_path = out_dir / f"{stem}.parquet"
    clean_df.to_parquet(parquet_path, index=False)
    db_path = _append_parquet_database(table, clean_df, stem)
    try:
        _cde_write_single_sheet_workbook(
            excel_path,
            [("Data", clean_df)],
            qr_payload=_cde_default_download_context(excel_path),
            zero_decimal=False,
            sheet_name="Data",
            workbook_title="Climate Data Extractor Product",
        )
    except Exception:
        clean_df.to_excel(excel_path, index=False, sheet_name="Data")
    return excel_path, parquet_path, db_path


# Lakes are drawn only from the operational Tanzania GeoJSON. Ocean and rivers
# are intentionally not rendered and are not exposed as user options.
def _draw_hydrology(ax, *, ocean: bool = False, lakes: bool = False, rivers: bool = False, data_dir: Path | None = None) -> None:
    from matplotlib.collections import PatchCollection
    from matplotlib.lines import Line2D
    ax.set_facecolor("white")
    if not lakes or data_dir is None:
        return
    lake_path = Path(data_dir) / "shapefiles" / "hydrography" / "tanzania_lakes.geojson"
    polygons = _read_natural_layer(lake_path, polygons=True) if lake_path.exists() else []
    patches = [Polygon(arr, closed=True) for arr in polygons if len(arr) >= 3]
    if patches:
        ax.add_collection(PatchCollection(
            patches, facecolor="#b9dff2", edgecolor="#2b83ba", linewidths=0.55,
            alpha=0.98, zorder=9,
        ))
        ax.legend(handles=[Line2D([0], [0], color="#2b83ba", lw=4, label="Lakes")],
                  loc="lower left", fontsize=6.5, frameon=True, framealpha=0.92)


_EXTENDED_PLOTS = [
    ("step_plot", "Step Line Plot"),
    ("lollipop", "Lollipop Plot"),
    ("rolling_mean", "Rolling Mean Plot"),
    ("cumulative_total", "Cumulative Total Plot"),
    ("percentile_band", "Monthly Percentile Envelope Plot"),
    ("exceedance_curve", "Exceedance Probability Curve"),
    ("ecdf", "Empirical Cumulative Distribution Plot"),
    ("violin", "Monthly Violin Distribution Plot"),
    ("diurnal_cycle", "Average Diurnal Cycle Plot"),
    ("rank_plot", "Ranked Value Plot"),
]
_existing_plot_keys = {key for key, _ in PLOT_TYPES}
PLOT_TYPES.extend((key, label) for key, label in _EXTENDED_PLOTS if key not in _existing_plot_keys)
for family, items in PLOT_FAMILIES.items():
    for key in ["step_plot", "lollipop", "rolling_mean", "percentile_band", "exceedance_curve", "ecdf", "violin", "rank_plot"]:
        if key not in items:
            items.append(key)
    if family == "rainfall" and "cumulative_total" not in items:
        items.append("cumulative_total")
    if "diurnal_cycle" not in items:
        items.append("diurnal_cycle")


def _extended_plot_frame(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame[["time", "value"]].copy()
    out["time"] = pd.to_datetime(out["time"])
    out["value"] = pd.to_numeric(out["value"], errors="coerce")
    return out.dropna(subset=["time", "value"]).sort_values("time")


def _generate_extended_point_plot(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    plot_type = str(params.get("plot_type") or "step_plot")
    resolution = str(params.get("resolution") or "monthly").lower()
    variable = str(params.get("variable") or "auto")
    season = str(params.get("season") or "").strip().upper() or None
    location = str(params.get("location_name") or "Selected Location")
    start = str(params.get("start_date") or "")
    end = str(params.get("end_date") or "")
    frame, context = extract_point_series(
        Path(data_dir), dataset_key, resolution,
        float(params.get("latitude")), float(params.get("longitude")),
        start, end, variable=variable, season=season,
    )
    data = _extended_plot_frame(frame)
    if data.empty:
        raise ValueError("No values were found for the selected plot period.")
    element = variable_display_name(variable, context, dataset_key)
    unit = str(context.get("unit") or DATASETS.get(dataset_key, {}).get("unit") or "")
    colour = plot_color_for(dataset_key, variable)
    period = plot_period_text(resolution, start, end)
    label = dict(PLOT_TYPES).get(plot_type, plot_type.replace("_", " ").title())
    title = f"{label}: {element} for {location} ({period})"
    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"{plot_type}_{dataset_key}_{variable}_{resolution}_{location}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    fig, ax = plt.subplots(figsize=(11.2, 6.2))
    export = data.copy()

    if plot_type == "step_plot":
        ax.step(data["time"], data["value"], where="mid", color=colour, linewidth=1.4)
        ax.set_xlabel("Time")
    elif plot_type == "lollipop":
        work = data.copy()
        if len(work) > 120:
            work["period"] = work["time"].dt.to_period("M").dt.to_timestamp()
            work = work.groupby("period", as_index=False)["value"].mean().rename(columns={"period": "time"})
        markerline, stemlines, baseline = ax.stem(work["time"], work["value"], basefmt=" ")
        plt.setp(markerline, color=colour, markersize=4)
        plt.setp(stemlines, color=colour, linewidth=0.8)
        export = work
        ax.set_xlabel("Time")
    elif plot_type == "rolling_mean":
        window = {"hourly": 24, "daily": 30, "monthly": 12, "annual": 5, "seasonal": 5}.get(resolution, 12)
        work = data.copy()
        work["rolling_mean"] = work["value"].rolling(window, min_periods=max(2, window // 3)).mean()
        ax.plot(work["time"], work["value"], color=colour, alpha=0.35, linewidth=0.9, label=element)
        ax.plot(work["time"], work["rolling_mean"], color=colour, linewidth=2.2, label=f"{window}-period rolling mean")
        ax.legend()
        export = work
        ax.set_xlabel("Time")
    elif plot_type == "cumulative_total":
        work = data.copy()
        work["cumulative_total"] = work["value"].cumsum()
        ax.fill_between(work["time"], work["cumulative_total"], color=colour, alpha=0.25)
        ax.plot(work["time"], work["cumulative_total"], color=colour, linewidth=1.8)
        export = work
        ax.set_xlabel("Time")
        unit = unit or "mm"
    elif plot_type == "percentile_band":
        work = data.assign(month=data["time"].dt.month)
        stats = work.groupby("month")["value"].quantile([0.10, 0.25, 0.50, 0.75, 0.90]).unstack().reset_index()
        stats.columns = ["month", "p10", "p25", "median", "p75", "p90"]
        x = np.arange(1, 13)
        full = pd.DataFrame({"month": x}).merge(stats, on="month", how="left")
        ax.fill_between(x, full["p10"], full["p90"], color=colour, alpha=0.15, label="10th–90th percentile")
        ax.fill_between(x, full["p25"], full["p75"], color=colour, alpha=0.30, label="25th–75th percentile")
        ax.plot(x, full["median"], color=colour, marker="o", linewidth=2, label="Median")
        ax.set_xticks(x); ax.set_xticklabels(_month_labels(x)); ax.set_xlabel("Month"); ax.legend()
        export = full
    elif plot_type == "exceedance_curve":
        values = np.sort(data["value"].to_numpy(dtype=float))[::-1]
        exceedance = np.arange(1, len(values) + 1) / (len(values) + 1) * 100.0
        export = pd.DataFrame({"exceedance_probability_percent": exceedance, "value": values})
        ax.plot(exceedance, values, color=colour, linewidth=1.8)
        ax.set_xlabel("Exceedance probability (%)")
    elif plot_type == "ecdf":
        values = np.sort(data["value"].to_numpy(dtype=float))
        probability = np.arange(1, len(values) + 1) / len(values) * 100.0
        export = pd.DataFrame({"value": values, "cumulative_probability_percent": probability})
        ax.step(values, probability, where="post", color=colour, linewidth=1.8)
        ax.set_xlabel(y_axis_label(element, unit)); ax.set_ylabel("Cumulative probability (%)")
    elif plot_type == "violin":
        work = data.assign(month=data["time"].dt.month)
        groups = [work.loc[work["month"] == m, "value"].dropna().to_numpy() for m in range(1, 13)]
        safe_groups = [g if len(g) else np.asarray([np.nan]) for g in groups]
        parts = ax.violinplot(safe_groups, positions=np.arange(1, 13), showmeans=True, showmedians=True, showextrema=True)
        for body in parts["bodies"]:
            body.set_facecolor(colour); body.set_alpha(0.55)
        ax.set_xticks(np.arange(1, 13)); ax.set_xticklabels(_month_labels(range(1, 13))); ax.set_xlabel("Month")
        export = work
    elif plot_type == "diurnal_cycle":
        if resolution != "hourly":
            raise ValueError("Average Diurnal Cycle Plot requires hourly data.")
        work = data.assign(hour=data["time"].dt.hour)
        stats = work.groupby("hour")["value"].agg(mean="mean", standard_deviation="std", minimum="min", maximum="max").reset_index()
        ax.fill_between(stats["hour"], stats["mean"] - stats["standard_deviation"], stats["mean"] + stats["standard_deviation"], color=colour, alpha=0.20, label="Mean ± 1 standard deviation")
        ax.plot(stats["hour"], stats["mean"], color=colour, marker="o", linewidth=2, label="Hourly mean")
        ax.set_xticks(range(0, 24, 2)); ax.set_xlabel("Hour of day"); ax.legend()
        export = stats
    elif plot_type == "rank_plot":
        values = np.sort(data["value"].to_numpy(dtype=float))[::-1]
        export = pd.DataFrame({"rank": np.arange(1, len(values) + 1), "value": values})
        ax.plot(export["rank"], export["value"], color=colour, linewidth=1.6)
        ax.set_xlabel("Rank (highest to lowest)")
    else:
        raise ValueError("Unsupported extended plot type.")

    if plot_type not in {"ecdf"}:
        ax.set_ylabel(y_axis_label(element if plot_type != "cumulative_total" else f"Cumulative {element}", unit))
    ax.set_title(title, fontsize=14, fontweight="bold", pad=13)
    apply_plot_grids(ax)
    _finalize_plot(fig, plot_path)
    excel_path, parquet_path, db_path = _save_data(export, dirs["plots"], stem + "_data", table="plot_products")
    context = dict(context)
    context.update({"variable_label": element, "period_label": period, "plot_type": plot_type})
    return {"plot_path": plot_path, "excel_path": excel_path, "parquet_path": parquet_path, "db_path": db_path, "rows": len(export), "context": context}


# Expanded climate-index plot catalogue.
INDEX_PLOT_TYPES = [
    ("auto", "Automatic — Best Plot for Selected Time Scale"),
    ("line", "Time-Series Line Plot"),
    ("bar", "Period Comparison Bar Chart"),
    ("area", "Filled Area Plot"),
    ("step", "Step Line Plot"),
    ("lollipop", "Lollipop Plot"),
    ("rolling_mean", "Rolling Mean Plot"),
    ("cumulative", "Cumulative Index Plot"),
    ("anomaly_bar", "Positive–Negative Anomaly Bar Plot"),
    ("heatmap", "Time Heat Map"),
    ("box", "Distribution Box Plot"),
    ("violin", "Distribution Violin Plot"),
    ("histogram", "Frequency Distribution Histogram"),
    ("exceedance", "Exceedance Probability Curve"),
    ("ecdf", "Empirical Cumulative Distribution Plot"),
]
_INDEX_ALL_PLOTS = [key for key, _ in INDEX_PLOT_TYPES]
INDEX_PLOT_RULES = {
    "hourly": [k for k in _INDEX_ALL_PLOTS if k not in {"bar", "anomaly_bar"}],
    "daily": list(_INDEX_ALL_PLOTS),
    "monthly": list(_INDEX_ALL_PLOTS),
    "annual": [k for k in _INDEX_ALL_PLOTS if k != "heatmap"],
    "seasonal": [k for k in _INDEX_ALL_PLOTS if k != "heatmap"],
}


def _index_plot(values: pd.DataFrame, params: Dict[str, Any], output: Path, title: str, unit: str, resolution: str) -> str:
    plot_type = str(params.get("index_plot_type") or "auto").lower()
    allowed = INDEX_PLOT_RULES.get(resolution, ["auto", "line"])
    if plot_type not in allowed:
        plot_type = "auto"
    if plot_type == "auto":
        plot_type = "bar" if resolution in {"annual", "seasonal"} else "line"
    frame = values[["time", "value"]].copy()
    frame["time"] = pd.to_datetime(frame["time"])
    frame["value"] = pd.to_numeric(frame["value"], errors="coerce")
    frame = frame.dropna().sort_values("time")
    colour = plot_color_for(index_type=str(params.get("index_type") or ""))
    fig, ax = plt.subplots(figsize=(11.4, 6.2))
    numeric = frame["value"]
    if plot_type == "heatmap":
        if resolution == "hourly":
            frame["row"] = frame["time"].dt.strftime("%Y-%m-%d"); frame["bucket"] = frame["time"].dt.hour; x_label = "Hour of day"
        elif resolution == "daily":
            frame["row"] = frame["time"].dt.year; frame["bucket"] = frame["time"].dt.dayofyear; x_label = "Day of year"
        else:
            frame["row"] = frame["time"].dt.year; frame["bucket"] = frame["time"].dt.month; x_label = "Month"
        pivot = frame.pivot_table(index="row", columns="bucket", values="value", aggfunc="mean")
        im = ax.imshow(pivot.to_numpy(dtype=float), aspect="auto", interpolation="nearest", cmap="viridis")
        if len(pivot.columns) <= 24:
            ax.set_xticks(np.arange(len(pivot.columns))); ax.set_xticklabels([str(v) for v in pivot.columns])
        if len(pivot.index) <= 30:
            ax.set_yticks(np.arange(len(pivot.index))); ax.set_yticklabels([str(v) for v in pivot.index])
        ax.set_xlabel(x_label); ax.set_ylabel("Date / Year")
        fig.colorbar(im, ax=ax, shrink=0.84, pad=0.03, label=unit or "Index value")
    elif plot_type == "bar":
        labels = frame["time"].dt.strftime("%Y") if resolution in {"annual", "seasonal"} else frame["time"].dt.strftime("%Y-%m-%d")
        x = np.arange(len(frame)); ax.bar(x, numeric, color=colour)
        step = max(1, math.ceil(len(frame) / 20)); ticks = np.arange(0, len(frame), step)
        ax.set_xticks(ticks); ax.set_xticklabels(labels.iloc[::step], rotation=45, ha="right"); ax.set_xlabel("Period")
    elif plot_type == "area":
        ax.fill_between(frame["time"], numeric, color=colour, alpha=0.28); ax.plot(frame["time"], numeric, color=colour, linewidth=1.4); ax.set_xlabel("Time")
    elif plot_type == "step":
        ax.step(frame["time"], numeric, where="mid", color=colour, linewidth=1.5); ax.set_xlabel("Time")
    elif plot_type == "lollipop":
        work = frame if len(frame) <= 150 else frame.iloc[::max(1, len(frame)//150)]
        markerline, stemlines, baseline = ax.stem(work["time"], work["value"], basefmt=" ")
        plt.setp(markerline, color=colour, markersize=4); plt.setp(stemlines, color=colour, linewidth=0.8); ax.set_xlabel("Time")
    elif plot_type == "rolling_mean":
        window = {"hourly": 24, "daily": 30, "monthly": 12, "annual": 5, "seasonal": 5}.get(resolution, 12)
        rolling = numeric.rolling(window, min_periods=max(2, window//3)).mean()
        ax.plot(frame["time"], numeric, color=colour, alpha=0.30, linewidth=0.9, label="Index value")
        ax.plot(frame["time"], rolling, color=colour, linewidth=2.2, label=f"{window}-period rolling mean"); ax.legend(); ax.set_xlabel("Time")
    elif plot_type == "cumulative":
        cumulative = numeric.cumsum(); ax.fill_between(frame["time"], cumulative, color=colour, alpha=0.25); ax.plot(frame["time"], cumulative, color=colour); ax.set_xlabel("Time")
    elif plot_type == "anomaly_bar":
        anomaly = numeric - numeric.mean(); colours = np.where(anomaly >= 0, "#d73027", "#4575b4")
        ax.bar(frame["time"], anomaly, color=colours); ax.axhline(0, color="black", linewidth=0.9); ax.set_xlabel("Time")
    elif plot_type == "box":
        if resolution in {"hourly", "daily", "monthly"}:
            groups = [numeric[frame["time"].dt.month == month].dropna().to_numpy() for month in range(1, 13)]
            _boxplot_with_labels(ax, groups, [pd.Timestamp(2000, month, 1).strftime("%b") for month in range(1, 13)], showfliers=False); ax.set_xlabel("Month")
        else:
            _boxplot_with_labels(ax, [numeric.dropna().to_numpy()], [INDEX_RESOLUTION_LABELS.get(resolution, resolution.title())], showfliers=True)
    elif plot_type == "violin":
        if resolution in {"hourly", "daily", "monthly"}:
            groups = [numeric[frame["time"].dt.month == month].dropna().to_numpy() for month in range(1, 13)]
            groups = [g if len(g) else np.asarray([np.nan]) for g in groups]
            parts = ax.violinplot(groups, positions=np.arange(1,13), showmeans=True, showmedians=True)
            for body in parts["bodies"]: body.set_facecolor(colour); body.set_alpha(0.55)
            ax.set_xticks(np.arange(1,13)); ax.set_xticklabels(_month_labels(range(1,13))); ax.set_xlabel("Month")
        else:
            parts = ax.violinplot([numeric.dropna().to_numpy()], showmeans=True, showmedians=True)
            for body in parts["bodies"]: body.set_facecolor(colour); body.set_alpha(0.55)
    elif plot_type == "histogram":
        ax.hist(numeric, bins=min(30, max(8, int(math.sqrt(max(1, len(numeric)))))), color=colour, alpha=0.85); ax.set_xlabel(unit or "Index value"); ax.set_ylabel("Frequency")
    elif plot_type == "exceedance":
        ordered = np.sort(numeric.to_numpy(dtype=float))[::-1]; p = np.arange(1, len(ordered)+1)/(len(ordered)+1)*100
        ax.plot(p, ordered, color=colour); ax.set_xlabel("Exceedance probability (%)")
    elif plot_type == "ecdf":
        ordered = np.sort(numeric.to_numpy(dtype=float)); p = np.arange(1, len(ordered)+1)/len(ordered)*100
        ax.step(ordered, p, where="post", color=colour); ax.set_xlabel(unit or "Index value"); ax.set_ylabel("Cumulative probability (%)")
    elif plot_type == "rank":
        ordered = np.sort(numeric.to_numpy(dtype=float))[::-1]
        ax.plot(np.arange(1, len(ordered)+1), ordered, color=colour); ax.set_xlabel("Rank (highest to lowest)")
    else:
        ax.plot(frame["time"], numeric, color=colour, linewidth=1.4); ax.set_xlabel("Time")
    if plot_type not in {"histogram", "ecdf"}:
        ax.set_ylabel(unit or "Index value")
    ax.set_title(title, fontsize=14, fontweight="bold", pad=13)
    apply_plot_grids(ax); _finalize_plot(fig, output)
    return plot_type


# Normalize legacy return keys so the UI exposes PNG and Excel only while the
# generated table remains persisted as Parquet.
def _normalise_product_result(result: Dict[str, Any]) -> Dict[str, Any]:
    result = dict(result or {})
    legacy = result.pop("csv_path", None)
    if legacy:
        path = Path(legacy)
        if path.suffix.lower() == ".xlsx":
            result.setdefault("excel_path", path)
        elif path.exists() and path.suffix.lower() == ".csv":
            try:
                frame = pd.read_csv(path)
                excel_path, parquet_path, db_path = _save_data(frame, path.parent, path.stem, table="plot_products")
                result.setdefault("excel_path", excel_path); result.setdefault("parquet_path", parquet_path); result.setdefault("db_path", db_path)
            finally:
                path.unlink(missing_ok=True)
    for key in ("data_used_csv_path",):
        path_value = result.pop(key, None)
        if path_value:
            Path(path_value).unlink(missing_ok=True)
    # Some legacy plot generators return only an Excel workbook. Persist the
    # corresponding data table as Parquet without exposing Parquet as a user
    # download. This guarantees that every generated product is database-ready.
    if not result.get("parquet_path") and result.get("excel_path"):
        excel = Path(result["excel_path"])
        if excel.exists():
            parquet_path = excel.with_suffix(".parquet")
            if parquet_path.exists():
                result["parquet_path"] = parquet_path
            else:
                frame = None
                try:
                    frame = pd.read_excel(excel, sheet_name="Data", header=None)
                except Exception:
                    try:
                        sheets = pd.read_excel(excel, sheet_name=None, header=None)
                        frame = next((table for table in sheets.values() if table is not None and not table.empty), None)
                    except Exception:
                        frame = None
                if frame is not None:
                    # Workbook metadata can mix datetimes, strings and numeric
                    # values in one column. Store a lossless textual fallback
                    # when no direct Parquet table was returned by the generator.
                    safe = frame.copy()
                    for column in safe.columns:
                        if safe[column].dtype == object:
                            safe[column] = safe[column].map(lambda value: None if pd.isna(value) else str(value))
                    safe.to_parquet(parquet_path, index=False)
                    result["parquet_path"] = parquet_path
                    result.setdefault("db_path", _append_parquet_database("plot_products", safe, parquet_path.stem))
    return result


_CDE_FINAL_GENERATE_PLOT_BASE = generate_plot_product

def generate_plot_product(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    prepared = dict(params)
    plot_type = str(prepared.get("plot_type") or "time_series")
    if plot_type.startswith("spatial_") or plot_type == "spatial_map":
        prepared["show_ocean"] = False
        prepared["show_rivers"] = False
        prepared["show_lakes"] = _truthy(prepared.get("show_lakes", False))
    if plot_type in {key for key, _ in _EXTENDED_PLOTS}:
        return _normalise_product_result(_generate_extended_point_plot(prepared, Path(data_dir), Path(export_dir)))
    return _normalise_product_result(_CDE_FINAL_GENERATE_PLOT_BASE(prepared, Path(data_dir), Path(export_dir)))


_CDE_FINAL_GENERATE_INDICES_BASE = generate_indices

def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    result = dict(_CDE_FINAL_GENERATE_INDICES_BASE(params, Path(data_dir), Path(export_dir)))
    # Persist a Parquet copy of the generated Excel data and remove all CSVs.
    csv_path = result.pop("csv_path", None)
    frame = None
    if csv_path:
        path = Path(csv_path)
        if path.exists() and path.suffix.lower() == ".csv":
            try:
                frame = pd.read_csv(path)
            finally:
                path.unlink(missing_ok=True)
    for key in ("data_used_csv_path",):
        path_value = result.pop(key, None)
        if path_value:
            Path(path_value).unlink(missing_ok=True)
    if frame is None and result.get("excel_path"):
        try:
            frame = pd.read_excel(result["excel_path"], sheet_name="Data")
        except Exception:
            frame = None
    if frame is not None and not frame.empty:
        parquet_path = Path(result["excel_path"]).with_suffix(".parquet") if result.get("excel_path") else ensure_output_dirs(Path(export_dir))["indices"] / f"{slugify('index_data')}.parquet"
        frame.to_parquet(parquet_path, index=False)
        result["parquet_path"] = parquet_path
        result["db_path"] = _append_parquet_database("climate_indices", frame, parquet_path.stem)
    result["rainfall_csv_only"] = False
    return result

# Dataset-aware plot filtering: never present controls that cannot be fulfilled
# by the selected dataset.
def dataset_allowed_plots(dataset_key: str) -> list[str]:
    metadata = DATASETS.get(dataset_key, {})
    available = list(PLOT_FAMILIES.get(metadata.get("family", "rainfall"), []))
    resolutions = set((metadata.get("resolutions") or {}).keys())
    if "hourly" not in resolutions:
        available = [item for item in available if item != "diurnal_cycle"]
    if dataset_key not in {"era5_temperature_stats", "era5_pressure_cloud", "era5_wind"}:
        available = [item for item in available if item not in {"multi_line", "scatter"}]
    if dataset_key not in {"era5_temperature", "era5_temperature_stats"}:
        available = [item for item in available if item != "temperature_monthly_profile"]
    if dataset_key != "era5_wind":
        available = [item for item in available if item != "wind_rose"]
    return list(dict.fromkeys(available))

# ---------------------------------------------------------------------------
# 2026-07-18 TITLE, BORDER-CLIPPING AND TRANSBOUNDARY-LAKE UPDATE
# ---------------------------------------------------------------------------

def _clean_custom_plot_title(value: Any) -> str:
    """Return a safe, compact user-supplied title."""
    text = re.sub(r"[\r\n\t]+", " ", str(value or "")).strip()
    text = re.sub(r"\s{2,}", " ", text)
    return text[:180]


def _map_clip_patch(ax, polygons: list[np.ndarray]):
    """Build one compound clipping path from Tanzania administrative polygons."""
    from matplotlib.path import Path as MplPath
    from matplotlib.patches import PathPatch
    paths = []
    for polygon in polygons:
        array = np.asarray(polygon, dtype=float)
        if array.ndim != 2 or len(array) < 3:
            continue
        vertices = array[:, :2]
        if not np.allclose(vertices[0], vertices[-1]):
            vertices = np.vstack([vertices, vertices[0]])
        codes = np.full(len(vertices), MplPath.LINETO, dtype=np.uint8)
        codes[0] = MplPath.MOVETO
        codes[-1] = MplPath.CLOSEPOLY
        paths.append(MplPath(vertices, codes))
    if not paths:
        return None
    compound = MplPath.make_compound_path(*paths)
    return PathPatch(compound, transform=ax.transData, facecolor="none", edgecolor="none")


def _clip_map_artist(artist: Any, clip_patch: Any) -> None:
    if clip_patch is None or artist is None:
        return
    collections = getattr(artist, "collections", None)
    if collections:
        for collection in collections:
            collection.set_clip_path(clip_patch)
    elif hasattr(artist, "set_clip_path"):
        artist.set_clip_path(clip_patch)


def _tight_polygon_bounds(polygons: list[np.ndarray], padding: float = 0.08) -> tuple[float, float, float, float]:
    valid = [np.asarray(p, dtype=float)[:, :2] for p in polygons if np.asarray(p).ndim == 2 and len(p) >= 3]
    if not valid:
        return (TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"], TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    points = np.vstack(valid)
    return (
        float(np.nanmin(points[:, 0])) - padding,
        float(np.nanmax(points[:, 0])) + padding,
        float(np.nanmin(points[:, 1])) - padding,
        float(np.nanmax(points[:, 1])) + padding,
    )


def _normalise_lake_polygon(array: np.ndarray) -> np.ndarray | None:
    """Normalise GeoJSON coordinates and tolerate accidental lat/lon reversal."""
    polygon = np.asarray(array, dtype=float)
    if polygon.ndim != 2 or polygon.shape[1] < 2 or len(polygon) < 3:
        return None
    polygon = polygon[:, :2]
    x_med = float(np.nanmedian(polygon[:, 0])); y_med = float(np.nanmedian(polygon[:, 1]))
    # A Tanzania lake layer written as [latitude, longitude] is corrected here.
    if -15.0 <= x_med <= 2.0 and 25.0 <= y_med <= 45.0:
        polygon = polygon[:, [1, 0]]
    if not np.isfinite(polygon).all():
        polygon = polygon[np.isfinite(polygon).all(axis=1)]
    return polygon if len(polygon) >= 3 else None


def _lake_polygons_and_status(data_dir: Path | None) -> tuple[list[np.ndarray], dict[str, Any]]:
    status = {"file_found": False, "polygon_count": 0, "lake_nyasa_detected": False}
    if data_dir is None:
        return [], status
    lake_path = Path(data_dir) / "shapefiles" / "hydrography" / "tanzania_lakes.geojson"
    status["path"] = str(lake_path)
    if not lake_path.exists():
        return [], status
    status["file_found"] = True
    polygons = []
    for raw in _read_natural_layer(lake_path, polygons=True):
        polygon = _normalise_lake_polygon(raw)
        if polygon is None:
            continue
        # Retain only features that overlap the wider Tanzania plotting area.
        minx, maxx = float(np.nanmin(polygon[:, 0])), float(np.nanmax(polygon[:, 0]))
        miny, maxy = float(np.nanmin(polygon[:, 1])), float(np.nanmax(polygon[:, 1]))
        if maxx < 27.5 or minx > 42.0 or maxy < -15.0 or miny > 1.5:
            continue
        polygons.append(polygon)
        # Lake Nyasa / Lake Malawi intersects this well-known transboundary box.
        if maxx >= 34.15 and minx <= 35.95 and maxy >= -12.55 and miny <= -8.65:
            status["lake_nyasa_detected"] = True
    status["polygon_count"] = len(polygons)
    return polygons, status


def _draw_hydrology(ax, *, ocean: bool = False, lakes: bool = False, rivers: bool = False,
                     data_dir: Path | None = None, clip_polygons: list[np.ndarray] | None = None) -> dict[str, Any]:
    """Draw operational lake geometries above the weather surface.

    Lake polygons are never clipped to Tanzania's national boundary. A lake is
    kept when it intersects the Tanzania plotting window and its original full
    geometry is retained. Rendering is clipped only by the fixed map axes
    (latitude 0 to 12°S), so geometry outside that operational latitude window
    is not displayed. ``clip_polygons`` is retained for backwards compatibility
    and is deliberately ignored.
    """
    from matplotlib.collections import PatchCollection
    from matplotlib.lines import Line2D
    ax.set_facecolor("white")
    polygons, status = _lake_polygons_and_status(data_dir)
    status["full_geometry_rendering"] = True
    if not lakes or not polygons:
        return status
    patches = [Polygon(arr, closed=True) for arr in polygons if len(arr) >= 3]
    if patches:
        collection = PatchCollection(
            patches,
            facecolor="#9ecae1",
            edgecolor="#2171b5",
            linewidths=0.80,
            alpha=1.0,
            zorder=18,
            clip_on=True,
        )
        collection.set_gid("cde-lakes")
        ax.add_collection(collection)
        ax.legend(
            handles=[Line2D([0], [0], color="#2171b5", lw=4, label="Lakes")],
            loc="lower left", fontsize=6.5, frameon=True, framealpha=0.94,
        )
    return status


def _international_boundary_segments(polygons: list[np.ndarray], precision: int = 6) -> list[np.ndarray]:
    """Return only Tanzania's true exterior/coastline segments.

    Admin-1 GeoJSON rings are not always vertex-for-vertex identical along a
    shared regional boundary.  Exact segment counting can therefore mistake a
    regional line (notably the Kagera/Mara division in Lake Victoria) for an
    international boundary and redraw it above the lake.  Each candidate edge
    is now tested on both sides against the union of all administrative
    polygons.  A true exterior edge has Tanzania on exactly one side; a shared
    regional edge has Tanzania on both sides and is excluded.
    """
    from matplotlib.path import Path as MplPath

    prepared: list[np.ndarray] = []
    paths: list[MplPath] = []
    for polygon in polygons:
        array = np.asarray(polygon, dtype=float)
        if array.ndim != 2 or array.shape[1] < 2 or len(array) < 3:
            continue
        vertices = array[:, :2]
        if not np.allclose(vertices[0], vertices[-1]):
            vertices = np.vstack([vertices, vertices[0]])
        prepared.append(vertices)
        paths.append(MplPath(vertices))

    if not prepared:
        return []

    def inside_union(points: np.ndarray) -> np.ndarray:
        result = np.zeros(len(points), dtype=bool)
        for polygon_path in paths:
            # A small positive radius makes the probe robust to floating-point
            # noise without turning nearby internal borders into exteriors.
            result |= polygon_path.contains_points(points, radius=1e-10)
        return result

    exterior: list[np.ndarray] = []
    seen: set[tuple[tuple[float, float], tuple[float, float]]] = set()
    for vertices in prepared:
        for start, end in zip(vertices[:-1], vertices[1:]):
            delta = end - start
            length = float(np.hypot(delta[0], delta[1]))
            if not np.isfinite(length) or length <= 1e-12:
                continue

            a = (round(float(start[0]), precision), round(float(start[1]), precision))
            b = (round(float(end[0]), precision), round(float(end[1]), precision))
            key = (a, b) if a <= b else (b, a)
            if key in seen:
                continue
            seen.add(key)

            midpoint = (start + end) / 2.0
            normal = np.asarray([-delta[1], delta[0]], dtype=float) / length
            # Roughly 10–50 m in geographic coordinates, scaled for long edges.
            offset = min(5e-4, max(1e-4, length * 1e-3))
            probes = np.vstack([midpoint + normal * offset, midpoint - normal * offset])
            left_inside, right_inside = inside_union(probes)

            # Exactly one side within the Tanzania polygon union means that the
            # segment is part of the national exterior/coastline.  Both sides
            # inside identifies a regional boundary and removes it above lakes.
            if bool(left_inside) != bool(right_inside):
                exterior.append(np.asarray([start, end], dtype=float))

    return exterior


def _draw_international_boundary_on_top(ax, polygons: list[np.ndarray]) -> None:
    """Draw Tanzania's outer boundary above complete lake polygons."""
    from matplotlib.collections import LineCollection
    segments = _international_boundary_segments(polygons)
    boundary_data = segments if segments else polygons
    if not boundary_data:
        return
    collection = LineCollection(
        boundary_data,
        colors="#05090d",
        linewidths=1.35,
        alpha=1.0,
        zorder=22,
        clip_on=True,
    )
    collection.set_gid("cde-international-boundary")
    ax.add_collection(collection)


def _render_tanzania_map_axis(
    ax, *, x: np.ndarray, y: np.ndarray, values: np.ndarray, data_dir: Path,
    selected_boundaries: list[np.ndarray], level: int, dataset_key: str,
    show_ocean: bool, show_lakes: bool, show_rivers: bool, title: str,
    norm=None, show_labels: bool = True, show_cartographic_elements: bool = True,
    render_style: str = "grid",
):
    """Render a map clipped exactly to Tanzania so surfaces meet the boundary."""
    # Preserve every native 0.25-degree source grid cell. Smoothing and
    # spatial interpolation are intentionally disabled for all map requests.
    render_style = "grid"
    cmap = plot_cmap_for(dataset_key)
    clip_patch = _map_clip_patch(ax, selected_boundaries)
    image = ax.pcolormesh(x, y, values, shading="nearest", cmap=cmap, norm=norm, zorder=4)
    _clip_map_artist(image, clip_patch)

    _draw_admin_boundaries(ax, selected_boundaries, level)
    try:
        national_like, _ = _admin_polygons(data_dir, 1)
        from matplotlib.collections import LineCollection
        ax.add_collection(LineCollection(national_like, colors="#101820", linewidths=1.05, alpha=0.98, zorder=15))
    except Exception:
        national_like = selected_boundaries

    lake_polygons, _lake_pre_status = _lake_polygons_and_status(data_dir)
    lake_status = _draw_hydrology(
        ax, ocean=False, lakes=show_lakes, rivers=False, data_dir=data_dir,
        clip_polygons=None,
    )

    # The international boundary is deliberately rendered after the lakes so
    # it remains visible wherever a transboundary lake crosses Tanzania.
    _draw_international_boundary_on_top(ax, national_like)

    if show_labels and level == 1:
        for name, lon, lat in _admin_feature_labels(data_dir, 1):
            ax.text(lon, lat, name, fontsize=5.7, fontweight="bold", color="#16232d",
                    ha="center", va="center", zorder=19)

    # The meteorological surface uses the operational Tanzania latitude domain
    # from 0° to 12°S. Complete transboundary lake geometries are still drawn
    # with clipping disabled, so their original polygons are not cut to the
    # national boundary. Longitude may expand to include complete lakes, while
    # the map latitude scale remains exactly 0 to -12 as requested.
    extent_polygons = list(national_like)
    if show_lakes and lake_polygons:
        extent_polygons.extend(lake_polygons)
    minx, maxx, _miny, _maxy = _tight_polygon_bounds(extent_polygons, padding=0.08)
    ax.set_xlim(minx, maxx)
    ax.set_ylim(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    ax.set_aspect("equal", adjustable="box")
    ax.margins(x=0, y=0)
    ax.set_title(title, fontsize=9.0 if not show_cartographic_elements else 13.0, fontweight="bold", pad=8)
    if show_cartographic_elements:
        ax.set_xlabel("Longitude")
        ax.set_ylabel("Latitude")
        _draw_north_arrow(ax)
        _draw_scale_bar(ax)
        apply_plot_grids(ax)
    else:
        ax.set_xticks([])
        ax.set_yticks([])
    setattr(ax, "_cde_lake_status", lake_status)
    return image


def _generate_requested_spatial_map(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    layout, jobs = _prepare_spatial_jobs(params, Path(data_dir))
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    level = max(1, min(3, int(params.get("map_admin_level") or 1)))
    boundaries, boundary_path = _admin_polygons(Path(data_dir), level)
    national, _ = _admin_polygons(Path(data_dir), 1)
    admin_label = {1: "Admin 1 — Regions", 2: "Admin 2 — Districts", 3: "Admin 3 — Wards"}[level]
    prepared = []
    finite_parts = []
    for grid, context, label in jobs:
        work = grid[
            grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"], TANZANIA_BOUNDS["lon_max"])
            & grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
        ].copy()
        if work.empty:
            continue
        pivot = work.pivot_table(index="latitude", columns="longitude", values="value", aggfunc="mean").sort_index()
        x = pivot.columns.to_numpy(float)
        y = pivot.index.to_numpy(float)
        raw_values = pivot.to_numpy(float)
        inside_values = _mask_grid_to_tanzania(raw_values, x, y, national)
        prepared.append((x, y, raw_values, work, context, label))
        finite = inside_values[np.isfinite(inside_values)]
        if finite.size:
            finite_parts.append(finite)
    if not prepared:
        raise ValueError("No spatial values were found within the Tanzania map extent.")

    from matplotlib.colors import Normalize
    finite = np.concatenate(finite_parts) if finite_parts else np.asarray([0.0, 1.0])
    vmin, vmax = float(np.nanmin(finite)), float(np.nanmax(finite))
    if np.isclose(vmin, vmax):
        vmax = vmin + 1.0
    norm = Normalize(vmin=vmin, vmax=vmax)
    element = _explicit_weather_element(dataset_key, params.get("variable"), prepared[0][4])
    custom_title = _clean_custom_plot_title(params.get("custom_plot_title"))
    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    style = str(params.get("map_render_style") or "grid").lower()
    stem = _cde_prefixed_stem(
        f"spatial_{style}_{dataset_key}_{'panels' if layout == 'panel' else slugify(prepared[0][5])}_admin{level}_{stamp}"
    )
    plot_path = dirs["plots"] / f"{stem}.png"
    show_lakes = _truthy(params.get("show_lakes", True))
    image = None
    lake_status = _lake_polygons_and_status(Path(data_dir))[1]

    if layout == "panel":
        count = len(prepared)
        columns = min(6, max(2, int(math.ceil(math.sqrt(count * 1.15)))))
        rows = int(math.ceil(count / columns))
        fig, axes = plt.subplots(rows, columns, figsize=(2.55 * columns + 0.7, 2.60 * rows), squeeze=False)
        flat = axes.ravel()
        for idx, (x, y, values, _work, _context, label) in enumerate(prepared):
            image = _render_tanzania_map_axis(
                flat[idx], x=x, y=y, values=values, data_dir=Path(data_dir),
                selected_boundaries=boundaries, level=level, dataset_key=dataset_key,
                show_ocean=False, show_lakes=show_lakes, show_rivers=False,
                title=label, norm=norm, show_labels=False,
                show_cartographic_elements=False, render_style=style,
            )
            if idx == 0:
                _draw_north_arrow(flat[idx])
        for axis in flat[len(prepared):]:
            axis.axis("off")
        default_title = f"{element} Spatial Distribution over Tanzania\n{_compact_year_label(prepared[0][4].get('years', []))} · {admin_label}"
        fig.suptitle(custom_title or default_title, fontsize=15, fontweight="bold", y=0.995)
        if image is not None:
            cax = fig.add_axes([0.948, 0.12, 0.014, 0.75])
            cbar = fig.colorbar(image, cax=cax, extend="both")
            cbar.set_label(y_axis_label(element, prepared[0][4].get("unit")))
        fig.subplots_adjust(left=0.018, right=0.925, bottom=0.018, top=0.92, wspace=0.035, hspace=0.10)
        fig.savefig(plot_path, dpi=130, bbox_inches="tight", pad_inches=0.03, facecolor="white")
        plt.close(fig)
    else:
        x, y, values, _work, context, label = prepared[0]
        fig = plt.figure(figsize=(11.6, 9.0))
        gs = fig.add_gridspec(1, 2, width_ratios=[1.0, 0.040], wspace=0.08)
        ax = fig.add_subplot(gs[0, 0])
        cax = fig.add_subplot(gs[0, 1])
        heading = _spatial_map_heading(element, context)
        title = custom_title or f"{heading}\n{admin_label}"
        image = _render_tanzania_map_axis(
            ax, x=x, y=y, values=values, data_dir=Path(data_dir),
            selected_boundaries=boundaries, level=level, dataset_key=dataset_key,
            show_ocean=False, show_lakes=show_lakes, show_rivers=False,
            title=title, norm=norm, show_labels=True,
            show_cartographic_elements=True, render_style=style,
        )
        cbar = fig.colorbar(image, cax=cax, extend="both")
        cbar.set_label(y_axis_label(element, context.get("unit")))
        fig.subplots_adjust(left=0.055, right=0.925, bottom=0.065, top=0.935, wspace=0.08)
        fig.savefig(plot_path, dpi=175, bbox_inches="tight", pad_inches=0.03, facecolor="white")
        plt.close(fig)

    export_parts = []
    for _x, _y, _values, grid, context, label in prepared:
        part = grid.copy()
        part.insert(0, "map_period", label)
        part.insert(1, "administrative_level", admin_label)
        part.insert(2, "map_surface", style.title())
        export_parts.append(part)
    export_grid = pd.concat(export_parts, ignore_index=True)
    excel_path, _, _ = _save_data(export_grid, dirs["plots"], stem + "_data", table="plot_products")
    context = dict(prepared[0][4])
    context.update({
        "administrative_level": level,
        "administrative_level_label": admin_label,
        "boundary_file": boundary_path.name,
        "map_output_layout": layout,
        "map_count": len(prepared),
        "map_render_style": style,
        "variable_label": element,
        "custom_plot_title": custom_title,
        "lake_layer": lake_status,
    })
    summary_cards = [
        {"label": "Weather Element", "value": element, "note": context.get("period_label")},
        {"label": "Map Surface", "value": "Smooth" if style == "smooth" else "Grid cells", "note": f"{len(prepared)} map panel(s)"},
        {"label": "Boundary Level", "value": admin_label, "note": boundary_path.name},
    ]
    if show_lakes:
        summary_cards.append({
            "label": "Lake Nyasa",
            "value": "Detected" if lake_status.get("lake_nyasa_detected") else "Not found in GeoJSON",
            "note": f"{lake_status.get('polygon_count', 0)} lake polygon(s) loaded",
        })
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "rows": len(export_grid),
        "context": context,
        "period_label": "All selected periods" if layout == "panel" else prepared[0][5],
        "summary_cards": summary_cards,
    }


def _replace_png_title(path: Path, title: str) -> None:
    """Replace the existing top title band for non-map plots."""
    title = _clean_custom_plot_title(title)
    path = Path(path)
    if not title or not path.exists() or path.suffix.lower() != ".png":
        return
    from PIL import Image, ImageDraw, ImageFont
    image = Image.open(path).convert("RGB")
    width, height = image.size
    band = max(72, min(125, int(height * 0.105)))
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, width, band), fill="white")
    font_size = max(18, min(34, int(width / max(42, len(title) * 0.90))))
    font_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")
    try:
        font = ImageFont.truetype(str(font_path), font_size) if font_path.exists() else ImageFont.load_default()
    except Exception:
        font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), title, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    draw.text(((width - text_width) / 2, max(8, (band - text_height) / 2 - bbox[1])), title, fill="black", font=font)
    image.save(path, format="PNG")


_CDE_TITLE_BORDER_PLOT_BASE = generate_plot_product

def generate_plot_product(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    result = _normalise_product_result(_CDE_TITLE_BORDER_PLOT_BASE(dict(params), Path(data_dir), Path(export_dir)))
    custom_title = _clean_custom_plot_title(params.get("custom_plot_title"))
    plot_type = str(params.get("plot_type") or "")
    if custom_title and result.get("plot_path") and plot_type != "annual_trend" and (not plot_type.startswith("spatial") or plot_type in {"spatial_std_map", "spatial_cv_map"}):
        _replace_png_title(Path(result["plot_path"]), custom_title)
        context = dict(result.get("context") or {})
        context["custom_plot_title"] = custom_title
        result["context"] = context
    return result

# ---------------------------------------------------------------------------
# 2026-07-18 correctness and bounded-memory spatial rendering update
# ---------------------------------------------------------------------------
import gc

CDE_MAX_SPATIAL_PANELS = 24
CDE_PANEL_MAX_AXIS_CELLS = 96
CDE_SINGLE_MAP_MAX_AXIS_CELLS = 240


def _thin_spatial_array(da: xr.DataArray, lat_name: str, lon_name: str, max_axis_cells: int) -> xr.DataArray:
    """Preserve every native grid cell; no spatial thinning is permitted."""
    return da


def _resolve_spatial_source(data_dir: Path, params: Dict[str, Any]) -> dict[str, Any]:
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    variable = str(params.get("variable") or "auto")
    if dataset_key == "era5_temperature" and variable in {"all", "all_in_one", "max_min"}:
        variable = "ta"
    if dataset_key == "era5_wind" and variable in {"auto", "wind_speed_direction", ""}:
        variable = "wind_speed"
    period = _spatial_period_definition(params)
    preferred = period["preferred_resolution"]
    season = period["season"] if preferred == "seasonal" else None
    try:
        file_path = find_file(data_dir, dataset_key, preferred, season=season)
        source_resolution = preferred
    except Exception:
        source_resolution = str(params.get("resolution") or "monthly")
        file_path = find_file(data_dir, dataset_key, source_resolution, season=season if source_resolution == "seasonal" else None)
    return {
        "dataset_key": dataset_key,
        "variable": variable,
        "period": period,
        "file_path": Path(file_path),
        "source_resolution": source_resolution,
        "season": season,
    }


def _spatial_grid_from_open_dataset(
    ds: xr.Dataset,
    spec: dict[str, Any],
    *,
    max_axis_cells: int,
) -> tuple[pd.DataFrame, Dict[str, Any]]:
    dataset_key = spec["dataset_key"]
    variable = spec["variable"]
    period = spec["period"]
    source_resolution = spec["source_resolution"]
    file_path = spec["file_path"]
    meta = DATASETS[dataset_key]
    time_name = detect_time_coord(ds)
    lat_name, lon_name = detect_lat_lon(ds)
    var_name = pick_variable(ds, dataset_key, variable)
    source_da = ds[var_name]
    da = _select_product_statistic_dimension(source_da, variable or var_name, keep_dims={time_name, lat_name, lon_name})
    actual_variable_label = str(source_da.attrs.get("long_name") or source_da.attrs.get("standard_name") or var_name).replace("_", " ").title()
    variable_label = requested_element_label(dataset_key, variable) or actual_variable_label
    source_units = str(source_da.attrs.get("units") or meta.get("unit") or "")
    da, unit = convert_dataarray_units(da, meta["family"], var_name)
    selected = _select_time_period(
        da,
        time_name,
        period["years"],
        None if source_resolution in {"annual", "seasonal"} else period["months"],
    )
    count = int(selected.sizes.get(time_name, 1)) if time_name in selected.dims else 1
    if count < 1:
        raise ValueError(f"No values were found for {period['period_label']}.")
    aggregation = "Selected period"
    if time_name in selected.dims:
        if count == 1:
            aggregated = selected.isel({time_name: 0}, drop=True)
        elif meta["family"] == "rainfall":
            if source_resolution in {"annual", "seasonal"}:
                aggregated = selected.mean(time_name, skipna=True)
                aggregation = "Mean total across selected years"
            elif period["mode"] == "custom" and len(period["years"]) > 1:
                try:
                    by_year = selected.groupby(f"{time_name}.year").sum(time_name, skipna=True)
                    aggregated = by_year.mean("year", skipna=True)
                except Exception:
                    aggregated = selected.sum(time_name, skipna=True) / max(1, len(period["years"]))
                aggregation = "Mean selected-month total across years"
            else:
                aggregated = selected.sum(time_name, skipna=True)
                aggregation = "Selected-period total"
        else:
            aggregated = selected.mean(time_name, skipna=True)
            aggregation = "Selected-period mean"
    else:
        aggregated = selected
    aggregated = _thin_spatial_array(aggregated, lat_name, lon_name, max_axis_cells)
    frame = aggregated.to_dataframe(name="value").reset_index()
    if lat_name != "latitude":
        frame = frame.rename(columns={lat_name: "latitude"})
    if lon_name != "longitude":
        frame = frame.rename(columns={lon_name: "longitude"})
    frame = frame[["latitude", "longitude", "value"]].dropna(subset=["value"]).reset_index(drop=True)
    context = {
        "file": file_path.name,
        "source_path": str(file_path),
        "dataset_key": dataset_key,
        "dataset_label": meta["label"],
        "family": meta["family"],
        "variable": variable,
        "actual_variable": var_name,
        "variable_label": variable_label,
        "actual_variable_label": actual_variable_label,
        "unit": unit or source_units,
        "resolution": source_resolution,
        "period_label": period["period_label"],
        "aggregation": aggregation,
        "months": period["months"],
        "years": period["years"],
        "render_grid_reduced": bool(
            int(aggregated.sizes.get(lat_name, 0)) < int(selected.sizes.get(lat_name, aggregated.sizes.get(lat_name, 0)))
            or int(aggregated.sizes.get(lon_name, 0)) < int(selected.sizes.get(lon_name, aggregated.sizes.get(lon_name, 0)))
        ),
    }
    return frame, context


def _prepare_spatial_jobs(params: Dict[str, Any], data_dir: Path):
    """Prepare panel maps by opening each Zarr store only once per request."""
    layout = str(params.get("map_output_layout") or "single").lower()
    if layout != "panel":
        spec = _resolve_spatial_source(Path(data_dir), params)
        with open_data_store(spec["file_path"], decode_times=True) as ds:
            grid, context = _spatial_grid_from_open_dataset(ds, spec, max_axis_cells=CDE_SINGLE_MAP_MAX_AXIS_CELLS)
        return "single", [(grid, context, context.get("period_label") or "Selected period")]

    panel_specs = _spatial_panel_jobs(params)
    if len(panel_specs) > CDE_MAX_SPATIAL_PANELS:
        raise ValueError(
            f"A multi-panel spatial figure is limited to {CDE_MAX_SPATIAL_PANELS} maps to prevent server overload. "
            "Use Custom Years to split the request."
        )
    resolved: list[tuple[dict[str, Any], str]] = [(_resolve_spatial_source(Path(data_dir), item), label) for item, label in panel_specs]
    grouped: dict[tuple[str, str, str, str], list[tuple[dict[str, Any], str]]] = {}
    for spec, label in resolved:
        group_key = (
            str(spec["file_path"].resolve()), spec["dataset_key"], spec["variable"], str(spec.get("season") or "")
        )
        grouped.setdefault(group_key, []).append((spec, label))

    jobs: list[tuple[pd.DataFrame, Dict[str, Any], str]] = []
    for entries in grouped.values():
        file_path = entries[0][0]["file_path"]
        with open_data_store(file_path, decode_times=True) as ds:
            for spec, label in entries:
                grid, context = _spatial_grid_from_open_dataset(ds, spec, max_axis_cells=CDE_PANEL_MAX_AXIS_CELLS)
                jobs.append((grid, context, label))
        gc.collect()
    if len(jobs) <= 1:
        layout = "single"
    return layout, jobs


_CDE_BOUNDED_SPATIAL_BASE = _generate_requested_spatial_map


def _generate_requested_spatial_map(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate spatial maps with bounded panel count, memory and output size."""
    result = _CDE_BOUNDED_SPATIAL_BASE(params, Path(data_dir), Path(export_dir))
    # The base renderer closes figures. Force collection of large contour and
    # dataframe objects before Gunicorn accepts the next request.
    gc.collect()
    context = dict(result.get("context") or {})
    context["spatial_panel_limit"] = CDE_MAX_SPATIAL_PANELS
    context["bounded_memory_rendering"] = True
    result["context"] = context
    return result

# ---------------------------------------------------------------------------
# 2026-07-18 REQUEST-ONLY, PUBLICATION-STYLE SPATIAL MAP UPDATE
# ---------------------------------------------------------------------------
# The workspace does not call any of these functions on GET. They are reached
# only after the user explicitly submits Generate Plot/Run Analysis.
CDE_MAX_SPATIAL_PANELS = 24
CDE_PANEL_MAX_AXIS_CELLS = 80
CDE_SINGLE_MAP_MAX_AXIS_CELLS = 160


def _sampled_listed_cmap(name: str, count: int, start: float = 0.08, stop: float = 0.96):
    """Build a calm discrete palette from a standard Matplotlib colour map."""
    from matplotlib.colors import ListedColormap
    base = plt.get_cmap(name)
    return ListedColormap(base(np.linspace(start, stop, count)))


def _spatial_publication_style(dataset_key: str, element: str, context: Dict[str, Any]):
    """Return standard operational classes for the selected weather element."""
    from matplotlib.colors import BoundaryNorm, ListedColormap

    family = str(context.get("family") or DATASETS.get(dataset_key, {}).get("family") or "").lower()
    variable = str(context.get("variable") or "").lower()
    element_lower = str(element or "").lower()
    months = [int(v) for v in context.get("months", []) if str(v).strip()]
    all_months = months == list(range(1, 13))

    if family == "rainfall" or any(word in element_lower for word in ("rainfall", "precipitation")):
        if all_months:
            boundaries = np.asarray([0, 100, 250, 500, 750, 1000, 1250, 1500, 1750, 2000, 2500, 3000], dtype=float)
            cmap = _sampled_listed_cmap("YlGn", len(boundaries) - 1, 0.05, 0.98)
        elif len(months) == 1:
            boundaries = np.asarray([0, 1, 5, 10, 25, 50, 75, 100, 150, 200, 300, 450], dtype=float)
            cmap = _sampled_listed_cmap("YlGnBu", len(boundaries) - 1, 0.04, 0.98)
        else:
            boundaries = np.asarray([0, 10, 25, 50, 100, 150, 200, 300, 400, 600, 800, 1000], dtype=float)
            cmap = _sampled_listed_cmap("YlGn", len(boundaries) - 1, 0.05, 0.98)
        label = "Rainfall (mm)" if dataset_key == "chirps_rainfall" or "rainfall" in element_lower else "Precipitation (mm)"
        extend = "max"
    elif family == "temperature" or "temperature" in element_lower:
        if variable in {"tmin", "tn", "minimum_temperature"} or "minimum" in element_lower:
            boundaries = np.asarray([5, 10, 13, 15, 17, 19, 21, 23, 26], dtype=float)
        elif variable in {"tmax", "tx", "maximum_temperature"} or "maximum" in element_lower:
            boundaries = np.asarray([15, 20, 23, 26, 29, 32, 35, 38, 42], dtype=float)
        else:
            boundaries = np.asarray([10, 15, 18, 20, 22, 24, 26, 28, 32], dtype=float)
        colours = ["#1a1aff", "#4169e1", "#3ec1d3", "#76e3ad", "#d7ff3f", "#ffd400", "#ff8c00", "#ff0000"]
        cmap = ListedColormap(colours[: len(boundaries) - 1])
        label = "Temperature (°C)"
        extend = "both"
    elif family == "humidity" or "humidity" in element_lower:
        boundaries = np.asarray([0, 20, 30, 40, 50, 60, 70, 80, 90, 100], dtype=float)
        cmap = _sampled_listed_cmap("YlGnBu", len(boundaries) - 1, 0.05, 0.98)
        label = "Relative Humidity (%)"
        extend = "neither"
    elif family in {"soil_water", "soil_moisture"} or "soil moisture" in element_lower or "soil water" in element_lower:
        boundaries = np.asarray([0, 0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50], dtype=float)
        cmap = _sampled_listed_cmap("YlGnBu", len(boundaries) - 1, 0.05, 0.98)
        label = y_axis_label(element, context.get("unit"))
        extend = "max"
    elif family == "cloud" or "cloud" in element_lower:
        boundaries = np.asarray([0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0], dtype=float)
        cmap = _sampled_listed_cmap("Greys", len(boundaries) - 1, 0.10, 0.92)
        label = y_axis_label(element, context.get("unit"))
        extend = "neither"
    else:
        # For variables without internationally fixed classes, use robust
        # percentile classes so one outlier does not make the map unreadable.
        boundaries = None
        cmap = plt.get_cmap(plot_cmap_for(dataset_key))
        label = y_axis_label(element, context.get("unit"))
        extend = "both"

    norm = BoundaryNorm(boundaries, cmap.N, clip=False) if boundaries is not None else None
    return {"boundaries": boundaries, "cmap": cmap, "norm": norm, "label": label, "extend": extend}


def _format_class_value(value: float) -> str:
    if abs(value) >= 100 or float(value).is_integer():
        return f"{value:.0f}"
    if abs(value) >= 1:
        return f"{value:.1f}".rstrip("0").rstrip(".")
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _configure_discrete_colorbar(cbar, style_spec: dict[str, Any]) -> None:
    boundaries = style_spec.get("boundaries")
    if boundaries is not None and len(boundaries) >= 2:
        centres = (boundaries[:-1] + boundaries[1:]) / 2.0
        cbar.set_ticks(centres)
        cbar.set_ticklabels([
            f"{_format_class_value(a)} - {_format_class_value(b)}"
            for a, b in zip(boundaries[:-1], boundaries[1:])
        ])
        cbar.ax.tick_params(labelsize=8, length=0, pad=8)
    cbar.set_label(style_spec.get("label") or "Value", fontsize=10, fontweight="bold", labelpad=12)
    cbar.outline.set_linewidth(0.8)


def _draw_hydrology(ax, *, ocean: bool = False, lakes: bool = False, rivers: bool = False,
                     data_dir: Path | None = None, clip_polygons: list[np.ndarray] | None = None) -> dict[str, Any]:
    """Draw every complete lake geometry in a quiet water colour above data."""
    from matplotlib.collections import PatchCollection
    ax.set_facecolor("white")
    polygons, status = _lake_polygons_and_status(data_dir)
    status["full_geometry_rendering"] = True
    if not lakes or not polygons:
        return status
    patches = [Polygon(arr, closed=True) for arr in polygons if len(arr) >= 3]
    if patches:
        collection = PatchCollection(
            patches,
            facecolor="#b9d0e8",
            edgecolor="#7fa6c9",
            linewidths=0.35,
            alpha=1.0,
            zorder=18,
            clip_on=True,
        )
        collection.set_gid("cde-lakes")
        ax.add_collection(collection)
    return status


def _publication_map_title(element: str, context: Dict[str, Any]) -> str:
    """Use concise operational titles matching TMA publication examples."""
    family = str(context.get("family") or "").lower()
    years = [int(v) for v in context.get("years", []) if str(v).strip()]
    months = [int(v) for v in context.get("months", []) if str(v).strip()]
    year_text = _compact_year_label(years) if years else str(context.get("period_label") or "Selected period")
    clean = re.sub(r"^(CHIRPS|ERA5(?:-Land)?)\s+", "", str(element or "Weather Element"), flags=re.I).strip()
    if family == "rainfall" or "precipitation" in clean.lower() or "rainfall" in clean.lower():
        clean = "Rainfall"
        if months == list(range(1, 13)):
            if len(years) == 1:
                return f"Tanzania Total Annual {clean} — {years[0]}"
            return f"Tanzania Long-Term Mean Annual {clean} — {year_text}"
        if len(months) == 1:
            month = pd.Timestamp(2000, months[0], 1).strftime("%B")
            if len(years) == 1:
                return f"Tanzania Total Monthly {clean} — {month} {years[0]}"
            return f"Tanzania Long-Term Mean {month} {clean} — {year_text}"
        for season_name, season_months in SEASON_DEFINITIONS.items():
            if list(season_months) == months:
                if len(years) == 1:
                    return f"Tanzania {season_name} {clean} — {years[0]}"
                return f"Tanzania Long-Term Mean {season_name} {clean} — {year_text}"
        month_names = ", ".join(pd.Timestamp(2000, m, 1).strftime("%b") for m in months)
        return f"Tanzania Mean Total {clean} ({month_names}) — {year_text}"
    if family == "temperature" or "temperature" in clean.lower():
        if len(years) > 1:
            return f"Tanzania {clean} Climatology — {year_text}"
        return f"Tanzania {clean} — {year_text}"
    if len(years) > 1:
        return f"Tanzania {clean} Climatology — {year_text}"
    return f"Tanzania {clean} — {year_text}"


def _render_tanzania_map_axis(
    ax, *, x: np.ndarray, y: np.ndarray, values: np.ndarray, data_dir: Path,
    selected_boundaries: list[np.ndarray], level: int, dataset_key: str,
    show_ocean: bool, show_lakes: bool, show_rivers: bool, title: str,
    norm=None, show_labels: bool = True, show_cartographic_elements: bool = True,
    render_style: str = "grid", style_spec: dict[str, Any] | None = None,
):
    """Render a clean grid or smooth map only after an explicit request."""
    from matplotlib.collections import LineCollection
    from matplotlib.ticker import FuncFormatter

    style_spec = style_spec or {"cmap": plt.get_cmap(plot_cmap_for(dataset_key)), "norm": norm, "boundaries": None}
    cmap = style_spec["cmap"]
    active_norm = style_spec.get("norm") or norm
    boundaries = style_spec.get("boundaries")
    clip_patch = _map_clip_patch(ax, selected_boundaries)
    # Spatial products are drawn only as native grid cells.  Contouring,
    # interpolation, smoothing and spatial resampling would alter the supplied
    # 0.25-degree data and are therefore deliberately disabled.
    render_style = "grid"
    image = ax.pcolormesh(x, y, values, shading="nearest", cmap=cmap, norm=active_norm, zorder=4)
    _clip_map_artist(image, clip_patch)

    # Internal administrative boundaries are below lakes; only the true
    # international boundary is redrawn above transboundary lakes.
    _draw_admin_boundaries(ax, selected_boundaries, level)
    try:
        national_like, _ = _admin_polygons(data_dir, 1)
    except Exception:
        national_like = selected_boundaries
    ax.add_collection(LineCollection(national_like, colors="#263238", linewidths=0.85, alpha=0.95, zorder=15))
    lake_status = _draw_hydrology(ax, lakes=show_lakes, data_dir=data_dir)
    _draw_international_boundary_on_top(ax, national_like)

    if show_labels and level == 1:
        for name, lon, lat in _admin_feature_labels(data_dir, 1):
            ax.text(lon, lat, name, fontsize=6.0, fontweight="bold", color="#17252f",
                    ha="center", va="center", zorder=20)

    # Fixed operational domain: it removes stray geometries and prevents a
    # lake or invalid feature from expanding the plot unexpectedly.
    lake_polygons, _ = _lake_polygons_and_status(data_dir) if show_lakes else ([], {})
    extent_polygons = list(national_like) + list(lake_polygons)
    minx, maxx, _miny, _maxy = _tight_polygon_bounds(extent_polygons, padding=0.08)
    # Permit the complete western transboundary lakes while preventing a bad
    # or distant feature from expanding the map into an unreadable canvas.
    ax.set_xlim(max(27.75, minx), min(41.25, maxx))
    ax.set_ylim(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
    ax.set_aspect("equal", adjustable="box")
    ax.margins(0)
    ax.set_title(title, fontsize=9 if not show_cartographic_elements else 14, fontweight="bold", pad=10)
    if show_cartographic_elements:
        ax.set_xlabel("Longitude")
        ax.set_ylabel("Latitude")
        ax.xaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{abs(value):g}°{'E' if value >= 0 else 'W'}"))
        ax.yaxis.set_major_formatter(FuncFormatter(lambda value, _: "0°" if abs(value) < 1e-9 else f"{abs(value):g}°{'N' if value > 0 else 'S'}"))
        ax.set_xticks(np.arange(28, 42, 2))
        ax.set_yticks(np.arange(-12, 1, 2))
        _draw_north_arrow(ax)
        _draw_scale_bar(ax)
        ax.grid(True, linewidth=0.45, alpha=0.24, color="#78909c", zorder=1)
    else:
        ax.set_xticks([])
        ax.set_yticks([])
    setattr(ax, "_cde_lake_status", lake_status)
    return image


def _generate_requested_spatial_map(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    """Generate reference-style maps with bounded memory and no eager loading."""
    layout, jobs = _prepare_spatial_jobs(params, Path(data_dir))
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    level = max(1, min(3, int(params.get("map_admin_level") or 1)))
    boundaries, boundary_path = _admin_polygons(Path(data_dir), level)
    national, _ = _admin_polygons(Path(data_dir), 1)
    admin_label = {1: "Admin 1 — Regions", 2: "Admin 2 — Districts", 3: "Admin 3 — Wards"}[level]

    prepared = []
    for grid, context, label in jobs:
        work = grid[
            grid["longitude"].astype(float).between(TANZANIA_BOUNDS["lon_min"], 41.0)
            & grid["latitude"].astype(float).between(TANZANIA_BOUNDS["lat_min"], TANZANIA_BOUNDS["lat_max"])
        ].copy()
        if work.empty:
            continue
        pivot = work.pivot_table(index="latitude", columns="longitude", values="value", aggfunc="mean").sort_index()
        x = pivot.columns.to_numpy(float)
        y = pivot.index.to_numpy(float)
        values = pivot.to_numpy(float)
        prepared.append((x, y, values, work, context, label))
    if not prepared:
        raise ValueError("No spatial values were found within the Tanzania map extent.")

    element = _explicit_weather_element(dataset_key, params.get("variable"), prepared[0][4])
    style_spec = _spatial_publication_style(dataset_key, element, prepared[0][4])
    # Operational map colour bars use an upper extension only. Values below the
    # first class remain in the first colour; values above the last class are
    # indicated by the single upward triangular extension.
    style_spec["extend"] = "max"
    # If a generic variable has no fixed classes, derive a robust shared scale.
    if style_spec.get("boundaries") is None:
        finite_parts = []
        for x, y, values, *_ in prepared:
            inside = _mask_grid_to_tanzania(values, x, y, national)
            finite = inside[np.isfinite(inside)]
            if finite.size:
                finite_parts.append(finite)
        finite = np.concatenate(finite_parts) if finite_parts else np.asarray([0.0, 1.0])
        low, high = np.nanpercentile(finite, [2, 98])
        if np.isclose(low, high):
            high = low + 1.0
        from matplotlib.colors import Normalize
        style_spec["norm"] = Normalize(float(low), float(high))

    custom_title = _clean_custom_plot_title(params.get("custom_plot_title"))
    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    render_style = "grid"
    stem = _cde_prefixed_stem(f"spatial_{render_style}_{dataset_key}_{'panels' if layout == 'panel' else slugify(prepared[0][5])}_admin{level}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    show_lakes = _truthy(params.get("show_lakes", True))
    image = None

    if layout == "panel":
        count = len(prepared)
        columns = min(6, max(2, int(math.ceil(math.sqrt(count * 1.15)))))
        rows = int(math.ceil(count / columns))
        fig, axes = plt.subplots(rows, columns, figsize=(2.7 * columns + 1.2, 2.75 * rows), squeeze=False)
        flat = axes.ravel()
        for idx, (x, y, values, _work, context, label) in enumerate(prepared):
            image = _render_tanzania_map_axis(
                flat[idx], x=x, y=y, values=values, data_dir=Path(data_dir),
                selected_boundaries=boundaries, level=level, dataset_key=dataset_key,
                show_ocean=False, show_lakes=show_lakes, show_rivers=False,
                title=label, show_labels=False, show_cartographic_elements=False,
                render_style=render_style, style_spec=style_spec,
            )
            if idx == 0:
                _draw_north_arrow(flat[idx])
        for axis in flat[len(prepared):]:
            axis.axis("off")
        default_title = _publication_map_title(element, prepared[0][4])
        fig.suptitle(custom_title or default_title, fontsize=15, fontweight="bold", y=0.995)
        if image is not None:
            cax = fig.add_axes([0.945, 0.17, 0.012, 0.66])
            cbar = fig.colorbar(
                image, cax=cax, extend=style_spec.get("extend", "both"),
                boundaries=style_spec.get("boundaries"), spacing="uniform",
            )
            _configure_discrete_colorbar(cbar, style_spec)
        fig.subplots_adjust(left=0.018, right=0.91, bottom=0.02, top=0.92, wspace=0.04, hspace=0.11)
        fig.savefig(plot_path, dpi=125, bbox_inches="tight", pad_inches=0.04, facecolor="white")
        plt.close(fig)
    else:
        x, y, values, _work, context, label = prepared[0]
        fig = plt.figure(figsize=(12.6, 9.0))
        gs = fig.add_gridspec(1, 2, width_ratios=[1.0, 0.035], wspace=0.055)
        ax = fig.add_subplot(gs[0, 0])
        cax = fig.add_subplot(gs[0, 1])
        title = custom_title or _publication_map_title(element, context)
        image = _render_tanzania_map_axis(
            ax, x=x, y=y, values=values, data_dir=Path(data_dir),
            selected_boundaries=boundaries, level=level, dataset_key=dataset_key,
            show_ocean=False, show_lakes=show_lakes, show_rivers=False,
            title=title, show_labels=True, show_cartographic_elements=True,
            render_style=render_style, style_spec=style_spec,
        )
        cbar = fig.colorbar(
            image, cax=cax, extend=style_spec.get("extend", "both"),
            boundaries=style_spec.get("boundaries"), spacing="uniform",
        )
        _configure_discrete_colorbar(cbar, style_spec)
        fig.subplots_adjust(left=0.06, right=0.94, bottom=0.07, top=0.93, wspace=0.055)
        fig.savefig(plot_path, dpi=165, bbox_inches="tight", pad_inches=0.04, facecolor="white")
        plt.close(fig)

    export_parts = []
    for _x, _y, _values, grid, context, label in prepared:
        part = grid.copy()
        part.insert(0, "map_period", label)
        part.insert(1, "administrative_level", admin_label)
        part.insert(2, "map_surface", "Native grid cells — no interpolation")
        export_parts.append(part)
    export_grid = pd.concat(export_parts, ignore_index=True)
    excel_path, _, _ = _save_data(export_grid, dirs["plots"], stem + "_data", table="plot_products")
    lake_status = _lake_polygons_and_status(Path(data_dir))[1]
    context = dict(prepared[0][4])
    context.update({
        "administrative_level": level,
        "administrative_level_label": admin_label,
        "boundary_file": boundary_path.name,
        "map_output_layout": layout,
        "map_count": len(prepared),
        "map_render_style": "grid",
        "native_grid_preserved": True,
        "spatial_interpolation": "None",
        "spatial_resampling": "None",
        "variable_label": element,
        "custom_plot_title": custom_title,
        "lake_layer": lake_status,
        "request_only_processing": True,
        "publication_style": True,
        "bounded_memory_rendering": True,
        "spatial_panel_limit": CDE_MAX_SPATIAL_PANELS,
    })
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "rows": len(export_grid),
        "context": context,
        "period_label": "All selected periods" if layout == "panel" else prepared[0][5],
        "summary_cards": [
            {"label": "Weather Element", "value": element, "note": context.get("period_label")},
            {"label": "Map Surface", "value": "Native grid cells", "note": f"No interpolation · {len(prepared)} map panel(s)"},
            {"label": "Boundary Level", "value": admin_label, "note": boundary_path.name},
        ],
    }

# Keep the panel-builder limit aligned with the request renderer. This check is
# intentionally done before any Zarr store is opened.
_CDE_REQUEST_ONLY_PANEL_JOBS_BASE = _spatial_panel_jobs

def _spatial_panel_jobs(params: Dict[str, Any]) -> list[tuple[Dict[str, Any], str]]:
    jobs = _CDE_REQUEST_ONLY_PANEL_JOBS_BASE(params)
    if len(jobs) > CDE_MAX_SPATIAL_PANELS:
        raise ValueError(
            f"The all-maps figure is limited to {CDE_MAX_SPATIAL_PANELS} maps to prevent server overload. "
            "Reduce the selected years or generate a single aggregated map."
        )
    return jobs

# ---------------------------------------------------------------------------
# 2026-07-19 requested output, labelling, wind, diurnal and native-grid update
# ---------------------------------------------------------------------------
_AXIS_RESOLUTION_CONTEXT: ContextVar[str | None] = ContextVar("cde_axis_resolution", default=None)


def _normalise_display_unit(unit: str | None) -> str:
    text = str(unit or "").strip()
    lowered = text.lower().replace(" ", "")
    aliases = {
        "ms-1": "m/s", "ms**-1": "m/s", "m/s": "m/s", "m.s-1": "m/s",
        "degrees": "degrees", "degree": "degrees", "degrees_from_north": "degrees",
        "degc": "°C", "degree_celsius": "°C", "degrees_celsius": "°C",
        "percent": "%", "%": "%",
    }
    return aliases.get(lowered, text)


def y_axis_label(
    element: str | None,
    unit: str | None,
    resolution: str | None = None,
    aggregation: str | None = None,
) -> str:
    """Return a concise, operational axis label with time-aware units."""
    element_text = str(element or "Value").replace("Rainfall", "Precipitation")
    for prefix in ("CHIRPS ", "ERA5-Land ", "ERA5 Land ", "ERA5 "):
        if element_text.startswith(prefix):
            element_text = element_text[len(prefix):]
    display_unit = _normalise_display_unit(unit)
    scale = str(resolution or _AXIS_RESOLUTION_CONTEXT.get() or "").lower()
    lower_element = element_text.lower()
    lower_aggregation = str(aggregation or "").lower()
    precipitation = "precipitation" in lower_element or "rainfall" in lower_element
    cumulative = "cumulative" in lower_element or "cumulative" in lower_aggregation
    if precipitation and display_unit.lower() == "mm" and not cumulative:
        suffix = {
            "hourly": "hour",
            "daily": "day",
            "monthly": "month",
            "annual": "year",
            "seasonal": "season",
        }.get(scale)
        if suffix:
            display_unit = f"mm/{suffix}"
    return f"{element_text} ({display_unit})" if display_unit else element_text


def _find_wind_component(ds: xr.Dataset, component: str) -> str | None:
    component = component.lower()
    exact = {
        "u": ("u10", "u", "u_component_of_wind", "10m_u_component_of_wind", "u_component_of_wind_10m"),
        "v": ("v10", "v", "v_component_of_wind", "10m_v_component_of_wind", "v_component_of_wind_10m"),
    }[component]
    lowered = {str(name).lower(): str(name) for name in ds.data_vars}
    for candidate in exact:
        if candidate in lowered:
            return lowered[candidate]
    phrases = ("u component", "eastward wind") if component == "u" else ("v component", "northward wind")
    for name in data_variables(ds):
        da = ds[name]
        text = " ".join(
            str(value or "").lower()
            for value in (name, da.attrs.get("long_name"), da.attrs.get("standard_name"), da.attrs.get("short_name"))
        )
        if any(phrase in text for phrase in phrases):
            return name
    return None


def _wind_dataarray(ds: xr.Dataset, requested: str) -> tuple[xr.DataArray, str, str, str]:
    """Return a direct or derived 10 m wind DataArray and descriptive metadata."""
    requested = str(requested or "wind_speed").lower()
    direct_aliases = {
        "wind_speed": ("wind_speed", "ws", "wspd", "si10", "10m_wind_speed"),
        "wind_direction": ("wind_direction", "wd", "wdir", "10m_wind_direction"),
    }
    lowered = {str(name).lower(): str(name) for name in ds.data_vars}
    for alias in direct_aliases.get(requested, ()):
        if alias in lowered:
            name = lowered[alias]
            da = ds[name]
            label = "Wind Speed at 10 m" if requested == "wind_speed" else "Wind Direction at 10 m"
            unit = str(da.attrs.get("units") or ("m/s" if requested == "wind_speed" else "degrees"))
            return da, name, label, unit

    u_name = _find_wind_component(ds, "u")
    v_name = _find_wind_component(ds, "v")
    if not u_name or not v_name:
        raise ValueError(
            f"Could not derive '{requested}'. The selected wind store must contain both U and V components. "
            f"Available variables: {', '.join(data_variables(ds)) or 'none'}."
        )
    u = ds[u_name].astype(float)
    v = ds[v_name].astype(float)
    if requested == "wind_direction":
        da = (270.0 - np.degrees(np.arctan2(v, u))) % 360.0
        da.attrs = {"long_name": "Wind Direction at 10 m", "units": "degrees"}
        return da, f"derived_from_{u_name}_{v_name}", "Wind Direction at 10 m", "degrees"
    da = np.hypot(u, v)
    unit = str(u.attrs.get("units") or v.attrs.get("units") or "m/s")
    da.attrs = {"long_name": "Wind Speed at 10 m", "units": unit}
    return da, f"derived_from_{u_name}_{v_name}", "Wind Speed at 10 m", unit


# Replace strict point extraction only for wind-component stores. All other
# datasets keep the established lazy, nearest-grid workflow.
_CDE_POINT_SERIES_BEFORE_WIND_DERIVATION = extract_point_series


def extract_point_series(
    data_dir: Path,
    dataset_key: str,
    resolution: str,
    lat: float,
    lon: float,
    start_date: str | None,
    end_date: str | None,
    variable: str | None = None,
    season: str | None = None,
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    requested = str(variable or "auto").lower()
    if dataset_key != "era5_wind" or requested not in {"auto", "wind_speed", "wind_direction", "wind_speed_direction"}:
        return _CDE_POINT_SERIES_BEFORE_WIND_DERIVATION(
            data_dir, dataset_key, resolution, lat, lon, start_date, end_date, variable, season
        )
    requested = "wind_speed" if requested in {"auto", "wind_speed_direction"} else requested
    file_path = find_file(Path(data_dir), dataset_key, resolution, season=season)
    assert file_path
    cache = _PRODUCT_DATA_CACHE.get()
    key = (
        "point", str(Path(file_path).resolve()), dataset_key, resolution,
        round(float(lat), 8), round(float(lon), 8), str(start_date or ""),
        str(end_date or ""), requested, str(season or ""), "derived-wind-v2",
    )
    if cache is not None and key in cache["point"]:
        cache["stats"]["point_hits"] += 1
        cached_df, cached_context = cache["point"][key]
        return cached_df.copy(deep=False), dict(cached_context)
    if cache is not None:
        cache["stats"]["point_misses"] += 1

    meta = DATASETS[dataset_key]
    with open_data_store(file_path, decode_times=True) as ds:
        time_name = detect_time_coord(ds)
        lat_name, lon_name = detect_lat_lon(ds)
        da, actual_variable, actual_label, source_units = _wind_dataarray(ds, requested)
        if start_date or end_date:
            da, time_name = slice_time_range(da, time_name, start_date, end_date)
        point = da.sel({lat_name: float(lat), lon_name: float(lon)}, method="nearest")
        nearest_lat = float(np.asarray(point[lat_name].values).reshape(-1)[0])
        nearest_lon = float(np.asarray(point[lon_name].values).reshape(-1)[0])
        df = _point_dataarray_to_frame(point, time_name, actual_variable)
        df["value"], unit = convert_series_units(df["value"], source_units, meta["family"], requested)
        df = df[["time", "value"]].dropna(subset=["value"]).sort_values("time").reset_index(drop=True)
    context = {
        "file": Path(file_path).name,
        "storage_format": store_kind(Path(file_path)),
        "source_path": str(file_path),
        "dataset_key": dataset_key,
        "dataset_label": meta["label"],
        "family": meta["family"],
        "variable": requested,
        "actual_variable": actual_variable,
        "variable_label": actual_label,
        "actual_variable_label": actual_label,
        "unit": _normalise_display_unit(unit),
        "resolution": resolution,
        "season": season or "",
        "requested_latitude": lat,
        "requested_longitude": lon,
        "nearest_latitude": nearest_lat,
        "nearest_longitude": nearest_lon,
        "wind_components_derived": actual_variable.startswith("derived_from_"),
    }
    if cache is not None:
        cache["point"][key] = (df.copy(deep=False), dict(context))
    return df.copy(deep=False), context


def _native_coordinate_spacing(values: Any) -> float | None:
    array = np.asarray(values, dtype=float).reshape(-1)
    array = np.unique(array[np.isfinite(array)])
    if array.size < 2:
        return None
    diffs = np.abs(np.diff(np.sort(array)))
    diffs = diffs[diffs > 0]
    return float(np.median(diffs)) if diffs.size else None


def _spatial_grid_from_open_dataset(
    ds: xr.Dataset,
    spec: dict[str, Any],
    *,
    max_axis_cells: int,
) -> tuple[pd.DataFrame, Dict[str, Any]]:
    """Create a temporal aggregate while retaining every native spatial cell."""
    dataset_key = spec["dataset_key"]
    variable = spec["variable"]
    period = spec["period"]
    source_resolution = spec["source_resolution"]
    file_path = spec["file_path"]
    meta = DATASETS[dataset_key]
    time_name = detect_time_coord(ds)
    lat_name, lon_name = detect_lat_lon(ds)

    if dataset_key == "era5_wind" and str(variable).lower() in {"wind_speed", "wind_direction"}:
        requested = str(variable).lower()
        da, var_name, actual_variable_label, source_units = _wind_dataarray(ds, requested)
        variable_label = actual_variable_label
    else:
        var_name = pick_variable(ds, dataset_key, variable)
        source_da = ds[var_name]
        da = _select_product_statistic_dimension(source_da, variable or var_name, keep_dims={time_name, lat_name, lon_name})
        actual_variable_label = str(source_da.attrs.get("long_name") or source_da.attrs.get("standard_name") or var_name).replace("_", " ").title()
        variable_label = requested_element_label(dataset_key, variable) or actual_variable_label
        source_units = str(source_da.attrs.get("units") or meta.get("unit") or "")

    da, unit = convert_dataarray_units(da, meta["family"], str(variable or var_name))
    selected = _select_time_period(
        da,
        time_name,
        period["years"],
        None if source_resolution in {"annual", "seasonal"} else period["months"],
    )
    count = int(selected.sizes.get(time_name, 1)) if time_name in selected.dims else 1
    if count < 1:
        raise ValueError(f"No values were found for {period['period_label']}.")

    aggregation = "Selected period"
    if time_name in selected.dims:
        if count == 1:
            aggregated = selected.isel({time_name: 0}, drop=True)
        elif dataset_key == "era5_wind" and str(variable).lower() == "wind_direction":
            # Circularly meaningful direction: average U and V first, then
            # derive the meteorological FROM direction from the mean vector.
            u_name = _find_wind_component(ds, "u")
            v_name = _find_wind_component(ds, "v")
            if not u_name or not v_name:
                raise ValueError("Wind direction requires both U and V components.")
            u_selected = _select_time_period(ds[u_name].astype(float), time_name, period["years"], None if source_resolution in {"annual", "seasonal"} else period["months"])
            v_selected = _select_time_period(ds[v_name].astype(float), time_name, period["years"], None if source_resolution in {"annual", "seasonal"} else period["months"])
            u_mean = u_selected.mean(time_name, skipna=True)
            v_mean = v_selected.mean(time_name, skipna=True)
            aggregated = (270.0 - np.degrees(np.arctan2(v_mean, u_mean))) % 360.0
            unit = "degrees"
            aggregation = "Mean-vector wind direction"
        elif meta["family"] == "rainfall":
            if source_resolution in {"annual", "seasonal"}:
                aggregated = selected.mean(time_name, skipna=True)
                aggregation = "Mean total across selected years"
            elif period["mode"] == "custom" and len(period["years"]) > 1:
                try:
                    by_year = selected.groupby(f"{time_name}.year").sum(time_name, skipna=True)
                    aggregated = by_year.mean("year", skipna=True)
                except Exception:
                    aggregated = selected.sum(time_name, skipna=True) / max(1, len(period["years"]))
                aggregation = "Mean selected-month total across years"
            else:
                aggregated = selected.sum(time_name, skipna=True)
                aggregation = "Selected-period total"
        else:
            aggregated = selected.mean(time_name, skipna=True)
            aggregation = "Selected-period mean"
    else:
        aggregated = selected

    # No call to coarsen, interp, reindex, resample, rolling spatial filters or
    # stride-based isel is made here. The original latitude/longitude cells are
    # written and rendered exactly as stored.
    frame = aggregated.to_dataframe(name="value").reset_index()
    if lat_name != "latitude":
        frame = frame.rename(columns={lat_name: "latitude"})
    if lon_name != "longitude":
        frame = frame.rename(columns={lon_name: "longitude"})
    frame = frame[["latitude", "longitude", "value"]].dropna(subset=["value"]).reset_index(drop=True)
    lat_spacing = _native_coordinate_spacing(aggregated[lat_name].values)
    lon_spacing = _native_coordinate_spacing(aggregated[lon_name].values)
    resolution_parts = [value for value in (lon_spacing, lat_spacing) if value is not None]
    native_label = " × ".join(f"{value:.4g}°" for value in resolution_parts) if resolution_parts else "Native source grid"
    context = {
        "file": Path(file_path).name,
        "source_path": str(file_path),
        "dataset_key": dataset_key,
        "dataset_label": meta["label"],
        "family": meta["family"],
        "variable": variable,
        "actual_variable": var_name,
        "variable_label": variable_label,
        "actual_variable_label": actual_variable_label,
        "unit": _normalise_display_unit(unit or source_units),
        "resolution": source_resolution,
        "period_label": period["period_label"],
        "aggregation": aggregation,
        "months": period["months"],
        "years": period["years"],
        "render_grid_reduced": False,
        "native_grid_preserved": True,
        "native_latitude_resolution_degrees": lat_spacing,
        "native_longitude_resolution_degrees": lon_spacing,
        "native_grid_resolution_label": native_label,
        "spatial_interpolation": "None",
        "spatial_resampling": "None",
    }
    return frame, context


def _selection_period_descriptor(params: Dict[str, Any], resolution: str | None = None) -> str:
    resolution = str(resolution or params.get("resolution") or params.get("index_resolution") or "").lower()
    start = str(params.get("start_date") or params.get("index_start_date") or "")[:10]
    end = str(params.get("end_date") or params.get("index_end_date") or "")[:10]
    parts: list[str] = []
    custom_years = str(params.get("index_custom_years") or params.get("map_custom_years") or "").strip()
    if custom_years:
        parts.append(f"years {custom_years}")
    elif start and end:
        if resolution in {"annual", "seasonal", "monthly"} and len(start) >= 4 and len(end) >= 4:
            parts.append(start[:4] if start[:4] == end[:4] else f"{start[:4]}–{end[:4]}")
        else:
            parts.append(start if start == end else f"{start} to {end}")
    season = str(params.get("season") or params.get("index_season") or "").strip().upper()
    custom_months = str(params.get("custom_months") or params.get("map_custom_months") or "").strip()
    if season and season not in {"ANNUAL", "ALL", "ALL MONTHS"}:
        parts.append(f"{season} season" if season != "CUSTOM" else f"months {custom_months or 'custom selection'}")
    elif custom_months and str(params.get("index_season") or "").upper() == "CUSTOM":
        parts.append(f"months {custom_months}")
    return " · ".join(parts) or "Selected period"


def _compute_diurnal_cycle_stats(data: pd.DataFrame, *, temperature: bool = False) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Compute the 24-hour climatological cycle and daily temperature range."""
    work = data[["time", "value"]].copy()
    work["time"] = pd.to_datetime(work["time"], errors="coerce")
    work["value"] = pd.to_numeric(work["value"], errors="coerce")
    work = work.dropna().sort_values("time")
    work["hour"] = work["time"].dt.hour
    hourly = work.groupby("hour", as_index=False)["value"].agg(
        mean="mean", standard_deviation="std", minimum="min", maximum="max", observations="count"
    )
    hourly = pd.DataFrame({"hour": np.arange(24)}).merge(hourly, on="hour", how="left")
    daily = pd.DataFrame(columns=["date", "daily_minimum", "daily_maximum", "diurnal_temperature_range"])
    if temperature:
        daily_work = work.assign(date=work["time"].dt.floor("D")).groupby("date", as_index=False)["value"].agg(
            daily_minimum="min", daily_maximum="max", observations="count"
        )
        # A full daily range needs at least two different hourly observations.
        daily_work = daily_work[daily_work["observations"] >= 2].copy()
        daily_work["diurnal_temperature_range"] = daily_work["daily_maximum"] - daily_work["daily_minimum"]
        daily = daily_work[["date", "daily_minimum", "daily_maximum", "diurnal_temperature_range"]]
        hourly["mean_daily_diurnal_temperature_range"] = float(daily["diurnal_temperature_range"].mean()) if not daily.empty else np.nan
        hourly["days_used_for_diurnal_range"] = int(len(daily))
    return hourly, daily


_CDE_EXTENDED_BEFORE_DIURNAL_UPDATE = _generate_extended_point_plot


def _generate_extended_point_plot(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    if str(params.get("plot_type") or "") != "diurnal_cycle":
        return _CDE_EXTENDED_BEFORE_DIURNAL_UPDATE(params, data_dir, export_dir)
    dataset_key = str(params.get("dataset") or "era5_temperature")
    resolution = str(params.get("resolution") or "hourly").lower()
    if resolution != "hourly":
        raise ValueError("Diurnal temperature and hourly-cycle plots require hourly data.")
    variable = str(params.get("variable") or "auto")
    season = str(params.get("season") or "").strip().upper() or None
    location = str(params.get("location_name") or "Selected Location")
    start = str(params.get("start_date") or "")
    end = str(params.get("end_date") or "")
    frame, context = extract_point_series(
        Path(data_dir), dataset_key, resolution,
        float(params.get("latitude")), float(params.get("longitude")),
        start, end, variable=variable, season=season,
    )
    data = _extended_plot_frame(frame)
    if data.empty:
        raise ValueError("No hourly values were found for the selected period.")
    element = variable_display_name(variable, context, dataset_key)
    unit = str(context.get("unit") or DATASETS.get(dataset_key, {}).get("unit") or "")
    temperature = str(context.get("family") or "").lower() == "temperature" or "temperature" in element.lower()
    hourly, daily = _compute_diurnal_cycle_stats(data, temperature=temperature)
    colour = plot_color_for(dataset_key, variable)
    descriptor = _selection_period_descriptor(params, resolution)
    custom_title = _clean_custom_plot_title(params.get("custom_plot_title"))
    default_title = (
        f"Average Diurnal Temperature Cycle for {location} — {descriptor}"
        if temperature else f"Average Diurnal Cycle of {element} for {location} — {descriptor}"
    )
    title = custom_title or default_title
    dirs = ensure_output_dirs(Path(export_dir))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _cde_prefixed_stem(f"diurnal_cycle_{dataset_key}_{variable}_{location}_{stamp}")
    plot_path = dirs["plots"] / f"{stem}.png"
    fig, ax = plt.subplots(figsize=(11.2, 6.2))
    x = hourly["hour"].to_numpy(dtype=float)
    mean = pd.to_numeric(hourly["mean"], errors="coerce").to_numpy(dtype=float)
    minimum = pd.to_numeric(hourly["minimum"], errors="coerce").to_numpy(dtype=float)
    maximum = pd.to_numeric(hourly["maximum"], errors="coerce").to_numpy(dtype=float)
    standard_deviation = pd.to_numeric(hourly["standard_deviation"], errors="coerce").fillna(0).to_numpy(dtype=float)
    ax.fill_between(x, minimum, maximum, color=colour, alpha=0.12, label="Observed hourly minimum–maximum")
    ax.fill_between(x, mean - standard_deviation, mean + standard_deviation, color=colour, alpha=0.24, label="Mean ± 1 standard deviation")
    ax.plot(x, mean, color=colour, marker="o", linewidth=2.1, label="Hourly mean")
    if temperature and not daily.empty:
        mean_dtr = float(daily["diurnal_temperature_range"].mean())
        ax.text(
            0.985, 0.965, f"Mean daily diurnal temperature range: {mean_dtr:.1f} {_normalise_display_unit(unit)}",
            transform=ax.transAxes, ha="right", va="top", fontsize=9,
            bbox={"boxstyle": "round,pad=0.35", "facecolor": "white", "edgecolor": "#8aa1b1", "alpha": 0.92},
        )
    ax.set_xticks(range(0, 24, 2))
    ax.set_xlim(0, 23)
    ax.set_xlabel("Hour of Day (UTC)")
    ax.set_ylabel(y_axis_label(element, unit, "hourly"))
    ax.set_title(title, fontsize=14, fontweight="bold", pad=13)
    ax.legend(loc="best")
    apply_plot_grids(ax)
    _finalize_plot(fig, plot_path)

    export = hourly.copy()
    excel_path, parquet_path, db_path = _save_data(export, dirs["plots"], stem + "_data", table="plot_products")
    context = dict(context)
    context.update({
        "variable_label": element,
        "period_label": descriptor,
        "plot_type": "diurnal_cycle",
        "diurnal_temperature_implemented": temperature,
        "daily_diurnal_temperature_range_mean": float(daily["diurnal_temperature_range"].mean()) if temperature and not daily.empty else None,
        "daily_diurnal_temperature_range_days": int(len(daily)),
        "custom_plot_title": custom_title,
    })
    return {
        "plot_path": plot_path,
        "excel_path": excel_path,
        "parquet_path": parquet_path,
        "db_path": db_path,
        "rows": len(export),
        "context": context,
        "summary_cards": [
            {"label": "Hourly Bins", "value": "24", "note": "00:00–23:00 UTC"},
            {"label": "Daily DTR", "value": f"{context['daily_diurnal_temperature_range_mean']:.1f} {_normalise_display_unit(unit)}" if context["daily_diurnal_temperature_range_mean"] is not None else "Not applicable", "note": f"{len(daily)} complete day(s)"},
            {"label": "Period", "value": descriptor, "note": element},
        ],
    }


# Ranked climate-index plots are intentionally removed. A stale submitted
# value is safely redirected to the automatic plot rather than executed.
INDEX_PLOT_TYPES = [
    ("auto", "Automatic — Best Plot for Selected Time Scale"),
    ("line", "Time-Series Line Plot"),
    ("bar", "Period Comparison Bar Chart"),
    ("area", "Filled Area Plot"),
    ("step", "Step Line Plot"),
    ("lollipop", "Lollipop Plot"),
    ("rolling_mean", "Rolling Mean Plot"),
    ("cumulative", "Cumulative Index Plot"),
    ("anomaly_bar", "Positive–Negative Anomaly Bar Plot"),
    ("heatmap", "Time Heat Map"),
    ("box", "Distribution Box Plot"),
    ("violin", "Distribution Violin Plot"),
    ("histogram", "Frequency Distribution Histogram"),
    ("exceedance", "Exceedance Probability Curve"),
    ("ecdf", "Empirical Cumulative Distribution Plot"),
]
_INDEX_ALL_PLOTS = [key for key, _ in INDEX_PLOT_TYPES]
INDEX_PLOT_RULES = {
    "hourly": [key for key in _INDEX_ALL_PLOTS if key not in {"bar", "anomaly_bar"}],
    "daily": list(_INDEX_ALL_PLOTS),
    "monthly": list(_INDEX_ALL_PLOTS),
    "annual": [key for key in _INDEX_ALL_PLOTS if key != "heatmap"],
    "seasonal": [key for key in _INDEX_ALL_PLOTS if key != "heatmap"],
}


def _index_value_axis_label(params: Dict[str, Any], unit: str, resolution: str) -> str:
    index_type = str(params.get("index_type") or "")
    label = _INDEX_LABELS.get(index_type, index_type.replace("_", " ").title())
    if index_type in _INDEX_COUNT_TYPES | _INDEX_RUN_TYPES:
        return f"{label} (days)"
    if index_type in {"spi", "soil_moisture_index"}:
        return f"{label} (standard deviations)"
    return y_axis_label(label, unit, resolution)


def _index_plot(values: pd.DataFrame, params: Dict[str, Any], output: Path, title: str, unit: str, resolution: str) -> str:
    plot_type = str(params.get("index_plot_type") or "auto").lower()
    if plot_type == "rank":
        plot_type = "auto"
    allowed = INDEX_PLOT_RULES.get(resolution, ["auto", "line"])
    if plot_type not in allowed:
        plot_type = "auto"
    if plot_type == "auto":
        plot_type = "bar" if resolution in {"annual", "seasonal"} else "line"
    frame = values[["time", "value"]].copy()
    frame["time"] = pd.to_datetime(frame["time"])
    frame["value"] = pd.to_numeric(frame["value"], errors="coerce")
    frame = frame.dropna().sort_values("time")
    colour = plot_color_for(index_type=str(params.get("index_type") or ""))
    value_label = _index_value_axis_label(params, unit, resolution)
    time_label = {
        "hourly": "Date / Time (UTC)", "daily": "Date", "monthly": "Month / Year",
        "annual": "Year", "seasonal": "Season Year",
    }.get(resolution, "Date / Time")
    fig, ax = plt.subplots(figsize=(11.4, 6.2))
    numeric = frame["value"]
    if plot_type == "heatmap":
        if resolution == "hourly":
            frame["row"] = frame["time"].dt.strftime("%Y-%m-%d"); frame["bucket"] = frame["time"].dt.hour; x_label = "Hour of Day (UTC)"
        elif resolution == "daily":
            frame["row"] = frame["time"].dt.year; frame["bucket"] = frame["time"].dt.dayofyear; x_label = "Day of Year"
        else:
            frame["row"] = frame["time"].dt.year; frame["bucket"] = frame["time"].dt.month; x_label = "Month"
        pivot = frame.pivot_table(index="row", columns="bucket", values="value", aggfunc="mean")
        image = ax.imshow(pivot.to_numpy(dtype=float), aspect="auto", interpolation="nearest", cmap="viridis")
        if len(pivot.columns) <= 24:
            ax.set_xticks(np.arange(len(pivot.columns))); ax.set_xticklabels([str(value) for value in pivot.columns])
        if len(pivot.index) <= 30:
            ax.set_yticks(np.arange(len(pivot.index))); ax.set_yticklabels([str(value) for value in pivot.index])
        ax.set_xlabel(x_label); ax.set_ylabel("Date / Year")
        fig.colorbar(image, ax=ax, shrink=0.84, pad=0.03, label=value_label)
    elif plot_type == "bar":
        labels = frame["time"].dt.strftime("%Y") if resolution in {"annual", "seasonal"} else frame["time"].dt.strftime("%Y-%m-%d")
        x = np.arange(len(frame)); ax.bar(x, numeric, color=colour)
        step = max(1, math.ceil(len(frame) / 20)); ticks = np.arange(0, len(frame), step)
        ax.set_xticks(ticks); ax.set_xticklabels(labels.iloc[::step], rotation=45, ha="right"); ax.set_xlabel(time_label)
    elif plot_type == "area":
        ax.fill_between(frame["time"], numeric, color=colour, alpha=0.28); ax.plot(frame["time"], numeric, color=colour, linewidth=1.4); ax.set_xlabel(time_label)
    elif plot_type == "step":
        ax.step(frame["time"], numeric, where="mid", color=colour, linewidth=1.5); ax.set_xlabel(time_label)
    elif plot_type == "lollipop":
        work = frame if len(frame) <= 150 else frame.iloc[::max(1, len(frame)//150)]
        markerline, stemlines, _baseline = ax.stem(work["time"], work["value"], basefmt=" ")
        plt.setp(markerline, color=colour, markersize=4); plt.setp(stemlines, color=colour, linewidth=0.8); ax.set_xlabel(time_label)
    elif plot_type == "rolling_mean":
        window = {"hourly": 24, "daily": 30, "monthly": 12, "annual": 5, "seasonal": 5}.get(resolution, 12)
        rolling = numeric.rolling(window, min_periods=max(2, window//3)).mean()
        ax.plot(frame["time"], numeric, color=colour, alpha=0.30, linewidth=0.9, label="Index value")
        ax.plot(frame["time"], rolling, color=colour, linewidth=2.2, label=f"{window}-period rolling mean"); ax.legend(); ax.set_xlabel(time_label)
    elif plot_type == "cumulative":
        cumulative = numeric.cumsum(); ax.fill_between(frame["time"], cumulative, color=colour, alpha=0.25); ax.plot(frame["time"], cumulative, color=colour); ax.set_xlabel(time_label)
    elif plot_type == "anomaly_bar":
        anomaly = numeric - numeric.mean(); colours = np.where(anomaly >= 0, "#d73027", "#4575b4")
        ax.bar(frame["time"], anomaly, color=colours); ax.axhline(0, color="black", linewidth=0.9); ax.set_xlabel(time_label)
    elif plot_type == "box":
        if resolution in {"hourly", "daily", "monthly"}:
            groups = [numeric[frame["time"].dt.month == month].dropna().to_numpy() for month in range(1, 13)]
            _boxplot_with_labels(ax, groups, [pd.Timestamp(2000, month, 1).strftime("%b") for month in range(1, 13)], showfliers=False); ax.set_xlabel("Month")
        else:
            _boxplot_with_labels(ax, [numeric.dropna().to_numpy()], [INDEX_RESOLUTION_LABELS.get(resolution, resolution.title())], showfliers=True)
    elif plot_type == "violin":
        if resolution in {"hourly", "daily", "monthly"}:
            groups = [numeric[frame["time"].dt.month == month].dropna().to_numpy() for month in range(1, 13)]
            groups = [group if len(group) else np.asarray([np.nan]) for group in groups]
            parts = ax.violinplot(groups, positions=np.arange(1, 13), showmeans=True, showmedians=True)
            for body in parts["bodies"]: body.set_facecolor(colour); body.set_alpha(0.55)
            ax.set_xticks(np.arange(1, 13)); ax.set_xticklabels(_month_labels(range(1, 13))); ax.set_xlabel("Month")
        else:
            parts = ax.violinplot([numeric.dropna().to_numpy()], showmeans=True, showmedians=True)
            for body in parts["bodies"]: body.set_facecolor(colour); body.set_alpha(0.55)
    elif plot_type == "histogram":
        ax.hist(numeric, bins=min(30, max(8, int(math.sqrt(max(1, len(numeric)))))), color=colour, alpha=0.85); ax.set_xlabel(value_label); ax.set_ylabel("Frequency")
    elif plot_type == "exceedance":
        ordered = np.sort(numeric.to_numpy(dtype=float))[::-1]; probability = np.arange(1, len(ordered)+1)/(len(ordered)+1)*100
        ax.plot(probability, ordered, color=colour); ax.set_xlabel("Exceedance Probability (%)")
    elif plot_type == "ecdf":
        ordered = np.sort(numeric.to_numpy(dtype=float)); probability = np.arange(1, len(ordered)+1)/len(ordered)*100
        ax.step(ordered, probability, where="post", color=colour); ax.set_xlabel(value_label); ax.set_ylabel("Cumulative Probability (%)")
    else:
        ax.plot(frame["time"], numeric, color=colour, linewidth=1.4); ax.set_xlabel(time_label)
    if plot_type not in {"histogram", "ecdf"}:
        ax.set_ylabel(value_label)
    ax.set_title(title, fontsize=14, fontweight="bold", pad=13)
    apply_plot_grids(ax)
    _finalize_plot(fig, output)
    return plot_type


_CDE_USER_REQUESTED_PLOT_BASE = generate_plot_product


def generate_plot_product(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    prepared = dict(params)
    plot_type = str(prepared.get("plot_type") or "time_series")
    is_spatial = plot_type.startswith("spatial") or plot_type == "spatial_map"
    if is_spatial:
        prepared["map_render_style"] = "grid"
    resolution = str(prepared.get("resolution") or "")
    token = _AXIS_RESOLUTION_CONTEXT.set(resolution or None)
    try:
        result = dict(_CDE_USER_REQUESTED_PLOT_BASE(prepared, Path(data_dir), Path(export_dir)))
    finally:
        _AXIS_RESOLUTION_CONTEXT.reset(token)
    context = dict(result.get("context") or {})
    context.setdefault("requested_resolution", resolution)
    if is_spatial:
        context.update({
            "map_render_style": "grid",
            "native_grid_preserved": True,
            "spatial_interpolation": "None",
            "spatial_resampling": "None",
        })
    custom_title = _clean_custom_plot_title(prepared.get("custom_plot_title"))
    if result.get("plot_path") and not is_spatial:
        element = str(context.get("variable_label") or variable_display_name(prepared.get("variable"), context, prepared.get("dataset")))
        location = str(prepared.get("location_name") or "Selected Location")
        descriptor = _selection_period_descriptor(prepared, resolution)
        label = dict(PLOT_TYPES).get(plot_type, plot_type.replace("_", " ").title())
        display_title = custom_title or f"{label}: {element} for {location} — {descriptor}"
        if plot_type != "annual_trend":
            _replace_png_title(Path(result["plot_path"]), display_title)
        context["display_title"] = display_title
    context["custom_plot_title"] = custom_title
    result["context"] = context
    return result


_CDE_USER_REQUESTED_INDICES_BASE = generate_indices


def generate_indices(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    prepared = dict(params)
    if str(prepared.get("index_plot_type") or "").lower() == "rank":
        prepared["index_plot_type"] = "auto"
    resolution = str(prepared.get("index_resolution") or "annual").lower()
    token = _AXIS_RESOLUTION_CONTEXT.set(resolution)
    try:
        result = dict(_CDE_USER_REQUESTED_INDICES_BASE(prepared, Path(data_dir), Path(export_dir)))
    finally:
        _AXIS_RESOLUTION_CONTEXT.reset(token)
    context = dict(result.get("context") or {})
    index_type = str(prepared.get("index_type") or "")
    index_label = _INDEX_LABELS.get(index_type, index_type.replace("_", " ").title())
    source_label = str(context.get("dataset_label") or DATASETS.get(str(prepared.get("dataset") or ""), {}).get("label") or "Selected dataset")
    location = str(prepared.get("location_name") or "Selected Location")
    descriptor = _selection_period_descriptor(prepared, resolution)
    season_label = str(context.get("season_label") or "").strip()
    if season_label and season_label.lower() not in descriptor.lower() and season_label.upper() not in {"ANNUAL", "ALL MONTHS"}:
        descriptor = f"{descriptor} · {season_label}"
    custom_title = _clean_custom_plot_title(prepared.get("custom_plot_title"))
    display_title = custom_title or f"{INDEX_RESOLUTION_LABELS.get(resolution, resolution.title())} {index_label} from {source_label} for {location} — {descriptor}"
    if result.get("plot_path"):
        _replace_png_title(Path(result["plot_path"]), display_title)
    context.update({
        "custom_plot_title": custom_title,
        "display_title": display_title,
        "plot_type": str(context.get("plot_type") or prepared.get("index_plot_type") or "auto"),
        "ranked_index_plot_removed": True,
    })
    result["context"] = context
    return result
