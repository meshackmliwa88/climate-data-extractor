#!/usr/bin/env python3
"""Zarr point extraction engine for CHIRPS and ERA5 datasets.

Separate from PostgreSQL. It reads chunked Zarr stores, extracts data by
latitude/longitude, and writes Excel outputs using hourly, daily, monthly,
annual, or seasonal Zarr stores. Non-Zarr data files are not opened.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import posixpath
import zipfile
import xml.etree.ElementTree as ET
from tempfile import TemporaryDirectory
from zoneinfo import ZoneInfo
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import xarray as xr

from cde_store import (
    catalog_pattern_variants,
    default_data_dir,
    glob_store_paths,
    iter_data_stores,
    open_data_stores,
    slice_time_range,
    store_display_name,
    store_kind,
)
from cde_variable_selection import choose_data_variable, select_requested_statistic_dimension
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import qrcode
from PIL import ImageOps


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CATALOG_PATH = PROJECT_ROOT / "config" / "zarr_catalog.json"
DEFAULT_DATA_DIR = default_data_dir(PROJECT_ROOT)
DEFAULT_EXPORT_DIR = PROJECT_ROOT / "storage" / "exports"

MPS_TO_KNOTS = 1.9438444924406

SEASON_DEFINITIONS: Dict[str, List[int]] = {
    "DJF": [12, 1, 2],
    "JFM": [1, 2, 3],
    "FMA": [2, 3, 4],
    "MAM": [3, 4, 5],
    "AMJ": [4, 5, 6],
    "MJJ": [5, 6, 7],
    "JJA": [6, 7, 8],
    "JAS": [7, 8, 9],
    "ASO": [8, 9, 10],
    "SON": [9, 10, 11],
    "OND": [10, 11, 12],
    "NDJ": [11, 12, 1],
    "ONDJFM": [10, 11, 12, 1, 2, 3],
    "NDJFMA": [11, 12, 1, 2, 3, 4],
    "DJFMA": [12, 1, 2, 3, 4],
    "DJFM": [12, 1, 2, 3],
    "JJAS": [6, 7, 8, 9],
    "ANNUAL_ALL_MONTHS": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],
}


@dataclass
class StationPoint:
    station_id: str
    station_name: str
    latitude: float
    longitude: float


def load_catalog(catalog_path: Path | str = DEFAULT_CATALOG_PATH) -> Dict[str, Any]:
    with open(catalog_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _patterns_for_source(cfg: Dict[str, Any]) -> Dict[str, str]:
    if "file_patterns" in cfg:
        return dict(cfg["file_patterns"])
    if "file_pattern" in cfg:
        return {freq: cfg["file_pattern"] for freq in cfg.get("supported_frequencies", [])}
    return {}


def _season_codes_for_catalog_scan() -> List[str]:
    return ["DJF", "MAM", "JJA", "SON", "OND", "NDJ", "DJFMA", "NDJFMA"]


def _expand_catalog_pattern(pattern_name: str, season: str | None = None) -> str:
    pattern_name = str(pattern_name or "")
    if season:
        return pattern_name.replace("{season}", season.upper()).replace("{SEASON}", season.upper())
    return pattern_name


def _candidate_pattern_paths(data_dir: Path, pattern_name: str, frequency: str, season: str | None = None) -> List[Path]:
    """Return Zarr-first candidates for frequency folders and legacy storage."""
    data_dir = Path(data_dir)
    roots = [data_dir]

    candidates: List[Path] = []
    for raw_variant in catalog_pattern_variants(_expand_catalog_pattern(pattern_name, season)):
        expanded_path = Path(raw_variant)
        basename = expanded_path.name
        if expanded_path.is_absolute():
            candidates.append(expanded_path)
            continue
        for root in roots:
            candidates.append(root / expanded_path)
            if frequency:
                candidates.append(root / frequency / basename)
            candidates.append(root / basename)
            candidates.append(root / "**" / basename)

    seen = set()
    out: List[Path] = []
    for candidate in candidates:
        key = str(candidate)
        if key not in seen:
            seen.add(key)
            out.append(candidate)
    return out


def _glob_catalog_files(data_dir: Path | str, pattern_name: str, frequency: str, season: str | None = None) -> List[str]:
    files: List[str] = []
    for pattern_path in _candidate_pattern_paths(Path(data_dir), pattern_name, frequency, season):
        files.extend(glob_store_paths(pattern_path))
    return sorted(set(files))


def list_available_sources(catalog: Dict[str, Any], data_dir: Path | str = DEFAULT_DATA_DIR) -> Dict[str, Any]:
    data_dir = Path(data_dir)
    output: Dict[str, Any] = {}
    for key, cfg in catalog.items():
        file_patterns = _patterns_for_source(cfg)
        available_by_frequency: Dict[str, Dict[str, Any]] = {}
        total_count = 0
        for freq, pattern_name in file_patterns.items():
            if "{season}" in str(pattern_name) or "{SEASON}" in str(pattern_name):
                season_counts: Dict[str, int] = {}
                files_for_freq: List[str] = []
                for season in _season_codes_for_catalog_scan():
                    season_files = _glob_catalog_files(data_dir, pattern_name, freq, season=season)
                    season_counts[season] = len(season_files)
                    files_for_freq.extend(season_files)
                files = sorted(set(files_for_freq))
                available_by_frequency[freq] = {
                    "pattern": pattern_name,
                    "file_count": len(files),
                    "available": len(files) > 0,
                    "season_file_counts": season_counts,
                }
            else:
                files = _glob_catalog_files(data_dir, pattern_name, freq)
                available_by_frequency[freq] = {
                    "pattern": pattern_name,
                    "file_count": len(files),
                    "available": len(files) > 0,
                }
            total_count += len(files)
        supported = cfg.get("supported_frequencies", [])
        output[key] = {
            "label": cfg.get("label", key),
            "description": cfg.get("description", ""),
            "file_patterns": file_patterns,
            "file_pattern": cfg.get("file_pattern", ""),
            "file_count": total_count,
            "available": any(v["available"] for v in available_by_frequency.values()),
            "available_by_frequency": available_by_frequency,
            "supported_frequencies": supported,
            "variables": cfg.get("variables", {}),
        }
    return output


def safe_sheet_name(name: str, used: Optional[set[str]] = None) -> str:
    """Return a concise Excel tab name while retaining resolution and location.

    Dataset source prefixes are useful in metadata but consume scarce space in
    a 31-character worksheet tab. They are removed only from the tab label. If
    a location follows `` - ``, both the temporal-resolution prefix and the
    location suffix are preserved when truncation is required.
    """
    used = used if used is not None else set()
    cleaned = re.sub(r"[\/*?:\[\]]", " ", str(name))
    cleaned = re.sub(r"\s+", " ", cleaned).strip() or "Sheet"
    cleaned = re.sub(r"\b(?:CHIRPS|ERA5(?:-Land)?)\s+", "", cleaned, flags=re.I)
    cleaned = cleaned.replace("Precipitation", "Rainfall")

    if len(cleaned) <= 31:
        base = cleaned
    elif " - " in cleaned:
        prefix, suffix = cleaned.rsplit(" - ", 1)
        if len(prefix) <= 20:
            # Keep a concise resolution/element label intact and shorten only
            # a long location name.
            suffix = re.sub(r"\bLocation\b", "Loc.", suffix, flags=re.I)
            available_suffix = max(5, 31 - len(prefix) - 3)
            base = f"{prefix} - {suffix[:available_suffix].rstrip(' -_,')}"[:31].rstrip()
        else:
            available = max(8, 31 - len(suffix) - 3)
            short_prefix = prefix[:available].rstrip(" -_,")
            word_prefix = short_prefix.rsplit(" ", 1)[0] if " " in short_prefix else short_prefix
            if len(word_prefix) >= max(8, available - 8):
                short_prefix = word_prefix
            base = f"{short_prefix} - {suffix}"[:31].rstrip()
    else:
        base = cleaned[:31].rstrip()

    candidate = base
    n = 2
    lowered_used = {value.lower() for value in used}
    while candidate.lower() in lowered_used:
        suffix = f" {n}"
        candidate = base[: 31 - len(suffix)].rstrip() + suffix
        n += 1
    used.add(candidate)
    return candidate


def find_files_for_source_frequency(cfg: Dict[str, Any], frequency: str, data_dir: Path | str = DEFAULT_DATA_DIR, season: str | None = None) -> List[str]:
    data_dir = Path(data_dir)
    patterns = _patterns_for_source(cfg)
    pattern_name = patterns.get(frequency)
    if not pattern_name:
        raise FileNotFoundError(f"No file pattern configured for {cfg.get('label', '')} at {frequency} resolution.")

    files: List[str]
    if ("{season}" in str(pattern_name) or "{SEASON}" in str(pattern_name)) and not season:
        files = []
        for season_code in _season_codes_for_catalog_scan():
            files.extend(_glob_catalog_files(data_dir, pattern_name, frequency, season=season_code))
        files = sorted(set(files))
    else:
        files = _glob_catalog_files(data_dir, pattern_name, frequency, season=season)

    if not files:
        # Some operational stores contain harmless filename differences such as
        # an extra space, underscore or hyphen. Match the configured basename
        # after removing punctuation, while still restricting the search to the
        # requested temporal-resolution folder.
        expected_names = {
            re.sub(r"[^a-z0-9]+", "", Path(value).name.lower())
            for value in catalog_pattern_variants(_expand_catalog_pattern(pattern_name, season))
        }
        frequency_name = str(frequency or "").strip().lower()
        tolerant = []
        for store in iter_data_stores(data_dir):
            normalized = re.sub(r"[^a-z0-9]+", "", store.name.lower())
            if normalized not in expected_names:
                continue
            if frequency_name and frequency_name not in {part.lower() for part in store.parts}:
                continue
            tolerant.append(str(store))
        files = sorted(set(tolerant))

    if not files:
        example = _candidate_pattern_paths(data_dir, pattern_name, frequency, season=season)[0]
        extra = f" for season {season}" if season else ""
        raise FileNotFoundError(
            f"No Zarr store found for {cfg.get('label', '')} / {frequency}{extra}. "
            f"Expected pattern like: {example}."
        )
    return files


def open_dataset_for_source_frequency(cfg: Dict[str, Any], frequency: str, data_dir: Path | str = DEFAULT_DATA_DIR, season: str | None = None) -> xr.Dataset:
    stores = find_files_for_source_frequency(cfg, frequency, data_dir, season=season)
    return open_data_stores(stores, time_coord=cfg.get("time_coord", "time"), decode_times=True)


def squeeze_and_select_extra_dims(ds: xr.Dataset, cfg: Dict[str, Any]) -> xr.Dataset:
    time_coord = cfg["time_coord"]
    lat_coord = cfg["lat_coord"]
    lon_coord = cfg["lon_coord"]
    keep_dims = {time_coord, lat_coord, lon_coord, "time", "valid_time", "date", "datetime"}

    for dim_name, dim_value in cfg.get("extra_select", {}).items():
        if dim_name in ds.coords or dim_name in ds.dims:
            try:
                ds = ds.sel({dim_name: dim_value}, method="nearest")
            except Exception:
                ds = ds.sel({dim_name: dim_value})

    for dim_name in list(ds.dims):
        if dim_name in keep_dims:
            continue
        size = int(ds.sizes[dim_name])
        if size == 1:
            ds = ds.isel({dim_name: 0}, drop=True)
        elif dim_name in {"time", "number", "expver", "member", "surface"}:
            ds = ds.isel({dim_name: 0}, drop=True)
    return ds


def normalize_longitude_for_dataset(ds: xr.Dataset, lon_coord: str, longitude: float) -> float:
    try:
        lon_min = float(ds[lon_coord].min().values)
        lon_max = float(ds[lon_coord].max().values)
    except Exception:
        return longitude
    if lon_min >= 0 and longitude < 0:
        return longitude % 360
    if lon_max <= 180 and longitude > 180:
        return ((longitude + 180) % 360) - 180
    return longitude


def select_point(da: xr.DataArray, lat_coord: str, lon_coord: str, latitude: float, longitude: float) -> Tuple[xr.DataArray, float, float]:
    lon = normalize_longitude_for_dataset(da.to_dataset(name="tmp"), lon_coord, longitude)
    point = da.sel({lat_coord: latitude, lon_coord: lon}, method="nearest")
    try:
        grid_lat = float(point[lat_coord].values)
        grid_lon = float(point[lon_coord].values)
    except Exception:
        grid_lat = latitude
        grid_lon = lon
    return point, grid_lat, grid_lon


def units_text(da: xr.DataArray | None) -> str:
    if da is None:
        return ""
    attrs = getattr(da, "attrs", {}) or {}
    return str(attrs.get("units", attrs.get("unit", ""))).lower()


def apply_conversion(values: pd.Series, conversion: str, source_units: str = "") -> pd.Series:
    values = pd.to_numeric(values, errors="coerce")
    u = (source_units or "").lower()
    if conversion in {"none", "calculate_from_u10_v10_to_knots", "calculate_from_u10_v10_direction"}:
        return values
    if conversion == "m_to_mm":
        return values * 1000.0
    if conversion == "k_to_c" or conversion == "kelvin_to_celsius":
        return values - 273.15
    if conversion == "auto_kelvin_to_celsius":
        if u in {"k", "kelvin"} or "kelvin" in u:
            return values - 273.15
        return values
    if conversion == "auto_precip_to_mm":
        if any(x in u for x in ["meter", "metre", "m", "kg m**-2"]):
            if "mm" not in u and "mill" not in u:
                return values * 1000.0
        return values
    if conversion == "pa_to_hpa":
        return values / 100.0
    if conversion == "auto_pressure_to_hpa":
        try:
            med = float(pd.to_numeric(values, errors="coerce").dropna().median())
        except Exception:
            med = 0.0
        if ("pa" in u and "hpa" not in u) or med > 2000:
            return values / 100.0
        return values
    if conversion == "fraction_to_percent":
        return values * 100.0
    if conversion == "fraction_to_oktas":
        return values * 8.0
    if conversion == "auto_cloud_to_oktas":
        # ERA5 total cloud cover normally ranges 0-1. If already 0-8, keep. If 0-100, convert.
        smax = pd.to_numeric(values, errors="coerce").max(skipna=True)
        if pd.isna(smax):
            return values
        if smax <= 1.5:
            return values * 8.0
        if smax > 8.5:
            return (values / 100.0) * 8.0
        return values
    if conversion == "mps_to_knots":
        return values * MPS_TO_KNOTS
    if conversion == "auto_wind_speed_to_knots":
        if "knot" in u or "kt" == u.strip():
            return values
        return values * MPS_TO_KNOTS
    return values


def aggregate_series(series: pd.Series, freq: str, agg: str) -> pd.Series:
    series = pd.to_numeric(series, errors="coerce").sort_index()
    if series.empty:
        return series
    if agg == "sum":
        return series.resample(freq).sum(min_count=1)
    if agg == "min":
        return series.resample(freq).min()
    if agg == "max":
        return series.resample(freq).max()
    if agg == "median":
        return series.resample(freq).median()
    return series.resample(freq).mean()


def _select_requested_statistic_dimension(point: xr.DataArray, variable_code: str) -> xr.DataArray:
    """Compatibility wrapper around the shared strict statistic selector."""
    time_like = {"time", "valid_time", "date", "datetime"}
    return select_requested_statistic_dimension(point, variable_code, keep_dims=time_like)


def dataarray_to_series(point: xr.DataArray, time_coord: str, value_name: str = "value", variable_code: str = "") -> pd.Series:
    """Convert a point DataArray to a clean DatetimeIndex series.

    This avoids ``DataFrame.reset_index`` ambiguity when a Zarr store contains
    both an index level and a coordinate column named ``time``.
    """
    point = _select_requested_statistic_dimension(point, variable_code)
    if time_coord not in point.coords and time_coord not in point.dims:
        for candidate in ("time", "valid_time", "date", "datetime"):
            if candidate in point.coords or candidate in point.dims:
                time_coord = candidate
                break
    if time_coord not in point.coords and time_coord not in point.dims:
        raise ValueError(f"Time coordinate '{time_coord}' not found after extraction.")

    time_values = np.asarray(point[time_coord].values).reshape(-1)
    values = np.asarray(point.values)
    # Collapse any residual non-time dimensions deterministically.
    if values.size != time_values.size:
        time_dim = point[time_coord].dims[0] if point[time_coord].dims else None
        if time_dim and time_dim in point.dims:
            axis = point.get_axis_num(time_dim)
            values = np.moveaxis(values, axis, 0).reshape(len(time_values), -1)[:, 0]
        else:
            values = values.reshape(-1)[: len(time_values)]
    else:
        values = values.reshape(-1)

    times = pd.to_datetime(pd.Series(time_values), errors="coerce")
    numeric = pd.to_numeric(pd.Series(values), errors="coerce")
    frame = pd.DataFrame({"time": times.to_numpy(), value_name: numeric.to_numpy()})
    frame = frame.dropna(subset=["time"]).drop_duplicates(subset=["time"], keep="first")
    frame = frame.sort_values("time").reset_index(drop=True)
    return pd.Series(frame[value_name].to_numpy(), index=pd.DatetimeIndex(frame["time"], name="time"), name=value_name)


def meteorological_wind_direction_degrees(u: pd.Series, v: pd.Series) -> pd.Series:
    u = pd.to_numeric(u, errors="coerce")
    v = pd.to_numeric(v, errors="coerce")
    degrees = (270.0 - np.degrees(np.arctan2(v, u))) % 360.0
    return pd.Series(degrees, index=u.index)


def _match_var_name(ds: xr.Dataset, variable_code: str, var_cfg: Dict[str, Any]) -> str | None:
    """Strictly match the requested weather element in operational Zarr stores."""
    try:
        return choose_data_variable(ds, variable_code, var_cfg.get("candidate_names", []))
    except ValueError:
        return None


def _extract_direct_variable(
    ds_time: xr.Dataset,
    cfg: Dict[str, Any],
    var_cfg: Dict[str, Any],
    variable_code: str,
    latitude: float,
    longitude: float,
) -> Dict[str, Any] | None:
    actual_name = _match_var_name(ds_time, variable_code, var_cfg)
    if not actual_name:
        return None
    lat_coord = cfg["lat_coord"]
    lon_coord = cfg["lon_coord"]
    time_coord = cfg["time_coord"]
    da = ds_time[actual_name]
    point, grid_lat, grid_lon = select_point(da, lat_coord, lon_coord, latitude, longitude)
    series = dataarray_to_series(point, time_coord, variable_code=variable_code)
    series = apply_conversion(series, var_cfg.get("conversion", "none"), units_text(da))
    return {
        "series": series,
        "u_series": None,
        "v_series": None,
        "grid_latitude": grid_lat,
        "grid_longitude": grid_lon,
        "actual_variable": actual_name,
    }


def get_selected_series(
    ds: xr.Dataset,
    cfg: Dict[str, Any],
    variable_code: str,
    latitude: float,
    longitude: float,
    start_date: str,
    end_date: str,
) -> Dict[str, Any]:
    time_coord = cfg["time_coord"]
    lat_coord = cfg["lat_coord"]
    lon_coord = cfg["lon_coord"]
    var_cfg = cfg["variables"][variable_code]

    ds = squeeze_and_select_extra_dims(ds, cfg)
    if time_coord not in ds.coords and time_coord not in ds.dims:
        # Some pre-processed files still use time instead of valid_time.
        for candidate in ["time", "valid_time", "date"]:
            if candidate in ds.coords or candidate in ds.dims:
                cfg = dict(cfg)
                cfg["time_coord"] = candidate
                time_coord = candidate
                break
    if time_coord not in ds.coords and time_coord not in ds.dims:
        raise ValueError(f"Time coordinate '{time_coord}' not found. Dataset coordinates: {list(ds.coords)}")

    ds_time, resolved_time = slice_time_range(ds, time_coord, start_date, end_date)
    if resolved_time != time_coord:
        cfg = dict(cfg)
        cfg["time_coord"] = resolved_time
        time_coord = resolved_time

    # Wind files may contain direct wind_speed/wind_direction variables OR u10/v10 components.
    if variable_code in {"wind_speed", "wind_direction"}:
        direct = _extract_direct_variable(ds_time, cfg, var_cfg, variable_code, latitude, longitude)
        if direct is not None:
            return direct

        u_name = cfg.get("wind_u_variable", "u10")
        v_name = cfg.get("wind_v_variable", "v10")
        if u_name in ds_time.data_vars and v_name in ds_time.data_vars:
            u_point, grid_lat, grid_lon = select_point(ds_time[u_name], lat_coord, lon_coord, latitude, longitude)
            v_point, _, _ = select_point(ds_time[v_name], lat_coord, lon_coord, latitude, longitude)
            u_series = dataarray_to_series(u_point, time_coord, "u10", variable_code="u10")
            v_series = dataarray_to_series(v_point, time_coord, "v10", variable_code="v10")
            if variable_code == "wind_speed":
                series = np.sqrt((u_series ** 2) + (v_series ** 2)) * MPS_TO_KNOTS
            else:
                series = meteorological_wind_direction_degrees(u_series, v_series)
            return {
                "series": pd.Series(series, index=u_series.index),
                "u_series": u_series,
                "v_series": v_series,
                "grid_latitude": grid_lat,
                "grid_longitude": grid_lon,
                "actual_variable": variable_code,
            }

    direct = _extract_direct_variable(ds_time, cfg, var_cfg, variable_code, latitude, longitude)
    if direct is not None:
        return direct
    available = ", ".join(list(ds_time.data_vars))
    raise ValueError(f"Variable '{variable_code}' not found. Available variables: {available}")


def build_frequency_series(raw: Dict[str, Any], var_cfg: Dict[str, Any], variable_code: str, frequency: str) -> pd.Series:
    # For hourly/daily/monthly/annual we prefer the precomputed file values exactly as selected.
    series: pd.Series = raw["series"]
    if frequency in {"hourly", "daily", "monthly", "annual"}:
        return series
    raise ValueError(f"Unsupported frequency: {frequency}")


def parse_custom_season(season_code: str) -> Tuple[str, List[int]]:
    if season_code.startswith("CUSTOM:"):
        raw = season_code.split(":", 1)[1]
        # Accept formats such as: 1,2,3 or 1 2 3 or 1, 2 3
        parts = [p for p in re.split(r"[\s,;|/]+", raw.strip()) if p]
        months: List[int] = []
        for part in parts:
            if not part.isdigit():
                raise ValueError("Custom season months must be numbers from 1 to 12, for example 1,2,3.")
            month = int(part)
            if month < 1 or month > 12:
                raise ValueError("Custom season months must be between 1 and 12.")
            if month not in months:
                months.append(month)
        if not months:
            raise ValueError("Custom season must contain at least one month.")
        return "CUSTOM", months
    if season_code not in SEASON_DEFINITIONS:
        raise ValueError(f"Unknown season: {season_code}")
    return season_code, SEASON_DEFINITIONS[season_code]


def season_year_for_timestamp(ts: pd.Timestamp, months: List[int]) -> int:
    month = ts.month
    wraps = months[-1] < months[0]
    if wraps and month < months[0]:
        return ts.year - 1
    return ts.year


def aggregate_seasonal_series(raw: Dict[str, Any], var_cfg: Dict[str, Any], variable_code: str, season_code: str) -> pd.Series:
    _, months = parse_custom_season(season_code)
    period_agg = var_cfg.get("period_aggregation", "mean")
    series = raw["series"].sort_index()
    series = pd.to_numeric(series, errors="coerce")
    series = series[series.index.month.isin(months)]
    if series.empty:
        return series

    # If u/v are available, produce vector mean direction for seasonal wind direction.
    if variable_code == "wind_direction" and raw.get("u_series") is not None:
        u = raw["u_series"].sort_index()
        v = raw["v_series"].sort_index()
        mask = u.index.month.isin(months)
        df = pd.DataFrame({"u": u[mask], "v": v[mask]})
        df["season_year"] = [season_year_for_timestamp(pd.Timestamp(i), months) for i in df.index]
        grouped = df.groupby("season_year")[["u", "v"]].mean()
        out = meteorological_wind_direction_degrees(grouped["u"], grouped["v"])
        out.index = pd.to_datetime([f"{int(y)}-01-01" for y in out.index])
        return out

    df = series.to_frame("value")
    df["season_year"] = [season_year_for_timestamp(pd.Timestamp(i), months) for i in df.index]
    if period_agg == "sum":
        out = df.groupby("season_year")["value"].sum(min_count=1)
    elif period_agg == "min":
        out = df.groupby("season_year")["value"].min()
    elif period_agg == "max":
        out = df.groupby("season_year")["value"].max()
    elif period_agg == "median":
        out = df.groupby("season_year")["value"].median()
    else:
        out = df.groupby("season_year")["value"].mean()
    out.index = pd.to_datetime([f"{int(y)}-01-01" for y in out.index])
    return out


def decimal_places_for_variable(variable_code: str, value_label: str = "") -> int:
    text = f"{variable_code} {value_label}".lower()
    if any(k in text for k in ["relative humidity", "wind speed", "wind direction", "wind_speed", "wind_direction"]) or variable_code in {"r"}:
        return 0
    return 1


def make_dataframe_for_excel(series: pd.Series, frequency: str, value_label: str, variable_code: str, season_code: str | None = None) -> pd.DataFrame:
    """Create clean Excel data tables.

    Monthly output is arranged horizontally: Year | Jan | Feb | ... | Dec.
    Annual and seasonal outputs do not include a Date column.
    """
    decimals = decimal_places_for_variable(variable_code, value_label)
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    df = series.reset_index()
    if df.empty:
        if frequency == "hourly":
            columns = ["Date", "Year", "Month", "Day", "Hour", value_label]
        elif frequency == "daily":
            columns = ["Date", "Year", "Month", "Day", value_label]
        elif frequency == "monthly":
            columns = ["Year"] + month_names
        elif frequency == "annual":
            columns = ["Year", value_label]
        elif frequency == "seasonal":
            columns = ["Season Year", "Season", "Months Used", value_label]
        else:
            columns = [value_label]
        return pd.DataFrame(columns=columns)

    time_col = df.columns[0]
    df[time_col] = pd.to_datetime(df[time_col])
    value_col = df.columns[-1]
    df[value_col] = pd.to_numeric(df[value_col], errors="coerce").round(decimals)

    if frequency == "hourly":
        out = pd.DataFrame({
            "Date/Time": df[time_col].dt.strftime("%Y-%m-%d %H:%M"),
            value_label: df[value_col],
        })
    elif frequency == "daily":
        out = pd.DataFrame({
            "Date": df[time_col].dt.strftime("%Y-%m-%d"),
            value_label: df[value_col],
        })
    elif frequency == "monthly":
        tmp = pd.DataFrame({
            "Year": df[time_col].dt.year,
            "Month": df[time_col].dt.month,
            "Value": df[value_col],
        })
        out = tmp.pivot_table(index="Year", columns="Month", values="Value", aggfunc="first").reset_index()
        out.columns = ["Year"] + [month_names[int(c) - 1] for c in out.columns[1:]]
        for m in month_names:
            if m not in out.columns:
                out[m] = np.nan
        out = out[["Year"] + month_names].sort_values("Year").reset_index(drop=True)
    elif frequency == "annual":
        out = pd.DataFrame({
            "Year": df[time_col].dt.year,
            value_label: df[value_col],
        })
    elif frequency == "seasonal":
        _, months = parse_custom_season(season_code or "MAM")
        out = pd.DataFrame({
            "Season Year": df[time_col].dt.year,
            "Season": season_code or "Season",
            "Months Used": ",".join(map(str, months)),
            value_label: df[value_col],
        })
    else:
        out = pd.DataFrame({"Date": df[time_col].dt.strftime("%Y-%m-%d"), value_label: df[value_col]})
    return out


def _number_format_for_sheet(sheet_title: str) -> str | None:
    title = sheet_title.lower()
    if "wind speed" in title or "wind direction" in title or "relative humidity" in title or "cloud cover" in title:
        return "0"
    if any(k in title for k in ["rainfall", "precipitation", "temperature", "pressure"]):
        return "0.0"
    return None


def data_type_label(freq: str, season_code: str | None = None) -> str:
    if freq == "hourly":
        return "Hourly Time Series Data"
    if freq == "daily":
        return "Daily Time Series Data"
    if freq == "monthly":
        return "Monthly Data"
    if freq == "annual":
        return "Annual Data"
    if freq == "seasonal":
        return f"Seasonal {season_code} Data" if season_code else "Seasonal Data"
    return f"{freq.title()} Data"


def default_download_context(output_path: Path | str) -> Dict[str, Any]:
    """Fallback context used when the extractor is run from the command line."""
    now = datetime_now_eat()
    return {
        "institution": "Climate Data Extractor",
        "system": "Climate Data Extractor",
        "download_id": f"CDE-{now.strftime('%Y%m%d')}-CLI",
        "downloaded_by": "Command Line User",
        "user_station": "",
        "downloaded_at": now.strftime("%Y-%m-%d %H:%M EAT"),
        "file_name": Path(output_path).name,
        "verification_url": "",
        "data_url": "",
    }


def datetime_now_eat():
    return pd.Timestamp.now(tz=ZoneInfo("Africa/Dar_es_Salaam")).to_pydatetime()


def build_qr_payload(
    *,
    context: Dict[str, Any],
    station: StationPoint,
    var_label: str,
    unit: str,
    frequency: str,
    start_date: str,
    end_date: str,
    output_file: str,
    season_code: str | None = None,
    source_label: str | None = None,
) -> Dict[str, Any]:
    download_id = str(context.get("download_id") or "")
    downloaded_at = str(context.get("downloaded_at") or "")
    report_date = str(context.get("date") or "").strip()
    if not report_date:
        try:
            clean_downloaded_at = re.sub(r"\s+EAT$", "", downloaded_at).strip()
            report_date = pd.Timestamp(clean_downloaded_at).strftime("%d %B, %Y") if clean_downloaded_at else datetime_now_eat().strftime("%d %B, %Y")
        except Exception:
            report_date = datetime_now_eat().strftime("%d %B, %Y")
    payload = {
        "institution": context.get("institution") or context.get("system") or "Climate Data Extractor",
        "document_type": "Data Delivery Report",
        "system": context.get("system") or context.get("institution") or "Climate Data Extractor",
        "reference_no": context.get("reference_no") or "CD533/620/01",
        "request_no": context.get("request_no") or download_id.replace("CDE-", "")[:12],
        "download_id": download_id,
        "station_name": station.station_name,
        "latitude": station.latitude,
        "longitude": station.longitude,
        "element": var_label,
        "data_type": data_type_label(frequency, season_code),
        "start_date": start_date,
        "end_date": end_date,
        "period": f"{start_date} to {end_date}",
        "mode_of_delivery": context.get("mode_of_delivery") or "Electronic copy",
        "units": unit,
        "customer_name": context.get("customer_name") or "",
        "customer_organization": context.get("customer_organization") or "",
        "customer_phone": context.get("customer_phone") or "",
        "customer_email": context.get("customer_email") or "",
        "customer_address": context.get("customer_address") or "",
        "served_by": context.get("served_by") or context.get("downloaded_by") or "",
        "downloaded_by": context.get("served_by") or context.get("downloaded_by") or "",
        "issued_by": context.get("issued_by") or "",
        "downloaded_at": context.get("downloaded_at") or "",
        "date": report_date,
        "file_name": output_file,
        "verification_url": context.get("verification_url") or "",
        "data_url": context.get("data_url") or "",
    }
    if source_label:
        payload["source"] = source_label
    if season_code:
        payload["season"] = season_code
    return payload


def qr_payload_to_plain_text(payload: Dict[str, Any]) -> str:
    """Encode the same human-readable details displayed in the source document.

    The text is kept to one line per field (no blank spacer lines) so the QR
    modules remain large enough to scan reliably in Excel and PDF outputs.
    """
    def pick(*keys):
        for key in keys:
            value = payload.get(key)
            if value not in (None, ""):
                return value
        return ""

    document_type = str(pick("document_type", "Document Type") or "Data Delivery Report")
    institution = str(pick("institution", "system") or "Climate Data Extractor")
    lines = [institution, f"Document Type: {document_type}"]

    if document_type == "Proforma Invoice":
        rows = [
            ("Ref. No.", pick("reference_no")),
            ("Date", pick("date")),
            ("Customer / Organization", pick("customer_name")),
            ("Customer Address", pick("customer_address")),
            ("Customer Category", pick("customer_category")),
            ("Data Type", pick("data_type", "temporal_resolution")),
            ("Stations", pick("stations")),
            ("Parameters", pick("parameters")),
            ("Period", pick("period") or (f"{pick('years')} year(s)" if pick("years") != "" else "")),
            ("Description of Requested Service", pick("description")),
            ("Total cost recovery fee", pick("total_fee")),
            ("Payment instruction", pick("payment_instruction") or "For payment, please request a control number from TMA."),
            ("Attended by", pick("served_by", "prepared_by", "issued_by")),
        ]
    elif document_type == "Proposed Cost-Recovery Proforma":
        rows = [
            ("Ref. No.", pick("reference_no")),
            ("Date", pick("date")),
            ("Customer / Organization", pick("customer_name")),
            ("Customer Address", pick("customer_address")),
            ("Customer Category", pick("customer_category")),
            ("Temporal Resolution", pick("temporal_resolution", "data_type")),
            ("Stations x Parameters x Years", pick("selection_basis")),
            ("Current Formula Group", pick("current_formula_group")),
            ("CURRENT FORMULA FEE", pick("current_formula_fee")),
            ("CDE cost-recovery adjustment", pick("recovery_adjustment")),
            ("Additional professional services", pick("additional_service_fee")),
            ("FINAL PROPOSED FEE", pick("total_fee")),
            ("INCREASE FROM CURRENT", pick("difference_from_current")),
            ("Description of data/service", pick("description")),
            ("Issued by", pick("issued_by", "prepared_by", "served_by")),
            ("Notice", pick("notice")),
        ]
    else:
        element = pick("element", "weather_element")
        data_type = pick("data_type")
        parameter_text = " - ".join(str(v) for v in (element, data_type) if v not in (None, ""))
        period = pick("period")
        if not period:
            start_date = pick("start_date")
            end_date = pick("end_date")
            period = f"{start_date} to {end_date}".strip(" to")
        rows = [
            ("Ref. No.", pick("reference_no") or "CD533/620/01"),
            ("Date", pick("date", "downloaded_at")),
            ("Request No. (yymmno)", pick("request_no") or str(pick("download_id")).replace("CDE-", "")[:12]),
            ("Customer Name", pick("customer_name", "customer_organization")),
            ("Customer Address", pick("customer_address")),
            ("Phone number", pick("customer_phone")),
            ("Email Address", pick("customer_email")),
            ("Parameter(s) provided", parameter_text),
            ("Station(s) provided", pick("station_name")),
            ("Period", period),
            ("Mode of Delivery", pick("mode_of_delivery") or "Electronic copy"),
            ("Attended by", pick("served_by", "issued_by", "downloaded_by")),
            ("File Name", pick("file_name")),
            ("Verification URL", pick("verification_url")),
        ]

    for label, value in rows:
        if value not in (None, ""):
            lines.append(f"{label}: {value}")
    return "\n".join(lines)

def _paste_logo_on_qr(qr_img: Any, logo_path: Path | str | None = None) -> Any:
    """Return a plain QR image; public builds do not embed organization logos."""
    return qr_img.convert("RGB")


def qr_payload_to_compact_text(payload: Dict[str, Any]) -> str:
    """Return a compact verification payload suitable for easy QR scanning.

    This helper is retained only for legacy compact-link uses. Current Excel
    outputs use the full document-detail payload generated above.
    """
    for key in ("verification_url", "data_url", "receipt_url"):
        value = str(payload.get(key) or "").strip()
        if value:
            return value
    request_no = str(payload.get("download_id") or payload.get("request_no") or "").strip()
    file_name = str(payload.get("file_name") or "").strip()
    lines = ["Climate Data Extractor"]
    if request_no:
        lines.append(f"Request: {request_no}")
    if file_name:
        lines.append(f"File: {file_name}")
    return "\n".join(lines)


def make_qr_png(
    payload: Dict[str, Any],
    output_png: Path | str,
    logo_path: Path | str | None = None,
    *,
    compact: bool = False,
) -> None:
    qr_text = qr_payload_to_compact_text(payload) if compact else qr_payload_to_plain_text(payload)
    qr = qrcode.QRCode(
        version=None,
        # Use high error correction for reliable scanning from PDFs, screenshots and office printers.
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=(10 if compact else 12),
        border=(4 if compact else 6),
        mask_pattern=(None if compact else 4),
    )
    qr.add_data(qr_text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    img = _paste_logo_on_qr(img, logo_path=logo_path)
    img.save(output_png, optimize=True)


_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CONTENT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"


def _xlsx_resolve_target(source_part: str, target: str) -> str:
    target = str(target or "")
    if target.startswith("/"):
        return posixpath.normpath(target.lstrip("/"))
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def _next_relationship_id(root: ET.Element) -> str:
    used = {node.get("Id", "") for node in root}
    number = 1
    while f"rId{number}" in used:
        number += 1
    return f"rId{number}"


def _insert_sheet_child_in_order(root: ET.Element, child: ET.Element, before_tags: tuple[str, ...]) -> None:
    before = {f"{{{_MAIN_NS}}}{tag}" for tag in before_tags}
    for index, existing in enumerate(list(root)):
        if existing.tag in before:
            root.insert(index, child)
            return
    root.append(child)


def ensure_excel_qr_on_every_page(
    path: Path | str,
    sheet_payloads: Dict[str, Dict[str, Any]] | None = None,
    context: Dict[str, Any] | None = None,
) -> None:
    """Remove frozen panes and place a large, compact QR in every printed page header.

    Header/footer images are stored as VML parts in XLSX. This package-level
    implementation works for workbooks produced by XlsxWriter, pandas or
    openpyxl and prevents a later library save from silently dropping the QR.
    """
    xlsx_path = Path(path)
    if not xlsx_path.exists() or xlsx_path.suffix.lower() != ".xlsx":
        return
    payloads = sheet_payloads or {}
    base_context = dict(context or {})
    base_context.setdefault("file_name", xlsx_path.name)

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        parts = {name: zin.read(name) for name in zin.namelist()}

    workbook_xml = ET.fromstring(parts["xl/workbook.xml"])
    workbook_rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    rel_targets = {node.get("Id"): node.get("Target") for node in workbook_rels}
    sheet_parts: list[tuple[str, str]] = []
    for sheet in workbook_xml.findall(f".//{{{_MAIN_NS}}}sheet"):
        name = sheet.get("name") or "Sheet"
        rid = sheet.get(f"{{{_DOC_REL_NS}}}id")
        target = rel_targets.get(rid or "")
        if target:
            sheet_parts.append((name, _xlsx_resolve_target("xl/workbook.xml", target)))

    # Ensure the package recognises PNG and VML parts.
    content_types = ET.fromstring(parts["[Content_Types].xml"])
    defaults = {node.get("Extension") for node in content_types.findall(f"{{{_CONTENT_NS}}}Default")}
    if "png" not in defaults:
        ET.SubElement(content_types, f"{{{_CONTENT_NS}}}Default", Extension="png", ContentType="image/png")
    if "vml" not in defaults:
        ET.SubElement(content_types, f"{{{_CONTENT_NS}}}Default", Extension="vml", ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing")
    parts["[Content_Types].xml"] = ET.tostring(content_types, encoding="utf-8", xml_declaration=True)

    with TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        for index, (sheet_name, sheet_part) in enumerate(sheet_parts, start=1):
            if sheet_part not in parts:
                continue
            payload = dict(base_context)
            payload.update(payloads.get(sheet_name, {}))
            payload.setdefault("file_name", xlsx_path.name)
            qr_file = tmpdir_path / f"qr_header_{index}.png"
            make_qr_png(payload, qr_file, compact=True)
            image_part = f"xl/media/cde_qr_header_{index}.png"
            parts[image_part] = qr_file.read_bytes()

            sheet_root = ET.fromstring(parts[sheet_part])
            # Remove all frozen/split panes and pane-specific selections.
            for view in sheet_root.findall(f".//{{{_MAIN_NS}}}sheetView"):
                for pane in list(view.findall(f"{{{_MAIN_NS}}}pane")):
                    view.remove(pane)
                for selection in view.findall(f"{{{_MAIN_NS}}}selection"):
                    selection.attrib.pop("pane", None)

            header_footer = sheet_root.find(f"{{{_MAIN_NS}}}headerFooter")
            if header_footer is None:
                header_footer = ET.Element(f"{{{_MAIN_NS}}}headerFooter")
                _insert_sheet_child_in_order(sheet_root, header_footer, ("rowBreaks", "colBreaks", "customProperties", "cellWatches", "ignoredErrors", "smartTags", "drawing", "legacyDrawing", "legacyDrawingHF", "picture", "oleObjects", "controls", "webPublishItems", "tableParts", "extLst"))
            header_footer.set("differentOddEven", "0")
            header_footer.set("differentFirst", "0")
            odd_header = header_footer.find(f"{{{_MAIN_NS}}}oddHeader")
            if odd_header is None:
                odd_header = ET.SubElement(header_footer, f"{{{_MAIN_NS}}}oddHeader")
            odd_header.text = "&R&G"

            page_margins = sheet_root.find(f"{{{_MAIN_NS}}}pageMargins")
            if page_margins is None:
                page_margins = ET.Element(f"{{{_MAIN_NS}}}pageMargins", left="0.25", right="0.25", top="1.65", bottom="0.45", header="0.15", footer="0.2")
                _insert_sheet_child_in_order(sheet_root, page_margins, ("pageSetup", "headerFooter", "rowBreaks", "colBreaks", "customProperties", "cellWatches", "ignoredErrors", "smartTags", "drawing", "legacyDrawing", "legacyDrawingHF", "picture", "oleObjects", "controls", "webPublishItems", "tableParts", "extLst"))
            else:
                try:
                    page_margins.set("top", str(max(float(page_margins.get("top", "0")), 1.65)))
                except Exception:
                    page_margins.set("top", "1.65")
                page_margins.set("header", "0.15")

            # Replace any prior header image reference with a fresh compact QR.
            for old in list(sheet_root.findall(f"{{{_MAIN_NS}}}legacyDrawingHF")):
                sheet_root.remove(old)

            rels_part = posixpath.join(posixpath.dirname(sheet_part), "_rels", posixpath.basename(sheet_part) + ".rels")
            if rels_part in parts:
                rels_root = ET.fromstring(parts[rels_part])
            else:
                rels_root = ET.Element(f"{{{_PKG_REL_NS}}}Relationships")
            header_rid = _next_relationship_id(rels_root)
            vml_part = f"xl/drawings/vmlDrawingCdeQr{index}.vml"
            vml_target = posixpath.relpath(vml_part, posixpath.dirname(sheet_part))
            ET.SubElement(
                rels_root,
                f"{{{_PKG_REL_NS}}}Relationship",
                Id=header_rid,
                Type=f"{_DOC_REL_NS}/vmlDrawing",
                Target=vml_target,
            )
            legacy = ET.Element(f"{{{_MAIN_NS}}}legacyDrawingHF")
            legacy.set(f"{{{_DOC_REL_NS}}}id", header_rid)
            _insert_sheet_child_in_order(sheet_root, legacy, ("picture", "oleObjects", "controls", "webPublishItems", "tableParts", "extLst"))
            parts[rels_part] = ET.tostring(rels_root, encoding="utf-8", xml_declaration=True)

            # About 37 mm square on the printed page: large enough for common
            # phone cameras while remaining clear of the data table.
            vml = (
                '<xml xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:x="urn:schemas-microsoft-com:office:excel">'
                '<o:shapelayout v:ext="edit"><o:idmap v:ext="edit" data="1"/></o:shapelayout>'
                '<v:shapetype id="_x0000_t75" coordsize="21600,21600" o:spt="75" o:preferrelative="t" path="m@4@5l@4@11@9@11@9@5xe" filled="f" stroked="f">'
                '<v:stroke joinstyle="miter"/><v:formulas><v:f eqn="if lineDrawn pixelLineWidth 0"/><v:f eqn="sum @0 1 0"/><v:f eqn="sum 0 0 @1"/><v:f eqn="prod @2 1 2"/><v:f eqn="prod @3 21600 pixelWidth"/><v:f eqn="prod @3 21600 pixelHeight"/><v:f eqn="sum @0 0 1"/><v:f eqn="prod @6 1 2"/><v:f eqn="prod @7 21600 pixelWidth"/><v:f eqn="sum @8 21600 0"/><v:f eqn="prod @7 21600 pixelHeight"/><v:f eqn="sum @10 21600 0"/></v:formulas>'
                '<v:path o:extrusionok="f" gradientshapeok="t" o:connecttype="rect"/><o:lock v:ext="edit" aspectratio="t"/></v:shapetype>'
                '<v:shape id="RH" o:spid="_x0000_s1025" type="#_x0000_t75" style="position:absolute;margin-left:0;margin-top:0;width:105pt;height:105pt;z-index:1">'
                '<v:imagedata o:relid="rId1" o:title="CDE verification QR"/><o:lock v:ext="edit" rotation="t"/></v:shape></xml>'
            )
            parts[vml_part] = vml.encode("utf-8")
            vml_rels_part = posixpath.join(posixpath.dirname(vml_part), "_rels", posixpath.basename(vml_part) + ".rels")
            vml_rels = ET.Element(f"{{{_PKG_REL_NS}}}Relationships")
            ET.SubElement(
                vml_rels,
                f"{{{_PKG_REL_NS}}}Relationship",
                Id="rId1",
                Type=f"{_DOC_REL_NS}/image",
                Target=posixpath.relpath(image_part, posixpath.dirname(vml_part)),
            )
            parts[vml_rels_part] = ET.tostring(vml_rels, encoding="utf-8", xml_declaration=True)
            parts[sheet_part] = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)

    temp_path = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
    with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for name, data in parts.items():
            zout.writestr(name, data)
    temp_path.replace(xlsx_path)


def _unique_sheet_name(wb, base_name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", str(base_name)).strip() or "Sheet"
    cleaned = cleaned[:31]
    if cleaned not in wb.sheetnames:
        return cleaned
    n = 2
    while True:
        suffix = f" {n}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        if candidate not in wb.sheetnames:
            return candidate
        n += 1


def _strip_print_header_qr(path: Path | str) -> None:
    """Remove legacy printed-page QR header parts from an existing workbook."""
    xlsx_path = Path(path)
    if not xlsx_path.exists() or xlsx_path.suffix.lower() != ".xlsx":
        return
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zin:
            parts = {name: zin.read(name) for name in zin.namelist()}
        for name in list(parts):
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                root = ET.fromstring(parts[name])
                for node in list(root.findall(f"{{{_MAIN_NS}}}legacyDrawingHF")):
                    root.remove(node)
                header_footer = root.find(f"{{{_MAIN_NS}}}headerFooter")
                if header_footer is not None:
                    for tag in ("oddHeader", "evenHeader", "firstHeader"):
                        node = header_footer.find(f"{{{_MAIN_NS}}}{tag}")
                        if node is not None and node.text:
                            node.text = node.text.replace("&G", "")
                parts[name] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            elif "vmlDrawingCdeQr" in name or "cde_qr_header_" in name:
                del parts[name]
        temp = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
        with zipfile.ZipFile(temp, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
            for name, data in parts.items():
                zout.writestr(name, data)
        temp.replace(xlsx_path)
    except Exception:
        # A visible in-sheet QR is still present even if an old optional print
        # header could not be removed from an unusual workbook package.
        return


def _worksheet_data_last_column(ws) -> int:
    markers = {"CDE QR Verification", "QR Verification", "Scan to verify"}
    marker_columns = {
        cell.column
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20))
        for cell in row
        if str(cell.value or "").strip() in markers
    }
    last = 1
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.column in marker_columns:
                continue
            if cell.value not in (None, ""):
                last = max(last, cell.column)
    return last


def add_qr_codes_to_workbook(path: Path | str, sheet_payloads: Dict[str, Dict[str, Any]], context: Dict[str, Any]) -> None:
    """Add one clearly visible verification QR inside every worksheet.

    The QR is kept in the sheet itself only. It is not repeated in printed-page
    headers, and no separate verification or summary sheet is created.
    """
    wb = load_workbook(path)
    if "Verification" in wb.sheetnames:
        del wb["Verification"]

    with TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        for n, ws in enumerate(wb.worksheets, start=1):
            payload = dict(context or {})
            payload.update(sheet_payloads.get(ws.title, {}))
            payload.setdefault("file_name", Path(path).name)
            qr_path = tmpdir_path / f"sheet_qr_{n}.png"
            make_qr_png(payload, qr_path, compact=False)

            ws.freeze_panes = None
            ws._images = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
                for cell in row:
                    if str(cell.value or "").strip() in {"CDE QR Verification", "QR Verification", "Scan to verify"}:
                        cell.value = None
                        cell.fill = PatternFill(fill_type=None)
                        cell.font = Font()

            data_last_col = _worksheet_data_last_column(ws)
            qr_col = data_last_col + 2
            # Never place the QR label inside an existing merged title row.
            # This can happen when a workbook contains a wide metadata table
            # followed by a narrower data table.
            for merged in ws.merged_cells.ranges:
                if merged.min_row <= 1 <= merged.max_row and merged.min_col <= qr_col <= merged.max_col:
                    qr_col = max(qr_col, merged.max_col + 2)
            label = ws.cell(row=1, column=qr_col, value="Scan to verify")
            label.font = Font(bold=True, size=11)
            label.alignment = Alignment(horizontal="center", vertical="center")
            img = XLImage(str(qr_path))
            img.width = 480
            img.height = 480
            ws.add_image(img, f"{get_column_letter(qr_col)}2")
            ws.column_dimensions[get_column_letter(qr_col)].width = 10
            ws.column_dimensions[get_column_letter(qr_col + 1)].width = 10

            # Keep ordinary page setup; no image in a printed header/footer.
            for section in (ws.oddHeader, ws.evenHeader, ws.firstHeader):
                section.left.text = None
                section.center.text = None
                section.right.text = None
            ws.page_margins.top = 0.5
            ws.page_margins.header = 0.2

        wb.save(path)

    _strip_print_header_qr(path)


def _compact_excel_column_width(ws, col_idx: int, header_row: int | None, data_last_col: int) -> float:
    """Return a compact width based on table content only.

    Metadata and titles are deliberately ignored because they are merged across
    the complete data table and must never force every monthly column to become
    excessively wide.
    """
    header = str(ws.cell(header_row, col_idx).value or "").strip() if header_row else ""
    key = header.lower().replace("_", " ")

    if key in {"year", "month", "day", "hour", "season year"}:
        return 9
    if header in {"Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"}:
        return 9
    if key in {"date", "observation date"}:
        return 12
    if key in {"date/time", "date / time", "date time", "datetime", "date_time"}:
        return 17
    if "latitude" in key or "longitude" in key or key in {"lat", "lon"}:
        return 12
    if key == "unit":
        return 8
    if key == "season":
        return 11
    if "temporal resolution" in key or key == "resolution":
        return 14
    if "location" in key or "station" in key:
        return 18

    first_row = header_row or 1
    max_len = len(header)
    numeric_only = True
    checked = 0
    for row_number in range(first_row + (1 if header_row else 0), min(ws.max_row, first_row + 3000) + 1):
        value = ws.cell(row_number, col_idx).value
        if value in (None, ""):
            continue
        checked += 1
        max_len = max(max_len, len(str(value)))
        if not isinstance(value, (int, float, np.integer, np.floating)) or isinstance(value, bool):
            numeric_only = False
        if checked >= 3000:
            break

    if numeric_only and checked:
        return min(max(max_len + 1, 9), 12)
    return min(max(max_len + 1, 9), 18)


def _merge_excel_metadata_rows(ws, header_row: int | None, data_last_col: int) -> None:
    """Merge each title/metadata line across the complete data table width."""
    if not header_row or header_row <= 1 or data_last_col <= 1:
        return

    # Preserve text from any existing merged cells before normalising the area.
    existing = list(ws.merged_cells.ranges)
    for merged in existing:
        if merged.min_row < header_row and merged.max_row < header_row:
            ws.unmerge_cells(str(merged))

    # The row immediately above the table may be blank. Merge only simple
    # title/metadata rows containing one or two values. Rows with three or more
    # populated cells are genuine table headers/data (for example the analysis
    # metadata table) and must remain untouched.
    for row_number in range(1, header_row):
        values = []
        for col in range(1, data_last_col + 1):
            value = ws.cell(row_number, col).value
            if value not in (None, ""):
                values.append(str(value).strip())
        if not values or len(values) > 2:
            continue

        if len(values) == 1:
            combined = values[0]
        else:
            label = values[0].rstrip().rstrip(":")
            combined = f"{label} : {values[1]}"

        for col in range(1, data_last_col + 1):
            ws.cell(row_number, col).value = None
        ws.cell(row_number, 1).value = combined
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=data_last_col)
        ws.cell(row_number, 1).alignment = Alignment(vertical="center", wrap_text=True)
        if row_number == 1:
            ws.cell(row_number, 1).font = Font(name="Calibri", size=12, bold=True, color="000000")
        else:
            ws.cell(row_number, 1).font = Font(name="Calibri", size=11, bold=False, color="000000")
        # Modest row height keeps long metadata visible without making the
        # otherwise simple sheet look like a report dashboard.
        ws.row_dimensions[row_number].height = 20


def style_excel(path: Path | str) -> None:
    """Apply a simple, compact Excel layout to every worksheet."""
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="E7E6E6")
    no_fill = PatternFill(fill_type=None)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_tokens = {"Date/Time", "Date / Time", "Date", "Year", "Season Year", "date_time"}

    for ws in wb.worksheets:
        ws.freeze_panes = None
        ws.auto_filter.ref = None
        ws.sheet_view.showGridLines = True
        ws.sheet_properties.tabColor = None

        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100)):
            values = {str(cell.value or "").strip() for cell in row}
            if values.intersection(header_tokens):
                header_row = row[0].row
                break

        data_last_col = _worksheet_data_last_column(ws)
        _merge_excel_metadata_rows(ws, header_row, data_last_col)

        # Remove dashboard-like fills/fonts while preserving simple metadata,
        # table headers and values.
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=data_last_col):
            for cell in row:
                if cell.coordinate in ws.merged_cells:
                    continue
                cell.fill = no_fill
                cell.font = Font(name="Calibri", size=11, bold=False, color="000000")
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.border = Border()

        if ws.cell(1, 1).value not in (None, ""):
            ws.cell(1, 1).font = Font(name="Calibri", size=12, bold=True)
            ws.cell(1, 1).alignment = Alignment(vertical="center", wrap_text=True)

        if header_row:
            for col in range(1, data_last_col + 1):
                cell = ws.cell(header_row, col)
                cell.fill = header_fill
                cell.font = Font(name="Calibri", size=11, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            for row_number in range(header_row + 1, ws.max_row + 1):
                for col in range(1, data_last_col + 1):
                    cell = ws.cell(row_number, col)
                    cell.border = border
                    header = str(ws.cell(header_row, col).value or "").strip().lower()
                    if isinstance(cell.value, (int, float, np.integer, np.floating)) and not isinstance(cell.value, bool):
                        if header in {"year", "month", "day", "hour", "season year"} or any(token in header for token in ("count", "records", "observations")):
                            cell.number_format = "0"
                        elif "latitude" in header or "longitude" in header or header in {"lat", "lon"}:
                            cell.number_format = "0.0000"
                        elif any(token in header for token in ("wind speed", "wind direction", "relative humidity")):
                            cell.number_format = "0"
                        else:
                            cell.number_format = "0.0"
            # Plain headers only: no dropdown arrows.
            ws.auto_filter.ref = None

        for col_idx in range(1, data_last_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = _compact_excel_column_width(
                ws, col_idx, header_row, data_last_col
            )

        ws.print_title_rows = None
        ws.oddFooter.left.text = None
        ws.oddFooter.center.text = None
        ws.oddFooter.right.text = None
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.4

    wb.save(path)
    _strip_print_header_qr(path)



def build_single_station(station_id: str, station_name: str, lat: float, lon: float) -> List[StationPoint]:
    return [StationPoint(station_id=station_id, station_name=station_name, latitude=lat, longitude=lon)]


def frequency_label(freq: str) -> str:
    return {"hourly": "Hourly", "daily": "Daily", "monthly": "Monthly", "annual": "Annual", "seasonal": "Seasonal"}.get(freq, freq.title())


def metadata_rows(title: str, station: StationPoint, source_label: str, var_label: str, unit: str, raw: Dict[str, Any], start_date: str, end_date: str, aggregation: str, extra: list[list[Any]] | None = None) -> list[list[Any]]:
    """Metadata rows for Excel exports.

    The user-facing layout is one merged row per metadata item, written as:
    ``Label : Value``. This avoids separated label/value columns and keeps long
    text wrapped within the width of the exported data table.
    """
    rows = [
        [title],
        [f"Location : {station.station_name}"],
        [f"Dataset : {source_label}"],
        [f"Weather Element : {var_label}"],
        [f"Unit : {unit}"],
        [f"Period : {start_date} to {end_date}"],
    ]
    if extra:
        for item in extra:
            if len(item) >= 2:
                rows.append([f"{item[0]} : {item[1]}"])
            elif item:
                rows.append([str(item[0])])
    return rows



def make_combined_wind_dataframe(speed: pd.Series, direction: pd.Series, frequency: str, season_code: str | None = None) -> pd.DataFrame:
    """Put wind speed and direction in one sheet with consecutive columns."""
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    # Use arrays rather than indexed Series so ``time`` cannot simultaneously
    # become an index level and a column label.
    count = min(len(speed), len(direction))
    df = pd.DataFrame({
        "time": pd.to_datetime(np.asarray(speed.index)[:count], errors="coerce"),
        "Wind Speed": pd.to_numeric(np.asarray(speed)[:count], errors="coerce").round(0),
        "Wind Direction": pd.to_numeric(np.asarray(direction)[:count], errors="coerce").round(0),
    }).reset_index(drop=True)
    df = df.dropna(how="all", subset=["Wind Speed", "Wind Direction"]).sort_values("time").reset_index(drop=True)
    if frequency == "hourly":
        return pd.DataFrame({"Date/Time": df["time"].dt.strftime("%Y-%m-%d %H:%M"), "Wind Speed": df["Wind Speed"], "Wind Direction": df["Wind Direction"]})
    if frequency == "daily":
        return pd.DataFrame({"Date": df["time"].dt.strftime("%Y-%m-%d"), "Wind Speed": df["Wind Speed"], "Wind Direction": df["Wind Direction"]})
    if frequency == "monthly":
        tmp = df.copy(); tmp["Year"] = tmp["time"].dt.year; tmp["Month"] = tmp["time"].dt.month
        rows = []
        for y, g in tmp.groupby("Year"):
            row = {"Year": int(y)}
            for m in range(1, 13):
                gm = g[g["Month"] == m]
                row[f"{month_names[m-1]} Wind Speed"] = gm["Wind Speed"].iloc[0] if not gm.empty else np.nan
                row[f"{month_names[m-1]} Wind Direction"] = gm["Wind Direction"].iloc[0] if not gm.empty else np.nan
            rows.append(row)
        cols = ["Year"] + [f"{mon} {name}" for mon in month_names for name in ["Wind Speed", "Wind Direction"]]
        return pd.DataFrame(rows, columns=cols)
    if frequency == "annual":
        return pd.DataFrame({"Year": df["time"].dt.year, "Wind Speed": df["Wind Speed"], "Wind Direction": df["Wind Direction"]})
    if frequency == "seasonal":
        _, months = parse_custom_season(season_code or "MAM")
        return pd.DataFrame({"Season Year": df["time"].dt.year, "Season": season_code or "Season", "Months Used": ",".join(map(str, months)), "Wind Speed": df["Wind Speed"], "Wind Direction": df["Wind Direction"]})
    return pd.DataFrame({"Date": df["time"].dt.strftime("%Y-%m-%d"), "Wind Speed": df["Wind Speed"], "Wind Direction": df["Wind Direction"]})

def _write_dataframe_to_workbook(
    writer: pd.ExcelWriter,
    frame: pd.DataFrame,
    *,
    sheet_name: str,
    startrow: int = 0,
    header: bool = True,
    index: bool = False,
) -> None:
    """Write a dataframe correctly in XlsxWriter constant-memory mode."""
    if writer.engine != "xlsxwriter":
        frame.to_excel(
            writer, sheet_name=sheet_name, index=index, header=header, startrow=startrow
        )
        return

    data = frame.reset_index() if index else frame
    ws = writer.sheets.get(sheet_name)
    if ws is None:
        ws = writer.book.add_worksheet(sheet_name)
        writer.sheets[sheet_name] = ws

    row_cursor = int(startrow)
    if header:
        ws.write_row(row_cursor, 0, [str(column) for column in data.columns])
        row_cursor += 1

    def clean(value):
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        if isinstance(value, np.generic):
            return value.item()
        return value

    for values in data.itertuples(index=False, name=None):
        ws.write_row(row_cursor, 0, [clean(value) for value in values])
        row_cursor += 1


def write_excel_output(
    output_path: Path | str,
    source_key: str,
    source_cfg: Dict[str, Any],
    stations: List[StationPoint],
    variables: List[str],
    start_date: str,
    end_date: str,
    frequencies: List[str],
    seasons: List[str] | None = None,
    data_dir: Path | str = DEFAULT_DATA_DIR,
    download_context: Dict[str, Any] | None = None,
    append: bool = False,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    seasons = seasons or []
    download_context = download_context or default_download_context(output_path)
    download_context.setdefault("file_name", output_path.name)
    sheet_qr_payloads: Dict[str, Dict[str, Any]] = {}

    supported = set(source_cfg.get("supported_frequencies", []))
    unsupported = sorted(set(frequencies) - supported)
    if unsupported:
        raise ValueError(f"Selected frequency is not supported for {source_cfg.get('label', source_key)}: {', '.join(unsupported)}")
    if "seasonal" in frequencies and not seasons:
        seasons = ["DJF", "MAM", "JJA", "SON", "OND", "NDJ", "DJFMA", "NDJFMA"]

    used_sheet_names: set[str] = set()
    if append and output_path.exists():
        existing_wb = load_workbook(output_path, read_only=True)
        used_sheet_names.update(existing_wb.sheetnames)
        existing_wb.close()
    ds_cache: Dict[str, xr.Dataset] = {}

    try:
        fast_new_workbook = not (append and output_path.exists())
        if fast_new_workbook:
            writer_options: Dict[str, Any] = {
                "engine": "xlsxwriter",
                "engine_kwargs": {
                    "options": {
                        "constant_memory": True,
                        "strings_to_urls": False,
                        "nan_inf_to_errors": True,
                    }
                },
            }
        else:
            writer_options = {"engine": "openpyxl", "mode": "a", "if_sheet_exists": "overlay"}
        with TemporaryDirectory() as fast_qr_dir:
            with pd.ExcelWriter(output_path, **writer_options) as writer:
                sheet_count = 0
                if not fast_new_workbook:
                    # Existing workbooks use openpyxl append mode.  No temporary
                    # information sheet is needed for new streamed workbooks.
                    pass
                for freq in frequencies:
                    if freq != "seasonal":
                        if freq not in ds_cache:
                            ds_cache[freq] = open_dataset_for_source_frequency(source_cfg, freq, data_dir)
                        ds = ds_cache[freq]

                    for station in stations:
                        # If wind speed and direction are both selected, export them together on one sheet.
                        if source_key == "ERA5_WIND" and {"wind_speed", "wind_direction"}.issubset(set(variables)):
                            if freq == "seasonal":
                                for season_code in seasons:
                                    seasonal_cache_key = f"seasonal:{season_code}"
                                    monthly_cache_key = "monthly:fallback_for_custom_seasons"
                                    try:
                                        if not season_code.startswith("CUSTOM:"):
                                            if seasonal_cache_key not in ds_cache:
                                                ds_cache[seasonal_cache_key] = open_dataset_for_source_frequency(source_cfg, "seasonal", data_dir, season=season_code)
                                            ds_season = ds_cache[seasonal_cache_key]
                                            raw_speed = get_selected_series(ds_season, source_cfg, "wind_speed", station.latitude, station.longitude, start_date, end_date)
                                            raw_dir = get_selected_series(ds_season, source_cfg, "wind_direction", station.latitude, station.longitude, start_date, end_date)
                                            speed_series = raw_speed["series"].sort_index()
                                            dir_series = raw_dir["series"].sort_index()
                                            aggregation_label = "from precomputed seasonal file"
                                        else:
                                            raise FileNotFoundError("Custom season requires monthly fallback.")
                                    except Exception:
                                        if monthly_cache_key not in ds_cache:
                                            ds_cache[monthly_cache_key] = open_dataset_for_source_frequency(source_cfg, "monthly", data_dir)
                                        ds_monthly = ds_cache[monthly_cache_key]
                                        raw_speed = get_selected_series(ds_monthly, source_cfg, "wind_speed", station.latitude, station.longitude, start_date, end_date)
                                        raw_dir = get_selected_series(ds_monthly, source_cfg, "wind_direction", station.latitude, station.longitude, start_date, end_date)
                                        speed_series = aggregate_seasonal_series(raw_speed, source_cfg["variables"]["wind_speed"], "wind_speed", season_code)
                                        dir_series = aggregate_seasonal_series(raw_dir, source_cfg["variables"]["wind_direction"], "wind_direction", season_code)
                                        aggregation_label = "mean / vector mean direction"
                                    df = make_combined_wind_dataframe(speed_series, dir_series, freq, season_code)
                                    title = f"Seasonal {season_code} Wind Speed and Direction - {station.station_name}"
                                    sheet_name = safe_sheet_name(title, used_sheet_names)
                                    _, months = parse_custom_season(season_code)
                                    meta = metadata_rows(title, station, source_cfg.get("label", source_key), "Wind Speed and Direction", "knots / degrees", raw_speed, start_date, end_date, aggregation_label, [["Season Months", ",".join(map(str, months))]])
                                    _write_dataframe_to_workbook(writer, pd.DataFrame(meta), sheet_name=sheet_name, index=False, header=False, startrow=0)
                                    _write_dataframe_to_workbook(writer, df, sheet_name=sheet_name, index=False, startrow=len(meta) + 1)
                                    sheet_qr_payloads[sheet_name] = build_qr_payload(context=download_context, station=station, var_label="Wind Speed and Direction", unit="knots / degrees", frequency=freq, season_code=season_code, start_date=start_date, end_date=end_date, output_file=output_path.name, source_label=source_cfg.get("label", source_key))
                                    sheet_count += 1
                                continue
                            raw_speed = get_selected_series(ds, source_cfg, "wind_speed", station.latitude, station.longitude, start_date, end_date)
                            raw_dir = get_selected_series(ds, source_cfg, "wind_direction", station.latitude, station.longitude, start_date, end_date)
                            speed_series = build_frequency_series(raw_speed, source_cfg["variables"]["wind_speed"], "wind_speed", freq)
                            dir_series = build_frequency_series(raw_dir, source_cfg["variables"]["wind_direction"], "wind_direction", freq)
                            df = make_combined_wind_dataframe(speed_series, dir_series, freq)
                            title = f"{frequency_label(freq)} Wind Speed and Direction - {station.station_name}"
                            sheet_name = safe_sheet_name(title, used_sheet_names)
                            meta = metadata_rows(title, station, source_cfg.get("label", source_key), "Wind Speed and Direction", "knots / degrees", raw_speed, start_date, end_date, "from selected file")
                            _write_dataframe_to_workbook(writer, pd.DataFrame(meta), sheet_name=sheet_name, index=False, header=False, startrow=0)
                            _write_dataframe_to_workbook(writer, df, sheet_name=sheet_name, index=False, startrow=len(meta) + 1)
                            sheet_qr_payloads[sheet_name] = build_qr_payload(context=download_context, station=station, var_label="Wind Speed and Direction", unit="knots / degrees", frequency=freq, start_date=start_date, end_date=end_date, output_file=output_path.name, source_label=source_cfg.get("label", source_key))
                            sheet_count += 1
                            continue

                        for variable_code in variables:
                            var_cfg = source_cfg["variables"][variable_code]
                            var_label = var_cfg.get("label", variable_code)
                            unit = var_cfg.get("unit", "")
                            source_label = source_cfg.get("label", source_key)

                            if freq == "seasonal":
                                for season_code in seasons:
                                    seasonal_cache_key = f"seasonal:{season_code}"
                                    monthly_cache_key = "monthly:fallback_for_custom_seasons"
                                    try:
                                        # Preferred: use the precomputed seasonal stores under storage/zarr/seasonal.
                                        if not season_code.startswith("CUSTOM:"):
                                            if seasonal_cache_key not in ds_cache:
                                                ds_cache[seasonal_cache_key] = open_dataset_for_source_frequency(source_cfg, "seasonal", data_dir, season=season_code)
                                            ds_season = ds_cache[seasonal_cache_key]
                                            raw = get_selected_series(ds_season, source_cfg, variable_code, station.latitude, station.longitude, start_date, end_date)
                                            series = raw["series"].sort_index()
                                            aggregation_label = "from precomputed seasonal file"
                                        else:
                                            raise FileNotFoundError("Custom season requires monthly fallback.")
                                    except Exception:
                                        # Backward-compatible fallback: derive the season from monthly data.
                                        if monthly_cache_key not in ds_cache:
                                            ds_cache[monthly_cache_key] = open_dataset_for_source_frequency(source_cfg, "monthly", data_dir)
                                        ds_monthly = ds_cache[monthly_cache_key]
                                        raw = get_selected_series(ds_monthly, source_cfg, variable_code, station.latitude, station.longitude, start_date, end_date)
                                        series = aggregate_seasonal_series(raw, var_cfg, variable_code, season_code)
                                        aggregation_label = var_cfg.get("period_aggregation", "mean")

                                    df = make_dataframe_for_excel(series, freq, var_label, variable_code, season_code)
                                    title = f"Seasonal {season_code} {var_label} - {station.station_name}"
                                    sheet_name = safe_sheet_name(title, used_sheet_names)
                                    _, months = parse_custom_season(season_code)
                                    meta = metadata_rows(title, station, source_label, var_label, unit, raw, start_date, end_date, aggregation_label, [["Season Months", ",".join(map(str, months))]])
                                    _write_dataframe_to_workbook(writer, pd.DataFrame(meta), sheet_name=sheet_name, index=False, header=False, startrow=0)
                                    _write_dataframe_to_workbook(writer, df, sheet_name=sheet_name, index=False, startrow=len(meta) + 1)
                                    sheet_qr_payloads[sheet_name] = build_qr_payload(
                                        context=download_context, station=station, var_label=var_label, unit=unit,
                                        frequency=freq, season_code=season_code, start_date=start_date, end_date=end_date,
                                        output_file=output_path.name, source_label=source_label,
                                    )
                                    sheet_count += 1
                                continue

                            raw = get_selected_series(ds, source_cfg, variable_code, station.latitude, station.longitude, start_date, end_date)
                            series = build_frequency_series(raw, var_cfg, variable_code, freq)
                            df = make_dataframe_for_excel(series, freq, var_label, variable_code)
                            title = f"{frequency_label(freq)} {var_label} - {station.station_name}"
                            sheet_name = safe_sheet_name(title, used_sheet_names)
                            agg_label = "from selected file" if freq in {"hourly", "daily", "monthly", "annual"} else var_cfg.get("period_aggregation", "mean")
                            meta = metadata_rows(title, station, source_label, var_label, unit, raw, start_date, end_date, agg_label)
                            _write_dataframe_to_workbook(writer, pd.DataFrame(meta), sheet_name=sheet_name, index=False, header=False, startrow=0)
                            _write_dataframe_to_workbook(writer, df, sheet_name=sheet_name, index=False, startrow=len(meta) + 1)
                            sheet_qr_payloads[sheet_name] = build_qr_payload(
                                context=download_context, station=station, var_label=var_label, unit=unit,
                                frequency=freq, start_date=start_date, end_date=end_date,
                                output_file=output_path.name, source_label=source_label,
                            )
                            sheet_count += 1

                if fast_new_workbook:
                    workbook = writer.book
                    title_fmt = workbook.add_format({
                        "bold": True, "align": "center", "valign": "vcenter",
                    })
                    text_fmt = workbook.add_format({"valign": "top"})
                    one_fmt = workbook.add_format({"num_format": "0.0"})
                    zero_fmt = workbook.add_format({"num_format": "0"})
                    for number, (sheet_name, payload) in enumerate(sheet_qr_payloads.items(), start=1):
                        ws = writer.sheets.get(sheet_name)
                        if ws is None:
                            continue
                        qr_path = Path(fast_qr_dir) / f"qr_{number}.png"
                        make_qr_png(payload, qr_path, compact=False)
                        qr_col = 30
                        ws.write(0, qr_col, "Scan to verify", title_fmt)
                        ws.insert_image(1, qr_col, str(qr_path), {
                            "x_scale": 0.34, "y_scale": 0.34, "object_position": 1
                        })
                        ws.set_column(qr_col, qr_col + 1, 10)
                        number_format = _number_format_for_sheet(sheet_name)
                        ws.set_column(0, 25, 12, zero_fmt if number_format == "0" else one_fmt)
                elif sheet_count > 0 and "Generation Info" in writer.book.sheetnames:
                    writer.book.remove(writer.book["Generation Info"])
        if not fast_new_workbook:
            style_excel(output_path)
            add_qr_codes_to_workbook(output_path, sheet_qr_payloads, download_context)
    finally:
        for ds in ds_cache.values():
            try:
                ds.close()
            except Exception:
                pass
    return output_path



# ---------------------------------------------------------------------------
# Multi-sheet Data Extractor workbook wrapper.
# The original extractor reads only the requested Zarr stores. Each distinct
# location, variable, temporal resolution and season remains on a separate,
# simple worksheet tab with an in-sheet verification QR.
# ---------------------------------------------------------------------------
_CDE_MULTI_SHEET_WRITE_EXCEL_OUTPUT = write_excel_output


def _copy_generated_workbook_into_one_sheet(
    source_path: Path,
    output_path: Path,
    *,
    download_context: Dict[str, Any],
    append: bool,
) -> None:
    """Stream existing and newly generated sections into one bordered sheet."""
    input_paths = []
    if append and output_path.exists():
        input_paths.append((output_path, True))
    input_paths.append((source_path, False))

    def visible_max_col(ws) -> int:
        first = next(ws.iter_rows(min_row=1, max_row=1), ())
        qr_limit = ws.max_column
        for cell in first:
            if str(cell.value or "").strip() in {"CDE QR Verification", "QR Verification"}:
                qr_limit = max(cell.column - 1, 1)
                break
        last_used = 1
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), max_col=qr_limit):
            for cell in row:
                if cell.value not in (None, ""):
                    last_used = max(last_used, cell.column)
        return last_used

    max_cols = 2
    for path, existing in input_paths:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = [wb["Data"]] if existing and "Data" in wb.sheetnames else wb.worksheets
        for ws in sheets:
            max_cols = max(max_cols, visible_max_col(ws))
        wb.close()

    temporary_output = output_path.with_name(output_path.stem + ".writing.xlsx")
    if temporary_output.exists():
        temporary_output.unlink()
    options = {"constant_memory": True, "strings_to_urls": False, "nan_inf_to_errors": True}
    workbook = __import__("xlsxwriter").Workbook(str(temporary_output), options)
    ws_out = workbook.add_worksheet("Data")
    ws_out.hide_gridlines(2)

    title_fmt = workbook.add_format({"bold": True, "font_color": "white", "bg_color": "#073B5C", "border": 1, "valign": "vcenter"})
    section_fmt = workbook.add_format({"bold": True, "font_color": "white", "bg_color": "#0B8F6F", "border": 1, "valign": "vcenter"})
    header_fmt = workbook.add_format({"bold": True, "font_color": "white", "bg_color": "#0B5E7A", "border": 1, "align": "center", "valign": "vcenter", "text_wrap": True})
    meta_fmt = workbook.add_format({"bold": True, "bg_color": "#F7FBFF", "border": 1, "valign": "top", "text_wrap": True})
    text_fmt = workbook.add_format({"border": 1, "valign": "top", "text_wrap": True})
    one_fmt = workbook.add_format({"border": 1, "num_format": "0.0", "valign": "top"})
    zero_fmt = workbook.add_format({"border": 1, "num_format": "0", "valign": "top"})
    coord_fmt = workbook.add_format({"border": 1, "num_format": "0.0000", "valign": "top"})
    date_fmt = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd hh:mm", "valign": "top"})
    blank_fmt = workbook.add_format({"border": 0})

    widths = [12] * max_cols
    row_cursor = 0
    ws_out.merge_range(row_cursor, 0, row_cursor, max_cols - 1, "TMA Climate Data Extraction", title_fmt)
    row_cursor += 2

    def clean(value):
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        if isinstance(value, np.generic):
            return value.item()
        return value

    for path, existing in input_paths:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = [wb["Data"]] if existing and "Data" in wb.sheetnames else wb.worksheets
        for src_ws in sheets:
            data_max_col = visible_max_col(src_ws)
            if existing:
                # Existing output already contains section headings. Copy its body,
                # but skip its workbook title and QR columns.
                source_rows = src_ws.iter_rows(min_row=3, max_col=data_max_col)
            else:
                first_row = next(src_ws.iter_rows(min_row=1, max_row=1, max_col=data_max_col), ())
                section_title = str(first_row[0].value or src_ws.title) if first_row else src_ws.title
                ws_out.merge_range(row_cursor, 0, row_cursor, max(data_max_col - 1, 1), section_title, section_fmt)
                row_cursor += 1
                source_rows = src_ws.iter_rows(min_row=2, max_col=data_max_col)

            active_headers: Dict[int, str] = {}
            before_header = True
            for source_row in source_rows:
                values = [clean(cell.value) for cell in source_row]
                if all(v in (None, "") for v in values):
                    row_cursor += 1
                    continue
                strings = [str(v or "") for v in values]
                is_header = any(v in strings for v in ("Date/Time", "Date", "Year", "Season Year", "date_time"))
                # Preserve styled section/header rows from a previously combined file.
                fill = ""
                try:
                    fill = str(source_row[0].fill.fgColor.rgb or source_row[0].fill.fgColor.indexed or "")
                except Exception:
                    fill = ""
                if "0B8F6F" in fill:
                    row_kind = "section"
                elif "0B5E7A" in fill:
                    row_kind = "header"
                    is_header = True
                elif "073B5C" in fill:
                    row_kind = "section"
                elif is_header:
                    row_kind = "header"
                elif before_header:
                    row_kind = "meta"
                else:
                    row_kind = "data"

                if row_kind == "header":
                    active_headers = {idx: str(v or "") for idx, v in enumerate(values)}
                    before_header = False
                elif row_kind == "section":
                    active_headers = {}
                    before_header = True

                for col, value in enumerate(values):
                    if row_kind == "section":
                        fmt = section_fmt
                    elif row_kind == "header":
                        fmt = header_fmt
                    elif row_kind == "meta":
                        fmt = meta_fmt
                    elif isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
                        h = active_headers.get(col, "").strip().lower()
                        if h in {"year", "month", "day", "hour", "season year"} or any(token in h for token in ("year", "month", "day", "hour", "count", "records", "observations")):
                            value = int(round(float(value)))
                            fmt = zero_fmt
                        elif "latitude" in h or "longitude" in h or h in {"lat", "lon"}:
                            fmt = coord_fmt
                        elif any(token in h for token in ("relative humidity", "wind speed", "wind direction")):
                            value = int(round(float(value)))
                            fmt = zero_fmt
                        else:
                            value = round(float(value), 1)
                            fmt = one_fmt
                    elif hasattr(value, "year") and hasattr(value, "month") and not isinstance(value, str):
                        fmt = date_fmt
                    else:
                        fmt = text_fmt
                    ws_out.write(row_cursor, col, value, fmt)
                    widths[col] = min(max(widths[col], len(str(value or "")) + 1), 24)
                row_cursor += 1
            row_cursor += 2
        wb.close()

    for col, width in enumerate(widths):
        ws_out.set_column(col, col, width)

    with TemporaryDirectory() as qr_tmp:
        qr_path = Path(qr_tmp) / "verification.png"
        payload = dict(download_context or {})
        payload.setdefault("file_name", output_path.name)
        make_qr_png(payload, qr_path, compact=False)
        qr_col = max_cols + 1
        qr_title_fmt = workbook.add_format({"bold": True, "font_color": "white", "bg_color": "#073B5C", "border": 1, "align": "center"})
        ws_out.merge_range(0, qr_col, 0, qr_col + 1, "CDE QR Verification", qr_title_fmt)
        ws_out.insert_image(1, qr_col, str(qr_path), {"x_scale": 0.30, "y_scale": 0.30, "object_position": 1})
        ws_out.set_column(qr_col, qr_col + 1, 10)
        ws_out.set_margins(left=0.3, right=0.3, top=0.5, bottom=0.4)
        workbook.close()

    temporary_output.replace(output_path)


def write_excel_output(
    output_path: Path | str,
    source_key: str,
    source_cfg: Dict[str, Any],
    stations: List[StationPoint],
    variables: List[str],
    start_date: str,
    end_date: str,
    frequencies: List[str],
    seasons: List[str] | None = None,
    data_dir: Path | str = DEFAULT_DATA_DIR,
    download_context: Dict[str, Any] | None = None,
    append: bool = False,
) -> Path:
    """Extract records into separate simple worksheet tabs.

    Every station × weather element × temporal resolution × season is kept on
    its own worksheet. Appending another saved request adds new worksheets
    rather than stacking unlike temporal resolutions in one long table.
    """
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    context = download_context or default_download_context(output)
    context.setdefault("file_name", output.name)
    result = _CDE_MULTI_SHEET_WRITE_EXCEL_OUTPUT(
        output_path=output,
        source_key=source_key,
        source_cfg=source_cfg,
        stations=stations,
        variables=variables,
        start_date=start_date,
        end_date=end_date,
        frequencies=frequencies,
        seasons=seasons,
        data_dir=data_dir,
        download_context=context,
        append=append,
    )
    # The streamed first request and any appended requests receive the same
    # simple formatting, filters and no frozen panes.
    style_excel(result)
    names_wb = load_workbook(result, read_only=True)
    final_payloads = {name: dict(context) for name in names_wb.sheetnames}
    names_wb.close()
    add_qr_codes_to_workbook(result, final_payloads, context)
    return Path(result)



_CDE_PRIMARY_WRITE_EXCEL_OUTPUT = write_excel_output
_CDE_PRIMARY_OPEN_DATASET = open_dataset_for_source_frequency


def _open_dataset_eager(cfg: Dict[str, Any], frequency: str, data_dir: Path | str = DEFAULT_DATA_DIR, season: str | None = None) -> xr.Dataset:
    stores = find_files_for_source_frequency(cfg, frequency, data_dir, season=season)
    return open_data_stores(stores, time_coord=cfg.get("time_coord", "time"), chunks="eager", decode_times=True)


def _compatibility_single_sheet_writer(
    output_path: Path | str,
    source_key: str,
    source_cfg: Dict[str, Any],
    stations: List[StationPoint],
    variables: List[str],
    start_date: str,
    end_date: str,
    frequencies: List[str],
    seasons: List[str] | None = None,
    data_dir: Path | str = DEFAULT_DATA_DIR,
    download_context: Dict[str, Any] | None = None,
) -> Path:
    """Eager compatibility writer that still preserves separate sheet tabs.

    This fallback is used only for operational Zarr stores whose chunk metadata
    cannot be opened lazily. It mirrors the normal workbook arrangement: one
    worksheet per temporal resolution/selection, with metadata and QR details.
    """
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    seasons = list(seasons or [])
    context = dict(download_context or default_download_context(output))
    context.setdefault("file_name", output.name)
    sections: list[tuple[str, pd.DataFrame, pd.DataFrame, Dict[str, Any]]] = []
    datasets: Dict[str, xr.Dataset] = {}
    try:
        for frequency in frequencies:
            for station in stations:
                for variable_code in variables:
                    var_cfg = source_cfg["variables"][variable_code]
                    var_label = var_cfg.get("label", variable_code)
                    unit = var_cfg.get("unit", "")
                    source_label = source_cfg.get("label", source_key)
                    seasonal_codes = seasons if frequency == "seasonal" else [None]
                    if frequency == "seasonal" and not seasonal_codes:
                        seasonal_codes = ["MAM"]
                    for season_code in seasonal_codes:
                        if frequency == "seasonal":
                            try:
                                if str(season_code).startswith("CUSTOM:"):
                                    raise FileNotFoundError
                                cache_key = f"seasonal:{season_code}"
                                if cache_key not in datasets:
                                    datasets[cache_key] = _open_dataset_eager(source_cfg, "seasonal", data_dir, season=season_code)
                                raw = get_selected_series(datasets[cache_key], source_cfg, variable_code, station.latitude, station.longitude, start_date, end_date)
                                series = raw["series"].sort_index()
                                aggregation_label = "from precomputed seasonal file"
                            except Exception:
                                cache_key = "monthly:seasonal-fallback"
                                if cache_key not in datasets:
                                    datasets[cache_key] = _open_dataset_eager(source_cfg, "monthly", data_dir)
                                raw = get_selected_series(datasets[cache_key], source_cfg, variable_code, station.latitude, station.longitude, start_date, end_date)
                                series = aggregate_seasonal_series(raw, var_cfg, variable_code, str(season_code or "MAM"))
                                aggregation_label = var_cfg.get("period_aggregation", "mean")
                        else:
                            cache_key = frequency
                            if cache_key not in datasets:
                                datasets[cache_key] = _open_dataset_eager(source_cfg, frequency, data_dir)
                            raw = get_selected_series(datasets[cache_key], source_cfg, variable_code, station.latitude, station.longitude, start_date, end_date)
                            series = build_frequency_series(raw, var_cfg, variable_code, frequency)
                            aggregation_label = "from selected file"
                        table = make_dataframe_for_excel(series, frequency, var_label, variable_code, season_code)
                        title = f"{frequency_label(frequency)}{' ' + str(season_code) if season_code else ''} {var_label} - {station.station_name}"
                        extra = []
                        if season_code:
                            _, months = parse_custom_season(str(season_code))
                            extra = [["Season Months", ",".join(map(str, months))]]
                        meta = pd.DataFrame(metadata_rows(title, station, source_label, var_label, unit, raw, start_date, end_date, aggregation_label, extra))
                        qr = build_qr_payload(
                            context=context, station=station, var_label=var_label, unit=unit,
                            frequency=frequency, season_code=season_code, start_date=start_date,
                            end_date=end_date, output_file=output.name, source_label=source_label,
                        )
                        sections.append((title, meta, table, qr))
        if not sections:
            raise ValueError("No data sections were generated for the selected request.")

        used_names: set[str] = set()
        sheet_payloads: Dict[str, Dict[str, Any]] = {}
        with pd.ExcelWriter(
            output,
            engine="xlsxwriter",
            engine_kwargs={"options": {"constant_memory": True, "strings_to_urls": False, "nan_inf_to_errors": True}},
        ) as writer:
            for title, meta, table, qr in sections:
                sheet_name = safe_sheet_name(title, used_names)
                _write_dataframe_to_workbook(writer, meta, sheet_name=sheet_name, index=False, header=False, startrow=0)
                _write_dataframe_to_workbook(writer, table, sheet_name=sheet_name, index=False, startrow=len(meta) + 1)
                sheet_payloads[sheet_name] = qr

        style_excel(output)
        add_qr_codes_to_workbook(output, sheet_payloads, context)
        return output
    finally:
        for dataset in datasets.values():
            try:
                dataset.close()
            except Exception:
                pass


def write_excel_output(
    output_path: Path | str,
    source_key: str,
    source_cfg: Dict[str, Any],
    stations: List[StationPoint],
    variables: List[str],
    start_date: str,
    end_date: str,
    frequencies: List[str],
    seasons: List[str] | None = None,
    data_dir: Path | str = DEFAULT_DATA_DIR,
    download_context: Dict[str, Any] | None = None,
    append: bool = False,
) -> Path:
    """Generate Excel with lazy reads, eager compatibility retry and safe fallback."""
    try:
        return _CDE_PRIMARY_WRITE_EXCEL_OUTPUT(
            output_path, source_key, source_cfg, stations, variables, start_date,
            end_date, frequencies, seasons, data_dir, download_context, append,
        )
    except TypeError as first_error:
        if "cannot be interpreted as an integer" not in str(first_error):
            raise
        # Retry with Dask bypassed. Some older operational Zarr stores encode
        # chunk metadata in a way that is accepted eagerly but rejected lazily.
        globals()["open_dataset_for_source_frequency"] = _open_dataset_eager
        try:
            return _CDE_PRIMARY_WRITE_EXCEL_OUTPUT(
                output_path, source_key, source_cfg, stations, variables,
                start_date, end_date, frequencies, seasons, data_dir,
                download_context, append,
            )
        except Exception:
            if append:
                raise first_error
            return _compatibility_single_sheet_writer(
                output_path, source_key, source_cfg, stations, variables,
                start_date, end_date, frequencies, seasons, data_dir,
                download_context,
            )
        finally:
            globals()["open_dataset_for_source_frequency"] = _CDE_PRIMARY_OPEN_DATASET


def parse_list(value: str) -> List[str]:
    return [x.strip() for x in value.split(",") if x.strip()]


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract CHIRPS/ERA5 Zarr data to Excel")
    parser.add_argument("--source", required=True, help="Source key from zarr_catalog.json (legacy filename is also supported)")
    parser.add_argument("--variables", required=True, help="Comma-separated variables")
    parser.add_argument("--lat", type=float, required=True, help="Latitude")
    parser.add_argument("--lon", type=float, required=True, help="Longitude")
    parser.add_argument("--location-name", default="Custom Location", help="Location name")
    parser.add_argument("--start", required=True, help="Start date YYYY-MM-DD")
    parser.add_argument("--end", required=True, help="End date YYYY-MM-DD")
    parser.add_argument("--frequencies", default="daily,monthly", help="Comma-separated: hourly,daily,monthly,annual,seasonal")
    parser.add_argument("--seasons", default="DJF,MAM,JJA,SON,OND,NDJ,DJFMA,NDJFMA", help="Comma-separated seasons, e.g. DJF,MAM,JJA,SON,OND,NDJ,DJFMA,NDJFMA")
    parser.add_argument("--output", required=True, help="Output Excel path")
    parser.add_argument("--catalog", default=str(DEFAULT_CATALOG_PATH))
    parser.add_argument("--data-dir", default=str(DEFAULT_DATA_DIR))
    parser.add_argument("--download-id", default="", help="Optional QR download ID, e.g. CDE-20260701-000001")
    parser.add_argument("--downloaded-by", default="Command Line User", help="Name to include in the QR payload")
    parser.add_argument("--user-station", default="", help="User station/office to include in the QR payload")
    parser.add_argument("--verification-url", default="", help="Verification URL to encode in the QR payload")
    parser.add_argument("--data-url", default="", help="Local hosted data URL to encode in the QR payload")
    args = parser.parse_args()

    catalog = load_catalog(args.catalog)
    if args.source not in catalog:
        raise ValueError(f"Source '{args.source}' not found in catalog.")
    now = datetime_now_eat()
    output_path = Path(args.output)
    download_id = args.download_id or f"CDE-{now.strftime('%Y%m%d')}-CLI"
    write_excel_output(
        output_path=output_path,
        source_key=args.source,
        source_cfg=catalog[args.source],
        stations=build_single_station("POINT_1", args.location_name, args.lat, args.lon),
        variables=parse_list(args.variables),
        start_date=args.start,
        end_date=args.end,
        frequencies=parse_list(args.frequencies),
        seasons=parse_list(args.seasons),
        data_dir=args.data_dir,
        download_context={
            "system": "Climate Data Extractor",
            "download_id": download_id,
            "downloaded_by": args.downloaded_by,
            "user_station": args.user_station,
            "downloaded_at": now.strftime("%Y-%m-%d %H:%M EAT"),
            "file_name": output_path.name,
            "verification_url": args.verification_url,
            "data_url": args.data_url,
        },
    )
    print(f"Created: {args.output}")


if __name__ == "__main__":
    main()
