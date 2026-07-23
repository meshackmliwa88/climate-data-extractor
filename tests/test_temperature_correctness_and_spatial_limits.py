from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

import numpy as np
import pandas as pd
import xarray as xr
from openpyxl import load_workbook


class TemperatureCorrectnessAndSpatialLimitsTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.data_dir = self.root / "zarr"
        self.export_dir = self.root / "exports"
        self.project = Path(__file__).resolve().parents[1]

        monthly = self.data_dir / "monthly" / "CDE_ERA5_Tanzania_Temperature_Mean_Min_Max_1940_2025_Operational.zarr"
        monthly.parent.mkdir(parents=True, exist_ok=True)
        times = pd.date_range("1991-01-01", periods=24, freq="MS")
        shape = (len(times), 2, 2)
        # Deliberately put maximum first and use operationally plausible aliases.
        ds = xr.Dataset(
            {
                "temperature_maximum": (("time", "latitude", "longitude"), np.full(shape, 305.0, dtype=np.float32)),
                "temperature_mean": (("time", "latitude", "longitude"), np.full(shape, 298.0, dtype=np.float32)),
                "temperature_minimum": (("time", "latitude", "longitude"), np.full(shape, 291.0, dtype=np.float32)),
            },
            coords={"time": times, "latitude": [-6.25, -6.0], "longitude": [35.5, 35.75]},
        )
        for variable in ds.data_vars:
            ds[variable].attrs["units"] = "K"
            ds[variable].attrs["long_name"] = variable.replace("_", " ")
        ds.to_zarr(monthly, mode="w", consolidated=True)

        rain = self.data_dir / "monthly" / "CDE_CHIRPS_Tanzania_Rainfall_1981_2025_Monthly_Total.zarr"
        rain.parent.mkdir(parents=True, exist_ok=True)
        rainfall = xr.Dataset(
            {"precip": (("time", "latitude", "longitude"), np.full(shape, 100.0, dtype=np.float32))},
            coords={"time": times, "latitude": [-6.25, -6.0], "longitude": [35.5, 35.75]},
        )
        rainfall["precip"].attrs.update(units="mm", long_name="Precipitation")
        rainfall.to_zarr(rain, mode="w", consolidated=True)

        shape_dir = self.data_dir / "shapefiles"
        shape_dir.mkdir(parents=True, exist_ok=True)
        boundary = {
            "type": "FeatureCollection",
            "features": [{
                "type": "Feature",
                "properties": {"NAME_1": "Synthetic Tanzania"},
                "geometry": {"type": "Polygon", "coordinates": [[[34, -7], [36, -7], [36, -5], [34, -5], [34, -7]]]},
            }],
        }
        for level in (1, 2, 3):
            (shape_dir / f"gadm41_TZA_{level}.json").write_text(json.dumps(boundary), encoding="utf-8")

    def tearDown(self):
        self.tmp.cleanup()

    def test_minimum_never_uses_maximum_variable(self):
        from cde_products import extract_point_series, generate_plot_product

        minimum, context = extract_point_series(
            self.data_dir, "era5_temperature_stats", "monthly", -6.1, 35.7,
            "1991-01-01", "1992-12-31", variable="tmin",
        )
        maximum, max_context = extract_point_series(
            self.data_dir, "era5_temperature_stats", "monthly", -6.1, 35.7,
            "1991-01-01", "1992-12-31", variable="tmax",
        )
        self.assertAlmostEqual(float(minimum["value"].iloc[0]), 17.85, places=2)
        self.assertAlmostEqual(float(maximum["value"].iloc[0]), 31.85, places=2)
        self.assertEqual(context["actual_variable"], "temperature_minimum")
        self.assertEqual(context["variable_label"], "Minimum Temperature")
        self.assertEqual(max_context["actual_variable"], "temperature_maximum")

        product = generate_plot_product({
            "plot_type": "time_series", "dataset": "era5_temperature_stats",
            "resolution": "monthly", "variable": "tmin", "location_name": "Dodoma",
            "latitude": -6.1, "longitude": 35.7,
            "start_date": "1991-01-01", "end_date": "1992-12-31",
        }, self.data_dir, self.export_dir)
        self.assertEqual(product["context"]["variable_label"], "Minimum Temperature")
        self.assertEqual(product["context"]["actual_variable"], "temperature_minimum")

    def test_main_data_extractor_temperature_workbook(self):
        from scripts.extractor import StationPoint, write_excel_output

        catalog = json.loads((self.project / "config" / "zarr_catalog.json").read_text(encoding="utf-8"))
        output = self.root / "temperature.xlsx"
        write_excel_output(
            output, "ERA5_TEMP_STATS", catalog["ERA5_TEMP_STATS"],
            [StationPoint("POINT_1", "Dodoma", -6.1, 35.7)],
            ["ta", "tmin", "tmax"], "1991-01-01", "1992-12-31",
            ["monthly"], [], self.data_dir,
        )
        self.assertTrue(output.is_file())
        workbook = load_workbook(output, data_only=True, read_only=True)
        self.assertEqual(len(workbook.sheetnames), 3)
        self.assertTrue(any("Monthly" in name for name in workbook.sheetnames))
        values = [
            cell.value
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
        ]
        workbook.close()
        self.assertTrue(
            any("Minimum Temperature" in str(value) for value in values if value is not None)
        )
        self.assertTrue(
            any("Maximum Temperature" in str(value) for value in values if value is not None)
        )
        numeric = [float(value) for value in values if isinstance(value, (int, float))]
        self.assertTrue(any(abs(value - 17.9) < 0.11 for value in numeric))
        self.assertTrue(any(abs(value - 31.9) < 0.11 for value in numeric))

    def test_spatial_panel_limit_prevents_overload(self):
        from cde_products import _spatial_panel_jobs, generate_plot_product

        with self.assertRaisesRegex(ValueError, "limited to 24 maps"):
            _spatial_panel_jobs({
                "dataset": "chirps_rainfall", "variable": "precip",
                "map_panel_basis": "year", "map_month_selection": "all",
                "map_year_selection": "range", "map_start_year": "1991", "map_end_year": "2015",
            })

        result = generate_plot_product({
            "plot_type": "spatial_map", "dataset": "chirps_rainfall", "variable": "precip",
            "resolution": "monthly", "map_output_layout": "panel", "map_panel_basis": "year",
            "map_month_selection": "all", "map_year_selection": "range",
            "map_start_year": "1991", "map_end_year": "1992", "map_admin_level": "1",
            "map_render_style": "grid", "show_lakes": False,
        }, self.data_dir, self.export_dir)
        self.assertTrue(Path(result["plot_path"]).is_file())
        self.assertTrue(result["context"]["bounded_memory_rendering"])
        self.assertEqual(result["context"]["spatial_panel_limit"], 24)


if __name__ == "__main__":
    unittest.main()
