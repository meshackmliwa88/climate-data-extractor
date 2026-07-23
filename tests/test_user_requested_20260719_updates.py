from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

import numpy as np
import pandas as pd
import xarray as xr
from openpyxl import load_workbook



def test_ranked_climate_index_plot_is_removed_and_axes_are_time_aware():
    import cde_products as products

    assert "rank" not in {value for value, _label in products.INDEX_PLOT_TYPES}
    assert all("rank" not in values for values in products.INDEX_PLOT_RULES.values())
    assert products.y_axis_label("ERA5 Precipitation", "mm", "hourly") == "Precipitation (mm/hour)"
    assert products.y_axis_label("CHIRPS Rainfall", "mm", "daily") == "Precipitation (mm/day)"
    assert products.y_axis_label("Minimum Temperature", "°C", "monthly") == "Minimum Temperature (°C)"



def test_diurnal_temperature_statistics_compute_hourly_cycle_and_daily_range():
    from cde_products import _compute_diurnal_cycle_stats

    times = pd.date_range("2020-01-01", periods=48, freq="h")
    # Two identical days ranging from 10 to 33 degrees: DTR = 23 °C.
    values = np.tile(np.arange(10.0, 34.0), 2)
    hourly, daily = _compute_diurnal_cycle_stats(
        pd.DataFrame({"time": times, "value": values}), temperature=True
    )

    assert len(hourly) == 24
    assert len(daily) == 2
    assert np.allclose(hourly["mean"].to_numpy(), np.arange(10.0, 34.0))
    assert np.allclose(daily["diurnal_temperature_range"].to_numpy(), [23.0, 23.0])
    assert float(hourly["mean_daily_diurnal_temperature_range"].iloc[0]) == 23.0



def test_u10_v10_are_derived_into_speed_and_meteorological_direction():
    from cde_products import _wind_dataarray

    ds = xr.Dataset(
        {
            "u10": (("time", "latitude", "longitude"), np.array([[[3.0]]] )),
            "v10": (("time", "latitude", "longitude"), np.array([[[4.0]]] )),
        },
        coords={"time": pd.date_range("2020-01-01", periods=1), "latitude": [-6.0], "longitude": [35.75]},
    )
    ds["u10"].attrs["units"] = "m s**-1"
    ds["v10"].attrs["units"] = "m s**-1"

    speed, speed_name, _speed_label, speed_unit = _wind_dataarray(ds, "wind_speed")
    direction, direction_name, _direction_label, direction_unit = _wind_dataarray(ds, "wind_direction")

    assert np.isclose(float(speed.values.squeeze()), 5.0)
    assert speed_name == "derived_from_u10_v10"
    assert speed_unit == "m s**-1"
    assert direction_name == "derived_from_u10_v10"
    assert direction_unit == "degrees"
    # Meteorological FROM direction for u=3, v=4 is about 216.87 degrees.
    assert np.isclose(float(direction.values.squeeze()), 216.86989765)



def test_spatial_grid_preserves_every_native_quarter_degree_cell():
    from cde_products import _spatial_grid_from_open_dataset

    latitudes = np.array([-6.50, -6.25, -6.00])
    longitudes = np.array([35.50, 35.75, 36.00])
    times = pd.to_datetime(["2000-01-01", "2000-02-01"])
    values = np.stack([
        np.arange(9, dtype=float).reshape(3, 3),
        np.arange(9, dtype=float).reshape(3, 3) + 10.0,
    ])
    ds = xr.Dataset(
        {"precip": (("time", "latitude", "longitude"), values)},
        coords={"time": times, "latitude": latitudes, "longitude": longitudes},
    )
    ds["precip"].attrs.update({"units": "mm", "long_name": "Precipitation"})
    spec = {
        "dataset_key": "chirps_rainfall",
        "variable": "precip",
        "period": {
            "years": [2000], "months": [1, 2], "period_label": "January–February 2000",
            "mode": "custom", "preferred_resolution": "monthly", "season": None,
        },
        "file_path": Path("synthetic_025.zarr"),
        "source_resolution": "monthly",
        "season": None,
    }

    frame, context = _spatial_grid_from_open_dataset(ds, spec, max_axis_cells=1)
    assert len(frame) == 9  # max_axis_cells must not thin the native 3×3 grid.
    assert context["native_grid_preserved"] is True
    assert context["render_grid_reduced"] is False
    assert context["native_latitude_resolution_degrees"] == 0.25
    assert context["native_longitude_resolution_degrees"] == 0.25
    assert context["spatial_interpolation"] == "None"
    assert context["spatial_resampling"] == "None"
    expected = values.sum(axis=0).reshape(-1)
    assert np.allclose(np.sort(frame["value"].to_numpy()), np.sort(expected))



def test_multiple_extraction_selections_are_written_to_separate_sheets():
    from cde_multi import generate_multi_extraction

    times = pd.date_range("1991-01-01", periods=2, freq="MS")

    def fake_extract(_data_dir, _dataset, _resolution, lat, lon, _start, _end, variable=None, season=None):
        base = {"ta": 25.0, "tmin": 18.0, "tmax": 32.0}[str(variable)]
        return (
            pd.DataFrame({"time": times, "value": [base, base + 1]}),
            {
                "nearest_latitude": lat,
                "nearest_longitude": lon,
                "unit": "°C",
                "variable_label": str(variable),
            },
        )

    with TemporaryDirectory() as tmp:
        with patch("cde_multi.extract_point_series", side_effect=fake_extract):
            result = generate_multi_extraction(
                {
                    "dataset": "era5_temperature_stats",
                    "variables": ["ta", "tmin", "tmax"],
                    "resolutions": ["monthly"],
                    "seasons": [],
                    "location_name": "Dodoma",
                    "latitude": -6.17,
                    "longitude": 35.74,
                    "start_date": "1991-01-01",
                    "end_date": "1991-02-28",
                },
                Path(tmp),
                Path(tmp),
            )
        workbook = load_workbook(result["excel_path"], read_only=True)
        assert len(workbook.sheetnames) == 3
        assert any("Mean Temperature" in name for name in workbook.sheetnames)
        assert any("Minimum Temperature" in name for name in workbook.sheetnames)
        assert any("Maximum Temperature" in name for name in workbook.sheetnames)
        workbook.close()
        assert result["context"]["separate_selection_sheets"] is True
        assert result["context"]["worksheet_count"] == 3



def test_combined_generated_outputs_keep_separate_worksheets():
    from cde_excel import combine_workbooks_to_separate_sheets, write_single_sheet_workbook

    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        first = write_single_sheet_workbook(
            root / "first.xlsx", [("Data", pd.DataFrame({"Year": [2000], "Value": [1.0]}))],
            workbook_title="First Selection",
        )
        second = write_single_sheet_workbook(
            root / "second.xlsx", [("Data", pd.DataFrame({"Year": [2001], "Value": [2.0]}))],
            workbook_title="Second Selection",
        )
        combined = combine_workbooks_to_separate_sheets(
            root / "combined.xlsx",
            [("Rainfall Monthly", first), ("Temperature Annual", second)],
        )
        workbook = load_workbook(combined, read_only=True)
        assert workbook.sheetnames == ["Rainfall Monthly", "Temperature Annual"]
        workbook.close()



def test_title_control_is_available_for_every_tab_and_smoothing_is_absent():
    template = (Path(__file__).resolve().parents[1] / "templates" / "plots.html").read_text(encoding="utf-8")
    assert 'name="custom_plot_title"' in template
    assert "Plot / Map Title" in template
    assert "Analysis Title" in template
    assert "Climate Index Title" in template
    assert "Native 0.25° Grid Cells — No Interpolation" in template
    assert "Smooth Filled Contours" not in template


def test_annual_trend_analysis_uses_yearly_totals_for_rainfall_and_means_otherwise():
    from cde_products import annual_trend_frame, _annual_trend_labels
    from cde_analysis import _long_term_trend_labels, _EXTRA_ANALYSIS_SCOPES

    times = pd.to_datetime([
        "2020-01-01", "2020-02-01", "2021-01-01", "2021-02-01",
    ])
    frame = pd.DataFrame({"time": times, "value": [100.0, 200.0, 150.0, 250.0]})
    rainfall, rainfall_stats = annual_trend_frame(frame, "rainfall")
    temperature, temperature_stats = annual_trend_frame(frame, "temperature")

    assert rainfall["annual_value"].tolist() == [300.0, 400.0]
    assert temperature["annual_value"].tolist() == [150.0, 200.0]
    assert rainfall_stats["aggregation"] == "sum"
    assert temperature_stats["aggregation"] == "mean"
    assert np.isclose(rainfall_stats["slope_per_year"], 100.0)
    assert np.isclose(rainfall_stats["r_squared"], 1.0)
    assert _annual_trend_labels("CHIRPS Precipitation", "mm/day", "rainfall")[2] == "Annual Rainfall (mm)"
    assert _long_term_trend_labels("ERA5 Precipitation", "rainfall", "mm/day")[2] == "Annual Rainfall (mm)"
    assert "long_term_annual_trend_analysis" in _EXTRA_ANALYSIS_SCOPES


def test_simple_excel_style_has_plain_tabs_and_light_header():
    from scripts.extractor import style_excel

    with TemporaryDirectory() as tmp:
        path = Path(tmp) / "styled.xlsx"
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for name in ("Monthly Rainfall - Dodoma", "Annual Rainfall - Dodoma"):
                metadata = pd.DataFrame([[name], ["Location : Dodoma"], ["Unit : mm"]])
                data = pd.DataFrame({"Year": [2020, 2021], "Rainfall": [850.4, 910.2]})
                metadata.to_excel(writer, sheet_name=name, index=False, header=False)
                data.to_excel(writer, sheet_name=name, index=False, startrow=4)
        style_excel(path)
        workbook = load_workbook(path)
        for ws in workbook.worksheets:
            assert ws.sheet_properties.tabColor is None
            assert ws.freeze_panes is None
            assert ws.sheet_view.showGridLines is True
            assert ws.auto_filter.ref is None
            assert ws[5][0].fill.fgColor.rgb[-6:] == "E7E6E6"
        workbook.close()


def test_spatial_colourbar_is_narrow_and_has_upper_extension_only():
    source = (Path(__file__).resolve().parents[1] / "cde_products.py").read_text(encoding="utf-8")
    assert 'style_spec["extend"] = "max"' in source
    assert "width_ratios=[1.0, 0.035]" in source
    assert "[0.945, 0.17, 0.012, 0.66]" in source


def test_long_term_annual_trend_is_available_in_analysis_tab():
    project = Path(__file__).resolve().parents[1]
    template = (project / "templates" / "plots.html").read_text(encoding="utf-8")
    app_source = (project / "app.py").read_text(encoding="utf-8")
    assert 'value="long_term_annual_trend_analysis"' in template
    assert "Long-Term Annual Trend Analysis (Equation and R²)" in template
    assert '"long_term_annual_trend_analysis"' in app_source
