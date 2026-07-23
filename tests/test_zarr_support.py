from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import numpy as np
import pandas as pd
import xarray as xr

from cde_analysis import generate_analysis_bundle
from cde_products import extract_point_series, find_file
from openpyxl import load_workbook
from cde_store import open_data_store
from scripts.extractor import (
    StationPoint,
    find_files_for_source_frequency,
    load_catalog,
    open_dataset_for_source_frequency,
    write_excel_output,
    make_dataframe_for_excel,
)


class ZarrSupportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.data_dir = self.root / "zarr"
        self.export_dir = self.root / "exports"
        store = self.data_dir / "monthly" / "CDE_CHIRPS_Tanzania_Rainfall_1981_2025_Monthly_Total.zarr"
        store.parent.mkdir(parents=True, exist_ok=True)

        time = pd.date_range("1991-01-01", "2000-12-01", freq="MS")
        latitude = np.array([-6.25, -6.00])
        longitude = np.array([35.50, 35.75])
        base = np.arange(len(time), dtype=np.float32)[:, None, None]
        values = 50.0 + (base % 12) * 2.0 + np.zeros((len(time), 2, 2), dtype=np.float32)
        ds = xr.Dataset(
            {"precip": (("time", "latitude", "longitude"), values)},
            coords={"time": time, "latitude": latitude, "longitude": longitude},
        )
        ds["precip"].attrs.update(long_name="Monthly precipitation", units="mm")
        ds.to_zarr(store, mode="w", consolidated=True)
        self.store = store

    def tearDown(self):
        self.tmp.cleanup()

    def test_shared_store_opener_and_discovery(self):
        found = find_file(self.data_dir, "chirps_rainfall", "monthly")
        self.assertEqual(found, self.store)
        with open_data_store(found) as ds:
            self.assertIn("precip", ds.data_vars)
            self.assertEqual(ds.sizes["time"], 120)

    def test_catalog_extractor_opens_zarr(self):
        cfg = {
            "label": "Synthetic CHIRPS",
            "time_coord": "time",
            "lat_coord": "latitude",
            "lon_coord": "longitude",
            "supported_frequencies": ["monthly"],
            "file_patterns": {"monthly": "monthly/CDE_CHIRPS_Tanzania_Rainfall_1981_2025_Monthly_Total.zarr"},
        }
        files = find_files_for_source_frequency(cfg, "monthly", self.data_dir)
        self.assertEqual([Path(p) for p in files], [self.store])
        with open_dataset_for_source_frequency(cfg, "monthly", self.data_dir) as ds:
            self.assertIn("precip", ds)

    def test_point_extraction_is_lazy_and_location_aware(self):
        frame, context = extract_point_series(
            self.data_dir,
            "chirps_rainfall",
            "monthly",
            -6.10,
            35.70,
            "1991-01-01",
            "1992-12-31",
            variable="precip",
        )
        self.assertEqual(len(frame), 24)
        self.assertEqual(context["storage_format"], "Zarr")
        self.assertAlmostEqual(float(context["nearest_latitude"]), -6.0)
        self.assertAlmostEqual(float(context["nearest_longitude"]), 35.75)
        self.assertEqual(context["unit"], "mm")

    def test_full_analysis_bundle(self):
        result = generate_analysis_bundle(
            {
                "dataset": "chirps_rainfall",
                "resolution": "monthly",
                "variable": "precip",
                "location_name": "Dodoma",
                "latitude": "-6.10",
                "longitude": "35.70",
                "start_date": "1991-01-01",
                "end_date": "2000-12-31",
                "baseline_start": "1991",
                "baseline_end": "2000",
                "include_plot": True,
            },
            self.data_dir,
            self.export_dir,
        )
        for key in ["excel_path", "time_series_plot_path", "anomaly_plot_path"]:
            self.assertTrue(Path(result[key]).is_file(), key)
        self.assertTrue(Path(result["excel_path"]).name.startswith("TMA_CDE_"))
        self.assertNotIn("csv_path", result)
        self.assertNotIn("zip_path", result)
        wb = load_workbook(result["excel_path"])
        self.assertEqual(wb.sheetnames, ["Data"])
        self.assertGreater(sum(len(ws._images) for ws in wb.worksheets), 0)
        ws = wb["Data"]
        date_header = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "date_time")
        self.assertIsNotNone(ws.cell(date_header + 1, 2).value)
        self.assertIsNotNone(ws.cell(date_header + 1, 1).border.left.style)
        wb.close()
        self.assertEqual(result["rows"], 120)
        self.assertIn(result["trend"]["trend"], {"increasing", "decreasing", "no significant trend"})

    def test_integrated_plot_product_analysis_scopes(self):
        scopes = {
            "data_extraction": 0,
            "statistical_summary": 1,
            "climatology_profile": 1,
            "trend_variability": 1,
            "extremes_analysis": 1,
            "comprehensive_analysis": 4,
        }
        for scope, minimum_previews in scopes.items():
            result = generate_analysis_bundle(
                {
                    "analysis_scope": scope,
                    "dataset": "chirps_rainfall",
                    "resolution": "monthly",
                    "variable": "precip",
                    "location_name": "Dodoma",
                    "latitude": "-6.10",
                    "longitude": "35.70",
                    "start_date": "1991-01-01",
                    "end_date": "2000-12-31",
                    "baseline_start": "1991",
                    "baseline_end": "2000",
                    "include_plot": True,
                },
                self.data_dir,
                self.export_dir / scope,
            )
            self.assertTrue(Path(result["excel_path"]).is_file(), scope)
            self.assertTrue(Path(result["excel_path"]).name.startswith("TMA_CDE_"), scope)
            self.assertNotIn("csv_path", result, scope)
            self.assertNotIn("zip_path", result, scope)
            self.assertGreaterEqual(len(result["preview_paths"]), minimum_previews, scope)
            self.assertNotIn("Zarr", result["product_title"])

    def test_temperature_indices_are_derived_from_authoritative_hourly_store(self):
        store = self.data_dir / "hourly" / "ERA5_Tanzania_Temperature_2M_Hourly_1940_2026.zarr"
        store.parent.mkdir(parents=True, exist_ok=True)
        time = pd.date_range("1991-01-01", "1992-12-31 23:00", freq="h")
        values = 295.0 + 3.0 * np.sin(2 * np.pi * time.hour.to_numpy() / 24.0)
        ds = xr.Dataset(
            {"t2m": (("valid_time", "latitude", "longitude"), values[:, None, None] + np.zeros((len(time), 2, 2)))},
            coords={"valid_time": time, "latitude": [-6.25, -6.0], "longitude": [35.5, 35.75]},
        )
        ds["t2m"].attrs.update(long_name="2 m Temperature", units="K")
        ds.to_zarr(store, mode="w", consolidated=True)

        from cde_products import generate_indices, DATASETS, INDEX_RESOLUTION_RULES
        self.assertEqual(DATASETS["era5_temperature"]["label"], "ERA5 2m Temperature")
        self.assertEqual(list(DATASETS["era5_temperature"]["resolutions"]), ["hourly", "daily", "monthly", "annual", "seasonal"])
        self.assertIn("monthly", INDEX_RESOLUTION_RULES["mean_temperature"])
        result = generate_indices(
            {
                "index_type": "mean_temperature", "dataset": "era5_temperature",
                "index_resolution": "monthly", "index_plot_type": "line", "season": "ANNUAL",
                "start_year": "1991", "end_year": "1992", "start_date": "1991-01-01", "end_date": "1992-12-31",
                "baseline_start": "1991", "baseline_end": "1992",
                "latitude": "-6.10", "longitude": "35.70", "location_name": "Dodoma",
            },
            self.data_dir, self.export_dir,
        )
        self.assertEqual(result["rows"], 24)
        self.assertTrue(Path(result["plot_path"]).is_file())
        self.assertTrue(Path(result["excel_path"]).is_file())
        self.assertTrue(Path(result["parquet_path"]).is_file())
        self.assertNotIn("csv_path", result)
        wb = load_workbook(result["excel_path"])
        self.assertGreater(sum(len(ws._images) for ws in wb.worksheets), 0)
        wb.close()

    def test_unlimited_year_validation_and_excel_only_ui(self):
        from app import _bounded_product_years
        self.assertEqual(_bounded_product_years("1940", "2026", "hourly", "plot"), (1940, 2026))
        project = Path(__file__).resolve().parents[1]
        app_text = (project / "app.py").read_text(encoding="utf-8")
        page = (project / "templates" / "plots.html").read_text(encoding="utf-8")
        self.assertIn('"index_type": "total_rainfall"', app_text)
        self.assertNotIn("Download CSV", page)
        self.assertNotIn('"csv_url"', app_text)
        self.assertIn("Download Excel", page)
        self.assertIn("TMA_CDE_", app_text)


    def test_multiple_requests_append_to_one_excel_with_qr(self):
        output = self.export_dir / "TMA_CDE_MULTI_REQUESTS.xlsx"
        cfg = load_catalog()["CHIRPS"]
        context = {"download_id": "TEST-MULTI", "file_name": output.name}
        write_excel_output(
            output, "CHIRPS", cfg, [StationPoint("1", "Dodoma", -6.0, 35.75)], ["precip"],
            "1991-01-01", "1992-12-31", ["monthly"], data_dir=self.data_dir, download_context=context,
        )
        write_excel_output(
            output, "CHIRPS", cfg, [StationPoint("2", "Second Location", -6.25, 35.5)], ["precip"],
            "1991-01-01", "1992-12-31", ["monthly"], data_dir=self.data_dir,
            download_context=context, append=True,
        )
        wb = load_workbook(output)
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertTrue(all("Monthly" in name for name in wb.sheetnames))
        self.assertEqual(sum(len(ws._images) for ws in wb.worksheets), 2)
        for sheet in wb.worksheets:
            year_headers = [row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Year"]
            self.assertEqual(len(year_headers), 1)
            header_row = year_headers[0]
            self.assertEqual(sheet.cell(header_row + 1, 2).value, 50.0)
            self.assertEqual(sheet.cell(header_row + 1, 1).number_format, "0")
            self.assertIsNotNone(sheet.cell(header_row + 1, 1).border.left.style)
            self.assertIsNone(sheet.freeze_panes)
            self.assertIsNone(sheet.auto_filter.ref)
        numeric_formats = {
            cell.number_format for ws in wb.worksheets for row in ws.iter_rows() for cell in row
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
        }
        self.assertIn("0.0", numeric_formats)
        wb.close()


    def test_operational_precision_rules(self):
        index = pd.date_range("2000-01-01", periods=2, freq="D")
        rainfall = make_dataframe_for_excel(pd.Series([1.24, 2.26], index=index), "daily", "Rainfall", "precip")
        humidity = make_dataframe_for_excel(pd.Series([71.6, 72.4], index=index), "daily", "Relative Humidity", "r")
        wind_speed = make_dataframe_for_excel(pd.Series([4.4, 5.6], index=index), "daily", "Wind Speed", "wind_speed")
        wind_direction = make_dataframe_for_excel(pd.Series([182.2, 359.7], index=index), "daily", "Wind Direction", "wind_direction")
        self.assertEqual(rainfall["Rainfall"].tolist(), [1.2, 2.3])
        self.assertEqual(humidity["Relative Humidity"].tolist(), [72.0, 72.0])
        self.assertEqual(wind_speed["Wind Speed"].tolist(), [4.0, 6.0])
        self.assertEqual(wind_direction["Wind Direction"].tolist(), [182.0, 360.0])

    def test_runtime_is_tuned_for_bounded_memory(self):
        project = Path(__file__).resolve().parents[1]
        service = (project / "deployment" / "netcdf-extractor.service").read_text(encoding="utf-8")
        self.assertIn("--workers 1", service)
        self.assertIn("--threads 2", service)
        self.assertIn('Environment="CDE_DASK_WORKERS=2"', service)
        self.assertIn("constant_memory", (project / "scripts" / "extractor.py").read_text(encoding="utf-8"))

    def test_non_zarr_file_is_rejected(self):
        old_file = self.root / "legacy.nc"
        old_file.write_bytes(b"not a zarr store")
        with self.assertRaises(ValueError):
            open_data_store(old_file)

    def test_sidebar_has_one_unified_products_menu(self):
        project = Path(__file__).resolve().parents[1]
        base = (project / "templates" / "base.html").read_text(encoding="utf-8")
        page = (project / "templates" / "plots.html").read_text(encoding="utf-8")
        extractor_page = (project / "templates" / "extractor.html").read_text(encoding="utf-8")
        tabs = (project / "templates" / "partials" / "workspace_tabs.html").read_text(encoding="utf-8")
        app_text = (project / "app.py").read_text(encoding="utf-8")
        self.assertEqual(base.count("Data Extractor &amp; Products"), 1)
        self.assertNotIn(">Data Extractor</a>", base)
        self.assertNotIn(">Plots &amp; Products</a>", base)
        self.assertIn('name="workspace_service" value="{{ active_service }}"', page)
        self.assertNotIn('id="workspaceService"', page)
        self.assertIn('partials/workspace_tabs.html', page)
        self.assertIn('partials/workspace_tabs.html', extractor_page)
        self.assertIn("Extract Data", tabs)
        self.assertIn("Create Plot or Map", tabs)
        self.assertIn("Climate Analysis", tabs)
        self.assertIn("Climate Indices", tabs)
        self.assertLess(page.index('id="plotDataset"'), page.index('id="visualType"'))
        self.assertLess(page.index('id="plotDataset"'), page.index('id="analysisType"'))
        self.assertLess(page.index('id="plotDataset"'), page.index('id="indexType"'))
        self.assertNotIn("Generate All Products", page)
        self.assertIn("Map Period", page)
        self.assertIn("Admin 1 — Regions", page)
        self.assertIn("Admin 2 — Districts", page)
        self.assertIn("Admin 3 — Wards", page)
        self.assertIn(">All Months<", page)
        self.assertNotIn(">Ocean<", page)
        self.assertIn("Lakes", page)
        self.assertNotIn(">Rivers<", page)
        self.assertIn("tanzania_lakes.geojson", (project / "cde_products.py").read_text(encoding="utf-8"))
        self.assertIn('INDEX_OPTIONS_BY_DATASET', app_text)
        self.assertIn('if service == "extraction" and request.method == "GET"', app_text)
        plots_route = app_text[app_text.index('def plots():'):app_text.index('@app.route("/parquet-records")')]
        self.assertIn("sources = product_form_sources()", plots_route)
        self.assertNotIn("available_datasets(DEFAULT_DATA_DIR)", plots_route)
        self.assertIn("@single_product_request", app_text)

    def test_custom_spatial_map_and_all_month_heatmap(self):
        import json
        from cde_products import generate_plot_product

        shape_dir = self.data_dir / "shapefiles"
        shape_dir.mkdir(parents=True, exist_ok=True)
        boundary = {
            "type": "FeatureCollection",
            "features": [{
                "type": "Feature",
                "properties": {"NAME": "Synthetic Tanzania"},
                "geometry": {"type": "Polygon", "coordinates": [[[34.0, -7.0], [36.0, -7.0], [36.0, -5.0], [34.0, -5.0], [34.0, -7.0]]]},
            }],
        }
        for level in (1, 2, 3):
            (shape_dir / f"gadm41_TZA_{level}.json").write_text(json.dumps(boundary), encoding="utf-8")

        spatial = generate_plot_product(
            {
                "plot_type": "spatial_map", "dataset": "chirps_rainfall", "resolution": "hourly", "variable": "auto",
                "map_period_mode": "custom", "map_custom_months": "1,2,3", "map_custom_years": "1991-2000",
                "start_year": "1991", "end_year": "2000", "map_admin_level": "2",
                "show_ocean": True, "show_lakes": True, "show_rivers": True, "map_output_layout": "panel",
            },
            self.data_dir, self.export_dir,
        )
        self.assertTrue(Path(spatial["plot_path"]).is_file())
        self.assertTrue(Path(spatial["excel_path"]).is_file())
        self.assertEqual(spatial["context"]["administrative_level"], 2)
        self.assertIn("January", spatial["context"]["period_label"])
        self.assertEqual(spatial["context"]["map_output_layout"], "panel")
        self.assertEqual(spatial["context"]["map_count"], 10)

        heatmap = generate_plot_product(
            {
                "plot_type": "heatmap", "dataset": "chirps_rainfall", "resolution": "hourly", "variable": "auto",
                "latitude": "-6.0", "longitude": "35.75", "location_name": "Dodoma",
                "start_year": "1991", "end_year": "2000", "heatmap_month_mode": "all", "heatmap_year_mode": "range",
            },
            self.data_dir, self.export_dir,
        )
        self.assertTrue(Path(heatmap["plot_path"]).is_file())
        self.assertTrue(Path(heatmap["excel_path"]).is_file())
        self.assertEqual(heatmap["context"]["months"], list(range(1, 13)))
        self.assertIn("All months", heatmap["context"]["period_label"])

    def test_removed_rainy_season_indices_and_multiple_selection_ui(self):
        project = Path(__file__).resolve().parents[1]
        page = (project / "templates" / "plots.html").read_text(encoding="utf-8")
        from cde_products import RAINFALL_INDICES
        keys = {key for key, _ in RAINFALL_INDICES}
        self.assertNotIn("rainy_season_onset", keys)
        self.assertNotIn("rainy_season_cessation", keys)
        self.assertNotIn("length_of_rainy_season", keys)
        self.assertIn("plot_variables", page)
        self.assertIn('name="plot_resolutions"', page)
        self.assertNotIn("This catalogue is loaded", page)
        self.assertNotIn("Only the selected plot", page)
        self.assertNotIn("Extraction is not repeated", page)

    def test_multi_extraction_one_sheet(self):
        from cde_multi import generate_multi_extraction
        result = generate_multi_extraction(
            {
                "dataset": "chirps_rainfall",
                "variables": ["precip"],
                "resolutions": ["monthly"],
                "seasons": ["MAM"],
                "location_name": "Dodoma",
                "latitude": -6.10,
                "longitude": 35.70,
                "start_date": "1991-01-01",
                "end_date": "2000-12-31",
            },
            self.data_dir,
            self.export_dir,
        )
        self.assertTrue(Path(result["excel_path"]).is_file())
        wb = load_workbook(result["excel_path"], read_only=False)
        self.assertEqual(wb.sheetnames, ["Data"])
        self.assertGreater(len(wb["Data"]._images), 0)
        self.assertEqual(result["rows"], 120)
        wb.close()

    def test_era5_temperature_all_resolutions_and_prepared_statistics(self):
        hourly_store = self.data_dir / "hourly" / "ERA5_Tanzania_Temperature_2M_Hourly_1940_2026.zarr"
        hourly_store.parent.mkdir(parents=True, exist_ok=True)
        hourly_time = pd.date_range("1991-01-01", "1991-01-02 23:00", freq="h")
        hourly_shape = (len(hourly_time), 2, 2)
        hourly = xr.Dataset(
            {"t2m": (("valid_time", "latitude", "longitude"), np.full(hourly_shape, 298.0, dtype=np.float32))},
            coords={"valid_time": hourly_time, "latitude": [-6.25, -6.0], "longitude": [35.5, 35.75]},
        )
        hourly["t2m"].attrs["units"] = "K"
        hourly.to_zarr(hourly_store, mode="w", consolidated=True)

        monthly_store = self.data_dir / "monthly" / "CDE_ERA5_Tanzania_Temperature_Mean_Min_Max_1940_2025_Monthly.zarr"
        monthly_store.parent.mkdir(parents=True, exist_ok=True)
        monthly_time = pd.date_range("1991-01-01", "1991-02-01", freq="MS")
        base = np.full((len(monthly_time), 2, 2), 298.0, dtype=np.float32)
        prepared = xr.Dataset(
            {
                "ta": (("time", "latitude", "longitude"), base),
                "tmin": (("time", "latitude", "longitude"), base - 5),
                "tmax": (("time", "latitude", "longitude"), base + 5),
            },
            coords={"time": monthly_time, "latitude": [-6.25, -6.0], "longitude": [35.5, 35.75]},
        )
        for variable in prepared.data_vars:
            prepared[variable].attrs["units"] = "K"
        prepared.to_zarr(monthly_store, mode="w", consolidated=True)

        from cde_multi import generate_multi_extraction
        hourly_result = generate_multi_extraction(
            {
                "dataset": "era5_temperature", "variables": ["ta"], "resolutions": ["hourly"],
                "seasons": [], "location_name": "Dodoma", "latitude": -6.10, "longitude": 35.70,
                "start_date": "1991-01-01", "end_date": "1991-01-02 23:59:59",
            }, self.data_dir, self.export_dir,
        )
        self.assertEqual(hourly_result["rows"], 48)
        wb = load_workbook(hourly_result["excel_path"], read_only=True)
        values = [cell.value for row in wb["Data"].iter_rows() for cell in row]
        self.assertIn("Mean 2m Temperature", values)
        wb.close()

        monthly_result = generate_multi_extraction(
            {
                "dataset": "era5_temperature_stats", "variables": ["ta", "tmin", "tmax"], "resolutions": ["monthly"],
                "seasons": [], "location_name": "Dodoma", "latitude": -6.10, "longitude": 35.70,
                "start_date": "1991-01-01", "end_date": "1991-02-28",
            }, self.data_dir, self.export_dir,
        )
        self.assertEqual(monthly_result["rows"], 2)
        wb = load_workbook(monthly_result["excel_path"], read_only=True)
        values = [cell.value for ws in wb.worksheets for row in ws.iter_rows() for cell in row]
        self.assertIn("Minimum Temperature", values)
        self.assertIn("Maximum Temperature", values)
        wb.close()

    def test_guided_dataset_first_workflow_and_monthly_profile(self):
        project = Path(__file__).resolve().parents[1]
        extractor_page = (project / "templates" / "extractor.html").read_text(encoding="utf-8")
        plots_page = (project / "templates" / "plots.html").read_text(encoding="utf-8")
        catalog = (project / "config" / "zarr_catalog.json").read_text(encoding="utf-8")
        for step in ("Dataset and Data Type", "Location and Period", "Customer and Generate"):
            self.assertIn(step, extractor_page)
        self.assertNotIn("The dataset controls every option shown below", plots_page)
        self.assertIn("Average Monthly Total / Mean (Jan–Dec)", plots_page)
        self.assertIn("All Maps in One Multi-Panel Plot", plots_page)
        self.assertIn("ERA5_TEMP_STATS", catalog)

        from cde_products import generate_plot_product
        product = generate_plot_product(
            {
                "plot_type": "monthly_climatology", "dataset": "chirps_rainfall", "resolution": "monthly",
                "variable": "precip", "location_name": "Dodoma", "latitude": -6.10, "longitude": 35.70,
                "start_date": "1991-01-01", "end_date": "2000-12-31",
            }, self.data_dir, self.export_dir,
        )
        self.assertEqual(product["rows"], 12)
        self.assertTrue(Path(product["plot_path"]).is_file())
        self.assertTrue(Path(product["excel_path"]).is_file())

    def test_delivery_report_preview_saved_requests_and_page_qr(self):
        project = Path(__file__).resolve().parents[1]
        app_text = (project / "app.py").read_text(encoding="utf-8")
        extractor_page = (project / "templates" / "extractor.html").read_text(encoding="utf-8")
        result_page = (project / "templates" / "result.html").read_text(encoding="utf-8")
        excel_writer = (project / "cde_excel.py").read_text(encoding="utf-8")
        extractor_writer = (project / "scripts" / "extractor.py").read_text(encoding="utf-8")
        self.assertIn('/delivery-report/preview', app_text)
        self.assertNotIn('id="previewReportBtn"', extractor_page)
        self.assertIn('Preview Data Delivery Report', result_page)
        self.assertIn('Save Request and Add Another', extractor_page)
        self.assertIn('localStorage', extractor_page)
        self.assertNotIn('Download Data Delivery Report', result_page)
        self.assertIn('add_qr_codes_to_workbook', excel_writer)
        self.assertIn('inside every worksheet', extractor_writer)
        self.assertIn('Repeat the report verification QR on every page', app_text)



    def test_statistic_dimension_and_wind_time_label_regressions(self):
        store = self.data_dir / "monthly" / "CDE_ERA5_Tanzania_Temperature_Mean_Min_Max_1940_2025_Statistic.zarr"
        times = pd.date_range("2000-01-01", periods=2, freq="MS")
        values = np.zeros((2, 3, 2, 2), dtype=np.float32)
        values[:, 0, :, :] = 298.0
        values[:, 1, :, :] = 293.0
        values[:, 2, :, :] = 303.0
        ds = xr.Dataset(
            {"temperature": (("time", "statistic", "latitude", "longitude"), values)},
            coords={"time": times, "statistic": ["mean", "minimum", "maximum"], "latitude": [-6.25, -6.0], "longitude": [35.5, 35.75]},
        )
        ds["temperature"].attrs["units"] = "K"
        ds.to_zarr(store, mode="w", consolidated=True)
        minimum, _ = extract_point_series(self.data_dir, "era5_temperature_stats", "monthly", -6.1, 35.7, "2000-01-01", "2000-02-28", variable="tmin")
        maximum, _ = extract_point_series(self.data_dir, "era5_temperature_stats", "monthly", -6.1, 35.7, "2000-01-01", "2000-02-28", variable="tmax")
        self.assertAlmostEqual(float(minimum["value"].iloc[0]), 19.85, places=2)
        self.assertAlmostEqual(float(maximum["value"].iloc[0]), 29.85, places=2)

        from scripts.extractor import make_combined_wind_dataframe
        index = pd.date_range("2000-01-01", periods=3, freq="MS", name="time")
        speed = pd.Series([4.0, 5.0, 6.0], index=index, name="value")
        direction = pd.Series([90.0, 180.0, 270.0], index=index, name="value")
        output = make_combined_wind_dataframe(speed, direction, "monthly")
        self.assertEqual(output.iloc[0]["Jan Wind Speed"], 4.0)
        self.assertEqual(output.iloc[0]["Jan Wind Direction"], 90.0)

    def test_spatial_map_always_preserves_native_grid(self):
        import json
        from cde_products import generate_plot_product
        shape_dir = self.data_dir / "shapefiles"
        shape_dir.mkdir(parents=True, exist_ok=True)
        boundary = {"type":"FeatureCollection","features":[{"type":"Feature","properties":{"NAME_1":"Synthetic"},"geometry":{"type":"Polygon","coordinates":[[[34,-7],[36,-7],[36,-5],[34,-5],[34,-7]]]}}]}
        for level in (1,2,3):
            (shape_dir / f"gadm41_TZA_{level}.json").write_text(json.dumps(boundary), encoding="utf-8")
        for style in ("grid", "smooth"):
            result = generate_plot_product({
                "plot_type":"spatial_climatology", "dataset":"chirps_rainfall", "variable":"precip", "resolution":"monthly",
                "map_month_selection":"all", "map_year_selection":"range", "map_start_year":"1991", "map_end_year":"2000",
                "map_output_layout":"single", "map_render_style":style, "map_admin_level":"1",
                "show_ocean":True, "show_lakes":True, "show_rivers":True,
            }, self.data_dir, self.export_dir)
            self.assertTrue(Path(result["plot_path"]).is_file())
            self.assertEqual(result["context"]["map_render_style"], "grid")
            self.assertTrue(result["context"]["native_grid_preserved"])
            self.assertEqual(result["context"]["spatial_interpolation"], "None")
            self.assertEqual(result["context"]["spatial_resampling"], "None")
            self.assertIn("Precipitation", result["context"]["variable_label"])



if __name__ == "__main__":
    unittest.main()
