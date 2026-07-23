from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook

from cde_excel import combine_workbooks_to_separate_sheets, write_single_sheet_workbook


def _assert_no_print_header_qr(path: Path) -> None:
    with ZipFile(path) as zf:
        ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        for part in sorted(name for name in zf.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")):
            root = ET.fromstring(zf.read(part))
            assert root.find("m:sheetViews/m:sheetView/m:pane", ns) is None
            assert root.find("m:legacyDrawingHF", ns) is None
        assert not any("cde_qr_header_" in name or "vmlDrawingCdeQr" in name for name in zf.namelist())


def test_excel_is_simple_and_has_one_visible_qr_inside_sheet():
    with TemporaryDirectory() as tmp:
        path = Path(tmp) / "TMA_CDE_QR_test.xlsx"
        write_single_sheet_workbook(
            path,
            [("Annual Rainfall", pd.DataFrame({"Year": [2023, 2024], "Rainfall (mm)": [900.0, 1000.0]}))],
            qr_payload={
                "download_id": "CDE-TEST-001",
                "file_name": path.name,
                "verification_url": "https://example.test/verify/CDE-TEST-001",
            },
            sheet_name="Annual",
        )
        wb = load_workbook(path)
        ws = wb["Annual"]
        assert ws.freeze_panes is None
        assert ws.sheet_view.showGridLines is True
        assert len(ws._images) == 1
        assert any(str(cell.value or "") == "Scan to verify" for row in ws.iter_rows(min_row=1, max_row=2) for cell in row)
        wb.close()
        _assert_no_print_header_qr(path)


def test_combined_workbook_has_only_selection_tabs_and_qr_inside_each():
    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        first = write_single_sheet_workbook(
            root / "first.xlsx",
            [("Daily", pd.DataFrame({"Date": ["2026-01-01"], "Rainfall (mm/day)": [12.0]}))],
            qr_payload={"verification_url": "https://example.test/verify/first"},
            sheet_name="Daily",
        )
        second = write_single_sheet_workbook(
            root / "second.xlsx",
            [("Annual", pd.DataFrame({"Year": [2025], "Temperature (°C)": [26.0]}))],
            qr_payload={"verification_url": "https://example.test/verify/second"},
            sheet_name="Annual",
        )
        combined = combine_workbooks_to_separate_sheets(
            root / "combined.xlsx",
            [("Daily Rainfall", first), ("Annual Temperature", second)],
            qr_payload={"verification_url": "https://example.test/verify/combined"},
        )
        wb = load_workbook(combined)
        assert wb.sheetnames == ["Daily Rainfall", "Annual Temperature"]
        assert all(ws.freeze_panes is None for ws in wb.worksheets)
        assert all(len(ws._images) == 1 for ws in wb.worksheets)
        assert all(ws.sheet_view.showGridLines is True for ws in wb.worksheets)
        wb.close()
        _assert_no_print_header_qr(combined)
