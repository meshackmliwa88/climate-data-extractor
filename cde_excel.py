"""Simple Excel exports for CDE products.

Generated workbooks keep the data easy to read: a plain title, compact metadata,
a basic header row, light borders and one visible QR code inside each worksheet.
Different selections or temporal resolutions remain in separate sheet tabs.
"""
from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Mapping, Sequence

import numpy as np
import pandas as pd
import xlsxwriter

from scripts.extractor import (
    add_qr_codes_to_workbook,
    default_download_context,
    make_qr_png,
    style_excel,
)


STRUCTURAL_NAMES = {
    "year", "month", "day", "hour", "season year", "season_year",
    "count", "records", "observations", "frequency", "n", "s",
}
COORDINATE_NAMES = {
    "latitude", "longitude", "lat", "lon", "requested latitude",
    "requested longitude", "nearest grid latitude", "nearest grid longitude",
    "requested_latitude", "requested_longitude", "nearest_grid_latitude",
    "nearest_grid_longitude",
}


def _clean(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, np.generic):
        return value.item()
    return value


def _is_zero_decimal(header: str, global_zero: bool) -> bool:
    h = str(header or "").strip().lower().replace("_", " ")
    return global_zero or any(token in h for token in (
        "relative humidity", "wind speed", "wind direction",
    ))


def write_single_sheet_workbook(
    path: Path | str,
    sections: Sequence[tuple[str, pd.DataFrame]],
    *,
    qr_payload: Mapping[str, Any] | None = None,
    zero_decimal: bool = False,
    sheet_name: str = "Data",
    workbook_title: str = "TMA Climate Data Product",
) -> Path:
    """Write a simple data workbook with one visible verification QR.

    The worksheet intentionally avoids dashboard-like styling. It uses a plain
    bold title, simple section labels, a light-grey table header, light borders
    and normal gridlines. No frozen panes or printed-page QR headers are added.
    """
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    prepared = [
        (str(title or "Data"), frame.copy() if frame is not None else pd.DataFrame())
        for title, frame in sections
    ]
    max_cols = max([len(frame.columns) for _, frame in prepared] + [2])
    payload = dict(qr_payload or default_download_context(output))
    payload.setdefault("file_name", output.name)
    payload["document_type"] = "Data Delivery Report"

    options = {"constant_memory": True, "strings_to_urls": False, "nan_inf_to_errors": True}
    with TemporaryDirectory() as tmpdir:
        qr_path = Path(tmpdir) / "verification.png"
        make_qr_png(payload, qr_path, compact=False)

        workbook = xlsxwriter.Workbook(str(output), options)
        ws = workbook.add_worksheet((sheet_name[:31] or "Data"))
        ws.hide_gridlines(0)

        title_fmt = workbook.add_format({"bold": True, "font_size": 12})
        section_fmt = workbook.add_format({"bold": True})
        header_fmt = workbook.add_format({
            "bold": True, "bg_color": "#E7E6E6", "border": 1,
            "align": "center", "valign": "vcenter", "text_wrap": True,
        })
        text_fmt = workbook.add_format({"border": 1, "valign": "top"})
        one_fmt = workbook.add_format({"border": 1, "num_format": "0.0", "valign": "top"})
        zero_fmt = workbook.add_format({"border": 1, "num_format": "0", "valign": "top"})
        coord_fmt = workbook.add_format({"border": 1, "num_format": "0.0000", "valign": "top"})
        datetime_fmt = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd hh:mm", "valign": "top"})
        date_fmt = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd", "valign": "top"})
        qr_label_fmt = workbook.add_format({"bold": True, "align": "center"})

        ws.write(0, 0, workbook_title, title_fmt)
        row = 2
        widths = [9] * max_cols

        for section_title, frame in prepared:
            ws.write(row, 0, section_title, section_fmt)
            row += 1
            if frame.empty and not len(frame.columns):
                ws.write(row, 0, "No records available")
                row += 2
                continue

            headers = [str(c) for c in frame.columns]
            for col, header in enumerate(headers):
                ws.write(row, col, header, header_fmt)
                widths[col] = min(max(widths[col], len(header) + 1), 14)
            header_row = row
            row += 1

            for values in frame.itertuples(index=False, name=None):
                for col, raw in enumerate(values):
                    value = _clean(raw)
                    header = headers[col]
                    h = header.strip().lower().replace("_", " ")
                    if hasattr(value, "hour") and hasattr(value, "year") and not isinstance(value, str):
                        fmt = datetime_fmt if getattr(value, "hour", 0) else date_fmt
                    elif isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
                        if h in COORDINATE_NAMES or "latitude" in h or "longitude" in h:
                            fmt = coord_fmt
                        elif h in STRUCTURAL_NAMES or any(token in h for token in ("year", "month", "day", "hour", "count", "records", "observations")):
                            fmt = zero_fmt
                            value = int(round(float(value))) if value is not None else value
                        elif _is_zero_decimal(header, zero_decimal):
                            fmt = zero_fmt
                            value = int(round(float(value))) if value is not None else value
                        else:
                            fmt = one_fmt
                            value = round(float(value), 1) if value is not None else value
                    else:
                        fmt = text_fmt
                    ws.write(row, col, value, fmt)
                    widths[col] = min(max(widths[col], len(str(value or "")) + 1), 14)
                row += 1
            row += 2

        for col, width in enumerate(widths):
            ws.set_column(col, col, width)

        qr_col = max_cols + 1
        ws.write(0, qr_col, "Scan to verify", qr_label_fmt)
        ws.insert_image(1, qr_col, str(qr_path), {
            "x_scale": 0.40, "y_scale": 0.40, "object_position": 1,
        })
        ws.set_column(qr_col, qr_col + 1, 10)
        workbook.close()

    style_excel(output)
    add_qr_codes_to_workbook(output, {(sheet_name[:31] or "Data"): payload}, payload)
    return output


# ---------------------------------------------------------------------------
# Multi-selection workbook assembly
# ---------------------------------------------------------------------------
def _safe_combined_sheet_name(value: str, used: set[str]) -> str:
    import re as _re

    base = _re.sub(r"[\\/*?:\[\]]+", " ", str(value or "Data"))
    base = _re.sub(r"\s+", " ", base).strip() or "Data"
    base = base[:31]
    candidate = base
    counter = 2
    while candidate.lower() in used:
        suffix = f" {counter}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate.lower())
    return candidate


def combine_workbooks_to_separate_sheets(
    path: Path | str,
    workbooks: Sequence[tuple[str, Path | str]],
    *,
    workbook_title: str = "TMA CDE Multi-Selection Output",
    qr_payload: Mapping[str, Any] | None = None,
) -> Path:
    """Combine outputs into simple separate tabs without a summary dashboard."""
    from openpyxl import Workbook, load_workbook

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    destination = Workbook()
    destination.remove(destination.active)
    used: set[str] = set()
    copied = 0

    for selection_label, source_path in workbooks:
        source_path = Path(source_path)
        if not source_path.exists() or source_path.suffix.lower() != ".xlsx":
            continue
        source = load_workbook(source_path, data_only=False)
        visible = [ws for ws in source.worksheets if ws.sheet_state == "visible"] or list(source.worksheets)
        for source_ws in visible:
            suffix = "" if len(visible) == 1 or source_ws.title.lower() == "data" else f" - {source_ws.title}"
            target = destination.create_sheet(_safe_combined_sheet_name(f"{selection_label}{suffix}", used))
            copied += 1
            for row in source_ws.iter_rows():
                for cell in row:
                    # Skip old QR labels/columns; a fresh QR is added later.
                    if str(cell.value or "").strip() in {"QR Verification", "CDE QR Verification", "Scan to verify"}:
                        continue
                    target_cell = target.cell(cell.row, cell.column, cell.value)
                    target_cell.number_format = cell.number_format
            for key, dimension in source_ws.column_dimensions.items():
                if dimension.width:
                    target.column_dimensions[key].width = min(float(dimension.width), 18)
            target.freeze_panes = None
        source.close()

    if copied == 0:
        raise ValueError("No valid Excel workbooks were available to combine.")

    destination.save(output)
    destination.close()
    style_excel(output)
    payload = dict(qr_payload or default_download_context(output))
    payload["document_type"] = "Data Delivery Report"
    payload.setdefault("file_name", output.name)
    wb = load_workbook(output, read_only=True)
    sheet_payloads = {name: dict(payload) for name in wb.sheetnames}
    wb.close()
    add_qr_codes_to_workbook(output, sheet_payloads, payload)
    return output
