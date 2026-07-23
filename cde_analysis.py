"""Professional point-based climate extraction and analysis products."""
from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Dict, Iterable

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from cde_products import DATASETS, extract_point_series, slugify, _append_parquet_database, y_axis_label, annual_trend_frame
from cde_excel import write_single_sheet_workbook
from scripts.extractor import add_qr_codes_to_workbook, default_download_context, make_qr_png

MONTH_NAMES = {i: pd.Timestamp(2000, i, 1).strftime("%b") for i in range(1, 13)}
ANALYSIS_SCOPES = {
    "data_extraction": ("Climate Data Extraction", "Selected observations prepared for professional reuse."),
    "statistical_summary": ("Descriptive Statistical Summary", "Central tendency, spread, distribution and data-quality indicators."),
    "climatology_profile": ("Climatology and Seasonal Profile", "Monthly, seasonal and decadal climate behaviour for the selected period."),
    "variability_analysis": ("Standard Deviation and Variability Assessment", "Standard deviation, coefficient of variation, mean spread and standardized anomalies."),
    "trend_variability": ("Trend and Variability Assessment", "Long-term change, moving averages, anomalies and significance tests."),
    "extremes_analysis": ("Extremes and Percentile Assessment", "Observed extremes, percentile thresholds and annual climate indicators."),
    "comprehensive_analysis": ("Comprehensive Climate Analysis", "Integrated extraction, climatology, variability, trends and extremes."),
}


def _fmt(value: Any, digits: int = 1) -> str:
    try:
        number = float(value)
    except Exception:
        return str(value or "N/A")
    if not np.isfinite(number):
        return "N/A"
    return f"{number:,.{digits}f}"


def _mann_kendall(values: pd.Series) -> Dict[str, float | str]:
    y = pd.to_numeric(values, errors="coerce").dropna().to_numpy(dtype=float)
    n = len(y)
    if n < 3:
        return {"n": n, "s": np.nan, "z": np.nan, "p_value": np.nan, "trend": "insufficient data", "sen_slope": np.nan}
    s = 0
    slopes: list[float] = []
    for i in range(n - 1):
        diffs = y[i + 1:] - y[i]
        s += int(np.sign(diffs).sum())
        slopes.extend((diffs / np.arange(1, n - i)).tolist())
    _, counts = np.unique(y, return_counts=True)
    tie_term = sum(c * (c - 1) * (2 * c + 5) for c in counts if c > 1)
    var_s = (n * (n - 1) * (2 * n + 5) - tie_term) / 18.0
    if var_s <= 0:
        z = 0.0
    elif s > 0:
        z = (s - 1) / math.sqrt(var_s)
    elif s < 0:
        z = (s + 1) / math.sqrt(var_s)
    else:
        z = 0.0
    p = math.erfc(abs(z) / math.sqrt(2.0))
    trend = "increasing" if z > 0 and p < 0.05 else "decreasing" if z < 0 and p < 0.05 else "no significant trend"
    return {"n": n, "s": s, "z": z, "p_value": p, "trend": trend, "sen_slope": float(np.nanmedian(slopes))}


def _linear_trend(years: pd.Series, values: pd.Series) -> Dict[str, float]:
    x = pd.to_numeric(years, errors="coerce")
    y = pd.to_numeric(values, errors="coerce")
    mask = x.notna() & y.notna()
    x = x[mask].to_numpy(dtype=float)
    y = y[mask].to_numpy(dtype=float)
    if len(x) < 2:
        return {"slope_per_year": np.nan, "intercept": np.nan, "r_squared": np.nan}
    slope, intercept = np.polyfit(x, y, 1)
    fitted = slope * x + intercept
    ss_res = float(np.sum((y - fitted) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot else 0.0
    return {"slope_per_year": float(slope), "intercept": float(intercept), "r_squared": float(r2)}


def _annual_aggregation(df: pd.DataFrame, family: str, resolution: str) -> pd.DataFrame:
    tmp = df.copy()
    tmp["year"] = tmp["time"].dt.year
    if family == "rainfall" and resolution != "annual":
        annual = tmp.groupby("year", as_index=False)["value"].sum(min_count=1)
        annual["aggregation"] = "sum"
    else:
        annual = tmp.groupby("year", as_index=False)["value"].mean()
        annual["aggregation"] = "mean"
    annual = annual.rename(columns={"value": "annual_value"})
    annual["five_year_moving_average"] = annual["annual_value"].rolling(5, min_periods=3, center=True).mean()
    annual["year_on_year_change"] = annual["annual_value"].diff()
    annual["year_on_year_change_percent"] = annual["annual_value"].pct_change(fill_method=None) * 100.0
    return annual


def _monthly_climatology(df: pd.DataFrame) -> pd.DataFrame:
    tmp = df.copy()
    tmp["month"] = tmp["time"].dt.month
    out = tmp.groupby("month")["value"].agg(["count", "mean", "median", "std", "min", "max"]).reset_index()
    out.insert(1, "month_name", out["month"].map(MONTH_NAMES))
    out["coefficient_of_variation_percent"] = np.where(out["mean"] != 0, out["std"] / out["mean"] * 100.0, np.nan)
    return out


def _seasonal_profile(df: pd.DataFrame, family: str) -> pd.DataFrame:
    definitions = {
        "DJF": [12, 1, 2], "MAM": [3, 4, 5], "JJA": [6, 7, 8],
        "SON": [9, 10, 11], "OND": [10, 11, 12], "NDJFMA": [11, 12, 1, 2, 3, 4],
    }
    rows = []
    tmp = df.copy()
    tmp["month"] = tmp["time"].dt.month
    for name, months in definitions.items():
        values = pd.to_numeric(tmp.loc[tmp["month"].isin(months), "value"], errors="coerce").dropna()
        rows.append({
            "season": name,
            "months": ",".join(map(str, months)),
            "count": int(values.count()),
            "climatological_value": values.sum() / max(1, tmp["time"].dt.year.nunique()) if family == "rainfall" else values.mean(),
            "median": values.median(),
            "standard_deviation": values.std(),
            "minimum": values.min(),
            "maximum": values.max(),
        })
    return pd.DataFrame(rows)


def _decadal_summary(annual: pd.DataFrame) -> pd.DataFrame:
    out = annual.copy()
    out["decade"] = (out["year"] // 10) * 10
    return out.groupby("decade")["annual_value"].agg(["count", "mean", "median", "std", "min", "max"]).reset_index()


def _descriptive(df: pd.DataFrame, unit: str) -> pd.DataFrame:
    v = pd.to_numeric(df["value"], errors="coerce")
    mean = v.mean()
    q25, q75 = v.quantile(0.25), v.quantile(0.75)
    rows = [
        ("Valid observations", int(v.count()), "records"),
        ("Mean", mean, unit),
        ("Median", v.median(), unit),
        ("Standard deviation", v.std(), unit),
        ("Variance", v.var(), f"{unit}²" if unit else ""),
        ("Minimum", v.min(), unit),
        ("5th percentile", v.quantile(0.05), unit),
        ("25th percentile", q25, unit),
        ("75th percentile", q75, unit),
        ("95th percentile", v.quantile(0.95), unit),
        ("Maximum", v.max(), unit),
        ("Range", v.max() - v.min(), unit),
        ("Interquartile range", q75 - q25, unit),
        ("Sum", v.sum(min_count=1), unit),
        ("Coefficient of variation", (v.std() / mean) * 100 if mean else np.nan, "%"),
        ("Skewness", v.skew(), ""),
        ("Kurtosis", v.kurt(), ""),
    ]
    return pd.DataFrame(rows, columns=["statistic", "value", "unit"])


def _extremes(df: pd.DataFrame, unit: str) -> pd.DataFrame:
    tmp = df.dropna(subset=["value"]).sort_values("value")
    if tmp.empty:
        return pd.DataFrame(columns=["metric", "value", "date", "unit"])
    q01, q05, q95, q99 = [tmp["value"].quantile(q) for q in (0.01, 0.05, 0.95, 0.99)]
    low, high = tmp.iloc[0], tmp.iloc[-1]
    rows = [
        ("Lowest observation", low["value"], low["time"], unit),
        ("Highest observation", high["value"], high["time"], unit),
        ("1st percentile threshold", q01, pd.NaT, unit),
        ("5th percentile threshold", q05, pd.NaT, unit),
        ("95th percentile threshold", q95, pd.NaT, unit),
        ("99th percentile threshold", q99, pd.NaT, unit),
        ("Observations at/below 5th percentile", int((tmp["value"] <= q05).sum()), pd.NaT, "records"),
        ("Observations at/above 95th percentile", int((tmp["value"] >= q95).sum()), pd.NaT, "records"),
    ]
    return pd.DataFrame(rows, columns=["metric", "value", "date", "unit"])


def _annual_extremes(df: pd.DataFrame) -> pd.DataFrame:
    tmp = df.copy()
    tmp["year"] = tmp["time"].dt.year
    return tmp.groupby("year")["value"].agg(annual_minimum="min", annual_maximum="max", annual_mean="mean", observations="count").reset_index()


def _distribution(df: pd.DataFrame, unit: str) -> pd.DataFrame:
    values = pd.to_numeric(df["value"], errors="coerce").dropna().to_numpy(dtype=float)
    if not len(values):
        return pd.DataFrame(columns=["bin_start", "bin_end", "frequency", "unit"])
    bins = min(20, max(6, int(np.sqrt(len(values)))))
    counts, edges = np.histogram(values, bins=bins)
    return pd.DataFrame({"bin_start": edges[:-1], "bin_end": edges[1:], "frequency": counts, "unit": unit})


def _climate_indicators(df: pd.DataFrame, family: str, resolution: str) -> pd.DataFrame:
    tmp = df.copy()
    tmp["year"] = tmp["time"].dt.year
    rows: list[dict[str, Any]] = []
    overall = pd.to_numeric(tmp["value"], errors="coerce")
    low10, high90 = overall.quantile(0.10), overall.quantile(0.90)
    for year, group in tmp.groupby("year"):
        values = pd.to_numeric(group["value"], errors="coerce")
        row: dict[str, Any] = {
            "year": int(year), "mean": values.mean(), "minimum": values.min(), "maximum": values.max(),
            "below_10th_percentile": int((values <= low10).sum()),
            "above_90th_percentile": int((values >= high90).sum()),
        }
        if family == "rainfall":
            row.update({
                "total": values.sum(min_count=1),
                "wet_observations_ge_1mm": int((values >= 1.0).sum()),
                "dry_observations_lt_1mm": int((values < 1.0).sum()),
                "maximum_1_period": values.max(),
                "simple_intensity_index": values[values >= 1.0].mean(),
            })
            if resolution in {"daily", "hourly"}:
                row["maximum_5_period_total"] = values.rolling(5, min_periods=5).sum().max()
        rows.append(row)
    return pd.DataFrame(rows)


def _expected_observations(start: pd.Timestamp, end: pd.Timestamp, resolution: str) -> int | None:
    freq = {"hourly": "h", "daily": "D", "monthly": "MS", "annual": "YS"}.get(resolution)
    if not freq:
        return max(1, end.year - start.year + 1)
    try:
        return len(pd.date_range(start=start, end=end, freq=freq))
    except Exception:
        return None


def _data_quality(df: pd.DataFrame, resolution: str) -> pd.DataFrame:
    valid = int(pd.to_numeric(df["value"], errors="coerce").notna().sum())
    missing = int(len(df) - valid)
    first, last = pd.Timestamp(df["time"].min()), pd.Timestamp(df["time"].max())
    expected = _expected_observations(first, last, resolution)
    completeness = (valid / expected * 100.0) if expected else np.nan
    rows = [
        ("First observation", first, ""),
        ("Last observation", last, ""),
        ("Rows returned", int(len(df)), "records"),
        ("Valid values", valid, "records"),
        ("Missing values", missing, "records"),
        ("Duplicate timestamps", int(df["time"].duplicated().sum()), "records"),
        ("Expected observations", expected if expected is not None else np.nan, "records"),
        ("Estimated completeness", completeness, "%"),
    ]
    return pd.DataFrame(rows, columns=["indicator", "value", "unit"])


def _style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="0B5E7A")
    accent = PatternFill("solid", fgColor="EAF6F2")
    for ws in wb.worksheets:
        ws.freeze_panes = None
        ws.auto_filter.ref = None
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = accent
            for cell in ws[row_idx]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = col[0].column_letter
            width = min(18, max(8, max(len(str(c.value or "")) for c in col) + 1))
            ws.column_dimensions[letter].width = width
    wb.save(path)


def _save_time_series_plot(df: pd.DataFrame, title: str, y_label: str, path: Path) -> Path:
    fig, ax = plt.subplots(figsize=(11, 5.8))
    ax.plot(df["time"], df["value"], linewidth=1.1)
    ax.set_title(title); ax.set_xlabel("Time"); ax.set_ylabel(y_label); ax.grid(True, alpha=0.3)
    fig.tight_layout(); fig.savefig(path, dpi=180); plt.close(fig)
    return path


def _save_climatology_plot(monthly: pd.DataFrame, seasonal: pd.DataFrame, title: str, y_label: str, path: Path) -> Path:
    fig, axes = plt.subplots(1, 2, figsize=(12, 5.5))
    axes[0].plot(monthly["month_name"], monthly["mean"], marker="o", linewidth=2)
    axes[0].set_title("Monthly climatology"); axes[0].set_xlabel("Month"); axes[0].set_ylabel(y_label); axes[0].grid(True, alpha=0.3)
    axes[1].bar(seasonal["season"], seasonal["climatological_value"])
    axes[1].set_title("Seasonal profile"); axes[1].set_xlabel("Season"); axes[1].set_ylabel(y_label); axes[1].grid(True, axis="y", alpha=0.3)
    fig.suptitle(title, fontsize=14); fig.tight_layout(); fig.savefig(path, dpi=180); plt.close(fig)
    return path


def _save_trend_plot(annual: pd.DataFrame, trend: dict[str, Any], title: str, y_label: str, path: Path) -> Path:
    fig, axes = plt.subplots(2, 1, figsize=(11, 9), sharex=True)
    axes[0].plot(annual["year"], annual["annual_value"], marker="o", linewidth=1.4, label="Annual value")
    axes[0].plot(annual["year"], annual["five_year_moving_average"], linewidth=2.2, linestyle="--", label="5-year moving average")
    if np.isfinite(float(trend.get("slope_per_year", np.nan))):
        fit = trend["slope_per_year"] * annual["year"] + trend["intercept"]
        axes[0].plot(annual["year"], fit, linewidth=1.5, linestyle=":", label="Linear trend")
    axes[0].set_ylabel(y_label); axes[0].set_title("Annual trend and variability"); axes[0].legend(); axes[0].grid(True, alpha=0.3)
    anomalies = pd.to_numeric(annual["anomaly"], errors="coerce")
    axes[1].bar(annual["year"], anomalies)
    axes[1].axhline(0, linewidth=0.9)
    axes[1].set_xlabel("Year"); axes[1].set_ylabel(f"Anomaly ({y_label})"); axes[1].set_title("Annual anomaly"); axes[1].grid(True, axis="y", alpha=0.3)
    fig.suptitle(title, fontsize=14); fig.tight_layout(); fig.savefig(path, dpi=180); plt.close(fig)
    return path



def _save_variability_plot(monthly: pd.DataFrame, annual: pd.DataFrame, title: str, y_label: str, path: Path) -> Path:
    fig, axes = plt.subplots(2, 1, figsize=(11, 9))
    x = np.arange(len(monthly))
    mean = pd.to_numeric(monthly["mean"], errors="coerce").to_numpy(dtype=float)
    sd = pd.to_numeric(monthly["std"], errors="coerce").fillna(0).to_numpy(dtype=float)
    axes[0].plot(x, mean, marker="o", linewidth=2, label="Monthly mean")
    axes[0].fill_between(x, mean - sd, mean + sd, alpha=0.22, label="Mean ± 1 standard deviation")
    axes[0].set_xticks(x)
    axes[0].set_xticklabels(monthly["month_name"])
    axes[0].set_xlabel("Month")
    axes[0].set_ylabel(y_label)
    axes[0].set_title("Monthly mean and standard-deviation spread")
    axes[0].legend()
    axes[0].grid(True, alpha=0.3)
    axes[1].bar(annual["year"], pd.to_numeric(annual["standardized_anomaly"], errors="coerce"))
    axes[1].axhline(0, linewidth=0.9)
    axes[1].axhline(1, linestyle="--", linewidth=0.8)
    axes[1].axhline(-1, linestyle="--", linewidth=0.8)
    axes[1].set_xlabel("Year")
    axes[1].set_ylabel("Standardized anomaly (z-score)")
    axes[1].set_title("Annual standardized anomalies")
    axes[1].grid(True, axis="y", alpha=0.3)
    fig.suptitle(title, fontsize=14)
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)
    return path

def _save_extremes_plot(df: pd.DataFrame, annual_extremes: pd.DataFrame, title: str, y_label: str, path: Path) -> Path:
    fig, axes = plt.subplots(1, 2, figsize=(12, 5.5))
    axes[0].hist(pd.to_numeric(df["value"], errors="coerce").dropna(), bins=25, alpha=0.85)
    axes[0].set_title("Frequency distribution"); axes[0].set_xlabel(y_label); axes[0].set_ylabel("Frequency"); axes[0].grid(True, alpha=0.25)
    axes[1].plot(annual_extremes["year"], annual_extremes["annual_maximum"], marker="o", label="Annual maximum")
    axes[1].plot(annual_extremes["year"], annual_extremes["annual_minimum"], marker="o", label="Annual minimum")
    axes[1].set_title("Annual extremes"); axes[1].set_xlabel("Year"); axes[1].set_ylabel(y_label); axes[1].legend(); axes[1].grid(True, alpha=0.3)
    fig.suptitle(title, fontsize=14); fig.tight_layout(); fig.savefig(path, dpi=180); plt.close(fig)
    return path


def _records_for_preview(frame: pd.DataFrame, limit: int = 12, decimals: int = 1) -> tuple[list[str], list[dict[str, str]]]:
    preview = frame.head(limit).copy()
    for column in preview.columns:
        if pd.api.types.is_datetime64_any_dtype(preview[column]):
            preview[column] = preview[column].dt.strftime("%Y-%m-%d %H:%M")
        elif pd.api.types.is_numeric_dtype(preview[column]):
            name = str(column).strip().lower()
            structural = any(token in name for token in ["year", "month", "day", "hour", "count", "records", "observations", "frequency"])
            places = 0 if structural else decimals
            preview[column] = preview[column].map(lambda v: "" if pd.isna(v) else f"{float(v):.{places}f}")
        else:
            preview[column] = preview[column].fillna("").astype(str)
    return list(preview.columns), preview.to_dict("records")


def _analysis_zero_decimal(context: Dict[str, Any], variable_label: str) -> bool:
    text = " ".join([
        str(context.get("dataset_label") or ""),
        str(context.get("variable_label") or ""),
        str(context.get("variable") or ""),
        str(variable_label or ""),
    ]).lower()
    return any(token in text for token in ["relative humidity", "wind speed", "wind direction", "wind_speed", "wind_direction"])


def _round_analysis_frame(frame: pd.DataFrame, zero_decimal: bool = False) -> pd.DataFrame:
    out = frame.copy()
    structural = {"year", "month", "day", "hour", "count", "observations", "frequency", "records", "n", "s"}
    coordinate_names = {"latitude", "longitude", "requested_latitude", "requested_longitude", "nearest_grid_latitude", "nearest_grid_longitude"}
    for column in out.columns:
        if not pd.api.types.is_numeric_dtype(out[column]):
            continue
        name = str(column).strip().lower()
        values = pd.to_numeric(out[column], errors="coerce")
        if name in coordinate_names or "latitude" in name or "longitude" in name:
            out[column] = values.round(4)
        elif name in structural or any(token in name for token in ["observations", "records", "count", "days", "year", "month", "day", "hour"]):
            # Counts and temporal identifiers remain whole numbers.
            if values.dropna().apply(lambda x: float(x).is_integer()).all():
                out[column] = values.round(0).astype("Int64")
            else:
                out[column] = values.round(1)
        elif zero_decimal and any(token in name for token in ["value", "mean", "median", "minimum", "maximum", "percentile", "anomaly", "slope", "range", "sum", "standard_deviation"]):
            out[column] = values.round(0).astype("Int64")
        else:
            out[column] = values.round(1)
    return out


def _add_analysis_qr(path: Path, metadata: Dict[str, Any]) -> None:
    context = default_download_context(path)
    context.update({
        "document_type": "Data Delivery Report",
        "file_name": path.name,
        "reference_no": "CD533/620/01",
        "request_no": str(context.get("download_id") or "").replace("CDE-", "")[:12],
        "mode_of_delivery": "Electronic copy",
        "source": metadata.get("dataset", ""),
        "element": metadata.get("weather_element", ""),
        "data_type": metadata.get("resolution", ""),
        "period": f"{metadata.get('start_date', '')} to {metadata.get('end_date', '')}",
        "station_name": metadata.get("location", ""),
        "latitude": metadata.get("requested_latitude", ""),
        "longitude": metadata.get("requested_longitude", ""),
        "units": metadata.get("unit", ""),
        "description": metadata.get("description", ""),
    })
    from openpyxl import load_workbook as _load_workbook
    wb = _load_workbook(path, read_only=True)
    sheet_names = list(wb.sheetnames)
    wb.close()
    payloads = {name: dict(context) for name in sheet_names}
    add_qr_codes_to_workbook(path, payloads, context)


def _write_analysis_workbook_fast(
    path: Path,
    sheets: list[tuple[str, pd.DataFrame]],
    metadata: Dict[str, Any],
    *,
    zero_decimal: bool,
) -> None:
    """Write every analytical section into one complete bordered worksheet."""
    context = default_download_context(path)
    context.update({
        "document_type": "Data Delivery Report",
        "file_name": path.name,
        "reference_no": "CD533/620/01",
        "request_no": str(context.get("download_id") or "").replace("CDE-", "")[:12],
        "mode_of_delivery": "Electronic copy",
        "source": metadata.get("dataset") or "",
        "element": metadata.get("weather_element") or "",
        "data_type": metadata.get("resolution") or "",
        "period": f"{metadata.get('start_date', '')} to {metadata.get('end_date', '')}",
        "station_name": metadata.get("location") or "",
        "latitude": metadata.get("requested_latitude") or "",
        "longitude": metadata.get("requested_longitude") or "",
        "units": metadata.get("unit") or "",
        "description": metadata.get("description") or "",
    })
    write_single_sheet_workbook(
        path,
        sheets,
        qr_payload=context,
        zero_decimal=zero_decimal,
        sheet_name="Data",
        workbook_title=str(metadata.get("product") or "TMA Climate Data Product"),
    )


def generate_analysis_bundle(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    resolution = str(params.get("resolution") or "monthly")
    variable = str(params.get("variable") or "auto")
    season = str(params.get("season") or "").strip().upper() or None
    scope = str(params.get("analysis_scope") or params.get("plot_type") or "comprehensive_analysis")
    if scope not in ANALYSIS_SCOPES:
        scope = "comprehensive_analysis"
    lat = float(params.get("latitude"))
    lon = float(params.get("longitude"))
    location = str(params.get("location_name") or "Selected Location")
    start = str(params.get("start_date") or "")
    end = str(params.get("end_date") or "")
    baseline_start = int(params.get("baseline_start") or 1991)
    baseline_end = int(params.get("baseline_end") or 2020)

    if dataset_key not in DATASETS:
        raise ValueError("Unknown dataset.")
    if not start or not end or start > end:
        raise ValueError("Provide a valid start and end date.")

    df, context = extract_point_series(data_dir, dataset_key, resolution, lat, lon, start, end, variable, season)
    if df.empty:
        raise ValueError("No values were found for the selected period and location.")
    df = df.sort_values("time").copy()
    df["time"] = pd.to_datetime(df["time"])
    family = context.get("family", DATASETS[dataset_key]["family"])
    unit = context.get("unit", DATASETS[dataset_key].get("unit", ""))
    if dataset_key == "chirps_rainfall":
        variable_label = "CHIRPS Precipitation"
    elif dataset_key == "era5_total_precipitation":
        variable_label = "ERA5 Precipitation"
    else:
        variable_label = context.get("variable_label") or context.get("dataset_label") or DATASETS[dataset_key]["label"]

    descriptive = _descriptive(df, unit)
    monthly = _monthly_climatology(df)
    seasonal = _seasonal_profile(df, family)
    annual = _annual_aggregation(df, family, resolution)
    baseline = annual[(annual["year"] >= baseline_start) & (annual["year"] <= baseline_end)]
    if baseline.empty:
        baseline = annual
    baseline_mean = pd.to_numeric(baseline["annual_value"], errors="coerce").mean()
    baseline_sd = pd.to_numeric(baseline["annual_value"], errors="coerce").std(ddof=0)
    annual["anomaly"] = annual["annual_value"] - baseline_mean
    annual["percent_anomaly"] = np.where(baseline_mean != 0, annual["anomaly"] / baseline_mean * 100.0, np.nan)
    annual["standardized_anomaly"] = annual["anomaly"] / baseline_sd if baseline_sd and np.isfinite(baseline_sd) else np.nan

    linear = _linear_trend(annual["year"], annual["annual_value"])
    mk = _mann_kendall(annual["annual_value"])
    trend_row = {
        **linear, **mk,
        "baseline_start": baseline_start, "baseline_end": baseline_end,
        "baseline_mean": baseline_mean, "baseline_standard_deviation": baseline_sd, "unit": unit,
    }
    trend = pd.DataFrame([trend_row])
    extremes = _extremes(df, unit)
    annual_extremes = _annual_extremes(df)
    distribution = _distribution(df, unit)
    decadal = _decadal_summary(annual)
    indicators = _climate_indicators(df, family, resolution)
    quality = _data_quality(df, resolution)

    scope_title, scope_description = ANALYSIS_SCOPES[scope]
    period_text = f"{start[:4]}–{end[:4]}"
    custom_title = " ".join(str(params.get("custom_plot_title") or "").split())[:180]
    product_title = custom_title or f"{scope_title}: {variable_label} for {location}"
    product_subtitle = f"{resolution.capitalize()} assessment for {period_text}"
    if season:
        product_subtitle += f" · {season} season"

    metadata = pd.DataFrame([{
        "product": scope_title,
        "dataset": context.get("dataset_label"),
        "weather_element": variable_label,
        "resolution": resolution,
        "season": season or "",
        "location": location,
        "requested_latitude": lat,
        "requested_longitude": lon,
        "nearest_grid_latitude": context.get("nearest_latitude"),
        "nearest_grid_longitude": context.get("nearest_longitude"),
        "start_date": start,
        "end_date": end,
        "baseline_period": f"{baseline_start}-{baseline_end}",
        "unit": unit,
        "description": scope_description,
    }])

    export_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = slugify(f"CDE_{scope}_{dataset_key}_{variable}_{location}_{stamp}")
    xlsx_path = export_dir / f"{stem}.xlsx"

    export_ts = df.rename(columns={"time": "date_time", "value": f"value_{unit}" if unit else "value"})
    sheets: list[tuple[str, pd.DataFrame]] = [("Metadata", metadata), ("Data Quality", quality), ("Data", export_ts)]
    if scope in {"statistical_summary", "variability_analysis", "comprehensive_analysis"}:
        sheets += [("Descriptive Statistics", descriptive), ("Distribution", distribution)]
    if scope in {"climatology_profile", "variability_analysis", "comprehensive_analysis"}:
        sheets += [("Monthly Climatology", monthly), ("Seasonal Profile", seasonal), ("Decadal Summary", decadal)]
    if scope in {"variability_analysis", "trend_variability", "comprehensive_analysis"}:
        sheets += [("Annual Analysis", annual), ("Trend Tests", trend)]
    if scope in {"extremes_analysis", "statistical_summary", "comprehensive_analysis"}:
        sheets += [("Extreme Records", extremes), ("Annual Extremes", annual_extremes), ("Climate Indicators", indicators)]

    zero_decimal = _analysis_zero_decimal(context, variable_label)
    rounded_sheets = [(sheet_name, _round_analysis_frame(frame, zero_decimal=zero_decimal)) for sheet_name, frame in sheets]
    _write_analysis_workbook_fast(
        xlsx_path,
        rounded_sheets,
        metadata.iloc[0].to_dict(),
        zero_decimal=zero_decimal,
    )

    main_frames = {
        "data_extraction": export_ts,
        "statistical_summary": descriptive,
        "climatology_profile": monthly,
        "variability_analysis": monthly,
        "trend_variability": annual,
        "extremes_analysis": indicators if not indicators.empty else annual_extremes,
        "comprehensive_analysis": annual,
    }
    main_frame = _round_analysis_frame(main_frames[scope], zero_decimal=zero_decimal)

    y_label = y_axis_label(variable_label, unit, resolution)
    preview_paths: list[Path] = []
    include_plot = str(params.get("include_plot") or "").strip().lower() in {"1", "true", "yes", "on"}
    def preview_title(default: str, suffix: str) -> str:
        return f"{custom_title} — {suffix}" if custom_title else default
    if include_plot and scope in {"comprehensive_analysis"}:
        preview_paths.append(_save_time_series_plot(df, preview_title(f"{variable_label} Time Series for {location} ({period_text})", "Time Series"), y_label, export_dir / f"{stem}_time_series.png"))
    if include_plot and scope in {"climatology_profile", "comprehensive_analysis"}:
        preview_paths.append(_save_climatology_plot(monthly, seasonal, preview_title(f"{variable_label} Climatology for {location} ({period_text})", "Climatology"), y_label, export_dir / f"{stem}_climatology.png"))
    if include_plot and scope in {"variability_analysis", "comprehensive_analysis"}:
        preview_paths.append(_save_variability_plot(monthly, annual, preview_title(f"{variable_label} Standard Deviation and Variability for {location} ({period_text})", "Variability"), y_label, export_dir / f"{stem}_variability.png"))
    if include_plot and scope in {"trend_variability", "comprehensive_analysis"}:
        preview_paths.append(_save_trend_plot(annual, trend_row, preview_title(f"{variable_label} Trend and Variability for {location} ({period_text})", "Trend and Variability"), y_label, export_dir / f"{stem}_trend_variability.png"))
    if include_plot and scope in {"statistical_summary", "extremes_analysis", "comprehensive_analysis"}:
        preview_paths.append(_save_extremes_plot(df, annual_extremes, preview_title(f"{variable_label} Distribution and Extremes for {location} ({period_text})", "Distribution and Extremes"), y_label, export_dir / f"{stem}_distribution_extremes.png"))

    valid_values = pd.to_numeric(df["value"], errors="coerce").dropna()
    quality_map = dict(zip(quality["indicator"], quality["value"]))
    display_decimals = 0 if zero_decimal else 1
    summary_cards = [
        {"label": "Observations", "value": f"{len(df):,}", "note": f"{resolution.capitalize()} records"},
        {"label": "Completeness", "value": f"{_fmt(quality_map.get('Estimated completeness'), 1)}%", "note": "Estimated temporal coverage"},
        {"label": "Average", "value": f"{_fmt(valid_values.mean(), display_decimals)} {unit}".strip(), "note": "Selected-period mean"},
        {"label": "Observed Range", "value": f"{_fmt(valid_values.min(), display_decimals)} – {_fmt(valid_values.max(), display_decimals)} {unit}".strip(), "note": "Minimum to maximum"},
    ]
    if scope == "variability_analysis":
        overall_mean = float(valid_values.mean()) if not valid_values.empty else np.nan
        overall_sd = float(valid_values.std(ddof=0)) if not valid_values.empty else np.nan
        overall_cv = overall_sd / abs(overall_mean) * 100.0 if np.isfinite(overall_mean) and abs(overall_mean) > 1e-12 else np.nan
        summary_cards = [
            {"label": "Standard Deviation", "value": f"{_fmt(overall_sd, display_decimals)} {unit}".strip(), "note": "Selected-period spread"},
            {"label": "Coefficient of Variation", "value": f"{_fmt(overall_cv, 1)}%", "note": "Spread relative to the mean"},
            {"label": "Baseline Standard Deviation", "value": f"{_fmt(baseline_sd, display_decimals)} {unit}".strip(), "note": f"{baseline_start}–{baseline_end}"},
            {"label": "Observations", "value": f"{len(valid_values):,}", "note": resolution.capitalize()},
        ]
    elif scope in {"trend_variability", "comprehensive_analysis"}:
        summary_cards = [
            {"label": "Trend", "value": str(trend_row["trend"]).title(), "note": f"Mann–Kendall p={_fmt(trend_row['p_value'], 1)}"},
            {"label": "Sen's Slope", "value": f"{_fmt(trend_row['sen_slope'], display_decimals)} {unit}/year".strip(), "note": "Robust annual change"},
            {"label": "Linear Slope", "value": f"{_fmt(trend_row['slope_per_year'], display_decimals)} {unit}/year".strip(), "note": f"R²={_fmt(trend_row['r_squared'], 1)}"},
            {"label": "Baseline Mean", "value": f"{_fmt(baseline_mean, display_decimals)} {unit}".strip(), "note": f"{baseline_start}–{baseline_end}"},
        ]

    preview_columns, preview_rows = _records_for_preview(main_frame, decimals=display_decimals)
    time_series_plot_path = next((path for path in preview_paths if path.name.endswith("_time_series.png")), None)
    trend_plot_path = next((path for path in preview_paths if path.name.endswith("_trend_variability.png")), None)
    return {
        "excel_path": xlsx_path,
        "preview_paths": preview_paths,
        "time_series_plot_path": time_series_plot_path,
        "anomaly_plot_path": trend_plot_path,
        "context": context,
        "rows": len(df),
        "trend": trend_row,
        "descriptive": descriptive.to_dict("records"),
        "summary_cards": summary_cards,
        "preview_columns": preview_columns,
        "preview_rows": preview_rows,
        "product_title": product_title,
        "product_subtitle": product_subtitle,
        "product_description": scope_description,
        "analysis_scope": scope,
        "_analysis_frame": export_ts,
    }



def _clean_analysis_element_name(value: str) -> str:
    import re
    clean = re.sub(r"^(CHIRPS|ERA5(?:-Land)?)\s+", "", str(value or "Weather Element"), flags=re.I).strip()
    return clean.replace("Precipitation", "Rainfall")


def _long_term_trend_labels(variable_label: str, family: str, unit: str) -> tuple[str, str, str]:
    clean = _clean_analysis_element_name(variable_label)
    if str(family).lower() == "rainfall":
        # Rainfall is aggregated to an annual total, therefore the displayed
        # annual unit is millimetres regardless of source time-step wording.
        return "Rainfall", "Annual rainfall", "Annual Rainfall (mm)"
    return clean, f"Annual mean {clean}", f"Annual Mean {clean}" + (f" ({unit})" if unit else "")


def _save_long_term_annual_trend_plot(
    annual: pd.DataFrame,
    trend: Dict[str, Any],
    *,
    title: str,
    series_label: str,
    y_label: str,
    family: str,
    path: Path,
) -> Path:
    """Create the one-panel trend plot shown in the supplied examples."""
    fig, ax = plt.subplots(figsize=(12.0, 7.0))
    marker = "o" if str(family).lower() == "rainfall" else None
    ax.plot(
        annual["year"], annual["annual_value"], marker=marker,
        linewidth=1.8, markersize=5 if marker else 0, label=series_label,
    )
    if len(annual) > 1 and np.isfinite(float(trend.get("slope_per_year", np.nan))):
        ax.plot(
            annual["year"], annual["linear_trend"], linestyle=":",
            linewidth=2.3, label="Linear trend",
        )
        slope = float(trend["slope_per_year"])
        intercept = float(trend["intercept"])
        sign = "+" if intercept >= 0 else "-"
        equation = f"y = {slope:.4f}x {sign} {abs(intercept):.3f}"
        ax.text(
            0.98, 0.955,
            f"{equation}\nR² = {float(trend['r_squared']):.4f}",
            transform=ax.transAxes, ha="right", va="top", fontsize=11,
            bbox=dict(boxstyle="round,pad=0.28", facecolor="white", edgecolor="none", alpha=0.9),
        )
    ax.set_title(title, fontsize=15, pad=14)
    ax.set_xlabel("Years")
    ax.set_ylabel(y_label)
    ax.legend(loc="upper left")
    ax.grid(True, alpha=0.32)
    fig.tight_layout()
    fig.savefig(path, dpi=190, facecolor="white")
    plt.close(fig)
    return path


def _write_long_term_trend_workbook(
    path: Path,
    annual: pd.DataFrame,
    trend_statistics: pd.DataFrame,
    metadata: pd.DataFrame,
) -> None:
    """Write a polished three-tab trend-analysis workbook."""
    with pd.ExcelWriter(path, engine="xlsxwriter", engine_kwargs={"options": {"strings_to_urls": False, "nan_inf_to_errors": True}}) as writer:
        workbook = writer.book
        title_fmt = workbook.add_format({
            "bold": True, "font_size": 15, "font_color": "white", "bg_color": "#073B5C",
            "align": "left", "valign": "vcenter", "border": 1,
        })
        header_fmt = workbook.add_format({
            "bold": True, "font_color": "white", "bg_color": "#0B5E7A",
            "align": "center", "valign": "vcenter", "text_wrap": True, "border": 1,
        })
        text_fmt = workbook.add_format({"border": 1, "valign": "top", "text_wrap": True})
        one_fmt = workbook.add_format({"border": 1, "num_format": "0.0", "valign": "top"})
        four_fmt = workbook.add_format({"border": 1, "num_format": "0.0000", "valign": "top"})
        int_fmt = workbook.add_format({"border": 1, "num_format": "0", "valign": "top"})
        alt_text_fmt = workbook.add_format({"border": 1, "valign": "top", "text_wrap": True, "bg_color": "#EAF6F2"})
        alt_one_fmt = workbook.add_format({"border": 1, "num_format": "0.0", "valign": "top", "bg_color": "#EAF6F2"})
        alt_four_fmt = workbook.add_format({"border": 1, "num_format": "0.0000", "valign": "top", "bg_color": "#EAF6F2"})
        alt_int_fmt = workbook.add_format({"border": 1, "num_format": "0", "valign": "top", "bg_color": "#EAF6F2"})

        def add_sheet(name: str, title: str, frame: pd.DataFrame, tab_colour: str) -> None:
            ws = workbook.add_worksheet(name[:31])
            writer.sheets[name[:31]] = ws
            ws.hide_gridlines(2)
            last_col = max(1, len(frame.columns) - 1)
            ws.merge_range(0, 0, 0, last_col, title, title_fmt)
            ws.set_row(0, 28)
            header_row = 2
            for col, column in enumerate(frame.columns):
                ws.write(header_row, col, str(column), header_fmt)
            for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=header_row + 1):
                alternating = row_index % 2 == 0
                for col, raw in enumerate(values):
                    value = raw.item() if isinstance(raw, np.generic) else raw
                    column_name = str(frame.columns[col]).lower()
                    if value is None or (isinstance(value, float) and np.isnan(value)):
                        ws.write_blank(row_index, col, None, alt_text_fmt if alternating else text_fmt)
                    elif isinstance(value, (int, np.integer)) or column_name == "year":
                        ws.write_number(row_index, col, int(value), alt_int_fmt if alternating else int_fmt)
                    elif isinstance(value, (float, np.floating)):
                        precise = any(token in column_name for token in ("slope", "intercept", "r_squared", "p_value", "z"))
                        fmt = (alt_four_fmt if alternating else four_fmt) if precise else (alt_one_fmt if alternating else one_fmt)
                        ws.write_number(row_index, col, float(value), fmt)
                    else:
                        ws.write(row_index, col, str(value), alt_text_fmt if alternating else text_fmt)
            for col, column in enumerate(frame.columns):
                max_len = max([len(str(column))] + [len(str(v)) for v in frame[column].head(500).fillna("")])
                ws.set_column(col, col, min(max(max_len + 1, 8), 24))
            ws.set_tab_color(tab_colour)
            ws.set_landscape()
            ws.fit_to_pages(1, 0)
            ws.set_margins(left=0.25, right=0.25, top=1.65, bottom=0.45)
            ws.set_footer("&LClimate Data Extractor&CTrend analysis&RPage &P of &N")

        add_sheet("Annual Trend Data", "Long-Term Annual Trend Data", annual, "7030A0")
        add_sheet("Trend Statistics", "Linear and Non-Parametric Trend Statistics", trend_statistics, "0B8F6F")
        add_sheet("Metadata", "Analysis Metadata", metadata, "2F75B5")


def _generate_long_term_annual_trend_analysis(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    dataset_key = str(params.get("dataset") or "chirps_rainfall")
    resolution = str(params.get("resolution") or "monthly")
    variable = str(params.get("variable") or "auto")
    season = str(params.get("season") or "").strip().upper() or None
    latitude = float(params.get("latitude"))
    longitude = float(params.get("longitude"))
    location = str(params.get("location_name") or "Selected Location")
    start = str(params.get("start_date") or "")
    end = str(params.get("end_date") or "")
    if not start or not end or start > end:
        raise ValueError("Provide a valid start and end date.")

    data, context = extract_point_series(
        Path(data_dir), dataset_key, resolution, latitude, longitude,
        start, end, variable, season,
    )
    if data.empty:
        raise ValueError("No values were found for the selected period and location.")
    data = data.sort_values("time").copy()
    data["time"] = pd.to_datetime(data["time"])
    family = str(context.get("family") or DATASETS[dataset_key].get("family") or "")
    unit = str(context.get("unit") or DATASETS[dataset_key].get("unit") or "")
    if dataset_key == "chirps_rainfall":
        variable_label = "CHIRPS Precipitation"
    elif dataset_key == "era5_total_precipitation":
        variable_label = "ERA5 Precipitation"
    else:
        variable_label = str(context.get("variable_label") or context.get("dataset_label") or DATASETS[dataset_key]["label"])

    annual, trend = annual_trend_frame(data, family)
    if annual.empty:
        raise ValueError("At least one valid annual value is required for trend analysis.")
    mk = _mann_kendall(annual["annual_value"])
    trend.update(mk)
    title_element, series_label, y_label = _long_term_trend_labels(variable_label, family, unit)
    annual_unit = "mm" if family == "rainfall" else unit
    start_year = int(annual["year"].min())
    end_year = int(annual["year"].max())
    default_title = (
        f"Long-Term Annual Rainfall Trend for {location} ({start_year}–{end_year})"
        if family == "rainfall"
        else f"Long-Term Annual Mean {title_element} Trend for {location} ({start_year}–{end_year})"
    )
    custom_title = " ".join(str(params.get("custom_plot_title") or "").split())[:180]
    plot_title = custom_title or default_title

    export_dir = Path(export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = slugify(f"CDE_long_term_annual_trend_{dataset_key}_{variable}_{location}_{stamp}")
    plot_path = export_dir / f"{stem}.png"
    xlsx_path = export_dir / f"{stem}.xlsx"
    _save_long_term_annual_trend_plot(
        annual, trend, title=plot_title, series_label=series_label,
        y_label=y_label, family=family, path=plot_path,
    )

    annual_export = annual.rename(columns={
        "year": "Year", "annual_value": y_label, "linear_trend": "Linear Trend",
    })
    annual_export["Departure from Trend"] = annual_export[y_label] - annual_export["Linear Trend"]
    trend_statistics = pd.DataFrame([
        ("Annual aggregation", "Total" if family == "rainfall" else "Mean", "Method"),
        ("Linear slope", trend.get("slope_per_year"), f"{annual_unit}/year" if annual_unit else "per year"),
        ("Linear intercept", trend.get("intercept"), annual_unit),
        ("Coefficient of determination", trend.get("r_squared"), "R²"),
        ("Mann–Kendall trend", trend.get("trend"), "Interpretation"),
        ("Mann–Kendall p-value", trend.get("p_value"), "Probability"),
        ("Sen's slope", trend.get("sen_slope"), f"{annual_unit}/year" if annual_unit else "per year"),
    ], columns=["Statistic", "Value", "Unit / Interpretation"])
    metadata = pd.DataFrame([
        ("Analysis", "Long-Term Annual Trend Analysis"),
        ("Dataset", context.get("dataset_label") or DATASETS[dataset_key]["label"]),
        ("Weather Element", variable_label),
        ("Input Resolution", resolution.capitalize()),
        ("Annual Aggregation", "Sum" if family == "rainfall" else "Mean"),
        ("Location", location),
        ("Requested Coordinates", f"{latitude:.4f}, {longitude:.4f}"),
        ("Nearest Native Grid", f"{float(context.get('nearest_latitude', latitude)):.4f}, {float(context.get('nearest_longitude', longitude)):.4f}"),
        ("Period", f"{start} to {end}"),
        ("Input Unit", unit),
        ("Annual Output Unit", annual_unit),
        ("Plot Title", plot_title),
    ], columns=["Field", "Value"])
    _write_long_term_trend_workbook(xlsx_path, annual_export, trend_statistics, metadata)
    _add_analysis_qr(xlsx_path, {
        "product": "Long-Term Annual Trend Analysis",
        "dataset": context.get("dataset_label") or DATASETS[dataset_key]["label"],
        "weather_element": variable_label,
        "resolution": resolution,
        "location": location,
        "requested_latitude": latitude,
        "requested_longitude": longitude,
        "start_date": start,
        "end_date": end,
        "unit": unit,
        "description": "Annual totals for precipitation or annual means for other variables, with a fitted linear trend, equation and R².",
    })

    parquet_path = xlsx_path.with_suffix(".parquet")
    annual_export.to_parquet(parquet_path, index=False)
    db_path = _append_parquet_database("climate_analysis", annual_export, parquet_path.stem)
    slope = float(trend.get("slope_per_year", np.nan))
    return {
        "excel_path": xlsx_path,
        "preview_paths": [plot_path],
        "time_series_plot_path": plot_path,
        "anomaly_plot_path": None,
        "context": {
            **context,
            "analysis_scope": "long_term_annual_trend_analysis",
            "variable_label": variable_label,
            "annual_aggregation": "sum" if family == "rainfall" else "mean",
            "slope_per_year": slope,
            "intercept": trend.get("intercept"),
            "r_squared": trend.get("r_squared"),
            "display_title": plot_title,
            "y_axis_label": y_label,
            "annual_output_unit": annual_unit,
        },
        "rows": len(annual_export),
        "trend": trend,
        "summary_cards": [
            {"label": "Annual Records", "value": f"{len(annual_export):,}", "note": f"{start_year}–{end_year}"},
            {"label": "Linear Slope", "value": f"{slope:.4f} {annual_unit}/year".strip(), "note": "Annual change"},
            {"label": "R²", "value": f"{float(trend.get('r_squared', np.nan)):.4f}", "note": "Linear fit"},
            {"label": "Mann–Kendall", "value": str(trend.get("trend") or "N/A").title(), "note": f"p={_fmt(trend.get('p_value'), 4)}"},
        ],
        "preview_columns": list(annual_export.columns),
        "preview_rows": annual_export.head(12).fillna("").to_dict("records"),
        "product_title": plot_title,
        "product_subtitle": f"{resolution.capitalize()} input aggregated to annual {'totals' if family == 'rainfall' else 'means'}",
        "product_description": "Annual series with linear trend equation, coefficient of determination, Mann–Kendall significance and Sen's slope.",
        "analysis_scope": "long_term_annual_trend_analysis",
        "parquet_path": parquet_path,
        "db_path": db_path,
    }

# ---------------------------------------------------------------------------
# 2026-07-18 expanded climate-analysis catalogue.
# ---------------------------------------------------------------------------
_EXTRA_ANALYSIS_SCOPES = {
    "data_quality_analysis": ("Data Quality and Completeness Assessment", "Completeness, missing periods, valid observations and temporal coverage."),
    "distribution_analysis": ("Distribution and Normality Assessment", "Distribution shape, quantiles, skewness, spread and outliers."),
    "percentile_analysis": ("Percentile and Threshold Assessment", "Percentile thresholds and the frequency of values above or below selected climate limits."),
    "monthly_variability_analysis": ("Monthly Variability Assessment", "Monthly mean, standard deviation, coefficient of variation and percentile spread."),
    "seasonal_comparison_analysis": ("Seasonal Comparison Assessment", "Comparison of climatological conditions and variability between seasons."),
    "decadal_change_analysis": ("Decadal Change Assessment", "Decadal averages, changes between decades and long-term evolution."),
    "anomaly_analysis": ("Anomaly and Standardized Anomaly Assessment", "Absolute, percentage and standardized anomalies relative to the selected baseline."),
    "trend_significance_analysis": ("Trend Significance Assessment", "Linear trend, Mann–Kendall significance, Sen's slope and moving averages."),
    "long_term_annual_trend_analysis": ("Long-Term Annual Trend Analysis", "Annual totals or means with a fitted linear trend, equation, R², Mann–Kendall test and Sen's slope."),
    "extreme_frequency_analysis": ("Extreme Event Frequency Assessment", "Annual frequencies above the 90th percentile and below the 10th percentile."),
}
ANALYSIS_SCOPES.update(_EXTRA_ANALYSIS_SCOPES)
_ANALYSIS_SCOPE_ENGINE = {
    "data_quality_analysis": "statistical_summary",
    "distribution_analysis": "statistical_summary",
    "percentile_analysis": "extremes_analysis",
    "monthly_variability_analysis": "variability_analysis",
    "seasonal_comparison_analysis": "climatology_profile",
    "decadal_change_analysis": "trend_variability",
    "anomaly_analysis": "trend_variability",
    "trend_significance_analysis": "trend_variability",
    "extreme_frequency_analysis": "extremes_analysis",
}
_CDE_ANALYSIS_BASE = generate_analysis_bundle

def generate_analysis_bundle(params: Dict[str, Any], data_dir: Path, export_dir: Path) -> Dict[str, Any]:
    requested = str(params.get("analysis_scope") or params.get("plot_type") or "comprehensive_analysis")
    if requested == "long_term_annual_trend_analysis":
        return _generate_long_term_annual_trend_analysis(dict(params), Path(data_dir), Path(export_dir))
    engine = _ANALYSIS_SCOPE_ENGINE.get(requested, requested)
    prepared = dict(params)
    prepared["analysis_scope"] = engine
    prepared["plot_type"] = engine
    result = dict(_CDE_ANALYSIS_BASE(prepared, Path(data_dir), Path(export_dir)))
    analysis_frame = result.pop("_analysis_frame", None)
    if isinstance(analysis_frame, pd.DataFrame) and result.get("excel_path"):
        parquet_path = Path(result["excel_path"]).with_suffix(".parquet")
        analysis_frame.to_parquet(parquet_path, index=False)
        result["parquet_path"] = parquet_path
        result["db_path"] = _append_parquet_database("climate_analysis", analysis_frame, parquet_path.stem)
    if requested in _EXTRA_ANALYSIS_SCOPES:
        title, description = _EXTRA_ANALYSIS_SCOPES[requested]
        variable = str((result.get("context") or {}).get("variable_label") or "Climate Variable")
        location = str(params.get("location_name") or "Selected Location")
        custom_title = " ".join(str(params.get("custom_plot_title") or "").split())[:180]
        result["product_title"] = custom_title or f"{title}: {variable} for {location}"
        result["product_description"] = description
        result["analysis_scope"] = requested
    return result
